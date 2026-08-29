#!/usr/bin/env python3
"""
Cedar Press - 99: Community Project Funding / earmarks  +  990 Schedule C.

Two advocacy-adjacent layers that sit on OPPOSITE sides of the influence chain
and must never be joined into a causal claim.

  LAYER 1  earmarks.csv          the OUTCOME side. A named member of Congress
                                 requested a named dollar amount for a named
                                 recipient. Requested and enacted are separate
                                 columns; a request that was not funded is a
                                 real record, kept with is_enacted = 0.

  LAYER 2  Schedule C fields     the SELF-REPORTED side. Lobbying expenditure,
                                 grassroots vs direct, the 501(h) election and
                                 political activity, read out of IRS 990
                                 e-file XML and APPENDED to np_financials.csv.

WHY LAYER 2 IS A FIELD GAP, NOT A DATASET GAP
---------------------------------------------
`np_financials.csv` already has a `lobbying_expenditure` column and a
`lobbying_field_basis` column. Measured 2026-08-07, before this script ran:

    lobbying_expenditure non-null ......... 0 of 8,507 rows
    lobbying_field_basis = not_exposed_by_api ......... 8,274
    lobbying_field_basis = 990pf_infleg_indicator_only ... 233

The column was created and honestly labelled empty. ProPublica's Nonprofit
Explorer API v2 does not expose Schedule C at all - its `filings_with_data`
array carries 46 fields and not one of them is a lobbying figure. So the fix is
not another API call to the same host; it is a different source. The IRS
publishes the raw e-file XML, and Schedule C is in it.

  index   https://apps.irs.gov/pub/epostcard/990/xml/{Y}/index_{Y}.csv
          EIN -> OBJECT_ID, one row per accepted return, submission year Y.
  return  https://s3.amazonaws.com/irs-form-990/{OBJECT_ID}_public.xml

Index files exist for submission years 2017-2026 ONLY (2009-2016 return 404 at
both apps.irs.gov and the S3 bucket root, probed 2026-08-07). That is the
coverage floor and it is recorded per row in `schedc_basis`, never smoothed.

THE CAVEAT THAT TRAVELS WITH EVERY FIGURE
-----------------------------------------
6,453 of 12,764 organisations in `np_orgs.csv` carry BMF FILING_REQ_CD = 02,
the 990-N e-Postcard. A 990-N filer reports gross receipts under $50,000 and
NOTHING ELSE - no Schedule C exists, and none is missing. Zero lobbying
expenditure there is the filing regime, not a finding.

That caveat is a COLUMN, not a footnote: `filing_regime` and
`schedc_expected` say, per row, whether a Schedule C could have existed. Any
denominator built without them is wrong by construction.

WHAT THIS SCRIPT REFUSES TO DO
------------------------------
- It never writes that an earmark RESULTED FROM lobbying. Both events are
  recorded with dates and the reader draws the line. `resulted_from` exists in
  the relationship vocabulary and is used only where a document says so.
- It never treats an organisation lobbying on Native issues as Native.
- It never conflates amount_requested with amount_enacted.
- It never rewrites a row of np_financials.csv. Columns are appended, matched
  on (ein, tax_period), and every pre-existing cell is asserted unchanged.

CONTAINMENT
-----------
Entity resolution imports `resolve_entity` from 33_apply_party_rulings.py - the
one resolver. On top of it this script applies the guards the AGENTS.md
containment defect demands: the record's name must be at least as specific as
the entity's, and where both carry a state they must agree. A containment match
on name alone lands at Tier B, never A.

Steps
-----
  probe            reachability, host locks, no writes
  irs-index        stream the IRS index CSVs, keep only our EINs
  irs-xml          pull the returns out of the IRS ZIP archives by HTTP range
  irs-deflate64    recover the six archives Python's zipfile cannot decompress
  schedc           parse Schedule C, APPEND columns to np_financials.csv
  earmarks-pull    retrieve House XLSX, Senate JSON and the enacted PDFs
  earmarks-stage   normalise them into one staging file with source quotes
  earmarks         resolve recipients -> data/clean/earmarks.csv
  crosscheck       990-reported lobbying vs LDA -> review/schedc_lda_gaps_<date>.csv
  codebook         variable entries only
  report           print the numbers

Run order: probe, irs-index, irs-xml, irs-deflate64, schedc,
earmarks-pull, earmarks-stage, earmarks, crosscheck, codebook, report.
Every step is resumable and every fetch step checkpoints before its first
request, so a killed run loses nothing.

Reads  data/clean/np_financials.csv, data/clean/np_orgs.csv
       data/clean/native_entity_lobbying_disclosures.csv
       data/spine/cedar_entity_spine.csv
       data/raw/external/earmarks/*            (retrieved disclosure tables)
Writes data/clean/earmarks.csv
       data/clean/np_financials.csv            (APPENDED COLUMNS ONLY)
       data/clean/codebook_master.csv          (variable rows only)
       review/schedc_lda_gaps_<date>.csv
       review/earmark_unresolved_<date>.csv
       docs/EARMARKS_SCHEDC_BUILD_LOG.md
"""

import argparse
import csv
import importlib.util
import io
import json
import os
import re
import sys
import time
import unicodedata
import urllib.error
import urllib.request
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from pathlib import Path

CEDAR = Path(__file__).resolve().parent.parent
CLEAN = CEDAR / "data" / "clean"
SPINE = CEDAR / "data" / "spine"
RAW = CEDAR / "data" / "raw" / "external"
REVIEW = CEDAR / "review"
LOGS = CEDAR / "logs"
DOCS = CEDAR / "docs"

SCHEDC_RAW = RAW / "irs990_schedc"
EARMARK_RAW = RAW / "earmarks"

TODAY = date.today().isoformat()
UA = ("CedarPress/1.0 (research data; elijahsamsonmoreno@gmail.com)")

csv.field_size_limit(min(sys.maxsize, 2**31 - 1))

sys.path.insert(0, str(CEDAR / "code"))
from cedar_keys import surrogate_id                            # noqa: E402

# --------------------------------------------------------------------------
# THE PRIMARY KEY OF earmarks.csv, AND WHAT IT IS MADE OF
#
# Three sites minted `earmark_id`, in three different broken ways:
#
#   House      f"EMK-H{fy}-{n+1:05d}"                     POSITIONAL
#   Senate     f"EMK-S{fy}-{n+1:05d}"                     POSITIONAL
#   Explanatory f"EMK-E{fy}-{abs(hash(p.stem)) % 10**6}-{n:05d}"
#                                                         PROCESS HASH
#
# The third is the SAME defect as `ferc_filing_id`: `hash()` on a string is
# randomised per process by PYTHONHASHSEED, so every one of those ids changes
# on every single run of this build. `cedar_keys.NON_DETERMINISTIC_COLUMNS`
# has recorded it since 2026-08-26 and it had never been written down
# anywhere else.
#
# All three now mint the SAME deterministic blake2b digest, because three
# branches writing one column must agree or the column holds three
# vocabularies. The columns are the ones `cedar_keys` already recorded as the
# join-instead, PLUS `source_url` and `source_quote`.
#
# WHY THOSE TWO EXTRA COLUMNS, measured rather than assumed: the six-column
# form leaves 7 collisions over 1,002 rows - one member requesting the same
# project twice in one year is a real thing and both requests are real rows.
# The eight-column form is unique with 0 blanks.
#
# THIS IS A WIDE KEY and it says so: every column in it is an attribute, so
# the id is an identity for the row's CONTENT and changes if any of those
# values is corrected. Load and diff on it; do not make it a foreign-key
# target. (Nothing in the repo references it today - proven by the full value
# scan in `327_migrate_class7_keys_to_digests.py`.)
# --------------------------------------------------------------------------
EARMARK_KEY_COLUMNS = ["fiscal_year", "chamber", "requesting_member",
                       "recipient_name", "project_title", "amount_enacted",
                       "source_url", "source_quote"]

# IRS e-file index coverage. Probed 2026-08-07: 2009-2016 404 at
# apps.irs.gov/pub/epostcard/990/xml/{Y}/ and at the S3 bucket root.
INDEX_YEARS = list(range(2017, 2027))
INDEX_URL = "https://apps.irs.gov/pub/epostcard/990/xml/{y}/index_{y}.csv"
XML_URL = "https://s3.amazonaws.com/irs-form-990/{oid}_public.xml"

IRS_NS = "{http://www.irs.gov/efile}"

# BMF FILING_REQ_CD -> filing regime. 17_build_nonprofit_990.py already rules
# 02 = 990-N; this reuses the same mapping rather than inventing a second one.
FILING_REGIME = {
    "01": "990_or_990EZ",
    "02": "990_N",              # e-Postcard. No Schedule C can exist.
    "03": "990_or_990EZ",
    "06": "990_PF",
    "07": "990_or_990EZ",
    "13": "990_or_990EZ",
    "14": "not_required",
    "00": "not_required",
}


# ---------------------------------------------------------------------------
# infrastructure
# ---------------------------------------------------------------------------

def log(msg):
    print(msg, flush=True)


def read_csv(p, encoding="utf-8-sig"):
    p = Path(p)
    if not p.exists():
        return []
    with open(p, encoding=encoding, newline="") as fh:
        return list(csv.DictReader(fh))


def write_csv(p, rows, fields=None):
    p = Path(p)
    p.parent.mkdir(parents=True, exist_ok=True)
    if fields is None:
        fields = list(rows[0].keys()) if rows else []
    with open(p, "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)


def load_m33():
    """The ONE resolver. Standing rule 8 - never write another name matcher.

    `norm` and `core` are memoised, and ONLY memoised. They are pure functions
    of a string, so caching cannot change an answer - and it has to happen,
    because `resolve_entity` re-normalises all 1,310 spine names on every call.
    Over 7,800 earmark rows that is roughly 16 million Unicode normalisations
    and the step ran for over ten minutes without finishing. The resolver's
    logic is untouched; re-implementing it is what rule 8 forbids.
    """
    from functools import lru_cache
    spec = importlib.util.spec_from_file_location(
        "m33", CEDAR / "code" / "33_apply_party_rulings.py")
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    _n, _c = m.norm, m.core
    m.norm = lru_cache(maxsize=None)(_n)
    m.core = lru_cache(maxsize=None)(lambda s: frozenset(_c(s)))
    return m


def load_domain():
    spec = importlib.util.spec_from_file_location(
        "cedar_domain", CEDAR / "code" / "cedar_domain.py")
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    return m


# ---- host locks (docs/PULL_DISCIPLINE.md rule 2) ---------------------------

def hostlock_path(host):
    return LOGS / f"_HOSTLOCK_{host}.json"


def claim_host(host, note):
    """Claim, or defer. Returns True if we may poll this host."""
    p = hostlock_path(host)
    if p.exists():
        try:
            cur = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            cur = {}
        pid = cur.get("pid") or (cur.get("holder") or {}).get("pid")
        if pid and pid != os.getpid() and _pid_alive(pid):
            cur.setdefault("queue", []).append(
                {"script": "code/99_build_earmarks_and_schedc.py",
                 "requested_at": datetime.now(timezone.utc).isoformat(),
                 "work": note})
            p.write_text(json.dumps(cur, indent=1), encoding="utf-8")
            log(f"  ! {host} held by pid {pid}; queued and DEFERRING")
            return False
    LOGS.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(
        {"host": host, "pid": os.getpid(),
         "script": "code/99_build_earmarks_and_schedc.py",
         "started": datetime.now(timezone.utc).isoformat(),
         "queue": [], "note": note}, indent=1), encoding="utf-8")
    log(f"  + claimed {host}")
    return True


def release_host(host):
    p = hostlock_path(host)
    if not p.exists():
        return
    try:
        cur = json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return
    if cur.get("pid") == os.getpid():
        cur["active"] = False
        cur["released"] = datetime.now(timezone.utc).isoformat()
        p.write_text(json.dumps(cur, indent=1), encoding="utf-8")


def _pid_alive(pid):
    """Win32_Process, never `ps` - PULL_DISCIPLINE says ps cannot answer this."""
    import subprocess
    try:
        out = subprocess.run(
            ["powershell", "-NoProfile", "-Command",
             f"(Get-CimInstance Win32_Process -Filter 'ProcessId={int(pid)}')"
             ".ProcessId"],
            capture_output=True, text=True, timeout=60).stdout.strip()
        return out.isdigit()
    except Exception:
        return False


class Fetcher:
    """Sequential, spaced, exponentially backing off. One per host.

    Distinguishes the three failure shapes PULL_DISCIPLINE names:
    an instant RemoteDisconnected is an EDGE BLOCK and we stop; a 429 is a
    throttle and we honour Retry-After; a slow timeout is just slow.
    """

    def __init__(self, gap=0.25, max_backoff=1800, give_up_after=7200):
        self.gap = gap
        self.max_backoff = max_backoff
        self.give_up_after = give_up_after
        self.last = 0.0
        self.blocked = False
        self.stats = Counter()

    def get(self, url, timeout=90, binary=True):
        """Returns (status, bytes|None). status 0 == transport failure."""
        if self.blocked:
            return 0, None
        wait = max(0.0, self.gap - (time.time() - self.last))
        if wait:
            time.sleep(wait)
        backoff = 60.0
        spent = 0.0
        while True:
            t0 = time.time()
            try:
                req = urllib.request.Request(url, headers={
                    "User-Agent": UA,
                    "Accept": "*/*",
                    "Accept-Language": "en-US,en;q=0.9"})
                with urllib.request.urlopen(req, timeout=timeout) as r:
                    body = r.read()
                self.last = time.time()
                self.stats["ok"] += 1
                return 200, body
            except urllib.error.HTTPError as e:
                self.last = time.time()
                if e.code == 429:
                    ra = e.headers.get("Retry-After")
                    sl = float(ra) if (ra or "").isdigit() else backoff
                    self.stats["throttled"] += 1
                    if spent + sl > self.give_up_after:
                        self.blocked = True
                        return 429, None
                    time.sleep(sl)
                    spent += sl
                    backoff = min(backoff * 2, self.max_backoff)
                    continue
                self.stats[f"http_{e.code}"] += 1
                return e.code, None
            except Exception as e:
                dt = time.time() - t0
                self.last = time.time()
                edge = dt < 1.0 and "timed out" not in str(e).lower()
                self.stats["edge_refused" if edge else "transport"] += 1
                if spent + backoff > self.give_up_after:
                    self.blocked = True
                    log(f"  !! giving up on {url} after {spent:.0f}s: {e}")
                    return 0, None
                time.sleep(backoff)
                spent += backoff
                backoff = min(backoff * 2, self.max_backoff)


# ---------------------------------------------------------------------------
# STEP: probe
# ---------------------------------------------------------------------------

PROBES = [
    ("apps.irs.gov", "https://apps.irs.gov/pub/epostcard/990/xml/2023/index_2023.csv"),
    ("s3.amazonaws.com", "https://s3.amazonaws.com/irs-form-990/index_2023.csv"),
    ("appropriations.house.gov", "https://appropriations.house.gov/"),
    ("www.appropriations.senate.gov", "https://www.appropriations.senate.gov/"),
]


def step_probe():
    log("=== 99 probe ===")
    for host, url in PROBES:
        t0 = time.time()
        try:
            req = urllib.request.Request(url, headers={"User-Agent": UA},
                                         method="HEAD")
            with urllib.request.urlopen(req, timeout=30) as r:
                log(f"  {host:32s} {r.status}  "
                    f"{r.headers.get('Content-Length','-')}  "
                    f"{time.time()-t0:.1f}s")
        except Exception as e:
            log(f"  {host:32s} ERR {type(e).__name__} {str(e)[:70]}  "
                f"{time.time()-t0:.1f}s")
    for p in sorted(LOGS.glob("_HOSTLOCK_*.json")):
        try:
            j = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            continue
        log(f"  lock {j.get('host')}: pid={j.get('pid')} "
            f"script={j.get('script')}")


# ---------------------------------------------------------------------------
# STEP: irs-index
# ---------------------------------------------------------------------------

def target_eins():
    """EINs we care about, and what each is wanted for.

    fin  = present in np_financials.csv (the append target, per tax period)
    univ = present in np_orgs.csv with a filing regime that CAN have a
           Schedule C (990/990-EZ/990-PF). 990-N and not-required are excluded
           HERE, not later, because including them would manufacture zeros.
    """
    fin = {r["ein"].strip().zfill(9) for r in read_csv(CLEAN / "np_financials.csv")
           if (r.get("ein") or "").strip()}
    univ = {}
    for r in read_csv(CLEAN / "np_orgs.csv"):
        ein = (r.get("EIN") or "").strip().replace("-", "").zfill(9)
        if not ein or not ein.isdigit():
            continue
        regime = FILING_REGIME.get((r.get("bmf_filing_req_cd") or "").strip(),
                                   "unknown")
        univ[ein] = regime
    return fin, univ


def step_irs_index(limit_years=None):
    log("=== 99 irs-index ===")
    SCHEDC_RAW.mkdir(parents=True, exist_ok=True)
    fin, univ = target_eins()
    schedc_possible = {e for e, g in univ.items()
                       if g in ("990_or_990EZ", "990_PF", "unknown")}
    wanted = fin | schedc_possible
    log(f"  np_financials EINs           {len(fin):,}")
    log(f"  np_orgs EINs                 {len(univ):,}")
    log(f"  of which Schedule C possible {len(schedc_possible):,}")
    log(f"  index filter set             {len(wanted):,}")

    if not claim_host("apps.irs.gov", "IRS 990 e-file index CSVs 2017-2026"):
        return

    years = limit_years or INDEX_YEARS
    out = SCHEDC_RAW / "_index_targets.csv"
    have = {}
    for r in read_csv(out):
        have.setdefault(int(r["index_year"]), 0)
        have[int(r["index_year"])] += 1
    rows = read_csv(out)
    done_years = {y for y, n in have.items() if n > 0}

    f = Fetcher(gap=1.0)
    for y in years:
        if y in done_years:
            log(f"  {y}: cached ({have[y]:,} rows)")
            continue
        url = INDEX_URL.format(y=y)
        t0 = time.time()
        status, body = f.get(url, timeout=600)
        if status != 200 or not body:
            log(f"  {y}: HTTP {status} - SKIPPED (recorded, not smoothed)")
            continue
        text = body.decode("utf-8", "replace")
        rdr = csv.DictReader(io.StringIO(text))
        kept = 0
        total = 0
        for rec in rdr:
            total += 1
            ein = (rec.get("EIN") or "").strip().replace("-", "").zfill(9)
            if ein not in wanted:
                continue
            rt = (rec.get("RETURN_TYPE") or "").strip()
            rows.append({
                "index_year": y,
                "ein": ein,
                "object_id": (rec.get("OBJECT_ID") or "").strip(),
                "return_type": rt,
                "tax_period": (rec.get("TAX_PERIOD") or "").strip(),
                "taxpayer_name": (rec.get("TAXPAYER_NAME") or "").strip(),
                "sub_date": (rec.get("SUB_DATE") or "").strip(),
                "dln": (rec.get("DLN") or "").strip(),
                "index_url": url,
                "fetched_date": TODAY,
            })
            kept += 1
        log(f"  {y}: {total:,} index rows -> {kept:,} ours "
            f"({time.time()-t0:.0f}s, {len(body)/1e6:.0f}MB streamed)")
        write_csv(out, rows)
    write_csv(out, rows)
    release_host("apps.irs.gov")
    log(f"  wrote {out}  ({len(rows):,} rows)")


# ---------------------------------------------------------------------------
# STEP: irs-xml
# ---------------------------------------------------------------------------

def step_irs_xml(max_fetch=None, priority_only=False):
    log("=== 99 irs-xml ===")
    idx = read_csv(SCHEDC_RAW / "_index_targets.csv")
    if not idx:
        log("  no index; run --steps irs-index first")
        return
    fin, univ = target_eins()

    xmldir = SCHEDC_RAW / "xml"
    xmldir.mkdir(parents=True, exist_ok=True)

    # Priority 1: every (ein, tax_period) that np_financials actually holds -
    # those are the rows we are appending onto. Priority 2: the latest return
    # per remaining EIN, which is what the universe-level cross-check needs.
    finkeys = {(r["ein"].strip().zfill(9), (r.get("tax_period") or "").strip())
               for r in read_csv(CLEAN / "np_financials.csv")}

    # RETURN_TYPE values observed in the 2017-2026 indexes: 990, 990EZ, 990O,
    # 990EO, 990PF, 990T, 990PR. 990T is the unrelated-business-income return
    # and carries no Schedule C; 990PR is Puerto Rico. Fetching them would cost
    # requests and return nothing, so they are excluded here rather than
    # parsed and discarded.
    SCHEDC_BEARING = {"990", "990EZ", "990O", "990EO", "990PF"}
    p1, p2 = [], []
    latest = {}
    for r in idx:
        if (r.get("return_type") or "") not in SCHEDC_BEARING:
            continue
        k = (r["ein"], r["tax_period"])
        if k in finkeys:
            p1.append(r)
        else:
            cur = latest.get(r["ein"])
            if cur is None or r["tax_period"] > cur["tax_period"]:
                latest[r["ein"]] = r
    p2 = [r for k, r in latest.items() if k not in {e for e, _ in finkeys}]

    queue = p1 + ([] if priority_only else p2)
    queue = [r for r in queue
             if not (xmldir / f"{r['object_id']}.xml").exists()]
    log(f"  priority-1 (np_financials rows)  {len(p1):,}")
    log(f"  priority-2 (latest per other EIN) {len(p2):,}")
    log(f"  to fetch (not cached)             {len(queue):,}")
    if max_fetch:
        queue = queue[:max_fetch]
    if not queue:
        return

    if not claim_host("apps.irs.gov", "IRS 990 e-file XML via ZIP range reads"):
        return

    want = {r["object_id"]: r for r in queue}
    fl = SCHEDC_RAW / "_xml_fetch_log.csv"
    flog = read_csv(fl)
    seen = {r["object_id"] for r in flog}

    f = Fetcher(gap=0.3)
    zips = zip_manifest(f)
    years_needed = sorted({r["index_year"] for r in queue})
    todo = [z for z in zips if z["year"] in years_needed]
    log(f"  zip archives to open: {len(todo)} (years {years_needed})")

    n_ok = 0
    for z in todo:
        if not want:
            break
        try:
            hf = HttpRangeFile(z["url"], f)
            import zipfile
            zf = zipfile.ZipFile(hf)
            names = zf.namelist()
        except Exception as e:
            log(f"  !! {z['name']}: cannot open ({type(e).__name__} {e})")
            continue
        # Members are named "<OBJECT_ID>_public.xml".
        bymember = {}
        for nm in names:
            base = nm.rsplit("/", 1)[-1]
            oid = base.split("_")[0]
            if oid in want:
                bymember[oid] = nm
        log(f"  {z['name']}: {len(names):,} members, {len(bymember):,} ours "
            f"({hf.bytes_read/1e6:.0f}MB read)")
        got = 0
        for oid, nm in bymember.items():
            try:
                body = zf.read(nm)
            except Exception as e:
                log(f"    !! {oid}: {type(e).__name__} {e}")
                continue
            (xmldir / f"{oid}.xml").write_bytes(body)
            r = want.pop(oid)
            got += 1
            n_ok += 1
            if oid not in seen:
                flog.append({"object_id": oid, "ein": r["ein"],
                             "tax_period": r["tax_period"],
                             "return_type": r["return_type"],
                             "url": z["url"], "zip_member": nm,
                             "http_status": 200, "fetched_date": TODAY})
                seen.add(oid)
        log(f"    extracted {got:,}  ({hf.bytes_read/1e6:.0f}MB total read; "
            f"{len(want):,} still wanted)")
        write_csv(fl, flog, ["object_id", "ein", "tax_period", "return_type",
                             "url", "zip_member", "http_status", "fetched_date"])
        if f.blocked:
            log("  !! host blocked; stopping (checkpoint written)")
            break

    # Anything still wanted was indexed by the IRS but is not in any archive.
    # That is a source disagreement and is recorded, not smoothed away.
    for oid, r in want.items():
        if oid not in seen:
            flog.append({"object_id": oid, "ein": r["ein"],
                         "tax_period": r["tax_period"],
                         "return_type": r["return_type"], "url": "",
                         "zip_member": "",
                         "http_status": "indexed_but_absent_from_archives",
                         "fetched_date": TODAY})
    write_csv(fl, flog, ["object_id", "ein", "tax_period", "return_type",
                         "url", "zip_member", "http_status", "fetched_date"])
    release_host("apps.irs.gov")
    log(f"  extracted ok={n_ok:,}  indexed-but-absent={len(want):,}  "
        f"stats={dict(f.stats)}")


def step_irs_deflate64():
    """Recover the returns Python's `zipfile` cannot decompress.

    Six of the 81 IRS archives are written with DEFLATE64 (compression method
    9). CPython's `zipfile` raises `NotImplementedError: That compression
    method is not supported`, and the pure-Python replacement needs a C build
    toolchain that is not on this machine. Measured: 1,282 returns lost that
    way, 95 of them rows of np_financials.

    Range reads cannot help here - the bytes arrive fine, it is the decoder
    that is missing. So these six archives are downloaded whole, opened with
    the system 7-Zip (which does implement DEFLATE64), the wanted members are
    extracted, and the archive is DELETED before the next one starts. Peak disk
    is one archive, about 500MB, not the 2.8GB the six would occupy together.

    If 7-Zip is absent this step does nothing and the affected rows keep
    `schedc_basis = efile_return_indexed_not_retrieved`, which is the truth.
    """
    log("=== 99 irs-deflate64 ===")
    import shutil
    import subprocess
    sevenzip = shutil.which("7z") or r"C:\Program Files\7-Zip\7z.exe"
    if not Path(sevenzip).exists():
        log("  7-Zip not found; DEFLATE64 archives cannot be read. Skipping.")
        return
    log(f"  using {sevenzip}")

    xmldir = SCHEDC_RAW / "xml"
    xmldir.mkdir(parents=True, exist_ok=True)
    idx = read_csv(SCHEDC_RAW / "_index_targets.csv")
    finkeys = {(r["ein"].strip().zfill(9), (r.get("tax_period") or "").strip())
               for r in read_csv(CLEAN / "np_financials.csv")}
    SCHEDC_BEARING = {"990", "990EZ", "990O", "990EO", "990PF"}
    want = {}
    latest = {}
    for r in idx:
        if r.get("return_type") not in SCHEDC_BEARING:
            continue
        if (xmldir / f"{r['object_id']}.xml").exists():
            continue
        if (r["ein"], r["tax_period"]) in finkeys:
            want[r["object_id"]] = r
        else:
            cur = latest.get(r["ein"])
            if cur is None or r["tax_period"] > cur["tax_period"]:
                latest[r["ein"]] = r
    for r in latest.values():
        want.setdefault(r["object_id"], r)
    log(f"  still missing: {len(want):,} returns")
    if not want:
        return

    zips = {z["name"]: z for z in zip_manifest(Fetcher(gap=0.3))}
    targets = ["2020_TEOS_XML_CT1.zip", "2025_TEOS_XML_05A.zip",
               "2025_TEOS_XML_05B.zip", "2025_TEOS_XML_11B.zip",
               "2026_TEOS_XML_05A.zip", "2026_TEOS_XML_05B.zip"]
    tmp = SCHEDC_RAW / "_tmp"
    tmp.mkdir(exist_ok=True)
    got = 0
    for name in targets:
        z = zips.get(name)
        if not z or not want:
            continue
        local = tmp / name
        log(f"  downloading {name} ...")
        try:
            req = urllib.request.Request(z["url"], headers={"User-Agent": UA})
            with urllib.request.urlopen(req, timeout=900) as r, \
                    open(local, "wb") as fh:
                shutil.copyfileobj(r, fh, 1 << 20)
        except Exception as e:
            log(f"    !! download failed: {type(e).__name__} {e}")
            local.unlink(missing_ok=True)
            continue
        out = tmp / "x"
        if out.exists():
            shutil.rmtree(out, ignore_errors=True)
        out.mkdir()
        listing = subprocess.run([sevenzip, "l", "-ba", "-slt", str(local)],
                                 capture_output=True, text=True).stdout
        members = re.findall(r"^Path = (.+\.xml)$", listing, re.M)
        mine = [mm for mm in members
                if mm.rsplit("\\", 1)[-1].rsplit("/", 1)[-1].split("_")[0] in want]
        log(f"    {len(members):,} members, {len(mine):,} ours")
        for chunk in [mine[i:i + 200] for i in range(0, len(mine), 200)]:
            subprocess.run([sevenzip, "e", "-y", f"-o{out}", str(local)]
                           + chunk, capture_output=True, text=True)
        for f in out.glob("*.xml"):
            oid = f.name.split("_")[0]
            if oid in want:
                shutil.move(str(f), str(xmldir / f"{oid}.xml"))
                want.pop(oid, None)
                got += 1
        shutil.rmtree(out, ignore_errors=True)
        local.unlink(missing_ok=True)
        log(f"    recovered {got:,} so far; {len(want):,} still missing")
    shutil.rmtree(tmp, ignore_errors=True)
    log(f"  recovered {got:,} returns from DEFLATE64 archives")


IRS_DOWNLOAD_PAGE = ("https://www.irs.gov/charities-non-profits/"
                     "form-990-series-downloads")


def zip_manifest(fetcher):
    """The authoritative list of e-file archives, from the IRS's own page.

    Guessing the filenames does not work: 2017-2020 use
    `download990xml_{Y}_{n}.zip`, 2021-2026 use `{Y}_TEOS_XML_{NN}{A..D}.zip`,
    and both patterns coexist in some years. The IRS publishes the list, so it
    is read rather than reconstructed - AGENTS.md: check the HTTP status, not
    the file, and never infer a URL that a source is willing to state.
    """
    cache = SCHEDC_RAW / "_zip_manifest.csv"
    rows = read_csv(cache)
    if rows:
        return rows
    status, body = fetcher.get(IRS_DOWNLOAD_PAGE, timeout=120)
    if status != 200 or not body:
        log(f"  !! IRS download page HTTP {status}")
        return []
    html = body.decode("utf-8", "replace")
    seen, rows = set(), []
    for m in re.finditer(r'(?:https?:)?//apps\.irs\.gov/pub/epostcard/990/'
                         r'xml/(\d{4})/([\w\.\-]+\.zip)', html):
        y, name = m.group(1), m.group(2)
        url = f"https://apps.irs.gov/pub/epostcard/990/xml/{y}/{name}"
        if url in seen:
            continue
        seen.add(url)
        rows.append({"year": y, "name": name, "url": url,
                     "basis": "listed_on_irs_download_page",
                     "source_url": IRS_DOWNLOAD_PAGE, "fetched_date": TODAY})

    # 2017 and 2018 archives EXIST but the IRS page no longer links them. They
    # follow the 2019/2020 naming the page does state, and each candidate below
    # was confirmed HTTP 200 with a real Content-Length before being added -
    # the status is checked, not the plausibility of the URL. Absent years are
    # left absent rather than guessed at.
    listed_years = {r["year"] for r in rows}
    for y in ("2017", "2018"):
        if y in listed_years:
            continue
        for n in range(1, 12):
            name = f"download990xml_{y}_{n}.zip"
            url = f"https://apps.irs.gov/pub/epostcard/990/xml/{y}/{name}"
            try:
                req = urllib.request.Request(url, headers={"User-Agent": UA},
                                             method="HEAD")
                with urllib.request.urlopen(req, timeout=60) as r:
                    if r.status != 200 or not r.headers.get("Content-Length"):
                        break
            except Exception:
                break
            rows.append({"year": y, "name": name, "url": url,
                         "basis": "probe_verified_http_200_not_page_listed",
                         "source_url": IRS_DOWNLOAD_PAGE,
                         "fetched_date": TODAY})
            time.sleep(0.3)
    rows.sort(key=lambda r: (r["year"], r["name"]))
    write_csv(cache, rows)
    log(f"  zip manifest: {len(rows)} archives from {IRS_DOWNLOAD_PAGE}")
    return rows


class HttpRangeFile(io.RawIOBase):
    """Seekable read-only file over HTTP Range requests.

    The IRS retired the per-return S3 objects
    (`s3.amazonaws.com/irs-form-990/<id>_public.xml` now 404s for every id
    tested on 2026-08-07) and publishes only multi-gigabyte ZIP archives. We
    need roughly 7,000 specific returns out of about six million, so
    downloading 30GB to keep 0.1% of it is the wrong trade.

    apps.irs.gov answers `Accept-Ranges: bytes` and returns 206, so
    `zipfile.ZipFile` can be pointed at this object: it reads the end-of-
    central-directory record, then the central directory, then only the local
    header and compressed bytes of each member we actually want.
    """

    WINDOW = 262144

    def __init__(self, url, fetcher):
        self.url = url
        self.f = fetcher
        self.pos = 0
        self.bytes_read = 0
        self._buf = b""
        self._buf_start = -1
        self.size = self._content_length()

    def _content_length(self):
        req = urllib.request.Request(self.url, headers={"User-Agent": UA},
                                     method="HEAD")
        with urllib.request.urlopen(req, timeout=90) as r:
            if r.headers.get("Accept-Ranges") != "bytes":
                raise OSError("host does not advertise byte ranges")
            return int(r.headers["Content-Length"])

    def _range(self, start, length):
        end = min(self.size - 1, start + length - 1)
        if end < start:
            return b""
        backoff, spent = 60.0, 0.0
        while True:
            try:
                req = urllib.request.Request(
                    self.url, headers={"User-Agent": UA,
                                       "Range": f"bytes={start}-{end}"})
                with urllib.request.urlopen(req, timeout=180) as r:
                    b = r.read()
                self.bytes_read += len(b)
                self.f.stats["range_ok"] += 1
                return b
            except Exception as e:
                self.f.stats["range_err"] += 1
                if spent + backoff > 1800:
                    raise
                time.sleep(backoff)
                spent += backoff
                backoff = min(backoff * 2, 600)

    def readable(self):
        return True

    def seekable(self):
        return True

    def tell(self):
        return self.pos

    def seek(self, off, whence=0):
        self.pos = (off if whence == 0 else
                    self.pos + off if whence == 1 else self.size + off)
        return self.pos

    def read(self, n=-1):
        if n is None or n < 0:
            n = self.size - self.pos
        n = max(0, min(n, self.size - self.pos))
        if n == 0:
            return b""
        if (self._buf_start >= 0 and self.pos >= self._buf_start
                and self.pos + n <= self._buf_start + len(self._buf)):
            o = self.pos - self._buf_start
            self.pos += n
            return self._buf[o:o + n]
        if n > self.WINDOW:
            b = self._range(self.pos, n)
            self.pos += len(b)
            return b
        b = self._range(self.pos, self.WINDOW)
        self._buf, self._buf_start = b, self.pos
        out = b[:n]
        self.pos += len(out)
        return out

    def readinto(self, buf):
        b = self.read(len(buf))
        buf[:len(b)] = b
        return len(b)


# ---------------------------------------------------------------------------
# STEP: schedc  - parse and append
# ---------------------------------------------------------------------------

def _txt(el):
    return (el.text or "").strip() if el is not None else ""


def _find(root, *paths):
    for p in paths:
        el = root.find(p.replace("{}", IRS_NS))
        if el is not None and (el.text or "").strip():
            return (el.text or "").strip()
    return ""


def parse_schedule_c(xml_bytes):
    """Read Schedule C out of one IRS e-file return.

    THE TAG NAMES BELOW WERE OBSERVED, NOT GUESSED. An earlier draft of this
    function invented plausible element names (`PaidStaffOrMgmtInd`,
    `LobbyingExpendituresGrp`, `Organization501hElectionInd`) and every one of
    them was wrong - the real schema says `PaidStaffOrManagementInd`,
    `TotalLobbyingExpendGrp`, and carries no election element at all. The
    inventory was taken across 2,647 retrieved returns before this was written.

    THREE REPORTING REGIMES, NEVER MERGED
    -------------------------------------
    Part II-A  501(h) ELECTING filers. Grassroots and direct are reported
               separately against a statutory dollar ceiling. Column (a) is the
               filing organisation; column (b) is an affiliated group and is a
               DIFFERENT entity's money, so only (a) is read.
    Part II-B  NON-ELECTING filers. Yes/no activity checkboxes plus one total.
               No grassroots/direct split exists, so those cells stay blank -
               blank because the form has no such line, not because we failed.
    Part III   501(c)(4)/(5)/(6) dues, proxy tax, and the carry-over.

    THE 501(h) ELECTION IS INFERRED, AND SAYS SO
    --------------------------------------------
    The election is made on Form 5768 and Schedule C carries no element for it.
    What the XML shows is which PART the filer completed, and only an electing
    organisation completes Part II-A. So the value is derived from that and
    `schedc_501h_basis` records the derivation rather than implying a checkbox
    was read.

    ABSENT IS NOT ZERO. A filer that answered "No" to the Form 990 Part IV
    lobbying question files no Schedule C at all. That is a reported fact and
    is distinct from a filer that completed the schedule and reported $0.
    """
    try:
        root = ET.fromstring(xml_bytes)
    except Exception as e:
        return {"schedc_parse_error": type(e).__name__}

    n = IRS_NS
    out = {}

    hdr = root.find(f"{n}ReturnHeader")
    if hdr is not None:
        for tag, col in ((f"{n}TaxPeriodEndDt", "schedc_tax_period_end"),
                         (f"{n}ReturnTypeCd", "schedc_return_type")):
            el = hdr.find(tag)
            if el is not None and (el.text or "").strip():
                out[col] = el.text.strip()

    data = root.find(f"{n}ReturnData")
    if data is None:
        return out

    # --- core-form trigger questions and Part IX lobbying fees -------------
    for tag in (f"{n}IRS990", f"{n}IRS990EZ", f"{n}IRS990PF"):
        core = data.find(tag)
        if core is None:
            continue
        for src, col in (
                ("PoliticalCampaignActyInd", "form990_political_activity_ind"),
                ("LobbyingActivitiesInd", "form990_lobbying_activities_ind"),
                ("InfluenceLegislationInd",
                 "form990pf_influence_legislation_ind"),
                ("LegislativePoliticalActyInd",
                 "form990pf_legislative_political_ind")):
            el = core.find(f"{n}{src}")
            if el is not None and (el.text or "").strip():
                out[col] = _norm_bool(el.text)
        # Form 990 Part IX line 11d - fees paid to OUTSIDE lobbyists. This is
        # a different measurement from Schedule C (which counts the
        # organisation's own lobbying expenditure) and it fills far more often,
        # so it gets its own column and is never added to a Schedule C total.
        g = core.find(f"{n}FeesForServicesLobbyingGrp")
        if g is not None:
            t = g.find(f"{n}TotalAmt")
            if t is not None and (t.text or "").strip():
                out["form990_part9_lobbying_fees"] = t.text.strip()

    sc = data.find(f"{n}IRS990ScheduleC")
    if sc is None:
        out["schedc_present"] = "0"
        return out
    out["schedc_present"] = "1"

    def grp_a(name):
        """Column (a), the FILING organisation's own figure. Column (b) is the
        affiliated group's and belongs to other legal persons."""
        g = sc.find(f"{n}{name}")
        if g is None:
            return ""
        for cand in ("FilingOrganizationsTotalAmt", "TotalAmt"):
            el = g.find(f"{n}{cand}")
            if el is not None and (el.text or "").strip():
                return el.text.strip()
        return ""

    def direct(name):
        el = sc.find(f"{n}{name}")
        return (el.text or "").strip() if el is not None else ""

    # --- Part II-A: 501(h) electing ---------------------------------------
    out["schedc_grassroots_lobbying"] = grp_a("TotalGrassrootsLobbyingGrp")
    out["schedc_direct_lobbying"] = grp_a("TotalDirectLobbyingGrp")
    out["schedc_total_lobbying"] = grp_a("TotalLobbyingExpendGrp")
    out["schedc_lobbying_nontaxable"] = grp_a("LobbyingNontaxableAmountGrp")
    out["schedc_grassroots_nontaxable"] = grp_a("GrassrootsNontaxableGrp")
    out["schedc_exempt_purpose_expend"] = (
        grp_a("TotalExemptPurposeExpendGrp")
        or grp_a("OtherExemptPurposeExpendGrp"))
    part2a = any(sc.find(f"{n}{t}") is not None for t in (
        "TotalLobbyingExpendGrp", "TotalDirectLobbyingGrp",
        "TotalGrassrootsLobbyingGrp", "LobbyingNontaxableAmountGrp",
        "AvgTotalLobbyingExpendGrp", "LobbyingCeilingAmt"))

    # --- Part II-B: non-electing ------------------------------------------
    ACTIVITY = (("VolunteersInd", "schedc_used_volunteers"),
                ("PaidStaffOrManagementInd", "schedc_used_paid_staff"),
                ("MediaAdvertisementsInd", "schedc_used_media"),
                ("MailingsMembersInd", "schedc_used_mailings"),
                ("PublicationsOrBroadcastInd", "schedc_used_publications"),
                ("GrantsOtherOrganizationsInd", "schedc_used_grants"),
                ("DirectContactLegislatorsInd", "schedc_used_direct_contact"),
                ("RalliesDemonstrationsInd", "schedc_used_rallies"),
                ("OtherActivitiesInd", "schedc_used_other"))
    part2b = False
    for src, col in ACTIVITY:
        v = direct(src)
        if v:
            out[col] = _norm_bool(v)
            part2b = True
    if part2b or direct("TotalAmt"):
        out["schedc_nonelecting_total"] = (
            direct("TotalLobbyingExpendituresAmt") or direct("TotalAmt"))

    if part2a:
        out["schedc_501h_election"] = "1"
        out["schedc_501h_basis"] = ("derived: Schedule C Part II-A completed, "
                                    "which only a 501(h) electing filer does")
    elif part2b:
        out["schedc_501h_election"] = "0"
        out["schedc_501h_basis"] = ("derived: Schedule C Part II-B completed, "
                                    "the non-electing regime")
    else:
        out["schedc_501h_basis"] = "not determinable from this return"

    # --- Part I: political campaign activity -------------------------------
    out["schedc_political_expenditure"] = (
        direct("PoliticalExpendituresAmt")
        or direct("TotalExemptFunctionExpendAmt"))
    tot527 = 0.0
    saw527 = False
    for g in sc.findall(f"{n}Section527PoliticalOrgGrp"):
        for tag in ("PaidInternalFundsAmt", "Expended527ActivitiesAmt"):
            el = g.find(f"{n}{tag}")
            v = _numf((el.text or "")) if el is not None else None
            if v is not None:
                tot527 += v
                saw527 = True
    if saw527:
        out["schedc_527_amount"] = f"{tot527:.0f}"

    # --- Part III: dues, proxy tax ----------------------------------------
    out["schedc_dues_received"] = (direct("DuesAssessmentsAmt")
                                   or direct("AggregateReportedDuesNtcAmt"))
    out["schedc_dues_lobbying_political"] = (
        direct("NonDeductibleLbbyngPltclTotAmt")
        or direct("NonDeductibleLbbyngPltclCYAmt"))

    return {k: v for k, v in out.items() if v != ""}


def _norm_bool(v):
    v = (v or "").strip().lower()
    if v in ("1", "true", "x", "yes", "y"):
        return "1"
    if v in ("0", "false", "no", "n"):
        return "0"
    return v


APPEND_COLS = [
    "filing_regime",
    "schedc_expected",
    "schedc_basis",
    "schedc_source_url",
    "schedc_object_id",
    "schedc_present",
    "schedc_501h_election",
    "schedc_501h_basis",
    "schedc_total_lobbying",
    "schedc_direct_lobbying",
    "schedc_grassroots_lobbying",
    "schedc_nonelecting_total",
    "schedc_lobbying_nontaxable",
    "schedc_grassroots_nontaxable",
    "schedc_exempt_purpose_expend",
    "schedc_political_expenditure",
    "schedc_527_amount",
    "schedc_dues_lobbying_political",
    "schedc_dues_received",
    "schedc_used_volunteers",
    "schedc_used_paid_staff",
    "schedc_used_media",
    "schedc_used_mailings",
    "schedc_used_publications",
    "schedc_used_grants",
    "schedc_used_direct_contact",
    "schedc_used_rallies",
    "schedc_used_other",
    "form990_lobbying_activities_ind",
    "form990_political_activity_ind",
    "form990_part9_lobbying_fees",
    "form990pf_influence_legislation_ind",
    "form990pf_legislative_political_ind",
    "schedc_lobbying_usd",
    "schedc_lobbying_basis",
    "schedc_built_date",
]


def step_schedc():
    log("=== 99 schedc ===")
    fin_path = CLEAN / "np_financials.csv"
    fin = read_csv(fin_path)
    orig_fields = list(fin[0].keys()) if fin else []
    # IDEMPOTENT. A re-run must not treat its own previous output as
    # pre-existing data, or the "nothing was rewritten" assertion below would
    # be comparing this build against the last one instead of against the
    # source. Columns this script owns are dropped and rebuilt; every other
    # column is untouchable.
    reran = [c for c in orig_fields if c in APPEND_COLS]
    if reran:
        log(f"  re-run: dropping {len(reran)} previously appended columns "
            f"before rebuilding them")
        orig_fields = [c for c in orig_fields if c not in APPEND_COLS]
        fin = [{k: v for k, v in r.items() if k in orig_fields} for r in fin]
    log(f"  np_financials rows {len(fin):,}  cols {len(orig_fields)}")

    # fill-rate BEFORE
    before = sum(1 for r in fin if (r.get("lobbying_expenditure") or "").strip())
    log(f"  lobbying_expenditure populated BEFORE: {before:,}")

    _, univ = target_eins()
    idx = {(r["ein"], r["tax_period"]): r
           for r in read_csv(SCHEDC_RAW / "_index_targets.csv")}
    xmldir = SCHEDC_RAW / "xml"

    parsed_cache = {}
    n_parsed = n_missing = 0
    out_rows = []
    for r in fin:
        ein = (r.get("ein") or "").strip().zfill(9)
        tp = (r.get("tax_period") or "").strip()
        regime = univ.get(ein, "unknown")
        new = {c: "" for c in APPEND_COLS}
        new["filing_regime"] = regime
        new["schedc_built_date"] = TODAY

        # THE CAVEAT AS A COLUMN. A 990-N filer files no schedule; a blank here
        # is the filing regime, not a missing observation, and must never be
        # counted as a zero.
        if regime == "990_N":
            new["schedc_expected"] = "0"
            new["schedc_basis"] = "990N_filer_no_schedule_exists"
            new["schedc_lobbying_basis"] = "990N_no_financial_detail_filed"
        elif regime == "not_required":
            new["schedc_expected"] = "0"
            new["schedc_basis"] = "bmf_filing_not_required"
            new["schedc_lobbying_basis"] = "no_filing_requirement"
        else:
            new["schedc_expected"] = "1"

        hit = idx.get((ein, tp))
        if hit and new["schedc_expected"] == "1":
            oid = hit["object_id"]
            p = xmldir / f"{oid}.xml"
            if p.exists():
                if oid not in parsed_cache:
                    parsed_cache[oid] = parse_schedule_c(p.read_bytes())
                sc = parsed_cache[oid]
                for k, v in sc.items():
                    if k in new and v != "":
                        new[k] = v
                new["schedc_object_id"] = oid
                new["schedc_source_url"] = XML_URL.format(oid=oid)
                if sc.get("schedc_parse_error"):
                    new["schedc_basis"] = "irs_efile_xml_parse_error"
                elif sc.get("schedc_present") == "1":
                    new["schedc_basis"] = "irs_efile_xml_schedule_c"
                    n_parsed += 1
                else:
                    new["schedc_basis"] = "irs_efile_xml_no_schedule_c_filed"
                    n_parsed += 1
            else:
                new["schedc_basis"] = "efile_return_indexed_not_retrieved"
                n_missing += 1
        elif new["schedc_expected"] == "1":
            ty = _num(r.get("tax_year"))
            if ty is not None and ty < 2015:
                new["schedc_basis"] = ("outside_efile_index_coverage_"
                                       "submission_years_2017_2026")
            else:
                new["schedc_basis"] = "no_efile_return_indexed_for_period"
            n_missing += 1

        # ONE headline lobbying number, with its own basis so the three very
        # different reporting regimes never collapse into one column.
        amt, basis = _consolidate_lobbying(new)
        new["schedc_lobbying_usd"] = amt
        if basis:
            new["schedc_lobbying_basis"] = basis

        merged = dict(r)
        merged.update(new)
        out_rows.append(merged)

    # --- APPEND, NEVER REWRITE. Assert every pre-existing cell is untouched.
    assert len(out_rows) == len(fin), "row count changed"
    for a, b in zip(fin, out_rows):
        for c in orig_fields:
            assert (a.get(c) or "") == (b.get(c) or ""), \
                f"pre-existing cell changed in column {c}"
    new_fields = orig_fields + [c for c in APPEND_COLS if c not in orig_fields]

    bak = fin_path.with_suffix(f".csv.bak_{TODAY}_pre99")
    if not bak.exists():
        bak.write_bytes(fin_path.read_bytes())
    write_csv(fin_path, out_rows, new_fields)
    log(f"  backup {bak.name}")
    log(f"  appended {len(new_fields)-len(orig_fields)} columns; "
        f"rows unchanged at {len(out_rows):,}")

    after = sum(1 for r in out_rows if (r.get("schedc_lobbying_usd") or "").strip())
    expected = sum(1 for r in out_rows if r["schedc_expected"] == "1")
    n990n = sum(1 for r in out_rows if r["filing_regime"] == "990_N")
    log(f"  schedule-C returns parsed          {n_parsed:,}")
    log(f"  indexed-but-unretrieved / no index {n_missing:,}")
    log(f"  rows where a Schedule C could exist {expected:,}")
    log(f"  rows that are 990-N (excluded, not zeroed) {n990n:,}")
    log(f"  schedc_lobbying_usd populated AFTER {after:,}")


def _num(v):
    try:
        return int(float(str(v).strip()))
    except Exception:
        return None


def _consolidate_lobbying(new):
    """One number, and the regime that produced it. Never a silent sum.

    Two regimes are genuinely different measurements and are ranked, never
    added:
      501(h) ELECTING  -> Part II-A total (direct + grassroots), a bright line
      NON-ELECTING     -> Part II-B total, no direct/grassroots split exists

    `form990_part9_lobbying_fees` is deliberately NOT a fallback here. Part IX
    line 11d counts fees paid to OUTSIDE lobbyists; Schedule C counts the
    organisation's own lobbying expenditure. They overlap without being the
    same quantity, and letting one stand in for the other would produce a
    column whose meaning changes from row to row. It keeps its own column.
    """
    for col, basis in (
            ("schedc_total_lobbying", "schedc_part2a_501h_electing_total"),
            ("schedc_nonelecting_total", "schedc_part2b_nonelecting_total")):
        v = (new.get(col) or "").strip()
        if v not in ("", "0") and _numf(v) is not None:
            return v, basis
    # An explicit reported zero is a fact and is kept as one.
    for col, basis in (("schedc_total_lobbying", "schedc_part2a_reported_zero"),
                       ("schedc_nonelecting_total",
                        "schedc_part2b_reported_zero")):
        if (new.get(col) or "").strip() == "0":
            return "0", basis
    if new.get("schedc_present") == "0":
        return "", "no_schedule_c_filed_with_return"
    if new.get("schedc_present") == "1":
        return "", "schedule_c_filed_no_expenditure_reported"
    return "", ""


def _numf(v):
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except Exception:
        return None


# ---------------------------------------------------------------------------
# STEP: earmarks
# ---------------------------------------------------------------------------

# The House posts ONE consolidated request spreadsheet per fiscal year. Each
# URL below was fetched and its sheet inspected before being written here; the
# schema is NOT stable across years, which is why the reader is per-year rather
# than a single column map.
HOUSE_REQUEST_FILES = [
    ("2022", "https://democrats-appropriations.house.gov/sites/evo-subsites/"
             "democrats-appropriations.house.gov/files/documents/"
             "Community%20Project%20Funding%20Request%20Table%20-%20"
             "06222021.csv_.xlsx"),
    ("2023", "https://democrats-appropriations.house.gov/sites/evo-subsites/"
             "democrats-appropriations.house.gov/files/documents/"
             "FY23%20CPF%20Requests%202022-06-06.xlsx"),
    ("2024", "https://appropriations.house.gov/sites/evo-subsites/"
             "republicans-appropriations.house.gov/files/"
             "FY%202024%20House%20CPF%20Requests%202023-04-27%20%28430pm%29.xlsx"),
    ("2025", "https://appropriations.house.gov/sites/evo-subsites/"
             "republicans-appropriations.house.gov/files/evo-media-document/"
             "fy25-house-cpfs-as-requested-06.28.2024.xlsx"),
    ("2026", "https://appropriations.house.gov/sites/evo-subsites/"
             "republicans-appropriations.house.gov/files/evo-media-document/"
             "fy26-house-cpf-consolidated.xlsx"),
]

# The Senate runs a DataTables grid over an undocumented ColdFusion endpoint.
# Method name -> fiscal year, confirmed against the certification-letter paths
# each row carries (cdsletters23/, cdsletters24/, ...). FY2025 uses the
# UNSUFFIXED method; getCDSTable25 does not exist. FY2022 is not in it at all.
SENATE_CDS_ENDPOINT = ("https://www.appropriations.senate.gov/"
                       "cfc_extensions/data/cds_requests.cfc")
SENATE_CDS_METHODS = [
    ("2023", "getCDSTable23"),
    ("2024", "getCDSTable24"),
    ("2025", "getCDSTable"),
    ("2026", "getCDSTable26"),
    ("2027", "getCDSTable27"),
]

# Landing pages carrying the ENACTED joint explanatory statement tables. The
# per-subcommittee PDF filenames are scraped from these rather than guessed:
# FY24 CJS alone exists in four revisions and picking the wrong one
# double-counts.
#
# FY2022 and FY2023 have no live landing page (both slugs 404), but the tables
# themselves are still served from the minority's document directory. Those
# filenames are listed explicitly below and each is status-checked and
# magic-byte-checked before it is written, so a stale name is recorded as a
# failure rather than shipped as an empty table.
ENACTED_DIRECT_FILES = [
    ("2022", "https://democrats-appropriations.house.gov/sites/evo-subsites/"
             "democrats-appropriations.house.gov/files/documents/{}")
    if False else None,
]
_DEM_DOCS = ("https://democrats-appropriations.house.gov/sites/evo-subsites/"
             "democrats-appropriations.house.gov/files/documents/")
ENACTED_DIRECT_FILES = (
    [("2022", _DEM_DOCS + n) for n in (
        "FY%2022%20Ag%20Funded%20CPF.pdf",
        "FY22%20CJS%20Funded%20CPF_0.pdf",
        "FY22%20Defense%20Funded%20CPF.pdf",
        "FY22%20E%26W%20Funded%20CPF.pdf",
        "FY22%20FSGG%20Funded%20CPF.pdf",
        "FY%2022%20Homeland%20Funded%20CPF.pdf",
        "FY22%20Interior%20Funded%20CPF.pdf",
        "FY%2022%20LHHS%20Funded%20CPF.pdf",
        "FY22%20MilConVA%20Funded%20CPF.pdf",
        "FY%2022%20THUD%20Funded%20CPF.pdf")]
    + [("2023", _DEM_DOCS + n) for n in (
        "AG_DISCL_FINAL.pdf",
        "CJS_DISCL_20221218.pdf",
        "DEFENSE_DISCL_20221208.pdf",
        "EW_DISCL_20221206_3.pdf",
        "FSGG_DISCL_20221206.pdf",
        "HOMELAND_DISCL_20221217_0.pdf",
        "INTERIOR_DISCL_20221218_0.pdf",
        "LHHS_DISCL_20221218_2_0.pdf",
        "MILCON_DISCL_20221217_0.pdf",
        "THUD_DISCL_20221216_0.pdf")])

ENACTED_LANDING_PAGES = [
    ("2024", "https://appropriations.house.gov/"
             "fiscal-year-2024-community-project-funding"),
    ("2025", "https://appropriations.house.gov/"
             "committee-activity/fy25-community-project-funding"),
    ("2026", "https://appropriations.house.gov/"
             "fy26-member-requests/fy26-community-project-funding"),
]

EARMARK_FIELDS = [
    "earmark_id", "fiscal_year", "chamber", "requesting_member",
    "member_state", "member_party", "recipient_name", "entity_id",
    "project_title", "project_description", "amount_requested",
    "amount_enacted", "agency", "account", "subcommittee", "is_enacted",
    "source_url", "source_quote", "fetched_date", "tier", "confidence",
    "built_date",
]


def step_earmarks_pull():
    """Retrieve the disclosure tables. Every byte lands on disk first.

    Three source families, all free and all published by the committees:
      House requests   one consolidated XLSX per fiscal year
      Senate requests  a JSON grid endpoint, paged
      Enacted amounts  per-subcommittee joint explanatory statement PDFs,
                       whose filenames are scraped from the landing pages
    """
    log("=== 99 earmarks-pull ===")
    EARMARK_RAW.mkdir(parents=True, exist_ok=True)
    man = read_csv(EARMARK_RAW / "_SOURCE_MANIFEST.csv")
    have = {r["file"] for r in man}

    # -- House request spreadsheets -------------------------------------
    if claim_host("appropriations.house.gov", "House CPF disclosure tables"):
        f = Fetcher(gap=1.2)
        for fy, url in HOUSE_REQUEST_FILES:
            fn = f"house_fy{fy}_cpf_requests.xlsx"
            if (EARMARK_RAW / fn).exists():
                log(f"  cached {fn}")
                continue
            st, body = f.get(url, timeout=180)
            log(f"  {fn}: HTTP {st} {len(body or b''):,}b")
            if st == 200 and body and body[:2] == b"PK":
                (EARMARK_RAW / fn).write_bytes(body)
                man.append({"file": fn, "url": url, "http_status": st,
                            "bytes": len(body), "fiscal_year": fy,
                            "kind": "house_requests", "fetched_date": TODAY})
            else:
                # appropriations.house.gov serves a full 105KB rendered page
                # with HTTP 404. Trusting "the file has content" would ship a
                # web page as a spreadsheet - AGENTS.md, BIA Southwest Region.
                man.append({"file": fn, "url": url, "http_status": st,
                            "bytes": len(body or b""), "fiscal_year": fy,
                            "kind": "house_requests_FAILED",
                            "fetched_date": TODAY})

        # -- enacted tables: the FY22/FY23 files, listed directly ---------
        for fy, u in ENACTED_DIRECT_FILES:
            base = u.rsplit("/", 1)[-1]
            fn = f"enacted_fy{fy}_{base}"[:150]
            if (EARMARK_RAW / fn).exists():
                continue
            st2, b2 = f.get(u, timeout=180)
            ok = st2 == 200 and b2 and b2[:5] == b"%PDF-"
            if ok:
                (EARMARK_RAW / fn).write_bytes(b2)
            man.append({"file": fn, "url": u, "http_status": st2,
                        "bytes": len(b2 or b""), "fiscal_year": fy,
                        "kind": "enacted_table" if ok
                        else "enacted_table_NOT_A_PDF",
                        "fetched_date": TODAY})
        # -- enacted tables: scrape the landing pages for PDF filenames ---
        for fy, page in ENACTED_LANDING_PAGES:
            st, body = f.get(page, timeout=120)
            if st != 200 or not body:
                log(f"  enacted FY{fy} landing page HTTP {st}")
                continue
            html = body.decode("utf-8", "replace")
            urls = set()
            for m in re.finditer(r'href="([^"]+\.pdf)"', html, re.I):
                u = m.group(1)
                # These pages emit PROTOCOL-RELATIVE hrefs (//host/path). A
                # naive "starts with / so prepend the root" produced
                # https://appropriations.house.gov//appropriations.house.gov/...
                # and 50 HTTP 404s that looked like missing documents.
                if u.startswith("//"):
                    u = "https:" + u
                elif u.startswith("/"):
                    root = re.match(r"https?://[^/]+", page).group(0)
                    u = root + u
                if u.startswith("http"):
                    urls.add(u)
            log(f"  enacted FY{fy}: {len(urls)} pdf links on landing page")
            for u in sorted(urls):
                base = u.rsplit("/", 1)[-1].split("?")[0]
                fn = f"enacted_fy{fy}_{base}"[:150]
                if not fn.lower().endswith(".pdf"):
                    fn += ".pdf"
                if (EARMARK_RAW / fn).exists():
                    continue
                st2, b2 = f.get(u, timeout=180)
                ok = st2 == 200 and b2 and b2[:5] == b"%PDF-"
                if ok:
                    (EARMARK_RAW / fn).write_bytes(b2)
                man.append({"file": fn, "url": u, "http_status": st2,
                            "bytes": len(b2 or b""), "fiscal_year": fy,
                            "kind": "enacted_table" if ok
                            else "enacted_table_NOT_A_PDF",
                            "fetched_date": TODAY})
        release_host("appropriations.house.gov")

    # -- Senate request grid --------------------------------------------
    if claim_host("www.appropriations.senate.gov",
                  "Senate congressionally directed spending grid"):
        import urllib.parse
        f = Fetcher(gap=1.5)
        for fy, method in SENATE_CDS_METHODS:
            fn = f"senate_fy{fy}_cds_requests.json"
            if (EARMARK_RAW / fn).exists():
                log(f"  cached {fn}")
                continue
            allrows, start, total = [], 0, None
            while True:
                data = urllib.parse.urlencode(
                    {"method": method, "start": start, "length": 5000,
                     "search": ""}).encode()
                req = urllib.request.Request(
                    SENATE_CDS_ENDPOINT, data=data,
                    headers={"User-Agent": UA,
                             "Content-Type":
                             "application/x-www-form-urlencoded",
                             "X-Requested-With": "XMLHttpRequest"})
                try:
                    with urllib.request.urlopen(req, timeout=180) as r:
                        j = json.loads(r.read().decode("utf-8", "replace"))
                except Exception as e:
                    log(f"  senate FY{fy} {method} start={start}: "
                        f"{type(e).__name__} {e}")
                    break
                total = j.get("RECORDSTOTAL", j.get("recordsTotal"))
                batch = j.get("DATA", j.get("data") or [])
                allrows.extend(batch)
                if not batch or len(allrows) >= (total or 0):
                    break
                start += len(batch)
                time.sleep(1.5)
            if allrows:
                (EARMARK_RAW / fn).write_text(
                    json.dumps({"method": method, "fiscal_year": fy,
                                "endpoint": SENATE_CDS_ENDPOINT,
                                "records_total": total,
                                "fetched_date": TODAY, "rows": allrows},
                               indent=1), encoding="utf-8")
                man.append({"file": fn, "url": SENATE_CDS_ENDPOINT,
                            "http_status": 200, "bytes": len(allrows),
                            "fiscal_year": fy, "kind": "senate_requests",
                            "fetched_date": TODAY})
            log(f"  senate FY{fy} {method}: {len(allrows):,} of "
                f"{total} records")
        release_host("www.appropriations.senate.gov")

    write_csv(EARMARK_RAW / "_SOURCE_MANIFEST.csv", man,
              ["file", "url", "http_status", "bytes", "fiscal_year", "kind",
               "fetched_date"])
    log(f"  manifest: {len(man)} entries")


# A CANDIDATE net, not a classifier. Everything caught here still has to
# resolve through the spine with the state and specificity guards before it
# becomes a row - which is what keeps Indian River County (FL), Pueblo County
# (CO), the Village of Indian Head Park (IL) and the Naval Surface Warfare
# Center Indian Head Division out of a Native dataset.
NATIVE_SCREEN = re.compile(
    r"\b(tribe|tribes|tribal|nation|nations|pueblo|rancheria|band|bands|"
    r"indian|native|alaska native|anvsa|village of|native village|"
    r"intertribal|inter-tribal|reservation|colony|community of|"
    r"confederated|chippewa|apache|navajo|cherokee|sioux|shoshone|paiute|"
    r"yakama|nez perce|hopi|zuni|oneida|seneca|mohawk|choctaw|chickasaw|"
    r"creek|seminole|osage|ute|crow|blackfeet|menominee|ho-chunk|winnebago|"
    r"lummi|makah|quinault|tulalip|swinomish|suquamish|colville|"
    r"salish|kootenai|arapaho|cheyenne|comanche|kiowa|caddo|wichita|"
    r"potawatomi|ottawa|miami|peoria|quapaw|modoc|wyandotte|shawnee|"
    r"tohono|akimel|gila river|salt river|ak-chin|yavapai|havasupai|"
    r"hualapai|kaibab|chemehuevi|cocopah|quechan|mohave|yaqui|"
    r"jicarilla|mescalero|acoma|laguna|isleta|jemez|santa clara|"
    r"taos|tesuque|nambe|pojoaque|picuris|zia|sandia|cochiti|"
    r"ohkay owingeh|san ildefonso|san felipe|santo domingo|kewa|"
    r"tribal college|urban indian|indian health)\b", re.I)


def step_earmarks_stage():
    """Normalise the retrieved tables into one staging file.

    Every staged row carries the URL it came from and a VERBATIM quote
    assembled from the source's own cells. A row without both is refused
    downstream - the prime directive, enforced in code rather than by care.
    """
    log("=== 99 earmarks-stage ===")
    import openpyxl
    staged = []
    man = {r["file"]: r for r in read_csv(EARMARK_RAW / "_SOURCE_MANIFEST.csv")}

    # ---- House ---------------------------------------------------------
    for fy, url in HOUSE_REQUEST_FILES:
        p = EARMARK_RAW / f"house_fy{fy}_cpf_requests.xlsx"
        if not p.exists():
            log(f"  house FY{fy}: file absent, skipped")
            continue
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        hdr = [str(c or "").strip() for c in next(rows)]
        low = [h.lower() for h in hdr]

        def col(*names, forbid=()):
            """Exact header match first, then substring - but never across a
            forbidden word.

            This guard is here because its absence produced a live
            misattribution. The FY2022 House table has NO `Recipient` column at
            all; a bare substring search for "recipient" matched
            `Recipient Address`, so every FY2022 recipient name became a postal
            address. Containment then resolved `2333 Biddle Ave, Wyandotte, MI`
            onto the Wyandotte Nation, `1654 West Onondaga Street, Syracuse NY`
            onto Onondaga, and `485 Gorman St, Shakopee, MN` onto the
            Shakopee Mdewakanton - 31 rows of pure place-name coincidence,
            carrying real dollars and a named member of Congress.
            """
            for nm in names:
                for i, h in enumerate(low):
                    if h == nm:
                        return i
            for nm in names:
                for i, h in enumerate(low):
                    if nm in h and not any(f in h for f in forbid):
                        return i
            return None

        i_last = col("member last name", "member lastname")
        i_first = col("member first name", "member firstname")
        i_memb = col("member")
        i_dist = col("district")
        i_party = col("party")
        i_sub = col("subcommittee")
        i_rec = col("recipient", forbid=("address",))
        i_proj = col("project purpose", "project purpose/description",
                     "program/language/project title", "request")
        i_addr = col("recipient address")
        i_amt = col("amount requested", "amount requested for fy22",
                    "amount")
        i_expl = col("explanation")
        n = 0
        for r in rows:
            if not r or all(c in (None, "") for c in r):
                continue
            g = lambda i: ("" if i is None or i >= len(r) or r[i] is None
                           else str(r[i]).strip())
            member = (f"{g(i_first)} {g(i_last)}".strip()
                      if (i_last is not None) else g(i_memb))
            recipient = g(i_rec)
            proj = g(i_proj)
            expl = g(i_expl)
            addr = g(i_addr)
            # FY22 has NO recipient column. Using the project title as a
            # recipient would invent a fact, so the field stays empty and the
            # row is screened on the title instead; nothing is manufactured.
            hay = " | ".join(x for x in (recipient, proj, expl, addr) if x)
            if not NATIVE_SCREEN.search(hay):
                continue
            dist = g(i_dist)
            st = ""
            m = re.match(r"([A-Z]{2})[-\s]?\d*$", dist.upper())
            if m:
                st = m.group(1)
            rst = _state_from_address(addr)
            quote = " | ".join(f"{hdr[i]}: {g(i)}" for i in range(len(hdr))
                               if i < len(r) and g(i))
            row = {
                # set below, from THIS row's own stated facts.
                "earmark_id": "",
                "fiscal_year": fy, "chamber": "House",
                "requesting_member": member, "member_state": st,
                "member_party": g(i_party), "recipient_name": recipient,
                "recipient_state": rst or st,
                "project_title": proj, "project_description": expl,
                "amount_requested": g(i_amt), "amount_enacted": "",
                "agency": "", "account": "", "subcommittee": g(i_sub),
                "source_url": url, "source_quote": quote[:1500],
                "fetched_date": man.get(p.name, {}).get("fetched_date", TODAY),
            }
            row["earmark_id"] = surrogate_id("EMK", row, EARMARK_KEY_COLUMNS)
            staged.append(row)
            n += 1
        log(f"  house FY{fy}: {n:,} candidate rows screened in")
        wb.close()

    # ---- Senate --------------------------------------------------------
    for fy, method in SENATE_CDS_METHODS:
        p = EARMARK_RAW / f"senate_fy{fy}_cds_requests.json"
        if not p.exists():
            log(f"  senate FY{fy}: file absent, skipped")
            continue
        j = json.loads(p.read_text(encoding="utf-8"))
        n = 0
        for row in j["rows"]:
            d = {k.lower(): ("" if v is None else str(v).strip())
                 for k, v in row.items()} if isinstance(row, dict) else {}
            if not d:
                continue
            hay = " | ".join(v for k, v in d.items()
                             if k in ("recipient", "project_purpose",
                                      "project_location"))
            if not NATIVE_SCREEN.search(hay):
                continue
            quote = " | ".join(f"{k}: {v}" for k, v in d.items() if v)
            row = {
                # set below, from THIS row's own stated facts.
                "earmark_id": "",
                "fiscal_year": fy, "chamber": "Senate",
                "requesting_member": d.get("member", ""),
                "member_state": d.get("member_state", ""),
                "member_party": "",
                "recipient_name": d.get("recipient", ""),
                "recipient_state": d.get("project_state", ""),
                "project_title": d.get("project_purpose", ""),
                "project_description": d.get("project_location", ""),
                "amount_requested": d.get("amount_requested", ""),
                "amount_enacted": "", "agency": "", "account": "",
                "subcommittee": d.get("subcommittee", ""),
                "source_url": f"{SENATE_CDS_ENDPOINT} (method={method})",
                "source_quote": quote[:1500],
                "fetched_date": j.get("fetched_date", TODAY),
            }
            row["earmark_id"] = surrogate_id("EMK", row, EARMARK_KEY_COLUMNS)
            staged.append(row)
            n += 1
        log(f"  senate FY{fy}: {n:,} candidate rows screened in")

    # ---- enacted tables ------------------------------------------------
    enacted = parse_enacted_tables()
    staged.extend(enacted)

    write_csv(EARMARK_RAW / "_staged_requests.csv", staged)
    log(f"  staged total {len(staged):,} -> "
        f"{EARMARK_RAW / '_staged_requests.csv'}")


STATE_NAMES = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR",
    "california": "CA", "colorado": "CO", "connecticut": "CT",
    "delaware": "DE", "florida": "FL", "georgia": "GA", "hawaii": "HI",
    "idaho": "ID", "illinois": "IL", "indiana": "IN", "iowa": "IA",
    "kansas": "KS", "kentucky": "KY", "louisiana": "LA", "maine": "ME",
    "maryland": "MD", "massachusetts": "MA", "michigan": "MI",
    "minnesota": "MN", "mississippi": "MS", "missouri": "MO",
    "montana": "MT", "nebraska": "NE", "nevada": "NV",
    "new hampshire": "NH", "new jersey": "NJ", "new mexico": "NM",
    "new york": "NY", "north carolina": "NC", "north dakota": "ND",
    "ohio": "OH", "oklahoma": "OK", "oregon": "OR", "pennsylvania": "PA",
    "rhode island": "RI", "south carolina": "SC", "south dakota": "SD",
    "tennessee": "TN", "texas": "TX", "utah": "UT", "vermont": "VT",
    "virginia": "VA", "washington": "WA", "west virginia": "WV",
    "wisconsin": "WI", "wyoming": "WY",
}


def _state_from_address(addr):
    if not addr:
        return ""
    parts = [p.strip() for p in addr.split(",")]
    for p in reversed(parts):
        if p.lower() in STATE_NAMES:
            return STATE_NAMES[p.lower()]
        if re.fullmatch(r"[A-Z]{2}", p):
            return p
    return ""


AMOUNT_RE = re.compile(r"^\$?[\d,]{3,}(?:\.\d{2})?$")


REV_MARKERS = re.compile(
    r"[-_]?(final|updated?|addendum|with[-_]technical.*|with[-_]managers.*|"
    r"\d{1,2}[\.\-]\d{1,2}[\.\-]\d{2,4}|_\d+|-\d+(?:\.\d+)?)", re.I)


def _latest_revision_per_table(paths):
    """One file per (fiscal year, subcommittee). Revisions double-count.

    FY24 Commerce-Justice-Science alone is published four times: a base table,
    `updated-10.31.23`, `updated-3.6.2024`, and `final`. Reading all four would
    report the same project up to four times with different amounts. The rule
    is deterministic - prefer `final`, then the latest embedded revision date,
    then the longest name - and the choice is logged so it is auditable.
    """
    groups = defaultdict(list)
    for p in paths:
        key = REV_MARKERS.sub("", p.stem.lower())
        key = re.sub(r"[-_]+", "-", key).strip("-")
        groups[key].append(p)
    chosen = []
    for key, ps in sorted(groups.items()):
        if len(ps) == 1:
            chosen.append(ps[0])
            continue

        def score(p):
            s = p.stem.lower()
            dates = re.findall(r"(\d{1,2})[\.\-](\d{1,2})[\.\-](\d{2,4})", s)
            d = max(((int(y) if int(y) > 100 else 2000 + int(y)),
                     int(mm), int(dd)) for dd, mm, y in dates) if dates else (0, 0, 0)
            return ("final" in s, d, len(s))
        best = max(ps, key=score)
        chosen.append(best)
        log(f"    revision pick [{key}]: {best.name}  "
            f"(dropped {len(ps)-1}: {', '.join(x.name for x in ps if x != best)})")
    return chosen


HEADER_KEYS = ("agency", "account", "state", "recipient", "project",
               "amount", "requestor", "requester", "origination", "location",
               "name")


def _header_columns(page):
    """Column x-boundaries taken from the table's own header words.

    The alternative - "the amount is the rightmost number on the line" - is
    what makes naive parses of these tables wrong. In the FY24 joint tables the
    columns run Agency | Account | Recipient | Project | Location | Amount |
    House Requestor | Senate | Origination, so "everything left of the number"
    is Agency+Account+Recipient+Project glued together, which no resolver can
    match and which would publish a recipient name that appears nowhere in the
    source. Reading the header gives real fields.
    """
    words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
    bands = defaultdict(list)
    for w in words:
        bands[round(w["top"] / 4.0)].append(w)
    for key in sorted(bands):
        ws = sorted(bands[key], key=lambda w: w["x0"])
        low = [w["text"].strip().lower().rstrip(":") for w in ws]
        if "amount" not in low:
            continue
        hits = sum(1 for t in low if t in HEADER_KEYS)
        if hits < 4:
            continue
        cols = []
        for w, t in zip(ws, low):
            if t in HEADER_KEYS:
                cols.append((w["x0"], t))
        cols.sort()
        return cols, words
    return None, words


def parse_enacted_tables():
    """Read the ENACTED joint explanatory statement tables.

    TWO THINGS THIS REFUSES TO CONFLATE
    -----------------------------------
    1. A joint table and a House-bill-stage table are different documents. The
       joint table carries a Senate requestor column and an `Origination`
       column; the House-stage table carries `House Amount` and `House
       Requestor(s)` only. A House-bill amount is neither what a member asked
       for nor what became law, so House-stage tables are READ, COUNTED and
       EXCLUDED - never written into `amount_enacted`. The test is on the
       document's own header text, not on its filename.

       This is why FY2022 and FY2025 contribute no enacted amounts. For FY2025
       that is the actual history: the full-year continuing resolution carried
       no community project funding, so no joint table exists to parse. For
       FY2022 the committee published only House-stage tables at these URLs.

    2. Revisions. `_latest_revision_per_table` picks one file per subcommittee
       and logs what it dropped.

    Rows whose amount or recipient cannot be read unambiguously are REFUSED and
    counted, never guessed at.
    """
    try:
        import pdfplumber
    except Exception:
        log("  pdfplumber unavailable; enacted tables skipped")
        return []
    out = []
    refused = 0
    skipped_stage = []
    no_header = []
    manifest = {r["file"]: r for r in read_csv(EARMARK_RAW / "_SOURCE_MANIFEST.csv")}
    files = _latest_revision_per_table(sorted(EARMARK_RAW.glob("enacted_fy*.pdf")))
    for p in files:
        m = re.match(r"enacted_fy(\d{4})_", p.name)
        fy = m.group(1) if m else ""
        url = manifest.get(p.name, {}).get("url", "")
        try:
            pdf = pdfplumber.open(p)
        except Exception as e:
            log(f"  {p.name}: cannot open ({e})")
            continue
        with pdf:
            head = (pdf.pages[0].extract_text() or "")[:1200]
            if "Origination" not in head:
                skipped_stage.append(p.name)
                continue
            n = 0
            for page in pdf.pages:
                cols, words = _header_columns(page)
                if not cols:
                    continue
                bounds = [c[0] for c in cols] + [1e9]
                names = [c[1] for c in cols]
                bands = defaultdict(list)
                for w in words:
                    bands[round(w["top"] / 4.0)].append(w)
                for key in sorted(bands):
                    ws = sorted(bands[key], key=lambda w: w["x0"])
                    line = " ".join(w["text"] for w in ws)
                    if not NATIVE_SCREEN.search(line):
                        continue
                    cell = defaultdict(list)
                    for w in ws:
                        for i in range(len(names)):
                            if bounds[i] - 3 <= w["x0"] < bounds[i + 1] - 3:
                                cell[names[i]].append(w["text"])
                                break
                    g = lambda k: " ".join(cell.get(k, [])).strip()
                    amt = ""
                    for t in g("amount").split():
                        if AMOUNT_RE.match(t) and len(
                                t.replace(",", "").replace("$", "")) >= 4:
                            amt = t
                    if not amt:
                        refused += 1
                        continue
                    recip = g("recipient") or g("name") or g("project")
                    if not recip:
                        refused += 1
                        continue
                    st = ""
                    for t in g("state").split():
                        if t.upper() in set(STATE_NAMES.values()):
                            st = t.upper()
                    n += 1
                    row = {
                        # WAS abs(hash(p.stem)) - a PROCESS HASH, so every
                        # explanatory-statement earmark changed id on every
                        # run. Now the same digest the other two branches use.
                        "earmark_id": "",
                        "fiscal_year": fy, "chamber": "Joint",
                        "requesting_member": (g("requestor") or
                                              g("requester")).strip(),
                        "member_state": "", "member_party": "",
                        "recipient_name": recip, "recipient_state": st,
                        "project_title": g("project"),
                        "project_description": g("location"),
                        "amount_requested": "", "amount_enacted": amt,
                        "agency": g("agency"), "account": g("account"),
                        "subcommittee": p.stem, "source_url": url,
                        "source_quote": line[:1500], "fetched_date": TODAY,
                    }
                    row["earmark_id"] = surrogate_id("EMK", row,
                                                     EARMARK_KEY_COLUMNS)
                    out.append(row)
            if n == 0:
                no_header.append(p.name)
            log(f"  {p.name}: {n:,} enacted rows")
    log(f"  House-bill-stage tables read and EXCLUDED from enacted: "
        f"{len(skipped_stage)}")
    for nm in skipped_stage:
        log(f"    excluded (no Origination column): {nm}")
    log(f"  joint tables yielding no parsable rows: {len(no_header)}")
    log(f"  rows refused for no unambiguous amount or recipient: {refused:,}")
    return out


def step_earmarks():
    """Build earmarks.csv from retrieved disclosure tables.

    Reads whatever the retrieval leg put in data/raw/external/earmarks/ as
    normalised staging rows (`_staged_requests.csv`). Every staged row must
    carry source_url and a verbatim source_quote or it is refused - the prime
    directive is not negotiable and a row without evidence is fabrication.
    """
    log("=== 99 earmarks ===")
    staged = read_csv(EARMARK_RAW / "_staged_requests.csv")
    if not staged:
        log("  no staged rows at data/raw/external/earmarks/_staged_requests.csv")
        log("  writing header-only earmarks.csv so the schema is fixed and the")
        log("  absence is explicit rather than implied")
        write_csv(CLEAN / "earmarks.csv", [], EARMARK_FIELDS)
        write_csv(REVIEW / f"earmark_unresolved_{TODAY}.csv", [],
                  ["earmark_id", "recipient_name", "state", "reason",
                   "source_url", "built_date"])
        return

    m = load_m33()
    spine = read_csv(SPINE / "cedar_entity_spine.csv")

    # ---- REQUEST -> ENACTED, joined only where it is unambiguous -----------
    #
    # No committee source links a request to its enacted outcome by any
    # identifier. The join has to be made on (fiscal year, recipient), and that
    # is only safe when exactly ONE request and exactly ONE enacted row share
    # the key. A tribe with two requests in one year and one enacted line is
    # ambiguous: assigning the enacted amount to either request would invent
    # which project got funded. Those stay as separate rows and are counted.
    #
    # A joined enacted row is REMOVED from the standalone set, so no dollar is
    # ever represented twice in the file.
    reqs = defaultdict(list)
    enas = defaultdict(list)
    for s in staged:
        k = (s.get("fiscal_year", ""), m.norm(s.get("recipient_name") or ""))
        if not k[1]:
            continue
        (enas if s.get("chamber") == "Joint" else reqs)[k].append(s)
    joined, ambiguous = {}, 0
    for k, es in enas.items():
        rs = reqs.get(k) or []
        if len(es) == 1 and len(rs) == 1:
            # lint-ok: class7 - `id()` here is PYTHON OBJECT IDENTITY for an in-memory object,
            # used as a key in a dict/set that lives and dies inside this one function. It is
            # never written to a file, nothing joins on it, and it is not a primary key.
            # 293_lint_bug_classes.py waives the same shape in its own source. Triaged LOW by
            # 326_triage_class7_key_risk.py on 2026-08-26: no clean or spine table carries it.
            joined[id(rs[0])] = es[0]
            # lint-ok: class7 - `id()` here is PYTHON OBJECT IDENTITY for an in-memory object,
            # used as a key in a dict/set that lives and dies inside this one function. It is
            # never written to a file, nothing joins on it, and it is not a primary key.
            # 293_lint_bug_classes.py waives the same shape in its own source. Triaged LOW by
            # 326_triage_class7_key_risk.py on 2026-08-26: no clean or spine table carries it.
            joined[id(es[0])] = None          # consumed; drop the standalone
        elif es and rs:
            ambiguous += 1
    log(f"  request<->enacted: {sum(1 for v in joined.values() if v):,} unique "
        f"joins; {ambiguous:,} recipient-years left unjoined as ambiguous")

    rows, unresolved = [], []
    refused = 0
    for i, s in enumerate(staged, 1):
        # lint-ok: class7 - `id()` here is PYTHON OBJECT IDENTITY for an in-memory object,
        # used as a key in a dict/set that lives and dies inside this one function. It is
        # never written to a file, nothing joins on it, and it is not a primary key.
        # 293_lint_bug_classes.py waives the same shape in its own source. Triaged LOW by
        # 326_triage_class7_key_risk.py on 2026-08-26: no clean or spine table carries it.
        if id(s) in joined and joined[id(s)] is None:
            continue                          # enacted row folded into its request
        url = (s.get("source_url") or "").strip()
        quote = (s.get("source_quote") or "").strip()
        if not url or not quote:
            refused += 1
            continue

        rec_name = (s.get("recipient_name") or "").strip()
        rec_state = (s.get("recipient_state") or s.get("member_state") or "").strip()
        eid, canon, how, tier, conf = resolve_recipient(m, spine, rec_name,
                                                        rec_state)
        # The FY2022 House table has no recipient column - the committee never
        # published one. Rather than manufacture a recipient from the address
        # (which is what the column-matching bug did), fall back to the PROJECT
        # TITLE, which often names the tribe outright, and cap the result at
        # tier B with the basis stated on the row.
        #
        # CONTAINMENT IS NOT ALLOWED ON A PROJECT TITLE. A title is a sentence,
        # not a name, and the entity-inside-record direction that containment
        # permits then fires on anything: `The NATIVE Project` matched
        # "Statewide Body Worn Camera Project", `Native Health` matched "UH
        # Rural Health Research Center", `Pueblo of Santa Clara` matched "Boys
        # and Girls Club of Santa Clara Valley". Only an exact, core-equal or
        # alias match is accepted here, which loses real rows and keeps the
        # file true. The losses go to review.
        if not eid and not rec_name:
            title = (s.get("project_title") or "").strip()
            e2, c2, h2, t2, cf2 = resolve_recipient(m, spine, title, rec_state)
            if e2 and h2 in ("exact", "core", "alias"):
                eid, canon, how, tier = e2, c2, h2, "B"
                conf = (cf2 + "; source table has NO recipient column - "
                        f"resolved on the project title by {h2} match only "
                        "(containment refused on free-text titles)")
            elif e2:
                how = ("project_title_containment_refused:free_text_title")
        req = _clean_money(s.get("amount_requested"))
        ena = _clean_money(s.get("amount_enacted"))

        # lint-ok: class7 - `id()` here is PYTHON OBJECT IDENTITY for an in-memory object,
        # used as a key in a dict/set that lives and dies inside this one function. It is
        # never written to a file, nothing joins on it, and it is not a primary key.
        # 293_lint_bug_classes.py waives the same shape in its own source. Triaged LOW by
        # 326_triage_class7_key_risk.py on 2026-08-26: no clean or spine table carries it.
        e = joined.get(id(s))
        if e:
            ena = _clean_money(e.get("amount_enacted"))
            quote = (quote + "  ||  ENACTED TABLE: "
                     + (e.get("source_quote") or ""))[:3000]
            url = url + "  ||  " + (e.get("source_url") or "")
            conf = (conf + "; enacted amount joined on unique "
                    "(fiscal year, recipient) match")
        # is_enacted = 1 means an enacted amount is ESTABLISHED for this row.
        # 0 means no enacted row was matched - which covers both "the request
        # was not funded" and "we hold no enacted table for that fiscal year".
        # FY2022 and FY2025 fall in the second case and the confidence column
        # says so, because reading a 0 as a rejection would be a fabricated
        # outcome.
        is_en = "1" if (ena not in ("", None)
                        and _numf(ena) not in (None, 0.0)) else "0"
        if is_en == "0" and (s.get("fiscal_year") in ("2022", "2025", "2027")):
            conf += ("; no enacted joint table held for FY"
                     + s.get("fiscal_year", "")
                     + " - is_enacted=0 means UNESTABLISHED, not rejected")

        eid_str = eid or ""
        row = {
            "earmark_id": s.get("earmark_id") or f"EMK-{i:06d}",
            "fiscal_year": (s.get("fiscal_year") or "").strip(),
            "chamber": (s.get("chamber") or "").strip(),
            "requesting_member": (s.get("requesting_member") or "").strip(),
            "member_state": (s.get("member_state") or "").strip(),
            "member_party": (s.get("member_party") or "").strip(),
            "recipient_name": rec_name,
            "entity_id": eid_str,
            "project_title": (s.get("project_title") or "").strip(),
            "project_description": (s.get("project_description") or "").strip(),
            "amount_requested": req,
            "amount_enacted": ena,
            "agency": (s.get("agency") or "").strip(),
            "account": (s.get("account") or "").strip(),
            "subcommittee": (s.get("subcommittee") or "").strip(),
            "is_enacted": is_en,
            "source_url": url,
            "source_quote": quote,
            "fetched_date": (s.get("fetched_date") or TODAY),
            "tier": tier,
            "confidence": conf,
            "built_date": TODAY,
        }
        # PRECISION OVER RECALL. `earmarks.csv` publishes rows that resolved to
        # a spine entity; a candidate that did not is NOT a Native earmark we
        # can assert, so it goes to review for a ruling instead of into the
        # dataset with a blank entity_id. The screen is a wide net - it catches
        # Indian River County, Pueblo County and the Naval Surface Warfare
        # Center Indian Head Division on purpose - and publishing its misses
        # would be exactly the false attribution this project forbids.
        if eid:
            rows.append(row)
        else:
            unresolved.append({
                "earmark_id": row["earmark_id"],
                "fiscal_year": row["fiscal_year"],
                "chamber": row["chamber"],
                "requesting_member": row["requesting_member"],
                "recipient_name": rec_name,
                "project_title": row["project_title"],
                "state": rec_state, "reason": how,
                "amount_requested": req, "amount_enacted": ena,
                "source_url": url, "source_quote": quote[:600],
                "built_date": TODAY})

    write_csv(CLEAN / "earmarks.csv", rows, EARMARK_FIELDS)
    write_csv(REVIEW / f"earmark_unresolved_{TODAY}.csv", unresolved,
              ["earmark_id", "fiscal_year", "chamber", "requesting_member",
               "recipient_name", "project_title", "state", "reason",
               "amount_requested", "amount_enacted", "source_url",
               "source_quote", "built_date"])
    log(f"  staged {len(staged):,}  refused for missing evidence {refused:,}")
    log(f"  wrote earmarks.csv {len(rows):,} resolved rows; "
        f"{len(unresolved):,} candidates to review")


_RESOLVE_CACHE = {}


def resolve_recipient(m, spine, name, state):
    """resolve_entity, then the guards AGENTS.md demands on top of it.

    Containment matched whenever one token set contained the other, in BOTH
    directions, and cost real money six independent ways. So:
      - the RECORD's name must be at least as specific as the entity's
        (entity tokens subset of record tokens), never the reverse;
      - where both carry a state they must agree;
      - a name-only match, however clean, is Tier B.
    """
    if not name:
        return None, None, "blank_name", "C", "no_name"
    ck = (name, state)
    if ck in _RESOLVE_CACHE:
        return _RESOLVE_CACHE[ck]
    out = _resolve_recipient_uncached(m, spine, name, state)
    _RESOLVE_CACHE[ck] = out
    return out


ADDRESS_RE = re.compile(r"^\s*(?:P\.?\s?O\.?\s+Box|\d+[A-Za-z]?\s+\S)"
                        r"|\b\d{5}(?:-\d{4})?\b")

# A record naming a COUNTY is never a tribe. Pueblo County (CO), Indian River
# County (FL), Taos County (NM) and Seminole County (FL) all sit on tribal
# words and all matched before this rule existed.
COUNTY_RE = re.compile(r"\bcount(?:y|ies)\b", re.I)

# Words that make a record self-identify as a tribal government or Native
# organisation. A one-token entity core may only match inside a longer record
# when one of these is present.
TRIBAL_STATUS_RE = re.compile(
    r"\b(tribe|tribes|tribal|nation|nations|pueblo|band|bands|rancheria|"
    r"native|indian|reservation|nsn|anvsa)\b", re.I)

# Entity cores made only of these are not distinctive enough to key on. This
# list exists because `The NATIVE Project` (a Spokane urban Indian
# organisation) has the core {project} once structural words are stripped, and
# containment then matched it to every appropriations request whose title ends
# in "Project" - 130-odd rows of pure noise, each carrying a dollar amount and
# a named member of Congress.
# Head nouns that make a record a DIFFERENT KIND OF THING from the entity.
# "Santa Ana College" is in Santa Ana, California; the spine's "Pueblo of Santa
# Ana" is in New Mexico - and the FY2023 House table publishes no district and
# no address, so the state guard has nothing to fire on. The word `college` is
# the only signal left that these are not the same body. Tribal colleges,
# clinics and airports do exist, so the rule only bites when the record carries
# no tribal status word at all.
INSTITUTION_HEAD = frozenset({
    "college", "university", "school", "schools", "academy", "hospital",
    "clinic", "airport", "museum", "library", "church", "camp", "park",
    "cooperative", "coop", "bank", "chamber", "seminary", "conservancy",
    "zoo", "aquarium", "symphony", "theatre", "theater", "ymca", "ywca",
    "police", "sheriff", "fire", "cemetery", "port", "railroad", "utilities",
    # civil governments that share a place name with a tribe
    "city", "town", "borough", "municipality", "municipal", "valley",
    "transportation", "transit", "metropolitan",
    # SUBORDINATE AND PROGRAM ENTITIES. AGENTS.md: every one of 148 tribal
    # housing entities resolved onto its own tribe, and the spine holds no
    # TDHE, so a "successful" match was guaranteed to be wrong. The same
    # applies to a tribe's development corporation or agricultural enterprise -
    # a different legal person receiving a different dollar. These go to review
    # so the relationship can be recorded without keying the money to the
    # tribe.
    "housing", "authority", "corporation", "enterprise", "enterprises",
    "ventures", "system", "supply", "department", "development",
})

# Recipient strings that are not names at all - wrapped cells from the PDF
# tables. "of Warm Springs for", "of Santa Clara Valley, San" and
# "of Hooper Bay Hooper Bay Beach" are the tail ends of a cell that ran over
# two lines. A leading preposition or a trailing comma is the tell.
FRAGMENT_RE = re.compile(r"^\s*(?:of|for|and|to|in|at|by|with|from|on)\b"
                         r"|,\s*$", re.I)

GENERIC_CORE = frozenset({
    "project", "projects", "health", "center", "centre", "services",
    "service", "community", "council", "association", "foundation",
    "corporation", "institute", "program", "programs", "authority",
    "development", "housing", "school", "college", "clinic", "hospital",
    "board", "agency", "office", "district", "society", "network", "fund",
})


def _resolve_recipient_uncached(m, spine, name, state):
    # A POSTAL ADDRESS IS NOT A NAME. Addresses carry place names, place names
    # are tribe names, and containment cannot tell the difference: "2333 Biddle
    # Ave, Wyandotte, MI" contains every token of "Wyandotte". Refuse before
    # the resolver is even asked.
    if ADDRESS_RE.search(name):
        return None, None, "record_is_a_postal_address_not_a_name", "C", \
            "refused_address_shaped"
    if FRAGMENT_RE.search(name):
        return None, None, "record_is_a_wrapped_table_cell_not_a_name", "C", \
            "refused_fragment"

    tid, canon, how = m.resolve_entity(name, spine)
    if not tid:
        return None, None, how, "C", "unresolved"

    ent = next((r for r in spine if r["tribe_id"] == tid), None)
    ent_state = ((ent or {}).get("state") or "").strip().upper()
    rec_state = (state or "").strip().upper()
    etok = set(m.core((ent or {}).get("canonical_name", "")) or set())
    ntok = set(m.core(name) or set())

    if how == "containment":
        # The record must be at least as specific as the entity, never the
        # reverse - AGENTS.md, the containment defect, six failures in one day.
        if not (etok and etok <= ntok):
            return None, None, ("containment_record_less_specific_than_entity"
                                f":{canon}"), "C", "refused_specificity"
        if COUNTY_RE.search(name):
            return None, None, "record_names_a_county_not_a_tribe", "C", \
                "refused_county"
        # The overlap must not consist entirely of trap tokens. "Creek",
        # "Indian", "United", "San" and the rest each cost a real
        # misattribution; a match resting only on those is a place-name
        # coincidence, not an entity.
        try:
            traps = load_domain().NAME_TRAPS
        except Exception:
            traps = frozenset()
        distinctive = [t for t in etok
                       if t not in traps and t not in GENERIC_CORE]
        if not distinctive:
            return None, None, (f"containment_on_generic_or_trap_tokens_only:"
                                f"{sorted(etok)}"), "C", "refused_trap_tokens"
        # A ONE-TOKEN entity core inside a longer record is only a match when
        # the record says it is tribal. "Nooksack Indian Tribe" qualifies;
        # "Camp Navajo" (an Arizona National Guard installation), "Taos County"
        # and "Seminole County Sheriff's Office" do not.
        if len(etok) < 2 and not TRIBAL_STATUS_RE.search(name):
            return None, None, ("single_token_entity_core_in_longer_record_"
                                f"with_no_tribal_status_word:{sorted(etok)}"), \
                "C", "refused_single_token"
        # And a one-token core must not pick up EXTRA identifying words. The
        # spine's "Arctic Village" (an Alaska Native village government) has
        # the core {arctic}, which sits inside "Arctic Slope Native Association
        # Ltd." - a regional health non-profit, a different legal person in a
        # different place. `slope` is the word that says so, and a guard that
        # ignores it books one organisation's money to another.
        extra = ntok - etok
        inst = extra & INSTITUTION_HEAD
        if inst and not TRIBAL_STATUS_RE.search(name):
            return None, None, (f"record_is_a_different_kind_of_institution:"
                                f"{sorted(inst)}"), "C", "refused_institution"
        if len(etok) < 2 and len(ntok - etok) > 1:
            return None, None, ("single_token_entity_core_but_record_carries_"
                                f"other_identifying_words:{sorted(ntok - etok)}"), \
                "C", "refused_extra_identifiers"

    if ent_state and rec_state and len(rec_state) == 2 and ent_state != rec_state:
        return None, None, (f"state_disagreement:entity={ent_state}"
                            f",record={rec_state}"), "C", "refused_state"

    if how == "exact" and ent_state and rec_state and ent_state == rec_state:
        return tid, canon, how, "A", "exact_name_and_state_agree"
    # Name-only is Tier B. Six containment failures in one day is the reason.
    return tid, canon, how, "B", f"{how}_name_only"


def _clean_money(v):
    v = (v or "").strip()
    if not v:
        return ""
    v = v.replace("$", "").replace(",", "").strip()
    if v.startswith("(") and v.endswith(")"):
        v = "-" + v[1:-1]
    f = _numf(v)
    return "" if f is None else (f"{f:.2f}" if f % 1 else f"{int(f)}")


# ---------------------------------------------------------------------------
# STEP: crosscheck  - Schedule C vs LDA
# ---------------------------------------------------------------------------

def step_crosscheck():
    """An org reporting lobbying on its 990 but absent from LDA is a discovery.

    Two possible explanations and the file must not choose between them:
      (a) the org is under the LDA registration threshold (the quarterly
          $14,000-ish expense floor, and the 20%-of-time test), so no filing is
          required and none is missing;
      (b) it is over the threshold and never registered - a registration gap.

    Schedule C and LDA also count DIFFERENT THINGS. Schedule C lobbying is the
    IRS definition and includes state and local legislative activity; LDA
    covers federal contacts only. An org whose lobbying is entirely at a state
    capitol is correctly on the 990 and correctly absent from LDA. So the file
    records the discrepancy and its size, and refuses to label it.
    """
    log("=== 99 crosscheck ===")
    fin = read_csv(CLEAN / "np_financials.csv")
    if not fin or "schedc_lobbying_usd" not in fin[0]:
        log("  np_financials has no Schedule C columns yet; run --steps schedc")
        return

    lda = read_csv(CLEAN / "native_entity_lobbying_disclosures.csv")
    m = load_m33()

    lda_names = set()
    lda_by_entity = defaultdict(list)
    for r in lda:
        for c in ("client_name", "registrant_name", "canonical_name"):
            v = (r.get(c) or "").strip()
            if v:
                lda_names.add(m.norm(v))
        eid = (r.get("entity_id") or "").strip()
        if eid:
            lda_by_entity[eid].append(r)

    orgs = {r["EIN"].strip().replace("-", "").zfill(9): r
            for r in read_csv(CLEAN / "np_orgs.csv")
            if (r.get("EIN") or "").strip()}

    # TWO reported lobbying figures, both from the Form 990 and both valid,
    # kept apart because they measure different things:
    #   Schedule C          the organisation's OWN lobbying expenditure
    #   Part IX line 11d    fees paid to OUTSIDE lobbyists
    # An org can report one and not the other. Either is enough to make its
    # absence from LDA a question worth asking, so both are swept - and the
    # row records which one produced the figure so nobody adds them together.
    per_ein = {}
    for r in fin:
        sc = _numf(r.get("schedc_lobbying_usd"))
        p9 = _numf(r.get("form990_part9_lobbying_fees"))
        amt = max((v for v in (sc, p9) if v is not None), default=None)
        if amt is None or amt <= 0:
            continue
        r = dict(r)
        r["_signal"] = ("schedule_c" if (sc or 0) > 0 else "") + \
            ("+" if (sc or 0) > 0 and (p9 or 0) > 0 else "") + \
            ("form990_part9_lobbying_fees" if (p9 or 0) > 0 else "")
        r["_amt"] = f"{amt:.0f}"
        ein = (r.get("ein") or "").strip().zfill(9)
        cur = per_ein.get(ein)
        if cur is None or (r.get("tax_year") or "") > (cur.get("tax_year") or ""):
            per_ein[ein] = r

    gaps = []
    ruled_out = 0
    for ein, r in sorted(per_ein.items()):
        name = (r.get("org_name") or "").strip()
        n = m.norm(name)
        in_lda = n in lda_names
        o = orgs.get(ein, {})
        eid = (o.get("tribe_id") or "").strip()
        if not in_lda and eid and eid in lda_by_entity:
            in_lda = True
        if in_lda:
            continue
        # AN ORGANISATION RULED OUT OF THE NATIVE UNIVERSE IS NOT A DISCOVERY.
        # np_orgs is a candidate funnel, not a roster: the first pass of this
        # file surfaced Yavapai Community Hospital Association, Pawnee Valley
        # Community Hospital and Wichita Downtown Development Corporation as
        # "Native organisations lobbying without registering". All three are
        # already ruled `place_name_coincidence` / tier X. Publishing them
        # would be exactly the false attribution the project forbids.
        if ((o.get("confidence_tier") or "").strip().upper() == "X"
                or (o.get("excluded_by_prior_ruling") or "") == "1"):
            ruled_out += 1
            continue
        amt = _numf(r.get("_amt")) or 0.0
        # The LDA registration floor is a quarterly expense test; annual 990
        # lobbying below roughly one quarter's floor cannot imply a gap.
        band = ("below_lda_registration_threshold_plausible" if amt < 14000
                else "above_lda_quarterly_threshold_registration_gap_candidate")
        gaps.append({
            "ein": ein,
            "org_name": name,
            "state": (r.get("state") or o.get("state") or "").strip(),
            "entity_id": eid,
            "np_org_confidence_tier": (o.get("confidence_tier") or "").strip(),
            "np_org_classification_ruling": (o.get("classification_ruling")
                                             or "").strip(),
            "placename_risk_flag": (o.get("placename_risk_flag") or "").strip(),
            "native_universe_status": (
                "SETTLED - ruled Native"
                if (o.get("classification_ruling") or "").startswith(
                    ("native_", "tribally_"))
                else "NOT SETTLED - np_orgs classification_ruling is "
                     f"'{(o.get('classification_ruling') or 'missing')}'. This "
                     "row is a question for review, not a finding, and must "
                     "not be reported as a Native organisation until ruled."),
            "tax_year": (r.get("tax_year") or "").strip(),
            "filing_regime": (r.get("filing_regime") or "").strip(),
            "reported_lobbying_usd": r.get("_amt", ""),
            "reported_lobbying_signal": r.get("_signal", ""),
            "form990_part9_lobbying_fees": r.get("form990_part9_lobbying_fees", ""),
            "schedc_lobbying_usd": r.get("schedc_lobbying_usd", ""),
            "schedc_lobbying_basis": r.get("schedc_lobbying_basis", ""),
            "schedc_501h_election": r.get("schedc_501h_election", ""),
            "schedc_direct_lobbying": r.get("schedc_direct_lobbying", ""),
            "schedc_grassroots_lobbying": r.get("schedc_grassroots_lobbying", ""),
            "in_lda_filings": "0",
            "interpretation_band": band,
            "caveat": ("Schedule C uses the IRS definition and includes STATE "
                       "and LOCAL legislative activity; LDA covers FEDERAL "
                       "contacts only. Absence from LDA is not evidence of "
                       "non-registration where the lobbying was not federal. "
                       "Form 990 Part IX line 11d counts fees paid to OUTSIDE "
                       "lobbyists and is a different quantity from Schedule C; "
                       "the two are never added."),
            "schedc_source_url": r.get("schedc_source_url", ""),
            "lda_source": ("data/clean/native_entity_lobbying_disclosures.csv "
                           f"({len(lda):,} filings, searched on normalised "
                           "client/registrant name and entity_id)"),
            "built_date": TODAY,
        })

    out = REVIEW / f"schedc_lda_gaps_{TODAY}.csv"
    write_csv(out, gaps)
    log(f"  orgs reporting >0 lobbying on a 990:   {len(per_ein):,}")
    log(f"    via Schedule C: "
        f"{sum(1 for r in per_ein.values() if 'schedule_c' in r['_signal']):,}"
        f"   via Part IX line 11d: "
        f"{sum(1 for r in per_ein.values() if 'part9' in r['_signal']):,}")
    log(f"  ruled OUT of the Native universe, dropped: {ruled_out:,}")
    log(f"  of which absent from LDA:              {len(gaps):,}")
    c = Counter(g["interpretation_band"] for g in gaps)
    for k, v in c.most_common():
        log(f"    {k}: {v:,}")
    log(f"  wrote {out}")


# ---------------------------------------------------------------------------
# STEP: codebook  (VARIABLES ONLY)
# ---------------------------------------------------------------------------

CODEBOOK_ENTRIES = [
    # (dataset, variable, description)
    ("14_earmarks", "earmark_id", "Cedar row identifier for one disclosed Community Project Funding / Congressionally Directed Spending request."),
    ("14_earmarks", "fiscal_year", "Appropriations fiscal year the request was made for."),
    ("14_earmarks", "chamber", "House or Senate. The two chambers run separate disclosure regimes with different tables and different fields."),
    ("14_earmarks", "requesting_member", "Member of Congress named in the disclosure as the requester. Verbatim from the table."),
    ("14_earmarks", "member_state", "State the requesting member represents, as printed in the disclosure."),
    ("14_earmarks", "member_party", "Party of the requesting member where the disclosure prints it. Blank where it does not; never inferred."),
    ("14_earmarks", "recipient_name", "Recipient organisation named in the disclosure, verbatim. Not necessarily a Native entity."),
    ("14_earmarks", "entity_id", "Cedar spine entity the recipient resolved to, or blank. Blank means unresolved, never 'not Native'."),
    ("14_earmarks", "project_title", "Project name as printed in the disclosure."),
    ("14_earmarks", "project_description", "Project description as printed. Blank where the table carries none."),
    ("14_earmarks", "amount_requested", "Dollar amount the member REQUESTED. Never the enacted amount. Blank where the table discloses no amount."),
    ("14_earmarks", "amount_enacted", "Dollar amount ENACTED in law, from the appropriations report. Blank means not established, which is different from zero."),
    ("14_earmarks", "agency", "Federal agency that would administer the funding, as printed."),
    ("14_earmarks", "account", "Appropriations account, as printed."),
    ("14_earmarks", "subcommittee", "Appropriations subcommittee whose bill carries the request."),
    ("14_earmarks", "is_enacted", "1 where a non-zero enacted amount is established; 0 otherwise. A request that was not funded is kept as a record, not dropped."),
    ("14_earmarks", "source_url", "URL of the disclosure table the row was read from."),
    ("14_earmarks", "source_quote", "Verbatim text from that source supporting the row."),
    ("14_earmarks", "fetched_date", "Date the source was retrieved."),
    ("14_earmarks", "tier", "A/B/C. A requires exact name and state agreement with the spine entity; a name-only match is B; unresolved is C."),
    ("14_earmarks", "confidence", "Why the row carries the tier it carries, in words."),
    ("14_earmarks", "built_date", "Date this row was written."),

    ("06_nonprofit", "filing_regime", "BMF filing requirement translated to a regime: 990_or_990EZ, 990_N, 990_PF, not_required, unknown. Determines whether a Schedule C could exist at all."),
    ("06_nonprofit", "schedc_expected", "1 where the filing regime permits a Schedule C; 0 where it does not (990-N e-Postcard filers and organisations with no filing requirement). Any lobbying denominator must use this, not the row count."),
    ("06_nonprofit", "schedc_basis", "Where the Schedule C reading came from, or why there is none: irs_efile_xml_schedule_c, irs_efile_xml_no_schedule_c_filed, 990N_filer_no_schedule_exists, outside_efile_index_coverage_submission_years_2017_2026, no_efile_return_indexed_for_period, efile_return_indexed_not_retrieved."),
    ("06_nonprofit", "schedc_source_url", "IRS e-file XML URL the Schedule C figures were read from."),
    ("06_nonprofit", "schedc_object_id", "IRS e-file OBJECT_ID identifying the exact accepted return."),
    ("06_nonprofit", "schedc_present", "1 where the return includes a Schedule C, 0 where it does not. A 990 filer answering No to the lobbying trigger question files none; that is a reported fact, not a gap."),
    ("06_nonprofit", "schedc_501h_election", "1 where the organisation has made the section 501(h) election (Form 5768). Electing filers face a bright-line dollar cap; non-electing filers face the vague 'no substantial part' test. The two regimes are not comparable."),
    ("06_nonprofit", "schedc_total_lobbying", "Schedule C Part II-A total lobbying expenditure, 501(h) electing filers only."),
    ("06_nonprofit", "schedc_direct_lobbying", "Schedule C Part II-A direct lobbying (contact with legislators). Electing filers only; non-electing filers report no split."),
    ("06_nonprofit", "schedc_grassroots_lobbying", "Schedule C Part II-A grassroots lobbying (attempts to influence public opinion). Electing filers only."),
    ("06_nonprofit", "schedc_nonelecting_total", "Schedule C Part II-B total lobbying expenditure, non-electing filers. No direct/grassroots split exists in this regime, so those columns are correctly blank."),
    ("06_nonprofit", "schedc_lobbying_nontaxable", "Schedule C Part II-A lobbying nontaxable amount - the statutory cap for this filer, not an amount spent."),
    ("06_nonprofit", "schedc_grassroots_nontaxable", "Schedule C Part II-A grassroots nontaxable amount - a cap, not a spend."),
    ("06_nonprofit", "schedc_exempt_purpose_expend", "Schedule C Part II-A total exempt purpose expenditures, the base the 501(h) cap is computed from."),
    ("06_nonprofit", "schedc_political_expenditure", "Schedule C Part I political campaign activity expenditure as reported."),
    ("06_nonprofit", "schedc_527_amount", "Amount reported as transferred to or spent through a section 527 political organisation."),
    ("06_nonprofit", "schedc_dues_lobbying_political", "Schedule C Part III lobbying and political expenditure allocable to member dues, for 501(c)(4)/(5)/(6) filers."),
    ("06_nonprofit", "schedc_dues_received", "Schedule C Part III dues and assessments received."),
    ("06_nonprofit", "form990_lobbying_activities_ind", "Core Form 990 Part IV trigger: did the organisation report lobbying activities. This is the question that determines whether a Schedule C exists."),
    ("06_nonprofit", "form990_political_activity_ind", "Core Form 990 Part IV trigger: did the organisation report political campaign activities."),
    ("06_nonprofit", "form990pf_influence_legislation_ind", "Form 990-PF Part VII-B: did the private foundation spend to influence legislation. 990-PF filers report here, not on Schedule C."),
    ("06_nonprofit", "form990pf_legislative_political_ind", "Form 990-PF Part VII-B: did the foundation spend on legislative or political activity of any kind."),
    ("06_nonprofit", "form990_part9_lobbying_fees", "Form 990 Part IX line 11d: fees paid to OUTSIDE lobbyists. A different quantity from Schedule C, which counts the organisation's own lobbying expenditure. The two overlap without being the same and are never added."),
    ("06_nonprofit", "schedc_501h_basis", "How schedc_501h_election was determined. Schedule C carries no election element; the value is derived from which part of the schedule the filer completed, and this column says so rather than implying a checkbox was read."),
    ("06_nonprofit", "schedc_used_volunteers", "Schedule C Part II-B: non-electing filer reported lobbying through volunteers."),
    ("06_nonprofit", "schedc_used_paid_staff", "Schedule C Part II-B: lobbying through paid staff or management."),
    ("06_nonprofit", "schedc_used_media", "Schedule C Part II-B: lobbying through media advertisements."),
    ("06_nonprofit", "schedc_used_mailings", "Schedule C Part II-B: lobbying through mailings to members, legislators or the public."),
    ("06_nonprofit", "schedc_used_publications", "Schedule C Part II-B: lobbying through publications or broadcast statements."),
    ("06_nonprofit", "schedc_used_grants", "Schedule C Part II-B: lobbying through grants to other organisations."),
    ("06_nonprofit", "schedc_used_direct_contact", "Schedule C Part II-B: direct contact with legislators, their staffs, government officials or a legislative body."),
    ("06_nonprofit", "schedc_used_rallies", "Schedule C Part II-B: rallies, demonstrations, seminars, conventions, speeches or lectures."),
    ("06_nonprofit", "schedc_used_other", "Schedule C Part II-B: other lobbying means reported."),
    ("06_nonprofit", "schedc_lobbying_usd", "One consolidated lobbying figure per filing. Blank is never zero - read schedc_lobbying_basis."),
    ("06_nonprofit", "schedc_lobbying_basis", "Which reporting regime produced schedc_lobbying_usd: schedc_part2a_501h_electing_total, schedc_part2b_nonelecting_total, form990pf_part7b_influence_legislation, an explicit reported zero, no_schedule_c_filed_with_return, or 990N_no_financial_detail_filed."),
    ("06_nonprofit", "schedc_built_date", "Date the Schedule C columns were appended."),
]


def _var_type_units(var):
    """Type and units, in the vocabulary codebook_master already uses."""
    v = var.lower()
    # Indicators are tested FIRST. `is_enacted` ends in "_enacted" and would
    # otherwise be typed as dollars, which is precisely the kind of quiet
    # metadata error that makes a codebook worse than none.
    if (v.endswith(("_ind", "_election", "_present", "_expected"))
            or v.startswith(("is_", "schedc_used_"))):
        return "integer", "0/1"
    if v.endswith(("_usd", "_amt", "_expend", "_requested", "_enacted",
                   "_lobbying", "_received", "_expenditure", "_nontaxable",
                   "_amount", "_political", "_total", "_fees")):
        return "numeric", "USD, nominal"
    if v.endswith(("_date",)):
        return "text", "YYYY-MM-DD"
    if v.endswith(("_url",)):
        return "text", "URL"
    if v in ("fiscal_year",):
        return "integer", "fiscal year"
    if v.endswith(("_id", "object_id")):
        return "text", "code"
    if v in ("member_state",):
        return "text", "2-letter code"
    return "text", "text"


def step_codebook():
    """VARIABLES ONLY. The codebook is a variable dictionary; adding dataset-
    level prose here would duplicate docs/datasets/*.md and drift from it."""
    log("=== 99 codebook ===")
    p = CLEAN / "codebook_master.csv"
    rows = read_csv(p)
    if not rows:
        log("  codebook_master.csv missing; skipping")
        return
    fields = list(rows[0].keys())

    # Fill rates recomputed from the data, never asserted (standing rule 10).
    sources = {"14_earmarks": read_csv(CLEAN / "earmarks.csv"),
               "06_nonprofit": read_csv(CLEAN / "np_financials.csv")}

    mine = {(ds.lower(), var.lower()) for ds, var, _ in CODEBOOK_ENTRIES}
    # A variable this script documented in an earlier run and then stopped
    # producing must not linger. `form990pf_influence_legislation_amt` was
    # documented against an element name that does not exist in the IRS schema;
    # leaving the entry would describe a column nobody can find.
    mine |= {("06_nonprofit", "form990pf_influence_legislation_amt")}
    # Standing rule 10: a number in a doc that is not recomputed from the data
    # is a claim, not a fact. So rows this script owns are REWRITTEN on a
    # re-run rather than skipped - otherwise pct_filled would freeze at
    # whatever the first build happened to produce.
    before = len(rows)
    rows = [r for r in rows
            if ((r.get("dataset") or "").strip().lower(),
                (r.get("variable") or "").strip().lower()) not in mine]
    refreshed = before - len(rows)

    have = {((r.get("dataset") or "").strip().lower(),
             (r.get("variable") or "").strip().lower()) for r in rows}
    added = 0
    for ds, var, desc in CODEBOOK_ENTRIES:
        if (ds.lower(), var.lower()) in have:
            continue
        src = sources.get(ds) or []
        n = len(src)
        filled = sum(1 for r in src if (r.get(var) or "").strip()) if n else 0
        t, u = _var_type_units(var)
        new = {c: "" for c in fields}
        new.update({
            "dataset": ds, "variable": var, "type": t, "units": u,
            "pct_filled": f"{100.0*filled/n:.1f}" if n else "0.0",
            "n_rows": str(n), "published": "1", "access_tier": "public",
            "description": desc, "generated": TODAY,
        })
        rows.append(new)
        have.add((ds.lower(), var.lower()))
        added += 1
    if added or refreshed:
        bak = p.with_suffix(f".csv.bak_{TODAY}_pre99")
        if not bak.exists():
            bak.write_bytes(p.read_bytes())
        write_csv(p, rows, fields)
    log(f"  {added} variable entries written "
        f"({refreshed} replaced from a previous run); {len(rows):,} total")


# ---------------------------------------------------------------------------
# STEP: report
# ---------------------------------------------------------------------------

def step_report():
    log("=== 99 report ===")
    em = read_csv(CLEAN / "earmarks.csv")
    log(f"\n-- earmarks: {len(em):,} rows")
    if em:
        by = defaultdict(lambda: [0, 0.0, 0.0, 0])
        for r in em:
            k = (r["fiscal_year"], r["chamber"])
            by[k][0] += 1
            by[k][1] += _numf(r["amount_requested"]) or 0.0
            by[k][2] += _numf(r["amount_enacted"]) or 0.0
            by[k][3] += 1 if r["is_enacted"] == "1" else 0
        log(f"  {'FY':6s} {'chamber':8s} {'rows':>6s} {'requested':>16s} "
            f"{'enacted':>16s} {'n_enacted':>9s}")
        for k in sorted(by):
            n, rq, en, ne = by[k]
            log(f"  {k[0]:6s} {k[1]:8s} {n:6,d} {rq:16,.0f} {en:16,.0f} {ne:9,d}")
        ents = {r["entity_id"] for r in em if r["entity_id"]}
        log(f"  distinct spine entities reached: {len(ents):,}")
        log(f"  tier: {dict(Counter(r['tier'] for r in em))}")

    fin = read_csv(CLEAN / "np_financials.csv")
    if fin and "schedc_expected" in fin[0]:
        log(f"\n-- Schedule C on np_financials: {len(fin):,} rows")
        exp = [r for r in fin if r["schedc_expected"] == "1"]
        n990n = sum(1 for r in fin if r["filing_regime"] == "990_N")
        nreq = sum(1 for r in fin if r["filing_regime"] == "not_required")
        read = sum(1 for r in fin if (r.get("schedc_present") or "") in ("0", "1"))
        filed = sum(1 for r in fin if r.get("schedc_present") == "1")
        amt = sum(1 for r in fin if (r.get("schedc_lobbying_usd") or "").strip())
        p9 = sum(1 for r in fin
                 if (r.get("form990_part9_lobbying_fees") or "").strip())
        trig = sum(1 for r in fin
                   if (r.get("form990_lobbying_activities_ind") or "").strip()
                   or (r.get("form990_political_activity_ind") or "").strip())
        log("  DENOMINATORS, which is the whole point of the caveat columns:")
        log(f"    rows total ....................... {len(fin):,}")
        log(f"    990-N filers - no schedule EXISTS  {n990n:,}  (excluded, not zeroed)")
        log(f"    no filing requirement ............ {nreq:,}  (excluded, not zeroed)")
        log(f"    rows where a Schedule C COULD exist {len(exp):,}")
        log("  OBSERVATIONS:")
        log(f"    returns retrieved and read ....... {read:,} "
            f"({100*read/max(1,len(exp)):.1f}% of possible)")
        log(f"    of which a Schedule C was filed .. {filed:,}")
        log(f"    with a lobbying expenditure figure {amt:,}")
        log(f"    Part IX line 11d lobbying fees ... {p9:,}")
        log(f"    core-form lobbying/political trigger answered {trig:,}")
        log(f"  basis: {dict(Counter(r['schedc_basis'] for r in fin).most_common(10))}")


# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--steps", default="report")
    ap.add_argument("--max-fetch", type=int, default=None)
    ap.add_argument("--priority-only", action="store_true")
    ap.add_argument("--years", default=None)
    a = ap.parse_args()
    steps = [s.strip() for s in a.steps.split(",") if s.strip()]
    years = [int(x) for x in a.years.split(",")] if a.years else None
    for s in steps:
        if s == "probe":
            step_probe()
        elif s == "irs-index":
            step_irs_index(years)
        elif s == "irs-xml":
            step_irs_xml(a.max_fetch, a.priority_only)
        elif s == "irs-deflate64":
            step_irs_deflate64()
        elif s == "schedc":
            step_schedc()
        elif s == "earmarks-pull":
            step_earmarks_pull()
        elif s == "earmarks-stage":
            step_earmarks_stage()
        elif s == "earmarks":
            step_earmarks()
        elif s == "crosscheck":
            step_crosscheck()
        elif s == "codebook":
            step_codebook()
        elif s == "report":
            step_report()
        else:
            log(f"unknown step {s}")


if __name__ == "__main__":
    main()
