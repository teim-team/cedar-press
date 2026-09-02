"""SHARD-G: build the institutional slice and mine every source ALREADY ON DISK.

ZERO NETWORK. Reads only. Writes only under
  data/staging/institution_registry/
  data/staging/tribe_harvest/shard_g/

Slice = cedar_identity_register.csv rows whose entity_class is one of
  BIE School | Native Community Development Financial Institution
  Tribal College or University | Native Financial Institution
joined to cedar_entity_spine.csv on cedar_uid.

On-disk authoritative caches mined here (pulled 2026-08-06 by scripts 73/75):
  data/raw/external/bie_uio/bie_schools_featureserver.json   BIE school directory
  data/raw/external/tcu_cdfi/cdfifund_certified_list_*.xlsx  CDFI Fund cert list
  data/raw/external/tcu_cdfi/cicd_nafi_map_data_*.xlsx       CICD NAFI map (FDIC cert,
                                                             NCUA charter, RSSD)
  data/raw/external/tcu_cdfi/_aihec_parsed.json              AIHEC TCU roster

Outputs
  data/staging/institution_registry/_slice.csv                the 315-entity slice
  data/staging/institution_registry/_mine_*_cacheonly.csv     zero-network baseline
  data/staging/tribe_harvest/shard_g/_mine_state.json

It deliberately does NOT write registry_crosswalk.csv or institution_facts.csv:
code/shard_g_build_crosswalk.py owns those and supersedes everything here.
"""
from __future__ import annotations

import csv, json, re, sys, unicodedata
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
SPINE = ROOT / "data" / "spine" / "cedar_entity_spine.csv"
REG = ROOT / "data" / "spine" / "cedar_identity_register.csv"
EXT = ROOT / "data" / "raw" / "external"
OUTREG = ROOT / "data" / "staging" / "institution_registry"
OUTH = ROOT / "data" / "staging" / "tribe_harvest" / "shard_g"
OUTREG.mkdir(parents=True, exist_ok=True)
OUTH.mkdir(parents=True, exist_ok=True)
csv.field_size_limit(10_000_000)

CLASSES = {
    "BIE School",
    "Native Community Development Financial Institution",
    "Tribal College or University",
    "Native Financial Institution",
}
TODAY = date.today().isoformat()


def norm(s: str) -> str:
    """Aggressive name key for joining across registries."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = s.lower()
    s = re.sub(r"[‘’'`]", "", s)
    s = re.sub(r"&", " and ", s)
    s = re.sub(r"[^a-z0-9]+", " ", s)
    for w in (" incorporated", " inc", " llc", " ltd", " corp", " corporation",
              " co", " the"):
        s = s.replace(w + " ", " ")
    s = re.sub(r"\b(inc|llc|ltd|corp|co)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def rows(p, enc="utf-8-sig"):
    with open(p, encoding=enc, newline="") as f:
        return list(csv.DictReader(f))


# ---------------------------------------------------------------- slice
register = rows(REG)
spine = rows(SPINE)
spine_by_uid = {r["cedar_uid"]: r for r in spine if r.get("cedar_uid")}

slice_rows = []
for r in register:
    if r["entity_class"] not in CLASSES:
        continue
    s = spine_by_uid.get(r["cedar_uid"], {})
    slice_rows.append({
        "cedar_uid": r["cedar_uid"],
        "handle": r.get("handle", ""),
        "canonical_name": r["canonical_name"],
        "entity_class": r["entity_class"],
        "tribe_id": s.get("tribe_id", ""),
        "state": s.get("state", ""),
        "city": s.get("city", ""),
        "spine_entity_website": s.get("entity_website", ""),
        "spine_recon_note": s.get("reconciliation_note", ""),
        "spine_source_url": s.get("source_url", ""),
        "spine_entity_source_url": s.get("entity_source_url", ""),
        "spine_entity_source_quote": s.get("entity_source_quote", ""),
        "bie_operation_type": s.get("bie_operation_type", ""),
        "parent_entity_id": s.get("parent_entity_id", ""),
        "parent_entity_name": s.get("parent_entity_name", ""),
        "parent_native_entity": s.get("parent_native_entity", ""),
        "aliases": s.get("aliases", ""),
        "in_spine": "1" if s else "0",
    })
print(f"slice rows: {len(slice_rows)}", file=sys.stderr)
by_key = {}
for r in slice_rows:
    by_key.setdefault(norm(r["canonical_name"]), []).append(r)

# alias index too
alias_key = {}
for r in slice_rows:
    for a in (r["aliases"] or "").split("|"):
        k = norm(a)
        if k:
            alias_key.setdefault(k, []).append(r)


def match(name):
    k = norm(name)
    hit = by_key.get(k) or alias_key.get(k)
    if hit and len(hit) == 1:
        return hit[0], "exact_name"
    if hit:
        return hit[0], "ambiguous_name"
    return None, ""


# ---------------------------------------------------------------- crosswalk sink
XW = []  # long: cedar_uid, canonical_name, entity_class, id_system, id_value, ...


def add_id(row, id_system, id_value, field, source_url, source_quote, method):
    if id_value is None:
        return
    v = str(id_value).strip()
    if not v or v.lower() in ("none", "nan", "not applicable", "unavailable", ""):
        return
    XW.append({
        "cedar_uid": row["cedar_uid"], "tribe_id": row["tribe_id"],
        "canonical_name": row["canonical_name"],
        "entity_class": row["entity_class"],
        "id_system": id_system, "id_value": v, "id_field": field,
        "source_url": source_url, "source_quote": source_quote[:600],
        "match_method": method, "captured_date": TODAY,
    })


FACTS = []  # long attribute table


def add_fact(row, attr, value, source_url, source_quote, method):
    if value is None:
        return
    v = str(value).strip()
    if not v or v.lower() in ("none", "nan", "unavailable", "not applicable"):
        return
    FACTS.append({
        "cedar_uid": row["cedar_uid"], "tribe_id": row["tribe_id"],
        "canonical_name": row["canonical_name"],
        "entity_class": row["entity_class"],
        "attribute": attr, "value": v,
        "source_url": source_url, "source_quote": source_quote[:600],
        "match_method": method, "captured_date": TODAY,
    })


unmatched = []

# ---------------------------------------------------------------- BIE directory
BIE_JSON = EXT / "bie_uio" / "bie_schools_featureserver.json"
BIE_SRC = ("https://www.bie.edu/schools -> https://services1.arcgis.com/"
           "UxqqIfhng71wUT9x/arcgis/rest/services/BIE_Schools_Directory/"
           "FeatureServer/0 (cached 2026-08-06)")
n_bie = 0
if BIE_JSON.exists():
    feats = json.load(open(BIE_JSON, encoding="utf-8"))["features"]
    for ft in feats:
        a = ft["attributes"]
        row, meth = match(a.get("School_Name", ""))
        if not row:
            unmatched.append({"registry": "BIE_directory",
                              "name": a.get("School_Name", ""), "note": "no slice match"})
            continue
        n_bie += 1
        q = (f"BIE school directory row: School_Name={a.get('School_Name')}; "
             f"Operation_Type={a.get('Operation_Type')}; "
             f"Grades_Served={a.get('Grades_Served')}; "
             f"ERC={a.get('Education_Resource_Center')}")
        for attr, key in (("bie_operation_type_raw", "Operation_Type"),
                          ("bie_navajo_operation", "Navajo_Operation"),
                          ("grades_served", "Grades_Served"),
                          ("bie_education_resource_center", "Education_Resource_Center"),
                          ("street_address", "Street_Address"),
                          ("city", "City"), ("state_name", "State"),
                          ("zip", "Zip_Code"), ("telephone", "telephone"),
                          ("latitude", "Latitude"), ("longitude", "Longitude")):
            add_fact(row, attr, a.get(key), BIE_SRC, q, meth)
        if a.get("website"):
            add_fact(row, "website_registry", a["website"], BIE_SRC, q, meth)

# ---------------------------------------------------------------- CDFI Fund list
try:
    import openpyxl
except ImportError:
    openpyxl = None

CDFI_XLSX = EXT / "tcu_cdfi" / "cdfifund_certified_list_2026-08-06.xlsx"
CDFI_SRC = "https://www.cdfifund.gov/media/8018641/download?inline (cached 2026-08-06)"
n_cdfi = 0
cdfi_native_total = None
if openpyxl and CDFI_XLSX.exists():
    wb = openpyxl.load_workbook(CDFI_XLSX, read_only=True, data_only=True)
    ws = wb["List of Certified CDFIs"]
    rr = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr_i = next(i for i, r in enumerate(rr)
                 if r and "Organization Name" in [str(c) for c in r if c])
    hdr = [str(c).strip() if c else "" for c in rr[hdr_i]]
    for r in rr[:hdr_i]:
        for c in r:
            if c and "Native CDFIs as of" in str(c):
                cdfi_native_total = str(c).strip()
    for raw in rr[hdr_i + 1:]:
        d = dict(zip(hdr, raw))
        nm = d.get("Organization Name")
        if not nm:
            continue
        nm = str(nm).strip().rstrip("`")
        row, meth = match(nm)
        if not row:
            continue
        n_cdfi += 1
        q = (f"CDFI Fund List of Certified CDFIs row: {nm}; "
             f"Cert Control Num={d.get('Cert Control Num')}; "
             f"Financial Institution Type={d.get('Financial Institution Type')}; "
             f"Native CDFI (Y/N)={d.get('Native CDFI (Y/N)')}; "
             f"RSSD ID={d.get('RSSD ID')}")
        add_id(row, "CDFI_FUND_CERT", d.get("Cert Control Num"),
               "Cert Control Num", CDFI_SRC, q, meth)
        add_id(row, "FRB_RSSD", d.get("RSSD ID"), "RSSD ID", CDFI_SRC, q, meth)
        dc = d.get("Date Certified")
        if dc is not None and str(dc).strip():
            try:  # excel serial
                from datetime import datetime, timedelta
                dcv = (datetime(1899, 12, 30) + timedelta(days=int(dc))).date().isoformat()
            except Exception:
                dcv = str(dc)
            add_fact(row, "cdfi_date_certified", dcv, CDFI_SRC, q, meth)
        add_fact(row, "cdfi_certification_status",
                 "certified (on CDFI Fund certified list)", CDFI_SRC, q, meth)
        add_fact(row, "financial_institution_type",
                 d.get("Financial Institution Type"), CDFI_SRC, q, meth)
        add_fact(row, "native_cdfi_flag", d.get("Native CDFI (Y/N)"),
                 CDFI_SRC, q, meth)
        add_fact(row, "website_registry", d.get("Organization Website"),
                 CDFI_SRC, q, meth)
        for attr, key in (("city", "City"), ("state_code", "State"),
                          ("zip", "Zipcode"), ("street_address", "Address")):
            add_fact(row, attr, d.get(key), CDFI_SRC, q, meth)

# ---------------------------------------------------------------- CICD NAFI map
NAFI_XLSX = EXT / "tcu_cdfi" / "cicd_nafi_map_data_2026-08-06.xlsx"
NAFI_SRC = ("https://github.com/frb-mpls-cde/nafi-map/raw/refs/heads/main/data/"
            "nafi-map-data_current.xlsx (cached 2026-08-06; last_updated 2025-08-20)")
n_nafi = 0
if openpyxl and NAFI_XLSX.exists():
    wb = openpyxl.load_workbook(NAFI_XLSX, read_only=True, data_only=True)
    ws = wb["data"]
    rr = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = [str(c).strip() if c else "" for c in rr[0]]
    for raw in rr[1:]:
        d = dict(zip(hdr, raw))
        nm = (d.get("name") or "").strip()
        if not nm:
            continue
        row, meth = match(nm)
        if not row:
            unmatched.append({"registry": "CICD_NAFI_map", "name": nm,
                              "note": "no slice match"})
            continue
        n_nafi += 1
        q = (f"Fed Minneapolis CICD Native American Financial Institutions map, "
             f"row: name={nm}; type={d.get('type')}; regulator={d.get('regulator')}; "
             f"rssd_id={d.get('rssd_id')}; bank_cert={d.get('bank_cert')}; "
             f"cu_number={d.get('cu_number')}")
        add_id(row, "FDIC_CERT", d.get("bank_cert"), "bank_cert", NAFI_SRC, q, meth)
        add_id(row, "NCUA_CHARTER", d.get("cu_number"), "cu_number", NAFI_SRC, q, meth)
        add_id(row, "FRB_RSSD", d.get("rssd_id"), "rssd_id", NAFI_SRC, q, meth)
        for attr, key in (("financial_institution_type", "type"),
                          ("federal_regulator", "regulator"),
                          ("native_mdi_flag", "nmdi"),
                          ("native_cdfi_flag", "ncdfi"),
                          ("cu_low_income_designated", "cu_lowinc"),
                          ("website_registry", "website"),
                          ("year_established", "yearest"),
                          ("employees", "employees"),
                          ("credit_union_members", "cu_members"),
                          ("total_assets", "total_assets"),
                          ("total_loans", "total_loans"),
                          ("cdfi_award_2021", "award_2021"),
                          ("cdfi_award_2022", "award_2022"),
                          ("cdfi_award_2024", "award_2024"),
                          ("financials_source", "detail_source"),
                          ("financials_period", "data_period"),
                          ("city", "city"), ("state_code", "state")):
            add_fact(row, attr, d.get(key), NAFI_SRC, q, meth)

# ---------------------------------------------------------------- AIHEC roster
AIHEC = EXT / "tcu_cdfi" / "_aihec_parsed.json"
AIHEC_SRC = "https://www.aihec.org/tcu-roster-and-profiles/ (cached 2026-08-06)"
n_aihec = 0
if AIHEC.exists():
    for d in json.load(open(AIHEC, encoding="utf-8")):
        row, meth = match(d.get("name", ""))
        if not row:
            unmatched.append({"registry": "AIHEC_roster", "name": d.get("name", ""),
                              "note": "no slice match"})
            continue
        n_aihec += 1
        ev = " ".join(d.get("evidence") or [])[:600]
        q = f"AIHEC TCU Roster, section={d.get('section')}: {d.get('name')}. {ev}"
        add_fact(row, "aihec_membership", d.get("section"), AIHEC_SRC, q, meth)
        add_fact(row, "acronym", d.get("acronym"), AIHEC_SRC, q, meth)
        add_fact(row, "chartered_year", d.get("chartered_year"), AIHEC_SRC, q, meth)
        add_fact(row, "state_code", d.get("state"), AIHEC_SRC, q, meth)

# ---------------------------------------------------------------- spine notes mine
# reconciliation_note carries "website=..." and "cert=..." for TCU/CDFI/NFI rows,
# written by code/73_add_tcu_and_cdfi.py.
n_note_web = n_note_cert = 0
for row in slice_rows:
    note = row["spine_recon_note"] or ""
    src = row["spine_entity_source_url"] or row["spine_source_url"]
    m = re.search(r"website[= ](\S+)", note)
    if m:
        u = m.group(1).strip().rstrip(";,")
        if u.startswith("http"):
            add_fact(row, "website_spine_note", u, src or "spine reconciliation_note",
                     f"cedar_entity_spine.csv reconciliation_note: {note[:300]}",
                     "spine_field")
            n_note_web += 1
    m = re.search(r"cert=([0-9A-Za-z]+)", note)
    if m:
        add_id(row, "CDFI_FUND_CERT", m.group(1), "reconciliation_note cert=",
               src or "spine reconciliation_note",
               f"cedar_entity_spine.csv reconciliation_note: {note[:300]}",
               "spine_field")
        n_note_cert += 1
    if row["spine_entity_website"]:
        add_fact(row, "website_spine", row["spine_entity_website"],
                 row["spine_source_url"] or "cedar_entity_spine.csv entity_website",
                 f"cedar_entity_spine.csv entity_website for {row['canonical_name']}",
                 "spine_field")

# ---------------------------------------------------------------- write
def write(path, recs, fields):
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(recs)
    print(f"wrote {path.relative_to(ROOT)}  rows={len(recs)}", file=sys.stderr)


# THIS SCRIPT DOES NOT WRITE registry_crosswalk.csv OR institution_facts.csv.
# code/shard_g_build_crosswalk.py owns both, and it re-mines every source this
# script reads (BIE directory, CDFI Fund list, CICD NAFI map, AIHEC roster, spine
# notes) from the FRESH pulls plus these caches. If both wrote the same two file
# names, running this one second - as a handoff verification does - would revert
# the crosswalk to the cache-only version and print a smaller row count that
# reads as a completed run. That is the rebuild-reverts-the-enricher shape
# (AGENTS.md concurrency rule 5), so the collision is removed rather than
# ordered around. The preliminary tables are kept under _mine_* names purely as
# a zero-network baseline to diff the enriched build against.
write(OUTREG / "_slice.csv", slice_rows, list(slice_rows[0].keys()))
XW.sort(key=lambda r: (r["entity_class"], r["canonical_name"], r["id_system"]))
write(OUTREG / "_mine_crosswalk_cacheonly.csv", XW,
      ["cedar_uid", "tribe_id", "canonical_name", "entity_class", "id_system",
       "id_value", "id_field", "source_url", "source_quote", "match_method",
       "captured_date"])
FACTS.sort(key=lambda r: (r["entity_class"], r["canonical_name"], r["attribute"]))
write(OUTREG / "_mine_facts_cacheonly.csv", FACTS,
      ["cedar_uid", "tribe_id", "canonical_name", "entity_class", "attribute",
       "value", "source_url", "source_quote", "match_method", "captured_date"])
write(OUTREG / "_mine_rows_unmatched_cacheonly.csv", unmatched,
      ["registry", "name", "note"])

state = {
    "script": "code/shard_g_slice_and_mine.py", "run_date": TODAY,
    "network_requests": 0,
    "slice_entities": len(slice_rows),
    "slice_by_class": {c: sum(1 for r in slice_rows if r["entity_class"] == c)
                       for c in sorted(CLASSES)},
    "matched_bie_directory": n_bie,
    "matched_cdfi_fund_list": n_cdfi,
    "matched_cicd_nafi_map": n_nafi,
    "matched_aihec_roster": n_aihec,
    "spine_note_websites": n_note_web,
    "spine_note_cert_ids": n_note_cert,
    "cdfi_fund_native_total_line": cdfi_native_total,
    "crosswalk_ids": len(XW),
    "ids_by_system": {s: sum(1 for r in XW if r["id_system"] == s)
                      for s in sorted({r["id_system"] for r in XW})},
    "facts": len(FACTS),
    "registry_rows_unmatched": len(unmatched),
}
(OUTH / "_mine_state.json").write_text(json.dumps(state, indent=2), encoding="utf-8")
print(json.dumps(state, indent=2))
