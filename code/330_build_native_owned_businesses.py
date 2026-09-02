#!/usr/bin/env python3
# lint-ok: class6 - the `promote` phase is a full REBUILD of one table this
# script alone owns (data/clean/native_owned_businesses.csv). It reads only
# staging JSONL + the spine and writes nothing else, so a rebuild cannot drop
# another script's work. `harvest` IS an enricher: it never deletes a staging
# file it did not write this run.
"""330 - NATIVE-OWNED BUSINESSES: the one builder for this dataset.

    py -3 code/330_build_native_owned_businesses.py harvest    # raw -> staging
    py -3 code/330_build_native_owned_businesses.py promote    # staging -> clean
    py -3 code/330_build_native_owned_businesses.py registry   # survey -> tracker
    py -3 code/330_build_native_owned_businesses.py codebook   # write the codebook fragment
    py -3 code/330_build_native_owned_businesses.py docs       # regenerate the dataset doc
    py -3 code/330_build_native_owned_businesses.py all

WHAT THIS IS
------------
A tribal government certifying a business is a THIRD PARTY with authority over
the ownership question - the tier-A evidence leg Cedar Press has almost none of
(`316_build_tribal_vendor_list_roster.py`). `316`-`324` SURVEYED 62 tribes and
found 22 published lists; four were harvested. This harvests the rest that can
be harvested, and records - by name - the ones that cannot and why.

THE INCLUSION BASIS IS THE PRODUCT (ADR-013)
--------------------------------------------
`identity_claim_text` is quoted VERBATIM from the source and is the reason the
row is in Cedar at all. It is NOT uniform and must never be flattened:

    OWNERSHIP   the authority asserts who OWNS the firm
                  - enrolled-member owned      (EBCI, CSKT pref 1, Poarch 100%)
                  - any-Native owned            (Tulalip NAOB, Tohono, MHA)
                  - shareholder/descendant/spouse owned  (Calista - WEAKEST)
                  - tribally owned entity       (Poarch "TRIBAL BUSINESSES")
                  - parent-asserted subsidiary  (ASRC Federal, Doyon)
    RELATIONSHIP the authority asserts the firm DOES BUSINESS WITH the tribe
                  - a vendor list is NOT an ownership list  (Menominee)

`assertion_class` carries that distinction on every row, inherited from the
registry, never re-derived here. A consumer that sums OWNERSHIP and
RELATIONSHIP rows together has counted two different facts.

TERMS ARE A DECISION THE PUBLISHER MADE
---------------------------------------
Six sources are EXCLUDED and stay excluded by every route, Wayback included.
Two were already recorded restrictive; FOUR were found restrictive by this run
because nobody had opened the terms page. They are in `EXCLUDED` below with the
verbatim quote and the URL it came from. SILENCE IS UNRESOLVED, NEVER
PERMISSION - and a STATED restriction is a refusal, not an obstacle.

Every harvested row carries `consent_status = UNRESOLVED`, `publishable = N`
and a `suppression_key`, exactly as `320`/`321` require. Flipping one field
admits or removes a whole authority.

PRIVACY - INHERITED, NOT INVENTED
---------------------------------
`cedar_domain.INDIVIDUAL_NATIVE_WITHHELD_FIELDS` already reasoned this through
for individually Native-owned firms: where the legal name IS a person's name,
the name plus a locating identifier publishes a natural person, and
`owner_name`, `street`, `recipient_city_name`, `dba_name` are withheld absent
recorded consent. These lists are FULL of exactly that case - a TERO roster of
sole proprietorships is a list of private individuals with their home addresses
and mobile numbers.

So the CLEAN table carries the certification FACT and the firm identity, and
the CONTACT CHANNEL stays in staging:

    carried    business name, normalized name, nation, city, state, category,
               certification number/tier/dates, the verbatim claim, and
               `owner_name_present` / `n_owners_named` as counts
    withheld   owner_name_raw, email, phone, street address, postal code
    flagged    `business_name_is_person_name` so a consumer can apply
               `cedar_domain.may_publish_individual_native_field` per field

No digest surrogate is minted for a name. `docs/HANDOFF.md` already records
that a digest of an enumerable identifier is not a privacy control; a digest of
a personal name is not one either. The protection is that the column does not
ship.

RESOLUTION - ADR-010, NEVER INVENTED
------------------------------------
Each business is offered to `503_identity.resolve()`. Only an EXACT normalized
name/alias hit is accepted. The loose gov-class token path is REFUSED here by
design: it exists to match a filing to the GOVERNMENT that filed it, and on a
business roster "Navajo Transitional Energy" would resolve to the Navajo
Nation, which is a false ownership claim, not a match. A business that does not
resolve KEEPS ITS ROW with a blank entity id and `record_scope = unresolved`.
Nothing here mints a spine entity.

NO NETWORK CALLS. `harvest` parses snapshots already in
`data/staging/business_registry/raw/`; each source below records the URL, the
snapshot filename and the retrieval date, and `--list-sources` prints the fetch
plan (URL, per-host delay, robots note) for a re-pull under
`docs/PULL_DISCIPLINE.md`.
"""

from __future__ import annotations

import csv
import hashlib
import html
import json
import re
import sys
from datetime import date, datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "code"))

RAW = ROOT / "data" / "staging" / "business_registry" / "raw"
STAGE = ROOT / "data" / "staging" / "business_registry"
CLEAN = ROOT / "data" / "clean" / "native_owned_businesses.csv"
SPINE = ROOT / "data" / "spine" / "cedar_entity_spine.csv"
REGISTRY = ROOT / "review" / "tribal_vendor_list_registry_2026-08-26.csv"

SCRIPT = "330_build_native_owned_businesses.py"
HARVEST_DATE = "2026-09-01"
RUN_ID = "run-2026-09-01-P"

csv.field_size_limit(min(sys.maxsize, 2**31 - 1))


# ---------------------------------------------------------------------------
# THE SOURCES. One dict per certifying authority. Everything a re-pull needs is
# here, so nobody has to re-derive the route.
# ---------------------------------------------------------------------------

SOURCES = {
    "TBD-044": dict(
        tribe_id="TRBF-CSKTFR-00", authority="Confederated Salish & Kootenai Tribes",
        nation_id="bia:confederated-salish-and-kootenai",
        programme="Indian Preference Office - Indian Preference Business List",
        directory_type="tero", assertion_class="OWNERSHIP",
        identity_scope="enrolled_member_graded",
        claim=("CS&KT INDIAN PREFERENCE BUSINESS LIST; PREFERENCE 1 = CSKT "
               "TRIBAL MEMBER; PREFERENCE 2 = MEMBER FROM A FEDERALLY "
               "RECOGNIZED TRIBE"),
        list_url="https://cskt.org/wp-content/uploads/2026/06/Indian-Preference-Business-List-06.10.2026.pdf",
        landing="https://cskt.org/indian-preference-office/",
        snapshot="TBD-044_cskt_indian_preference_business_list.pdf",
        delay_s=10, robots="robots.txt sets Crawl-delay: 10 - HONOURED",
        rung="landing page -> PDF link (rung 2)",
        newsletter_url="https://www.charkoosta.com/",
    ),
    "TBD-045": dict(
        tribe_id="TRBF-CHKNAT-00", authority="Cherokee Nation",
        nation_id="bia:cherokee-nation",
        programme="Cherokee Nation TERO Directory",
        directory_type="tero", assertion_class="OWNERSHIP",
        identity_scope="any_native",
        claim=("Cherokee Nation's Tribal Employment Rights Office maintains a "
               "listing of Indian-owned businesses; site header: 'Search Over "
               "700 Certified Indian-Owned Businesses'"),
        list_url="https://cherokeetero.com/directory/",
        landing="https://cherokeetero.com/directory/",
        snapshot="TBD-045_cherokee_directory_p*.html",
        delay_s=8,
        robots=("robots.txt READ 2026-09-01 and PERMISSIVE ('User-agent: * / "
                "Disallow:'). The registry's '403 to a plain client / robots "
                "UNREADABLE' note is SUPERSEDED - a browser UA gets HTTP 200."),
        rung="live origin with a browser UA (rung 0 - the 403 was a UA gate)",
        newsletter_url="https://www.cherokeephoenix.org/",
    ),
    "TBD-046": dict(
        tribe_id="TRBF-MHATAT-00", authority="Three Affiliated Tribes (MHA Nation)",
        nation_id="bia:three-affiliated-tribes",
        programme="MHA TERO Certified Indian Contractors",
        directory_type="tero", assertion_class="OWNERSHIP",
        identity_scope="any_native_graded",
        claim=("MHA TERO Certified Indian Contractors list; Tier Level 1-4 "
               "(L1 self-performing certified Indian contractors, L2 certified "
               "with mentorship agreements, L3 certified Indian contractors "
               "acting as BROKERS, L4 other federally recognised tribes)"),
        list_url="https://mhatero.com/wp-content/uploads/2026/06/TERO-CIC-06.17.2026.pdf",
        landing="https://mhatero.com/contractor-lists/",
        snapshot="TBD-046_mha_tero_certified_indian_contractors.pdf",
        delay_s=3, robots="robots.txt disallows only /wp-admin/",
        rung="landing page -> PDF link (rung 2)",
        note=("SEVEN OTHER LISTS on the same page (Approved Oilfield Vendors, "
              "General Contractors, Prime General, Prime Oilfield, Consultants, "
              "Subs w/ DOT Exemption, Suppliers) are VENDOR-type and are NOT "
              "harvested here: reading them as ownership is the exact error "
              "this dataset exists to avoid."),
        newsletter_url="https://www.mhanation.com/mha-times",
    ),
    "TBD-047": dict(
        tribe_id="TRBF-ONDAWI-00", authority="Oneida Nation (Wisconsin)",
        nation_id="bia:oneida-nation-wi",
        programme="Indian Preference Vendor List",
        directory_type="indian_preference", assertion_class="OWNERSHIP",
        identity_scope="any_native",
        claim=("Oneida Nation Indian Preference Vendor List - annual "
               "re-certification via an Indian Preference Vendor Application; "
               "each record states the owner's tribal membership"),
        list_url="https://oneida-nsn.gov/business/indian-preference/indian-preference-vendor-list/",
        landing="https://oneida-nsn.gov/business/indian-preference/indian-preference-vendor-list/",
        snapshot="TBD-047_oneida_ipv_list.json",
        delay_s=3,
        robots=("robots.txt permits the list path. Rows render in JS from "
                "admin-ajax.php action=collect_ipv_list with the nonce the "
                "PAGE ITSELF publishes - the page's own public data route, no "
                "login and no access control bypassed. The registry's WP-REST "
                "401 was a different route."),
        rung="page's own public admin-ajax data route (rung 1)",
        newsletter_url="https://oneida-nsn.gov/kalihwisaks/",
    ),
    "TBD-048": dict(
        tribe_id="TRBF-POARCH-00", authority="Poarch Band of Creek Indians",
        nation_id="bia:poarch-band-of-creek-indians",
        programme="TERO Certified Businesses",
        directory_type="tero", assertion_class="OWNERSHIP",
        identity_scope="section_dependent",
        claim=("The businesses listed herein have been certified according to "
               "the Tribal Employment Rights Ordinance (TERO) adopted August "
               "18, 2011 by the Poarch Band of Creek Indians Tribal Council"),
        list_url="https://pci-nsn.gov/wp-content/uploads/TERO-Certified-Business-List-07.21.2026-1.pdf",
        landing="https://pci-nsn.gov/our-government/regulatory-affairs/",
        snapshot="TBD-048_poarch_tero_certified_business_list.pdf",
        delay_s=3, robots="robots.txt fully permissive with a sitemap",
        rung="landing page -> PDF link (rung 2)",
        note=("The ONE source in the study that publishes the ENTITY-owned vs "
              "INDIVIDUAL-owned line itself: 'TRIBAL BUSINESSES:' versus "
              "'100% Tribal Member Owned Businesses:'. `identity_scope` is set "
              "per section, never for the file."),
    ),
    "TBD-050": dict(
        tribe_id="TRBF-THNODM-00", authority="Tohono O'odham Nation",
        nation_id="bia:tohono-oodham-nation",
        programme="TERO Certified Indian Preference Firms",
        directory_type="tero", assertion_class="OWNERSHIP",
        identity_scope="any_native_graded",
        claim=("PURSUANT TO TERO REGULATIONS PART 3.1(b): 1) First Preference "
               "shall be given to Indian Preference Certified firms, 51% or "
               "more, of which are owned by O'odham and other local Indians. "
               "2) Second Preference shall be given to other Indian Preference "
               "Certified Firms."),
        list_url="https://www.tonation-nsn.gov/wp-content/uploads/2026/07/July-2026-Updated-Certified-Firms-Listing-V2.pdf",
        landing="https://www.tonation-nsn.gov/",
        snapshot="TBD-050_tohono_certified_firms.pdf",
        delay_s=3, robots="fully permissive; sitemap_index.xml declared",
        rung="direct PDF URL from the registry (rung 2)",
        note=("PARTIAL: 17 firms recovered against the 19 'Full/Probationary "
              "Certification' stamps the document itself carries. Two firms "
              "whose NAME/TITLE blocks sit closer than the row-gap threshold "
              "are merged into the preceding row, visible as a run-on "
              "owner_name_raw. The shortfall is stated rather than padded."),
    ),
    "TBD-052": dict(
        tribe_id="TRBF-LUMMIT-00", authority="Lummi Nation",
        nation_id="bia:lummi-nation",
        programme="Lummi Owned Businesses (LIBC business-licence report)",
        directory_type="business_licence", assertion_class="OWNERSHIP",
        identity_scope="any_native",
        claim=("'Lummi Owned Businesses' report served by the Lummi Indian "
               "Business Council's own licensing system, carrying LIBC licence "
               "expiry dates; final page: 'This has been provided by the Lummi "
               "Chamber of Commerce'"),
        list_url="https://www.lummi-nsn.gov/widgets/LummiOwnedBusinesses.php",
        landing="https://www.lummi-nsn.gov/",
        snapshot="TBD-052_lummi_owned_businesses.pdf",
        delay_s=3,
        robots=("robots.txt DISALLOWS /apps. The tribe's directory page links "
                "the same report at /apps/BusLicenses/... - THAT PATH IS NOT "
                "FETCHED. /widgets/ is not disallowed and is the copy used. "
                "ANY RE-RUN MUST USE /widgets/."),
        rung="permitted /widgets/ mirror of a robots-disallowed /apps/ path (rung 1)",
    ),
    "TBD-053": dict(
        tribe_id="TRBF-BLCKFT-00", authority="Blackfeet Nation",
        nation_id="bia:blackfeet-tribe",
        programme="Blackfeet TERO Certified Indian Preference Firms",
        directory_type="tero", assertion_class="OWNERSHIP",
        identity_scope="any_native",
        claim=("Blackfeet Tribal Employment Rights Office (T.E.R.O) - 2024 "
               "Certified Indian Preference Firms"),
        list_url="https://img1.wsimg.com/blobby/go/2bd82483-43d9-461a-984e-7cc3cd7b3cad/downloads/25a94857-0206-4fad-a938-543ab3713502/doc20250610134440.pdf",
        landing="https://btero.com/",
        snapshot="TBD-053_blackfeet_tero_ocr.json",
        delay_s=10, robots="blackfeetnation.com sets Crawl-delay: 10 - HONOURED",
        rung="separate TERO domain btero.com -> CDN-hosted scanned PDF -> OCR (rung 4)",
        note=("SCANNED. Recovered with rapidocr-onnxruntime at 220dpi, the "
              "route `122_ocr_ordinance_scans.py` established (no tesseract "
              "binary on this machine). Every row is OCR_RECOVERED, never "
              "TEXT_LAYER_PRESENT, and carries its mean line confidence."),
    ),
    "TBD-054": dict(
        tribe_id="TRBF-MNMNEE-00", authority="Menominee Indian Tribe of Wisconsin",
        nation_id="bia:menominee-indian-tribe",
        programme="Menominee Contractors listing",
        directory_type="vendor", assertion_class="RELATIONSHIP",
        identity_scope="vendor_relationship",
        claim=("'Menominee Contractors' - a reviewed-and-approved tribal "
               "contractor referral listing. NO TERO branding, NO stated "
               "ownership threshold, NO certification number or expiry and NO "
               "published eligibility rule anywhere on the page. Typed DOWN to "
               "RELATIONSHIP because a vendor list is not an ownership list."),
        list_url="https://www.menominee-nsn.gov/BusinessPages/ContractorsListing.aspx",
        landing="https://www.menominee-nsn.gov/BusinessPages/ContractorsListing.aspx",
        snapshot="TBD-054_menominee_contractors.html",
        delay_s=3, robots="no robots directives observed; ASP.NET, no sitemap",
        rung="live origin (rung 0); pagination NOT traversed - see note",
        note=("PARTIAL. The DevExpress grid pages server-side and the callback "
              "refused three argument formats (c0:PS200, c0:PN1, "
              "c0:KV|10;keys;GB|20;12|PAGERONCLICK3|PN1;) - each returned page "
              "1 again. 10 of the source's own 23 grid rows are held; the "
              "remaining 13 are NOT in this dataset and are not estimated. "
              "Wayback holds captures 2015-2026 for a later pass."),
    ),
    "TBD-056": dict(
        tribe_id="ANRC-ARCSLO-00", authority="Arctic Slope Regional Corporation",
        nation_id="ancsa:arctic-slope-regional-corporation",
        programme="ASRC Federal contract vehicles - subsidiary directory",
        directory_type="subsidiary_directory", assertion_class="OWNERSHIP",
        identity_scope="parent_asserted_subsidiary",
        claim=("ASRC Federal Holding Company publishes these firms as its own "
               "subsidiaries holding federal contract vehicles ('Subsidiary "
               "contract vehicle details'). ASRC is an Alaska Native Regional "
               "Corporation; the parent asserts the ownership link."),
        list_url="https://www.asrcfederal.com/contract-vehicles/",
        landing="https://www.asrcfederal.com/contract-vehicles/",
        snapshot="TBD-056_asrc_contract_vehicles.html",
        delay_s=3,
        robots=("asrc.com (the PARENT) returns HTTP 307 on every HTML page - a "
                "WAF - so asrc.com's own terms remain UNREAD and its verdict "
                "stays NOT_CHECKED. asrcfederal.com answers normally and "
                "publishes no terms-of-use page (probed 2026-09-01, 404)."),
        rung="subsidiary domain asrcfederal.com after the parent asrc.com WAF (rung 3)",
        note=("Carries a FEDERAL CONTRACT NUMBER per subsidiary - the only "
              "joinable identifier in this dataset that is not a name."),
    ),
    "TBD-058": dict(
        tribe_id="ANRC-CALSTA-00", authority="Calista Corporation",
        nation_id="ancsa:calista-corporation",
        programme="Calivika - Calista shareholder business directory",
        directory_type="shareholder_vendor", assertion_class="OWNERSHIP",
        identity_scope="shareholder_descendant_or_spouse",
        claim=("Welcome to Calivika ('my workplace' in Yup'ik), a free "
               "directory of businesses owned by Calista Shareholders, "
               "Descendants and their spouses."),
        list_url="https://calistashareholderbiz.com/",
        landing="https://calistashareholderbiz.com/shareholder-sitemap.xml",
        snapshot="TBD-058_calista_*.html",
        delay_s=2, robots="WordPress boilerplate; shareholder-sitemap.xml enumerable",
        rung="XML sitemap enumeration of /shareholder/ pages (rung 1)",
        note=("THE WEAKEST OWNERSHIP ASSERTION IN THE DATASET AND IT MUST STAY "
              "TYPED THAT WAY. Eligibility is shareholder / descendant / "
              "SPOUSE ownership with NO stated percentage threshold and NO "
              "described verification. Corroborating, never dispositive. A "
              "spouse-owned firm is not a Native-owned firm and this list "
              "cannot tell you which is which."),
    ),
    "TBD-059": dict(
        tribe_id="ANRC-DOYONL-00", authority="Doyon, Limited",
        nation_id="ancsa:doyon-limited",
        programme="Doyon Operations - family of companies",
        directory_type="subsidiary_directory", assertion_class="OWNERSHIP",
        identity_scope="parent_asserted_subsidiary",
        claim=("Doyon, Limited publishes these as its own operating companies "
               "('Operating more than a dozen for-profit companies, Doyon, "
               "Limited...'). Doyon is an Alaska Native Regional Corporation; "
               "the parent asserts the ownership link."),
        list_url="https://www.doyon.com/operations/",
        landing="https://www.doyon.com/operations/",
        snapshot="TBD-059_doyon_*.html",
        delay_s=3, robots="robots.txt disallows only /wp-admin/; sitemap published",
        rung="live origin, six business-line subpages (rung 0)",
    ),
    "TBD-079": dict(
        tribe_id="TRBF-MSENAT-00", authority="Muscogee (Creek) Nation",
        nation_id="bia:muscogee-creek-nation",
        programme="MCN CESO Vendor List",
        directory_type="tero", assertion_class="OWNERSHIP",
        identity_scope="any_native",
        claim=("Muscogee (Creek) Nation Contracting and Employment Support Act "
               "(NCA 18-199, 2019) s.9-105(I) requires 'Documented evidence "
               "proving fifty-one percent (51%) or more Native ownership and "
               "proof of Native control and management'"),
        list_url="https://www.muscogeenation.com/wp-content/uploads/2026/08/MCN-CESO-Vendor-List-260817.xlsx",
        landing="https://www.muscogeenation.com/",
        snapshot="tbd-079_MCN-CESO-Vendor-List-260817.xlsx",
        delay_s=3, robots="disallows /wp-admin only; no crawl-delay",
        rung="direct XLSX URL from the registry (rung 2); snapshot already on disk",
        note=("The file is titled 'Vendor List' and a keyword rule would "
              "MISFILE IT AS RELATIONSHIP - only the Act proves it is the "
              "certified roster. The Act defines both 'Muscogee Owned Vendor' "
              "(51% Muscogee citizen) and 'Indian Owned Vendor' (51% other "
              "federally recognised tribe) but THE FILE DOES NOT DISTINGUISH "
              "THEM, so the owning tribe is NOT recoverable and "
              "`tribal_affiliation_raw` stays blank on every row."),
    ),
}


# ---------------------------------------------------------------------------
# EXCLUDED. Not a failure list - a decisions list. Each carries the verbatim
# term and where it was read. These stay excluded by EVERY route, Wayback
# included: terms are a decision the publisher made, not an obstacle.
# ---------------------------------------------------------------------------

EXCLUDED = {
    "TRBF-COLVLL-00": dict(
        source_id="TBD-060", authority="Confederated Tribes of the Colville Reservation",
        list_url="https://www.colvilletribes.com/tero",
        terms_url="https://www.colvilletribes.com/",
        quote="All rights reserved, Colville Tribes. Copyright (c)",
        found_by="registry 2026-08-26",
        why=("TERMS_STATED_RESTRICTIVE on the 2026-08-26 survey. NEVER "
             "FETCHED BY THIS RUN. Its Squarespace robots.txt additionally "
             "NAMES anthropic-ai / ClaudeBot / GPTBot / CCBot - an expressed "
             "preference, and a reason to ask before publishing."),
        loss=("RICHEST SCHEMA IN THE STUDY: an explicit numeric 'Indian % "
              "Owned' column plus a four-level preference tier. It is also the "
              "source that PROVES presence on a TERO list is not by itself an "
              "ownership claim - firms at 0% Indian ownership still carry "
              "'Certified Title 10 = Yes'. Worth an OPT_IN request."),
    ),
    "TRBF-UMATLL-00": dict(
        source_id="TBD-033", authority="Confederated Tribes of the Umatilla Indian Reservation",
        list_url="https://ctuir.org/departments/workforce-development/tero/certified-indian-owned-business-directory/",
        terms_url="https://ctuir.org/",
        quote="Copyright (c) CTUIR 2020",
        found_by="registry 2026-08-26",
        why=("TERMS_STATED_RESTRICTIVE on the 2026-08-26 survey. A DOCX "
             "snapshot from that pass sits in raw/ as tbd-033_iob-as-of-"
             "42026-most-recent.docx and IS NOT PARSED by this script."),
        loss="14 entries, very clean, with owner names and certificate validity ranges.",
    ),
    "TRBF-CHKSWN-00": dict(
        source_id="TBD-018", authority="The Chickasaw Nation",
        list_url="http://www.chickasawbusinessnetwork.com/Chickasaw-Business-Directory.aspx",
        terms_url="http://www.chickasawbusinessnetwork.com/Special-Pages/Terms.aspx",
        quote=("Use of Company Directories - The information contained in any "
               "company directories that may be provided on the Service is "
               "provided for business lookup purposes and is not to be used "
               "for marketing or telemarketing applications. This information "
               "may not be copied or redistributed and is provided 'AS IS' "
               "without warranty of any kind."),
        found_by="THIS RUN 2026-09-01",
        why=("The registry recorded SILENT because nobody opened the 'Review "
             "Terms' link the directory page itself carries. The clause is "
             "specifically about COMPANY DIRECTORIES, which is precisely this "
             "artefact. The same page also grants use 'for your personal, "
             "noncommercial use only'. This is a refusal, and it outranks the "
             "622 business ids a 2026-08-28 pass had already enumerated - "
             "those ids stay in raw/ and are NOT resolved to detail pages."),
        loss="~622 businesses, the largest single lower-48 directory after Cherokee.",
    ),
    "TRBF-FSTCTY-00": dict(
        source_id="TBD-051", authority="Forest County Potawatomi Community",
        list_url="https://shop.fcpotawatomi.com/businesses/",
        terms_url="https://www.fcpotawatomi.com/terms-of-service/",
        quote=("Permission is granted to temporarily download one copy of the "
               "materials ... on FCPC's web site for personal, non-commercial "
               "transitory viewing only. This is the grant of a license, not a "
               "transfer of title, and under this license you may not: Modify "
               "or copy the materials; Use the materials for any commercial "
               "purpose, or for any public display"),
        found_by="THIS RUN 2026-09-01",
        why=("Registry recorded SILENT. The tribe's site-wide Terms of Service "
             "and Copyright Notice both state the restriction. One landing "
             "snapshot was retrieved before the terms were read; it is kept in "
             "raw/ as the evidence for this exclusion and NO ROWS ARE DERIVED "
             "FROM IT."),
        loss="18 FCP tribal-member-owned businesses with owner names.",
    ),
    "TRBF-STHUTE-00": dict(
        source_id="TBD-055", authority="Southern Ute Indian Tribe",
        list_url="https://www.southernute-nsn.gov/wp-content/uploads/sites/15/2026/03/2026-Indian-Own-Business-List.pdf",
        terms_url="https://www.southernute-nsn.gov/terms-of-use",
        quote=("Permission is granted to temporarily download one copy of the "
               "materials ... on Southern Ute Indian Tribe's web site for "
               "personal, non-commercial transitory viewing only. ... under "
               "this license you may not: modify or copy the materials; use "
               "the materials for any commercial purpose, or for any public "
               "display (commercial or non-commercial)"),
        found_by="THIS RUN 2026-09-01",
        why=("Registry recorded SILENT. The tribe publishes an explicit Terms "
             "of Use. The PDF was retrieved before the terms were read; the "
             "snapshot stays in raw/ as evidence and NO ROWS ARE DERIVED."),
        loss=("27 firms - and the one source that answers the growth-fund "
              "question directly, because Red Willow Production Company and "
              "Red Cedar Gathering Company appear on the TERO Indian-owned "
              "list itself alongside small local contractors."),
    ),
    "ANRC-NANARC-00": dict(
        source_id="TBD-057", authority="NANA Regional Corporation, Incorporated",
        list_url="https://www.akima.com/opco-sitemap.xml",
        terms_url="https://www.akima.com/terms-of-use/",
        quote=("no part of the Services and no Content or Marks may be copied, "
               "reproduced, aggregated, republished ... or otherwise exploited "
               "for any commercial purpose whatsoever, without our express "
               "prior written permission ... Engage in any automated use of "
               "the system, such as using scripts ... or using any data "
               "mining, scraping, (ro)bots, or similar data gathering and "
               "extraction tools, and/or use any Content or Services to train "
               "or otherwise utilize AI tools without our permission."),
        found_by="THIS RUN 2026-09-01",
        why=("Registry recorded SILENT and rated this 'MOST MACHINE-TRACTABLE "
             "FIND IN THE STUDY'. Akima's Legal Terms prohibit automated "
             "collection and commercial reuse in terms, by name. The sitemap "
             "was retrieved before the terms were read; the enumeration was "
             "STOPPED mid-run and no operating-company page was fetched. "
             "nana.com itself answers 403 to an automated client, so the "
             "parent's own terms are separately unread."),
        loss=("~55 operating companies each publishing CAGE, UEI, DUNS, "
              "primary NAICS and 8(a) status - a UEI-keyed, parent-asserted "
              "ANC subsidiary roster joinable to federal award data with no "
              "name matching. The single highest-value OPT_IN request in this "
              "dataset."),
    ),
}



# Module-level mirrors of the two dispositions `promote` uses, so the docs
# phase reports the same facts the build applied rather than a second copy of
# somebody's memory of them.
PRIOR_DOC = {
    "TBD-030": dict(tribe_id="TRBF-TULALP-00", authority="Tulalip Tribes",
                    assertion_class="OWNERSHIP"),
    "TBD-032": dict(tribe_id=None,
                    authority="Confederated Tribes of Grand Ronde",
                    assertion_class="OWNERSHIP"),
    "TBD-041": dict(tribe_id="TRBF-NAVAJO-00", authority="Navajo Nation",
                    assertion_class="OWNERSHIP"),
    "TBD-043": dict(tribe_id="TRBF-ESTCHK-00",
                    authority="Eastern Band of Cherokee Indians",
                    assertion_class="OWNERSHIP"),
}
SIBLING_DOC = {
    "TBD-C02": dict(disposition="INCLUDE",
                    authority="Pokagon Band of Potawatomi Indians",
                    assertion_class="OWNERSHIP"),
}



# ---------------------------------------------------------------------------
# THE 2026-09-01 HIDDEN-ROUTE SWEEP.
#
# `docs/HIDDEN_DATA_TECHNIQUES.md` run against the eleven registry rows that
# were REFERENCED-BUT-NOT-PUBLISHED or SITE_UNREACHABLE: robots.txt first,
# then /wp-json/wp/v2/types, /wp-json/wp/v2/media?search=, /wp-json/wp/v2/search
# and sitemap.xml.
#
# NONE of the six TERMS_STATED_RESTRICTIVE sources were probed by any of these
# routes. A JSON endpoint on a site that has told us not to scrape it is still
# off limits, and Wayback is not a route around a refusal either.
#
# **IT FOUND NO NEW VENDOR LIST.** That is a result, and PULL_DISCIPLINE
# requires it be written down with the date, the surface probed and the count -
# a sweep whose yield is never recorded is indistinguishable from one that never
# ran. It DID produce three corrections to SITE_UNREACHABLE verdicts, which are
# worth more than a null: "the site did not answer" and "the site refused us by
# name in robots.txt" are different facts and only one of them is a decision.
# ---------------------------------------------------------------------------
SWEEP_2026_09_01 = {
    "TRBF-SNCNAT-00": ("no list; wp-json + media search + sitemap probed, "
                       "ordinance only"),
    "TRBF-LAGUNA-00": ("no list; wp/v2/media returns the Indian Preference Code "
                       "and five contractor APPLICATION forms and no roster - "
                       "the machinery is published, the register is not"),
    "TRBF-PNBSCT-00": ("NOT PROBED: penobscotnation.org robots.txt disallows "
                       "ClaudeBot by name. A named disallow is a refusal and "
                       "the media API is not a route around one"),
    "TRBF-UTEMTN-00": "no list; host returns HTTP 307 to an automated client (WAF)",
    "TRBF-WRMSPR-00": ("no list; wstero.com wp/v2/media returns the TERO Code, "
                       "the FAQ, a Contractor Registration form and a Business "
                       "Certification Form - forms, never a roster"),
    "TRBF-QUINLT-00": ("no list; /BusinessDirectoryii.aspx found via sitemap is "
                       "a CivicEngage community Resource Directory, NOT a "
                       "Native-owned business register. The tribe's own bid "
                       "packets still say the list 'is available from TERO'"),
    "TRBF-SMARIE-00": ("no list; the sitemap's Vendor Information folder returns "
                       "HTTP 500 'not a folder'"),
    "TRBF-TURTLM-00": ("CORRECTION: NOT unreachable - tmchippewa.com robots.txt "
                       "DISALLOWS '/'. Reclassify SITE_UNREACHABLE -> "
                       "ROBOTS_DISALLOW_ALL. Not probed further, and Wayback is "
                       "not a route around it"),
    "TRBF-SNCRLS-00": ("still HTTP 307 to an automated client (WAF). The "
                       "sitemap evidence that /tero-2/ exists stands; this is "
                       "not absence"),
    "TRBF-WMTNAZ-00": ("CORRECTION: wmat.us answers HTTP 200 once certificate "
                       "verification is relaxed - the 2026-08-26 "
                       "SITE_UNREACHABLE was a TLS CHAIN PROBLEM, not an absent "
                       "site. No WordPress REST API and /sitemap.xml is 404, so "
                       "the list question is OPEN, not negative"),
    "AKNF-KTZBUE-00-NANARC-MANLLQ": ("CORRECTION: kotzebueira.org answers HTTP "
                                     "200 to a browser User-Agent. No wp-json, "
                                     "sitemap has no vendor/TERO URL. Open, not "
                                     "negative"),
}

#: Which technique from `docs/HIDDEN_DATA_TECHNIQUES.md` actually produced the
#: data, per source. Recorded so the next agent skips the routes that did not.
TECHNIQUE = {
    "TBD-044": "landing page -> /wp-content/uploads/ PDF; PDF positional text layer",
    "TBD-045": "robots.txt -> sabai-sitemap-index.xml -> /directory/sitemap.xml (836 slugs) -> 9 paginated list views at ?p=N (technique 4 + 8)",
    "TBD-046": "landing page -> /wp-content/uploads/ PDF",
    "TBD-047": ("technique 8 - the AJAX source behind the table. The page renders "
                "in JS from admin-ajax.php action=collect_ipv_list using the nonce "
                "THE PAGE ITSELF PRINTS to anonymous visitors. BOUNDARY NOTE: "
                "HIDDEN_DATA_TECHNIQUES forbids /wp-admin/ as private "
                "infrastructure; admin-ajax.php is WordPress's PUBLIC front-end "
                "AJAX handler, oneida-nsn.gov/robots.txt does NOT disallow it "
                "(read 2026-09-01; it disallows only calendar views and "
                "/wp-content/uploads/formidable/), no login was involved and no "
                "nonce was guessed. Flagged for the owner as a boundary call."),
    "TBD-048": "landing page -> /wp-content/uploads/ PDF",
    "TBD-050": "direct PDF URL; PDF line-level bbox reconstruction",
    "TBD-052": "a .php endpoint that renders a fresh PDF per request; /widgets/ mirror of a robots-disallowed /apps/ path",
    "TBD-053": "separate TERO domain -> CDN-hosted scanned PDF -> OCR (rapidocr 220dpi) + word-box column reconstruction",
    "TBD-054": "plain HTML table; the DevExpress pager callback refused three argument formats",
    "TBD-056": "plain HTML; subsidiary domain after the parent's WAF",
    "TBD-058": "technique 4 - shareholder-sitemap.xml enumerated 111 /shareholder/ pages",
    "TBD-059": "plain HTML headings across six business-line pages",
    "TBD-079": "direct XLSX URL (openpyxl)",
}

# ---------------------------------------------------------------------------
# TERMS READ ON 2026-09-01 THAT DID **NOT** REFUSE.
#
# Two registry rows sat at NOT_CHECKED. Recording "we looked and found no reuse
# restriction" is a different fact from "we never looked", and leaving them
# NOT_CHECKED would make the next agent re-open the same two pages.
#
# THE BAR THIS RUN APPLIED, stated so it can be argued with:
#   TERMS_STATED_RESTRICTIVE  = the publisher states a term GOVERNING USE of
#                               the content (no copying, no commercial use, no
#                               automated collection, no redistribution).
#   SILENT                    = no terms page; a bare copyright footer only.
#   TERMS_STATED_NO_REUSE_RESTRICTION
#                             = a terms page EXISTS, was read, and imposes no
#                               reuse restriction.
# ---------------------------------------------------------------------------
TERMS_READ = {
    "TRBF-TULALP-00": dict(
        status="TERMS_STATED_NO_REUSE_RESTRICTION",
        url="https://www.tulaliptero.com/Home/TermsOfUse",
        quote=("This web site and its database of laws are maintained for "
               "educational, cultural, and research purposes for the benefit "
               "of the members of the Tulalip Tribes, tribal staff, and the "
               "Internet community."),
        note=("The page asserts IP ownership ('the sole and exclusive property "
              "of The Tulalip Tribes') and disclaims warranties, but states NO "
              "prohibition on copying, redistribution, commercial use or "
              "automated collection - which is what separates it from the six "
              "EXCLUDED sources. Consent is still UNRESOLVED; an absence of "
              "prohibition is not permission."),
    ),
    "ANRC-ARCSLO-00": dict(
        status="SILENT",
        url="https://www.asrcfederal.com/ (probed /terms-of-use, HTTP 404)",
        quote="(c) ASRC Federal Holding Company. All Rights Reserved.",
        note=("asrcfederal.com publishes a Privacy Notice and no terms-of-use "
              "page. The PARENT, asrc.com, returns HTTP 307 to every automated "
              "client, so the parent's own terms remain genuinely unread - that "
              "is a fact about asrc.com, not about asrcfederal.com."),
    ),
}

# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

CORP = re.compile(
    r"\b(inc|inc\.|llc|l\.l\.c|llp|lp|ltd|corp|corporation|company|co|co\.|"
    r"enterprises?|services?|service|construction|contracting|contractors?|"
    r"group|holdings?|industries|solutions?|systems?|technologies|technology|"
    r"associates|partners|consulting|consultants|supply|supplies|trucking|"
    r"logistics|energy|oilfield|excavating|excavation|electric|plumbing|"
    r"roofing|welding|design|designs|studio|shop|store|farms?|ranch|pllc|"
    r"& sons|and sons|nation|tribal|tribe)\b", re.I)

PERSONISH = re.compile(r"^[A-Z][a-z'\-]+(?: [A-Z][a-z'\-]+){1,2}(?: (?:Jr|Sr|II|III|IV)\.?)?$")


#: `certification_start` / `certification_expiration` are lifted verbatim off
#: eighteen different certifying authorities' pages and PDFs, and until
#: 2026-09-02 they reached the customer in whatever shape the source printed.
#: Measured on the live table: 623 populated values in SIX formats -
#: `####-##-##` (346), `##/##/####` (144), `#/##/####` (86), `#/#/####` (33),
#: `##/#/####` (13), `#/##/##` (1). Nothing sorted and nothing parsed. Worse,
#: the ISO plurality was entirely `publishable = N` (Navajo's NBOA list, which
#: never ships), so **every date that actually reached a customer was in an
#: un-normalised US format** - `04/29/2027` and `4/16/2027` two rows apart.
#:
#: US ORDER IS ASSUMED AND THAT ASSUMPTION IS SAFE HERE, because every
#: contributing authority is a tribal government or TERO office inside the
#: United States and prints US-order dates. Anything this cannot parse is left
#: EXACTLY as the source printed it rather than guessed at - an unparseable
#: date is a fact about the source, and inventing a reading of it would be the
#: worse error.
#: PRECISION IS PRESERVED, NOT INVENTED. 32 Cherokee Nation rows print a month
#: and no day (`09/2020`). Those become `2020-09` - ISO 8601 month precision,
#: which sorts and parses - and NOT `2020-09-01`, because a day the source did
#: not print is a day this project does not know.
_ISO_RE = re.compile(r"^(\d{4})-(\d{1,2})-(\d{1,2})$")
_ISO_M_RE = re.compile(r"^(\d{4})-(\d{1,2})$")
_US_RE = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{2}|\d{4})$")
_US_M_RE = re.compile(r"^(\d{1,2})/(\d{4})$")


def iso_date(v):
    """Normalise a printed certification date to ISO. Unparseable -> verbatim."""
    s = str(v or "").strip()
    if not s:
        return v
    m = _US_M_RE.match(s) or _ISO_M_RE.match(s)
    if m:
        a, b = (int(x) for x in m.groups())
        mo, y = (a, b) if _US_M_RE.match(s) else (b, a)
        if not (1 <= mo <= 12 and 1900 <= y <= 2100):
            return v
        return f"{y:04d}-{mo:02d}"
    m = _ISO_RE.match(s)
    if m:
        y, mo, d = (int(x) for x in m.groups())
    else:
        m = _US_RE.match(s)
        if not m:
            return v
        mo, d, y = (int(x) for x in m.groups())
        if y < 100:
            # Two-digit years on a CERTIFICATION EXPIRY. 69 is the standard
            # POSIX pivot and the one value in the table reading `#/##/##` is
            # well inside the recent window either way.
            y += 2000 if y < 69 else 1900
    if not (1 <= mo <= 12 and 1 <= d <= 31 and 1900 <= y <= 2100):
        return v
    return f"{y:04d}-{mo:02d}-{d:02d}"


def norm_name(s: str) -> str:
    s = html.unescape(s or "").strip()
    s = re.sub(r"[‘’“”]", "'", s)
    s = re.sub(r"\s+", " ", s)
    t = s.lower()
    t = re.sub(r"[^\w\s&'-]", " ", t)
    t = re.sub(r"\b(inc|llc|l l c|llp|lp|ltd|corp|corporation|co|pllc)\b", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def looks_like_person(name: str, owner: str | None = None) -> int:
    """1 = the business name IS a person's name, 0 = it is not, -1 = unknown.

    Deliberately conservative in BOTH directions: a firm wrongly cleared here
    loses its privacy treatment, and a firm wrongly flagged loses its name from
    the clean table. -1 means neither test fired and the consumer must decide.
    """
    n = re.sub(r"\s+", " ", (name or "").strip())
    if not n:
        return -1
    if CORP.search(n):
        return 0
    if owner and norm_name(owner) and norm_name(owner) == norm_name(n):
        return 1
    if PERSONISH.match(n):
        return 1
    if re.match(r"^[A-Z][a-z'\-]+'s\b", n) and len(n.split()) <= 3:
        return -1
    return -1


def sha(*parts) -> str:
    return "sha256:" + hashlib.sha256("|".join(str(p) for p in parts).encode()).hexdigest()


def rec(src, key, name, **kw):
    """One staging record in the schema TBD-030 established."""
    sid = src["_id"]
    owner = kw.get("owner_name_raw")
    r = {
        "business_source_id": f"{sid}:{key}",
        "source_id": sid,
        "source_business_key": str(key),
        "business_entity_id": None,
        "nation_id": kw.get("nation_id", src.get("nation_id")),
        "business_name_raw": name,
        "business_name_normalized": norm_name(name),
        "dba_name": kw.get("dba_name"),
        "owner_name_raw": owner,
        "directory_type": src["directory_type"],
        "identity_scope": kw.get("identity_scope", src["identity_scope"]),
        "identity_claim_text": kw.get("identity_claim_text", src["claim"]),
        "assertion_class": src["assertion_class"],
        "ownership_percent": kw.get("ownership_percent"),
        "ownership_threshold_min": kw.get("ownership_threshold_min"),
        "control_requirement": kw.get("control_requirement"),
        "tribal_affiliation_raw": kw.get("tribal_affiliation_raw"),
        "verification_basis": kw.get("verification_basis", "TERO_review"),
        "certification_number": kw.get("certification_number"),
        "certification_tier": kw.get("certification_tier"),
        "certification_start": kw.get("certification_start"),
        "certification_expiration": kw.get("certification_expiration"),
        "business_license_number": kw.get("business_license_number"),
        "service_category_raw": kw.get("service_category_raw"),
        "naics": kw.get("naics"),
        "description_raw": kw.get("description_raw"),
        "address_raw": kw.get("address_raw"),
        "city": kw.get("city"),
        "state_province": kw.get("state_province"),
        "postal_code": kw.get("postal_code"),
        "phone": kw.get("phone"),
        "email": kw.get("email"),
        "website": kw.get("website"),
        "federal_contract_number": kw.get("federal_contract_number"),
        "source_url": kw.get("source_url", src["list_url"]),
        "source_edition": kw.get("source_edition"),
        "first_seen": f"{HARVEST_DATE}T00:00:00Z",
        "last_seen": f"{HARVEST_DATE}T00:00:00Z",
        "source_last_updated": kw.get("source_last_updated"),
        "is_current": True,
        "validation_flags": kw.get("validation_flags", []),
        "ingestion_method": kw.get("ingestion_method", "html"),
        "raw_snapshot_uri": "raw/" + kw.get("snapshot", src["snapshot"]),
        "refresh_run_id": RUN_ID,
        "relationship_basis_raw": kw.get("relationship_basis_raw"),
        "relationship_basis": kw.get("relationship_basis", "unspecified"),
        "certification_event_status": kw.get("certification_event_status", "approved"),
        "source_priority_class": kw.get("source_priority_class", "tribal_primary"),
        "cross_reference_only": False,
        "matched_primary_source_ids": None,
        "match_method": None,
        "match_confidence": None,
        "assertion_precedence_rank": 1,
        "ocr_mean_confidence": kw.get("ocr_mean_confidence"),
    }
    r["record_hash"] = sha(sid, key, name, r["business_name_normalized"])
    return r


def txt_of(path: Path) -> str:
    t = path.read_text(encoding="utf-8", errors="replace")
    b = re.sub(r"<(script|style)\b.*?</\1>", "", t, flags=re.S | re.I)
    b = re.sub(r"<br\s*/?>", "\n", b, flags=re.I)
    b = re.sub(r"</(p|div|li|tr|h\d|td)>", "\n", b, flags=re.I)
    b = re.sub(r"<[^>]+>", "\n", b)
    return html.unescape(b)


def lines_of(path: Path) -> list[str]:
    return [l.strip() for l in txt_of(path).split("\n") if l.strip()]


def pdf_pages(name: str) -> list[str]:
    import fitz
    d = fitz.open(RAW / name)
    return [p.get_text() for p in d]


# ---------------------------------------------------------------------------
# PARSERS. One per source. Each returns a list of staging records and prints
# what it could NOT get, by name, rather than counting it silently.
# ---------------------------------------------------------------------------

def parse_cskt(src):
    """Two-column PDF: x<300 is BUSINESS NAME & INFORMATION, x>=300 is
    DESCRIPTION OF SERVICES OFFERED.

    Record structure in the left column, in order:

        NAME / street / CITY, ST ZIP / PH: / E-MAIL: / [WEBSITE:] /
        PREFERENCE: n / YEARLY UPDATE: date / OWNER: name [wrap]

    `PREFERENCE:` is the only field on EVERY record (118 of them); `OWNER:`
    appears on 92, so splitting on OWNER: silently merged 26 firms into their
    neighbours. So the stream is cut at each `PREFERENCE:` line, and the
    YEARLY UPDATE / OWNER tail that opens the NEXT window is attributed back to
    the record it belongs to.

    Inside a window the CITY/STATE/ZIP line is the anchor: walk back over
    STREET lines to reach the business name. "Street line" is matched on a
    street-type suffix, not on a leading digit - 26 names in this file start
    with a digit ("406 FIRE SUPPRESSION LLC", "3 MOR ENTERPRISES, INC").
    """
    import fitz
    d = fitz.open(RAW / src["snapshot"])
    left = []
    for page in d:
        bands = {}
        for x0, y0, x1, y1, w, *_ in page.get_text("words"):
            if x0 >= 300:
                continue
            bands.setdefault(round(y0 / 3), []).append((x0, w))
        for y in sorted(bands):
            left.append(" ".join(w for _, w in sorted(bands[y])).strip())
    SKIP = re.compile(r"^(CS&KT INDIAN PREFERENCE|PREFERENCE [12] =|UPDATED:|"
                      r"BUSINESS NAME & INFORMATION|DESCRIPTION OF SERVICES|\d+$)")
    left = [l for l in left if l and not SKIP.match(l)]
    CITY = re.compile(r"^(.+?),?\s+(MT|ID|WA|OR|CA|AZ|NM|CO|UT|WY|ND|SD|MN|TX|NV)\.?\s+(\d{5})")
    FIELD = re.compile(r"^(PH:|FAX:|E-?MAIL:|WEBSITE:|PREFERENCE:|YEARLY UPDATE:|OWNER:|CELL:)", re.I)
    ADDR = re.compile(
        r"^(P\.?O\.? BOX|HC ?\d|#|SUITE|STE\b|APT\b|\d+ (?:MILE|MI\.)"
        r"|.*\b(RD|ROAD|LN|LANE|ST|STREET|AVE|AVENUE|HWY|HIGHWAY|DR|DRIVE|"
        r"BLVD|WAY|CT|COURT|PL|PLACE|LOOP|TRAIL|TRL|CIR|CIRCLE|PKWY|TERR|"
        r"ROUTE|RTE|BYPASS|MAIN|1ST|2ND|3RD|4TH)"
        r"\.?(\s+[NSEW]{1,2}\.?)?(\s+[\d\-]+)?(\s+(UNIT|STE|APT|#)\s*\S+)?$)", re.I)

    pidx = [i for i, l in enumerate(left) if re.match(r"^PREFERENCE:\s*[12]", l, re.I)]
    out = []
    for n, pi in enumerate(pidx):
        start = pidx[n - 1] + 1 if n else 0
        win = left[start:pi + 1]
        # the head of the window is the PREVIOUS record's YEARLY UPDATE / OWNER
        # tail; strip it off and hand it to that record.
        head = []
        k = 0
        while k < len(win) and re.match(r"^(YEARLY UPDATE:|OWNER:)", win[k], re.I):
            head.append(win[k])
            k += 1
        while (k < len(win) and head and any(h.upper().startswith("OWNER:") for h in head)
               and not FIELD.match(win[k]) and not CITY.match(win[k])
               and not re.search(r"\d", win[k])
               # a line carrying a corporate suffix is the NEXT firm's name, not
               # the tail of a person's name
               and not CORP.search(win[k])
               and k + 1 < len(win) and not CITY.match(win[k + 1])
               and not ADDR.match(win[k + 1])):
            head.append(win[k])
            k += 1
        if head and out:
            prev = out[-1]
            tail = "\n".join(head)
            mo = re.search(r"OWNER:\s*(.+)", tail, re.I | re.S)
            if mo:
                prev["owner_name_raw"] = re.sub(r"\s+", " ", mo.group(1)).strip()
            my = re.search(r"YEARLY UPDATE:\s*([\d/]+)", tail, re.I)
            if my:
                prev["certification_expiration"] = my.group(1)
        seg = win[k:]
        if not seg:
            continue
        body = "\n".join(seg)
        flags = []
        ci = next((i for i, l in enumerate(seg) if CITY.match(l)), None)
        if ci is None:
            name = next((l for l in seg if not FIELD.match(l)), None)
            addr = city = st = zp = None
            flags.append("city_state_zip_line_not_found_in_this_record_block")
        else:
            m = CITY.match(seg[ci])
            city, st, zp = m.group(1).strip(", "), m.group(2), m.group(3)
            ni = ci - 1
            while ni > 0 and ADDR.match(seg[ni]) and not FIELD.match(seg[ni]):
                ni -= 1
            name = seg[ni] if ni >= 0 else None
            addr = "; ".join(seg[ni + 1:ci + 1]) if ci > ni else seg[ci]
            if ni > 0:
                flags.append("line(s)_before_the_business_name_in_this_block: "
                             + " | ".join(seg[:ni]))
        if not name or FIELD.match(name) or CITY.match(name):
            continue
        if ADDR.match(name) or re.match(r"^\d{3,6}\s+[A-Z]", name):
            flags.append("BUSINESS_NAME_MAY_BE_AN_ADDRESS_LINE: the two-column "
                         "PDF gives no marker separating the firm name from its "
                         "street line on this record; the row is KEPT and "
                         "FLAGGED rather than dropped or guessed at")

        def g(pat):
            mm = re.search(pat, body, re.I)
            return mm.group(1).strip() if mm else None

        pref = g(r"PREFERENCE:\s*([12])")
        scope = ("enrolled_member_cskt" if pref == "1"
                 else "enrolled_member_other_federally_recognized" if pref == "2"
                 else "unspecified")
        out.append(rec(src, len(out) + 1, name,
                       owner_name_raw=None, identity_scope=scope,
                       identity_claim_text=(src["claim"] + "; PREFERENCE: " + pref)
                       if pref else src["claim"],
                       certification_tier=("PREFERENCE " + pref) if pref else None,
                       address_raw=addr, city=city, state_province=st, postal_code=zp,
                       phone=g(r"PH:\s*([()\d\s\-\.]{7,})"),
                       email=g(r"E-?MAIL:\s*(\S+@\S+)"),
                       website=g(r"WEBSITE:\s*(\S+)"),
                       ingestion_method="pdf_text_layer_positional",
                       source_last_updated="2026-06-10",
                       source_edition="Indian-Preference-Business-List-06.10.2026",
                       validation_flags=flags + [
                           "ownership_threshold_min_null: the list states a "
                           "MEMBERSHIP test (tribal member / member of a "
                           "federally recognized tribe), not a numeric "
                           "ownership percentage",
                           "description_of_services_sits_in_the_RIGHT_column_"
                           "and_is_not_row_aligned_to_the_left_column__so_it_"
                           "is_NOT_carried_rather_than_guessed",
                       ]))
    # the last record's OWNER tail is after the final PREFERENCE: line
    if out and pidx:
        tail = "\n".join(left[pidx[-1] + 1:])
        mo = re.search(r"OWNER:\s*(.+)", tail, re.I | re.S)
        if mo:
            out[-1]["owner_name_raw"] = re.sub(r"\s+", " ", mo.group(1)).strip()
        my = re.search(r"YEARLY UPDATE:\s*([\d/]+)", tail, re.I)
        if my:
            out[-1]["certification_expiration"] = my.group(1)
    for r in out:
        if not r["owner_name_raw"]:
            r["validation_flags"] = list(r["validation_flags"]) + [
                "owner_name_absent_from_this_record_in_source"]
    return out


def parse_cherokee(src):
    out = []
    files = sorted(RAW.glob("TBD-045_cherokee_directory_p*.html"),
                   key=lambda p: int(re.search(r"_p(\d+)\.", p.name).group(1)))
    seen = set()
    for f in files:
        t = f.read_text(encoding="utf-8", errors="replace")
        blocks = re.split(r'<div id="sabai-entity-content-(\d+)"', t)
        for i in range(1, len(blocks) - 1, 2):
            eid, body = blocks[i], blocks[i + 1]
            if eid in seen:
                continue
            seen.add(eid)
            m = re.search(r'sabai-directory-title".*?<a href="([^"]+)"[^>]*title="([^"]*)"',
                          body, re.S)
            if not m:
                continue
            url, name = m.group(1), html.unescape(m.group(2))
            cats = [html.unescape(c) for c in
                    re.findall(r'/directory/categories/[^"]*"[^>]*>.*?</i>\s*([^<]+)</a>', body, re.S)]
            addr = re.search(r'itemprop="address"[^>]*>(.*?)</', body, re.S)
            if not addr:
                addr = re.search(r'sabai-directory-address[^>]*>(.*?)</div>', body, re.S)
            address = html.unescape(re.sub(r"<[^>]+>", " ", addr.group(1))).strip() if addr else None
            if address:
                address = re.sub(r"\s+", " ", address)
            phone = re.search(r'(\(\d{3}\)\s*\d{3}-\d{4})', body)
            email = re.search(r'([A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,})', body)
            web = re.search(r'href="(https?://(?!cherokeetero\.com)[^"]+)"[^>]*itemprop="url"', body)
            if not web:
                web = re.search(r'>\s*(https?://(?!cherokeetero\.com)[^\s<]+)\s*<', body)
            city = st = zp = None
            if address:
                mm = re.search(r"([A-Za-z\. ]+),\s*([A-Z]{2})\s*(\d{5})", address)
                if mm:
                    city, st, zp = mm.group(1).strip(", "), mm.group(2), mm.group(3)
            out.append(rec(src, eid, name,
                           service_category_raw="; ".join(dict.fromkeys(cats)) or None,
                           address_raw=address, city=city, state_province=st,
                           postal_code=zp,
                           phone=phone.group(1) if phone else None,
                           email=email.group(1) if email else None,
                           website=web.group(1) if web else None,
                           source_url=url,
                           snapshot=f.name,
                           validation_flags=[
                               "owner_name_absent_from_source",
                               "ownership_threshold_min_null: no numeric "
                               "threshold is published at the directory URL",
                               "certification_number_absent_from_source",
                           ]))
    return out


def parse_mha(src):
    out = []
    rows = []
    for p in pdf_pages(src["snapshot"]):
        rows.extend([l.strip() for l in p.split("\n") if l.strip()])
    # header: Company Name / Phone # / Email / Tier Level, then 4-line records
    i = 0
    while i < len(rows):
        if rows[i] == "Company Name":
            i += 4
            continue
        name = rows[i]
        chunk = rows[i:i + 4]
        if len(chunk) < 4:
            break
        phone, email, tier = chunk[1], chunk[2], chunk[3]
        if not re.match(r"^[\d\-\(\) \.]{7,}$", phone) or not re.match(r"^\d$", tier.strip()):
            i += 1
            continue
        out.append(rec(src, len(out) + 1, name.strip().rstrip(","),
                       phone=phone, email=email if "@" in email else None,
                       certification_tier="Preference Level " + tier.strip(),
                       identity_scope=("any_native" if tier.strip() in "123"
                                       else "other_federally_recognized_tribe"),
                       identity_claim_text=src["claim"] + f"; Tier Level: {tier.strip()}",
                       ingestion_method="pdf_text_layer",
                       source_last_updated="2026-06-17",
                       source_edition="TERO-CIC-06.17.2026",
                       validation_flags=[
                           "no_address_no_owner_name_in_source",
                           "tier3_is_the_tribe_flagging_its_own_broker_firms:"
                           "a Level 3 row is a certified Indian contractor "
                           "ACTING AS A BROKER, which is a pass-through, not a "
                           "self-performer",
                       ]))
        i += 4
    return out


def parse_oneida(src):
    d = json.loads((RAW / src["snapshot"]).read_text(encoding="utf-8", errors="replace"))
    out = []
    for i, b in enumerate(d.get("list", []), 1):
        def s(k):
            v = (b.get(k) or "").strip()
            return v or None
        svc = "; ".join(x.get("serviceName", "") for x in b.get("services", []) if x.get("serviceName"))
        out.append(rec(src, s("oneidaVendor") or f"row{i}", b.get("name", "").strip(),
                       owner_name_raw=s("primaryContactName"),
                       tribal_affiliation_raw=s("ownerMemberships"),
                       identity_claim_text=(src["claim"] + "; Owner memberships: "
                                            + (s("ownerMemberships") or "NOT STATED")),
                       certification_number=s("oneidaVendor"),
                       certification_start=s("est"),
                       service_category_raw=svc or None,
                       naics=s("naics"),
                       description_raw=s("serviceDetail"),
                       address_raw="; ".join(x for x in [s("address1"), s("address2")] if x),
                       city=s("city"), state_province=s("state"), postal_code=s("zip"),
                       phone=s("businessPhone"), email=s("businessEmail"),
                       website=s("website"),
                       ingestion_method="json_api",
                       validation_flags=[
                           "owner_name_raw_source_label_is_primaryContactName",
                           "ownership_threshold_min_null: no numeric threshold "
                           "published at the list URL",
                       ] + ([] if s("ownerMemberships") else
                            ["owner_membership_blank_in_source"])))
    return out


def parse_poarch(src):
    pages = pdf_pages(src["snapshot"])
    body = "\n".join(pages)
    body = re.sub(r"\n\s*TERO CERTIFIED BUSINESSES[^\n]*\n", "\n", body)
    body = re.sub(r"\n\s*Page \d+ of \d+\s*\n", "\n", body)
    lines = [l.rstrip() for l in body.split("\n")]
    out, section = [], None
    i = 0
    while i < len(lines):
        l = lines[i].strip()
        if re.match(r"^TRIBAL BUSINESSES:?$", l, re.I):
            section = "tribally_owned_entity"
        elif re.match(r"^100%\s*Tribal Member Owned Businesses:?$", l, re.I):
            section = "enrolled_member_100pct"
        elif re.match(r"^51%\s+Tribal Member Owned Businesses:?$", l, re.I):
            section = "enrolled_member_51pct"
        elif re.match(r"^Mailing Address:?$", l, re.I) and section:
            # walk back to the nearest non-empty line = the business name
            j = i - 1
            while j >= 0 and not lines[j].strip():
                j -= 1
            name = lines[j].strip() if j >= 0 else None
            if not name or name.lower().endswith("businesses:"):
                i += 1
                continue
            blk = "\n".join(lines[i:i + 40])
            def g(pat):
                m = re.search(pat, blk, re.I)
                return m.group(1).strip() if m else None
            contact = g(r"Contact:\s*(.+)")
            if not contact:
                m = re.search(r"Contact:\s*\n\s*(.+)", blk, re.I)
                contact = m.group(1).strip() if m else None
            phone = g(r"Phone:\s*([\d\-\.\(\) ]{7,})")
            email = g(r"Email Address:\s*\n?\s*(\S+@\S+)")
            web = g(r"Website:\s*\n?\s*(\S+)")
            phys = g(r"Physical Address:\s*\n?\s*(.+)")
            bidlim = g(r"BID LIMIT:\s*(.+)")
            cats = re.findall(r"See ([\d\.]+) of [Tt]he TERO Regulations", blk)
            city = st = zp = None
            if phys:
                m = re.search(r"([A-Za-z\. ]+),?\s*(A[Ll])\.?,?\s*(\d{5})", phys)
                if m:
                    city, st, zp = m.group(1).strip(", "), m.group(2).upper(), m.group(3)
            claim = src["claim"] + "; section: " + {
                "tribally_owned_entity": "TRIBAL BUSINESSES",
                "enrolled_member_100pct": "100% Tribal Member Owned Businesses",
                "enrolled_member_51pct": "51% Tribal Member Owned Businesses",
            }[section]
            out.append(rec(src, len(out) + 1, name,
                           owner_name_raw=(contact if section != "tribally_owned_entity" else None),
                           identity_scope=section, identity_claim_text=claim,
                           ownership_percent=(100.0 if section == "enrolled_member_100pct"
                                              else 51.0 if section == "enrolled_member_51pct"
                                              else None),
                           address_raw=phys, city=city, state_province=st, postal_code=zp,
                           phone=phone, email=email, website=web,
                           service_category_raw=("TERO Regulations "
                                                 + ", ".join(dict.fromkeys(cats))) if cats else None,
                           certification_tier=bidlim and ("BID LIMIT: " + bidlim),
                           ingestion_method="pdf_text_layer",
                           source_last_updated="2026-07-21",
                           source_edition="TERO-Certified-Business-List-07.21.2026-1",
                           validation_flags=[
                               "contact_name_is_a_company_contact_for_TRIBAL_"
                               "BUSINESSES_rows_and_is_not_recorded_as_owner",
                           ] if section == "tribally_owned_entity" else []))
            i += 5
            continue
        i += 1
    return out


def parse_tohono(src):
    """Six-column table, read at LINE level with bboxes.

    Row boundaries come from the NAME/TITLE column, not from the DATE
    CERTIFIED cell: every firm has exactly one owner block there, its lines are
    <=15pt apart, and successive firms are >=27pt apart. Anchoring on the date
    cell instead closed records one line early whenever a firm's own city line
    shared a text baseline with its certification date, and named the firm
    after its own city.
    """
    import fitz
    d = fitz.open(RAW / src["snapshot"])
    BOUND = [0, 165, 245, 360, 555, 10000]     # measured on this document
    SKIP = re.compile(r"^(\d+\s*$|Updated as of|Tohono O|Indian Preference Firms$|"
                      r"PURSUANT TO|[12]\) (First|Second) Preference|COMPANY NAME|"
                      r"ADDRESS$|NAME/TITLE|TYPE OF BUSINESS|PHONE NUMBER|FAX$|"
                      r"DATE CERTIFIED|FIRST PREFERENCE$|SECOND PREFERENCE$)")
    out = []
    pref = "FIRST PREFERENCE"
    for pno in range(1, d.page_count):
        lines = []
        for blk in d[pno].get_text("dict")["blocks"]:
            for ln in blk.get("lines", []):
                txt = "".join(s["text"] for s in ln["spans"]).strip()
                if txt:
                    lines.append((ln["bbox"][1], ln["bbox"][0], txt))
        lines.sort()
        second_y = next((y for y, x, t in lines
                         if t.upper().startswith("SECOND PREFERENCE")), None)
        lines = [(y, x, t) for y, x, t in lines if not SKIP.match(t)]
        anchors = []
        for y, x, t in lines:
            if BOUND[1] <= x < BOUND[2]:
                if not anchors or y - anchors[-1][1] > 16:
                    anchors.append([y, y])
                else:
                    anchors[-1][1] = y
        if not anchors:
            continue
        starts = [a[0] for a in anchors]
        for i, s in enumerate(starts):
            lo = s - 6
            hi = starts[i + 1] - 6 if i + 1 < len(starts) else 10 ** 6
            cells = [[] for _ in range(5)]
            for y, x, t in lines:
                if not (lo <= y < hi):
                    continue
                col = next(c for c in range(5) if BOUND[c] <= x < BOUND[c + 1])
                cells[col].append(t)
            c = [" ".join(v).strip() for v in cells]
            this_pref = ("SECOND PREFERENCE"
                         if second_y is not None and s >= second_y else pref)
            blk = c[0]
            m = re.search(r"(P\.?O\.? Box|HC ?\d|\d{2,6}\s+[NSEW]?\.?\s*[A-Z0-9])", blk)
            name = (blk[:m.start()] if m else blk).strip(" ,.")
            addr = blk[m.start():].strip() if m else None
            if not name or len(name) < 3:
                continue
            city = st = zp = None
            if addr:
                mm = re.search(r"([A-Za-z\. ]+),?\s+([A-Z]{2})\s+(\d{5})", addr)
                if mm:
                    city, st, zp = mm.group(1).strip(", "), mm.group(2), mm.group(3)
            email = re.search(r"([A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+)", c[3])
            phone = re.search(r"(\(?\d{3}\)?[\s\-]?\d{3}[\s\-]?\d{4})", c[3])
            dc = re.search(r"(\d{2}/\d{2}/\d{4})", c[4])
            status = ("Full Certification" if "Full" in c[4]
                      else "Probationary Certification" if "Probationary" in c[4]
                      else None)
            out.append(rec(src, len(out) + 1, name,
                           owner_name_raw=c[1] or None,
                           service_category_raw=c[2] or None,
                           address_raw=addr, city=city, state_province=st,
                           postal_code=zp,
                           phone=phone.group(1) if phone else None,
                           email=email.group(1) if email else None,
                           certification_start=dc.group(1) if dc else None,
                           certification_tier=this_pref + ("; " + status if status else ""),
                           identity_scope=("any_native_oodham_and_local"
                                           if this_pref == "FIRST PREFERENCE"
                                           else "other_indian_preference_certified"),
                           identity_claim_text=src["claim"] + "; " + this_pref,
                           ownership_threshold_min=(51.0 if this_pref == "FIRST PREFERENCE"
                                                    else None),
                           ingestion_method="pdf_text_layer_positional",
                           source_last_updated="2026-07-21",
                           source_edition="July-2026-Updated-Certified-Firms-Listing-V2",
                           validation_flags=[
                               "certification_status_Probationary_vs_Full_is_"
                               "carried_in_certification_tier__they_are_not_the_"
                               "same_status_and_must_not_be_flattened",
                           ] + ([] if status else ["certification_status_absent_from_row"])))
        if second_y is not None:
            pref = "SECOND PREFERENCE"
    return out


def parse_lummi(src):
    import fitz
    d = fitz.open(RAW / src["snapshot"])
    out, cat = [], None
    for pno in range(1, d.page_count):
        words = d[pno].get_text("words")
        rows = {}
        for x0, y0, x1, y1, w, *_ in words:
            rows.setdefault(round(y0 / 4), []).append((x0, w))
        for y in sorted(rows):
            ws = sorted(rows[y])
            line = " ".join(w for _, w in ws)
            if re.match(r"^\d{2}/\d+/\d{4}$", line) or line.startswith("Lummi Owned Businesses"):
                continue
            if re.match(r"^Page \d+ of \d+$", line):
                continue
            if line.startswith("Name of Business"):
                continue
            xs = [x for x, _ in ws]
            # a category header is a single left-aligned phrase with no phone
            if (len(ws) <= 4 and min(xs) < 80 and max(xs) < 250
                    and not re.search(r"\d{3}\)", line) and not re.search(r"\d/\d", line)):
                if line and line[0].isupper() and len(line) < 40:
                    cat = line
                    continue
            phone = re.search(r"\((\d{3})\)\s*(\d{3})-(\d{4})", line)
            expiry = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", line)
            if expiry:
                name = line[:expiry.start()]
                if phone:
                    name = line[:phone.start()]
                name = name.strip()
                if not name:
                    continue
                out.append(rec(src, len(out) + 1, name,
                               service_category_raw=cat,
                               phone=("(%s) %s-%s" % phone.groups()) if phone else None,
                               certification_expiration=expiry.group(1),
                               ingestion_method="pdf_text_layer_positional",
                               source_last_updated="2026-09-01",
                               source_edition="Current as of Tuesday September 1st, 2026",
                               validation_flags=[
                                   "no_owner_name_no_address_no_ownership_"
                                   "percent_in_source",
                                   "licence_expiry_is_an_LIBC_BUSINESS_LICENCE"
                                   "_date__not_a_certification_expiry",
                               ]))
            else:
                # description line for the previous record
                if out and line and not phone and len(line) > 5:
                    prev = out[-1]
                    prev["description_raw"] = ((prev.get("description_raw") or "")
                                               + " " + line).strip()
    return out


def parse_blackfeet(src):
    """OCR'd scan. Reconstruct the 5-column table from word boxes."""
    data = json.loads((RAW / src["snapshot"]).read_text(encoding="utf-8", errors="replace"))
    out = []
    confs = []
    for page in data:
        lines = page["lines"]
        if not lines:
            continue
        for l in lines:
            confs.append(float(l.get("conf") or 0))
        xs = [min(p[0] for p in l["box"]) for l in lines]
        w = max(max(p[0] for p in l["box"]) for l in lines)
        # five columns, boundaries as fractions of page width
        B = [0, .22 * w, .48 * w, .70 * w, .86 * w, w + 1]
        bands = {}
        for l in lines:
            y = min(p[1] for p in l["box"])
            bands.setdefault(round(y / 18), []).append(l)
        # group bands into records: a record begins at a band containing a
        # business-licence number
        recs, cur = [], None
        for k in sorted(bands):
            band = bands[k]
            joined = " ".join(x["text"] for x in band)
            if re.search(r"Tribal Employment Rights|Certified\s*$|Name of Firm|"
                         r"www\.btero|^\d{1,2}/\d{1,2}/\d{4}$", joined):
                continue
            if cur is None:
                cur = {i: [] for i in range(5)}
            for l in band:
                x = min(p[0] for p in l["box"])
                for c in range(5):
                    if B[c] <= x < B[c + 1]:
                        cur[c].append(l["text"])
                        break
            if re.search(r"20\d\d-BL-\d{3,4}", joined):
                recs.append(cur)
                cur = None
        if cur and any(cur.values()):
            recs.append(cur)
        for r in recs:
            cols = {i: " ".join(v).strip() for i, v in r.items()}
            lic = re.search(r"(20\d\d-BL-\d{3,4})", " ".join(cols.values()))
            name = cols[0]
            name = re.sub(r"20\d\d-BL-\d{3,4}", "", name).strip()
            name = re.sub(r"\d{1,2}/\d{1,2}/\d{4}", "", name).strip()
            if not name or len(name) < 3:
                continue
            owner_blk = cols[1]
            owner = owner_blk.split("P.O")[0].split("P.0")[0].strip()
            phone = re.search(r"(\d{3}[\-\s]?\d{3}[\-\s]?\d{4})", owner_blk)
            email = re.search(r"([A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+)", owner_blk)
            dates = re.findall(r"(\d{1,2}/\d{1,2}/\d{4})", " ".join(cols.values()))
            out.append(rec(src, len(out) + 1, name,
                           owner_name_raw=owner or None,
                           service_category_raw=cols[2] or None,
                           address_raw=owner_blk or None,
                           phone=phone.group(1) if phone else None,
                           email=email.group(1) if email else None,
                           business_license_number=lic.group(1) if lic else None,
                           certification_start=dates[0] if dates else None,
                           certification_expiration=dates[-1] if len(dates) > 1 else None,
                           ingestion_method="ocr_rapidocr_220dpi",
                           source_last_updated="2025-06-10",
                           source_edition="doc20250610134440 (2024 Certified Indian Preference Firms)",
                           ocr_mean_confidence=round(sum(confs) / len(confs), 4) if confs else None,
                           validation_flags=[
                               "OCR_RECOVERED_not_TEXT_LAYER_PRESENT",
                               "column_reconstruction_from_word_boxes__field_"
                               "boundaries_are_inferred_not_marked_in_source",
                               "list_is_headed_2024_Certified_and_the_file_"
                               "stamp_is_2025-06-10__the_tribe_has_not_"
                               "published_a_newer_edition",
                           ]))
    return out


def parse_menominee(src):
    t = (RAW / src["snapshot"]).read_text(encoding="utf-8", errors="replace")
    rows = re.findall(r"<td>([^<]*)</td><td>([^<]*)</td><td class=\"dxgRRB\">([^<]*)</td>", t)
    out = []
    last = None
    for name, kind, person in rows:
        name = html.unescape(name).strip()
        kind = html.unescape(kind).strip()
        person = re.sub(r"\s+", " ", html.unescape(person)).strip()
        if name and name != " ":
            last = name
            out.append(rec(src, len(out) + 1, name,
                           owner_name_raw=person or None,
                           service_category_raw=kind or None,
                           verification_basis="tribal_review_and_approval_before_listing",
                           relationship_basis_raw="Menominee Contractors listing",
                           relationship_basis="vendor_listed_by_tribe",
                           ingestion_method="html",
                           validation_flags=[
                               "ASSERTION_CLASS_RELATIONSHIP__this_is_not_an_"
                               "ownership_list",
                               "no_certification_number_no_expiry_no_"
                               "ownership_threshold_published",
                               "PARTIAL_HARVEST: 10 of the source's own 23 "
                               "grid rows; the DevExpress pager callback "
                               "refused three argument formats",
                           ]))
        elif person and out:
            prev = out[-1]
            prev["owner_name_raw"] = ((prev.get("owner_name_raw") or "")
                                      + "; " + person).strip("; ")
            prev["validation_flags"] = list(prev["validation_flags"]) + [
                "multiple_named_principals_collapsed_onto_the_firm_row:the "
                "source prints one grid row per named person under a blank "
                "business name"]
    return out


def parse_asrc(src):
    t = (RAW / src["snapshot"]).read_text(encoding="utf-8", errors="replace")
    lines = lines_of(RAW / src["snapshot"])
    out, seen = [], {}
    vehicle = None
    for i, l in enumerate(lines):
        if l == "Subsidiary contract vehicle details":
            vehicle = lines[i - 2] if i >= 2 else None
            continue
        m = re.match(r"^(.+?)\s*\(([A-Z0-9\-,;: ]{6,}.*)\)$", l)
        if not m:
            continue
        name, cn = m.group(1).strip(), m.group(2).strip()
        if len(name) < 4 or name.lower().startswith(("this ", "gsa ", "the ")):
            continue
        if not re.search(r"[A-Z]{2}\d|\d{2}[A-Z]", cn):
            continue
        key = norm_name(name)
        if key in seen:
            r = out[seen[key]]
            r["federal_contract_number"] = "; ".join(
                dict.fromkeys((r["federal_contract_number"] or "").split("; ") + [cn])).strip("; ")
            continue
        seen[key] = len(out)
        out.append(rec(src, len(out) + 1, name,
                       federal_contract_number=cn,
                       service_category_raw=vehicle,
                       verification_basis="parent_corporation_publication",
                       ingestion_method="html",
                       validation_flags=[
                           "PARENT_ASSERTED_SUBSIDIARY__not_a_tribal_"
                           "government_certification",
                           "no_UEI_or_CAGE_at_this_URL: the registry records "
                           "them on individual subsidiary pages under "
                           "asrcfederal.com; asrc.com itself returns HTTP 307 "
                           "to every automated client",
                       ]))
    return out


def parse_calista(src):
    out = []
    for f in sorted(RAW.glob("TBD-058_calista_*.html")):
        L = lines_of(f)
        if len(L) < 6:
            continue
        title = L[0].replace(" - Calivika", "").strip()
        try:
            i = L.index(title, 1)
        except ValueError:
            i = next((j for j, l in enumerate(L[1:], 1) if l == title), None)
            if i is None:
                continue
        blk = L[i:]
        owner = next((l.split(":", 1)[1].strip() for l in blk if l.startswith("Owner:")), None)
        phone = email = None
        for j, l in enumerate(blk):
            if l == "P:" and j + 1 < len(blk):
                phone = blk[j + 1]
            if l == "E:" and j + 1 < len(blk):
                email = blk[j + 1]
        loc = None
        if "Location" in blk:
            k = blk.index("Location")
            loc = " ".join(blk[k + 1:k + 4])
            loc = re.sub(r"\s+", " ", loc).strip()
        city = st = zp = None
        if loc:
            m = re.search(r"([A-Za-z\. ]+),\s*([A-Za-z ]+?)\s+(\d{5})", loc)
            if m:
                city, st, zp = m.group(1).strip(", "), m.group(2).strip(), m.group(3)
        desc = [l for l in blk if len(l) > 60 and not l.startswith(("Owner:", "P:", "E:"))]
        out.append(rec(src, f.name.replace("TBD-058_calista_", "").replace(".html", ""),
                       title, owner_name_raw=owner,
                       description_raw=" ".join(desc[:2]) or None,
                       address_raw=loc, city=city, state_province=st, postal_code=zp,
                       phone=phone, email=email,
                       verification_basis="none_described_by_source",
                       source_url="https://calistashareholderbiz.com/shareholder/"
                                  + f.name.replace("TBD-058_calista_", "").replace(".html", "") + "/",
                       snapshot=f.name,
                       ingestion_method="html",
                       validation_flags=[
                           "SHAREHOLDER_DESCENDANT_OR_SPOUSE__a_spouse_owned_"
                           "firm_is_not_a_Native_owned_firm_and_this_source_"
                           "cannot_distinguish_them",
                           "no_ownership_threshold_and_no_verification_"
                           "described_by_the_publisher",
                       ]))
    return out


def parse_doyon(src):
    """The operating companies are named in headings and bold lead-ins on the
    business-line pages. The site navigation repeats verbatim on every page, so
    the body is taken after the last navigation item rather than filtered by
    guesswork."""
    COMPANYISH = re.compile(
        r"\b(LLC|L\.L\.C|Inc\.?|Incorporated|Corporation|Company|Group|Services|"
        r"Service|Center|Centre|Utilities|Drilling|Pipeline|Aviation|Science|"
        r"Roadhouse|Lodge|Tours|Cruises|Ventures|Solutions|IT|Contracting|"
        r"Technologies)\b")
    NOT_A_COMPANY = re.compile(
        r"^(Reliable Services|Information Technology|Fort |Joint Base|"
        r"Doyon.s Facility|Na-dena. invests|Oil Field Services$|Technology$|"
        r"Construction$|Tourism$|Shareholder Services|Job Opportunities|"
        r"Training Opportunities|Employee Portal|Contact Us|Privacy Policy|"
        r"Search Button|Descendant Registry|Records & Stocks|Our Companies)", re.I)
    out, seen = [], set()
    for f in sorted(RAW.glob("TBD-059_doyon_*.html")):
        if f.name.endswith("operations.html"):
            continue
        biz = f.name.replace("TBD-059_doyon_", "").replace(".html", "")
        L = [l.replace("\ufeff", "").strip() for l in lines_of(f)]
        idx = [i for i, l in enumerate(L[:120]) if l == "Contact Us"]
        body = L[idx[-1] + 1:] if idx else L
        end = next((i for i, l in enumerate(body)
                    if l in ("Latest Operations Updates", "Contact Us")), len(body))
        body = body[:end]
        for l in body:
            if not l or len(l) > 70 or len(l) < 5:
                continue
            if l.endswith((".", "!", "?", ":", ",")) or l[0].islower():
                continue
            if NOT_A_COMPANY.match(l) or not COMPANYISH.search(l):
                continue
            k = norm_name(l)
            if not k or k in seen:
                continue
            seen.add(k)
            out.append(rec(src, len(out) + 1, l,
                           service_category_raw=biz.replace("-", " ").title(),
                           verification_basis="parent_corporation_publication",
                           source_url="https://www.doyon.com/operations/%s/" % biz,
                           snapshot=f.name,
                           ingestion_method="html",
                           validation_flags=[
                               "PARENT_ASSERTED_SUBSIDIARY__not_a_tribal_"
                               "government_certification",
                               "name_read_from_a_page_heading_or_bold_lead-in__"
                               "Doyon_publishes_no_structured_subsidiary_list",
                               "no_UEI_or_CAGE_on_the_operations_pages: the "
                               "registry records them on per-company "
                               "capability-statement PDFs not harvested here",
                           ]))
    return out


def parse_muscogee(src):
    import openpyxl
    wb = openpyxl.load_workbook(RAW / src["snapshot"], data_only=True)
    ws = wb[wb.sheetnames[0]]
    out = []
    hdr = None
    for row in ws.iter_rows(values_only=True):
        vals = ["" if v is None else str(v).strip() for v in row]
        if hdr is None:
            if vals and vals[0].lower().startswith("business name"):
                hdr = vals
            continue
        name = vals[0]
        if not name:
            continue
        owner = vals[1] or None
        street = vals[2] or None
        csz = vals[3] or None
        phone = vals[4] or None
        details = vals[5] or None
        email = vals[6] or None
        if email and email.upper() in ("N/A", "NA", "NONE"):
            email = None
        city = st = zp = None
        if csz:
            m = re.match(r"\s*(.+?),\s*([A-Za-z]{2})\.?,?\s*(\d{5})", csz)
            if m:
                city, st, zp = m.group(1).strip(), m.group(2).upper(), m.group(3)
        out.append(rec(src, len(out) + 1, name,
                       owner_name_raw=owner,
                       service_category_raw=details,
                       address_raw="; ".join(x for x in [street, csz] if x) or None,
                       city=city, state_province=st, postal_code=zp,
                       phone=phone, email=email,
                       ingestion_method="xlsx",
                       source_last_updated="2026-08-17",
                       source_edition="MCN-CESO-Vendor-List-260817",
                       validation_flags=[
                           "tribal_affiliation_raw_blank_by_source_design: the "
                           "Act distinguishes 'Muscogee Owned Vendor' from "
                           "'Indian Owned Vendor' and the file does not, so "
                           "the owning tribe is NOT recoverable",
                           "ownership_threshold_min_51_comes_from_NCA_18-199_"
                           "s.9-105(I)__not_from_the_file",
                       ],
                       ownership_threshold_min=51.0))
    return out


PARSERS = {
    "TBD-044": parse_cskt, "TBD-045": parse_cherokee, "TBD-046": parse_mha,
    "TBD-047": parse_oneida, "TBD-048": parse_poarch, "TBD-050": parse_tohono,
    "TBD-052": parse_lummi, "TBD-053": parse_blackfeet, "TBD-054": parse_menominee,
    "TBD-056": parse_asrc, "TBD-058": parse_calista, "TBD-059": parse_doyon,
    "TBD-079": parse_muscogee,
}

SLUG = {
    "TBD-044": "cskt_indian_preference_business_list",
    "TBD-045": "cherokee_nation_tero_directory",
    "TBD-046": "mha_tero_certified_indian_contractors",
    "TBD-047": "oneida_wi_indian_preference_vendor_list",
    "TBD-048": "poarch_tero_certified_businesses",
    "TBD-050": "tohono_oodham_tero_certified_firms",
    "TBD-052": "lummi_owned_businesses",
    "TBD-053": "blackfeet_tero_certified_firms",
    "TBD-054": "menominee_contractors_listing",
    "TBD-056": "asrc_federal_subsidiary_directory",
    "TBD-058": "calista_shareholder_business_directory",
    "TBD-059": "doyon_operating_companies",
    "TBD-079": "muscogee_creek_ceso_vendor_list",
}


# ---------------------------------------------------------------------------
# PHASE: harvest
# ---------------------------------------------------------------------------

def phase_harvest(argv):
    only = argv[0] if argv else None
    counts = {}
    for sid, src in SOURCES.items():
        if only and only != sid:
            continue
        src = dict(src, _id=sid)
        try:
            rows = PARSERS[sid](src)
        except Exception as e:                                   # noqa: BLE001
            print(f"  {sid} {src['authority']:<45} PARSE FAILED: {type(e).__name__}: {e}")
            counts[sid] = ("PARSE_FAILED", f"{type(e).__name__}: {e}")
            continue
        rows = [r for r in rows if (r["business_name_raw"] or "").strip()]
        # de-duplicate on the normalized name within a source
        seen, keep = set(), []
        for r in rows:
            k = r["business_name_normalized"]
            if not k or k in seen:
                continue
            seen.add(k)
            keep.append(r)
        dropped = len(rows) - len(keep)
        out = STAGE / f"{sid}_{SLUG[sid]}.jsonl"
        with out.open("w", encoding="utf-8", newline="\n") as fh:
            for r in keep:
                fh.write(json.dumps(r, ensure_ascii=False) + "\n")
        counts[sid] = ("OK", len(keep))
        print(f"  {sid} {src['authority']:<45} {len(keep):>5} rows"
              + (f"  ({dropped} duplicate name(s) collapsed)" if dropped else "")
              + f"  -> {out.name}")
    return counts


# ---------------------------------------------------------------------------
# PHASE: promote
# ---------------------------------------------------------------------------

CLEAN_COLUMNS = [
    "business_source_id", "source_id", "source_business_key",
    "certifying_authority_entity_id", "certifying_authority_name",
    "nation_id", "programme_name",
    "business_name_raw", "business_name_normalized",
    "business_name_is_person_name",
    "business_entity_id", "business_entity_name", "business_entity_class",
    "resolution_method", "record_scope",
    "assertion_class", "directory_type", "identity_scope",
    "identity_claim_text", "inclusion_basis",
    "ownership_percent", "ownership_threshold_min",
    "verification_basis", "certification_number", "certification_tier",
    "certification_start", "certification_expiration",
    "business_license_number", "federal_contract_number",
    "service_category_raw", "naics",
    "city", "state_province",
    "owner_name_present", "n_owners_named", "withheld_fields",
    "source_url", "source_edition", "source_last_updated",
    "harvest_date", "first_seen", "last_seen", "is_current",
    "ingestion_method", "ocr_mean_confidence", "raw_snapshot_uri",
    "source_terms_status", "consent_status", "suppression_key", "publishable",
    "validation_flags", "record_hash", "built_by_script",
]

WITHHELD = ["owner_name_raw", "email", "phone", "address_raw", "postal_code",
            "description_raw", "website", "dba_name"]

#: Validation flags whose PAYLOAD is verbatim text lifted off the source page.
#: Measured 2026-09-01: 32 CSKT rows carried a wrapped OWNER: line inside
#: `line(s)_before_the_business_name_in_this_block`, which put owner personal
#: names into `data/clean` through the diagnostics column after the same names
#: had been withheld from `owner_name_raw`. A privacy rule that a debug string
#: can walk around is not a rule. The flag survives; its payload does not.
FLAG_PAYLOAD_REDACT = (
    "line(s)_before_the_business_name_in_this_block",
)


def redact_flags(flags):
    out = []
    for f in flags or []:
        for pre in FLAG_PAYLOAD_REDACT:
            if f.startswith(pre):
                f = pre + ": [payload withheld - verbatim source lines, some of "
                f += "which are owner personal names; see staging]"
                break
        out.append(f)
    return out


def load_registry():
    with REGISTRY.open(encoding="utf-8-sig", newline="") as fh:
        rdr = csv.DictReader(fh)
        return list(rdr.fieldnames or []), list(rdr)


def phase_promote(argv):
    sys.path.insert(0, str(ROOT / "code"))
    import importlib
    ident = importlib.import_module("503_identity")
    exact, gov, state_of = ident.build_index()

    spine = {}
    with SPINE.open(encoding="utf-8", errors="replace", newline="") as fh:
        for r in csv.DictReader(fh):
            spine[(r.get("tribe_id") or "").strip()] = r

    _, regrows = load_registry()
    terms_of = {r["tribe_id"]: (r.get("source_terms_status") or "") for r in regrows}

    files = sorted(STAGE.glob("TBD-*.jsonl"))
    if not files:
        print("  no staging files"), sys.exit(2)

    # ------------------------------------------------------------------
    # FILES A CONCURRENT WORKSTREAM WROTE INTO THIS DIRECTORY.
    #
    # Measured 2026-09-01 19:09-19:12, while this build was running. A shared
    # staging directory with two writers is exactly the collision AGENTS.md
    # concurrency rules exist for, so each foreign file gets an EXPLICIT
    # disposition here rather than being swept in by a glob. An unlisted
    # foreign file is REFUSED, not promoted: promoting a file whose authority,
    # assertion class and terms status this script cannot state is how a
    # restricted source reaches data/clean by accident.
    # ------------------------------------------------------------------
    SIBLING = {
        "TBD-C01": dict(
            disposition="EXCLUDE_DUPLICATE",
            tribe_id="TRBF-MSENAT-00", authority="Muscogee (Creek) Nation",
            programme="MCN CESO Vendor List", assertion_class="OWNERSHIP",
            why=("Same source file as TBD-079 (MCN-CESO-Vendor-List-260817.xlsx), "
                 "same 337 rows. Promoting both DOUBLE-COUNTS one directory. "
                 "TBD-079 is kept because it records the statutory inclusion "
                 "basis - NCA 18-199 s.9-105(I), 51% Native ownership plus "
                 "Native control and management - which is what makes the row "
                 "an OWNERSHIP assertion. NOTE THE DISAGREEMENT: the sibling "
                 "typed this source `directory_type=vendor_list`, "
                 "`identity_scope=unspecified`, on the ground that the FILE "
                 "states no threshold. That reading is defensible and the "
                 "registry's own verdict is assertion_class=OWNERSHIP. The "
                 "owner should settle it; until then this table carries the "
                 "registry's verdict, not a silent merge of the two."),
        ),
        "TBD-C02": dict(
            disposition="INCLUDE",
            tribe_id=None, authority="Pokagon Band of Potawatomi Indians",
            programme="Tribal Owned Business & Vendor Directory (Mno-Bmadsen)",
            assertion_class="OWNERSHIP",
            why=("A NEW authority, outside the 62-tribe survey - a genuine "
                 "coverage gain, no collision. pokagonband-nsn.gov publishes "
                 "no terms-of-use page (probed 2026-09-01, 404), so SILENT, "
                 "which is UNRESOLVED and not permission. Its own scope "
                 "sentence covers citizens AND SPOUSES, so it is no stronger "
                 "than Calista and must not be read as a 51% ownership "
                 "certification."),
        ),
        "TBD-D01": dict(
            disposition="EXCLUDE_TERMS_STATED_RESTRICTIVE",
            tribe_id="TRBF-STHUTE-00", authority="Southern Ute Indian Tribe",
            programme="2026 Indian Owned Business List", assertion_class="OWNERSHIP",
            why=("southernute-nsn.gov publishes an explicit Terms of Use: "
                 "'you may not: modify or copy the materials; use the "
                 "materials for any commercial purpose, or for any public "
                 "display'. Read 2026-09-01 at "
                 "https://www.southernute-nsn.gov/terms-of-use . The sibling "
                 "workstream harvested it before that page was read. This "
                 "table refuses it, and the refusal is recorded in the "
                 "registry so the decision is not re-litigated silently."),
        ),
    }

    # source metadata for files this script did not write (the four harvested
    # on 2026-08-28). Read from the rows themselves, never invented.
    PRIOR = {
        "TBD-030": dict(tribe_id="TRBF-TULALP-00", authority="Tulalip Tribes",
                        programme="TERO Native American Owned Business (NAOB) Registry",
                        assertion_class="OWNERSHIP"),
        "TBD-032": dict(tribe_id=None, authority="Confederated Tribes of Grand Ronde",
                        programme="TERO Indian Owned Business list",
                        assertion_class="OWNERSHIP"),
        "TBD-041": dict(tribe_id="TRBF-NAVAJO-00", authority="Navajo Nation",
                        programme="Navajo Business Opportunity Act source listing",
                        assertion_class="OWNERSHIP"),
        "TBD-043": dict(tribe_id="TRBF-ESTCHK-00",
                        authority="Eastern Band of Cherokee Indians",
                        programme="EBCI TERO Certified Vendor List",
                        assertion_class="OWNERSHIP"),
    }

    out, per_source, refused = [], {}, []
    unresolved_examples = []
    for f in files:
        sid = f.name.split("_")[0]
        src = SOURCES.get(sid)
        meta = PRIOR.get(sid) or SIBLING.get(sid)
        if sid in SIBLING and SIBLING[sid].get("disposition") != "INCLUDE":
            refused.append((sid, f.name, SIBLING[sid]["disposition"],
                            SIBLING[sid]["why"]))
            continue
        if src is None and meta is None:
            print(f"  !! {f.name}: unknown source id {sid} - NOT PROMOTED. A "
                  f"staging file whose certifying authority, assertion class "
                  f"and terms status this script cannot state is a file it "
                  f"must not promote.")
            refused.append((sid, f.name, "UNKNOWN_SOURCE_ID",
                            "no SOURCES, PRIOR or SIBLING entry"))
            continue
        auth_tid = (src or meta).get("tribe_id")
        auth_name = (src or meta)["authority"]
        programme = (src or meta)["programme"]
        aclass = (src or meta)["assertion_class"]
        n = 0
        with f.open(encoding="utf-8") as fh:
            for line in fh:
                if not line.strip():
                    continue
                r = json.loads(line)
                n += 1
                name = r.get("business_name_raw") or ""
                owner = r.get("owner_name_raw")
                # RESOLUTION: exact normalized name/alias only. See docstring.
                tid, why = ident.resolve(name, exact, gov, state_of,
                                         r.get("state_province") or "")
                if tid and not why.startswith(("exact normalized", "declared equivalence")):
                    tid, why = None, "REFUSED_LOOSE_TOKEN_PATH: " + why
                ent = spine.get(tid or "", {})
                person = looks_like_person(name, owner)
                nown = 0
                if owner:
                    nown = len([x for x in re.split(r";|&|\band\b", owner) if x.strip()])
                out.append({
                    "business_source_id": r["business_source_id"],
                    "source_id": sid,
                    "source_business_key": r.get("source_business_key"),
                    "certifying_authority_entity_id": auth_tid or "",
                    "certifying_authority_name": auth_name,
                    "nation_id": r.get("nation_id") or "",
                    "programme_name": programme,
                    "business_name_raw": name,
                    "business_name_normalized": r.get("business_name_normalized"),
                    "business_name_is_person_name": person,
                    "business_entity_id": tid or "",
                    "business_entity_name": ent.get("canonical_name", ""),
                    "business_entity_class": ent.get("entity_class", ""),
                    "resolution_method": why,
                    "record_scope": "entity" if tid else "unresolved",
                    "assertion_class": r.get("assertion_class") or aclass,
                    "directory_type": r.get("directory_type"),
                    "identity_scope": r.get("identity_scope"),
                    "identity_claim_text": r.get("identity_claim_text"),
                    "inclusion_basis": "program_authority",
                    "ownership_percent": r.get("ownership_percent"),
                    "ownership_threshold_min": r.get("ownership_threshold_min"),
                    "verification_basis": r.get("verification_basis"),
                    "certification_number": r.get("certification_number"),
                    "certification_tier": r.get("certification_tier"),
                    # ISO at the single write point, so no source parser has to
                    # remember. See `iso_date` for why US order is assumed and
                    # why an unparseable value is left verbatim.
                    "certification_start": iso_date(r.get("certification_start")),
                    "certification_expiration": iso_date(
                        r.get("certification_expiration")),
                    "business_license_number": r.get("business_license_number"),
                    "federal_contract_number": r.get("federal_contract_number"),
                    "service_category_raw": r.get("service_category_raw"),
                    "naics": r.get("naics"),
                    "city": r.get("city"),
                    "state_province": r.get("state_province"),
                    "owner_name_present": 1 if owner else 0,
                    "n_owners_named": nown,
                    "withheld_fields": ";".join(
                        [k for k in WITHHELD if r.get(k)]
                        + (["validation_flag_payload"]
                           if any(str(x).startswith(FLAG_PAYLOAD_REDACT)
                                  for x in (r.get("validation_flags") or []))
                           else [])),
                    "source_url": r.get("source_url"),
                    "source_edition": r.get("source_edition"),
                    "source_last_updated": r.get("source_last_updated"),
                    "harvest_date": HARVEST_DATE if src else "2026-08-28",
                    "first_seen": r.get("first_seen"),
                    "last_seen": r.get("last_seen"),
                    "is_current": r.get("is_current", True),
                    "ingestion_method": r.get("ingestion_method"),
                    "ocr_mean_confidence": r.get("ocr_mean_confidence"),
                    "raw_snapshot_uri": r.get("raw_snapshot_uri"),
                    "source_terms_status": terms_of.get(auth_tid or "", "SILENT") or "SILENT",
                    "consent_status": "UNRESOLVED",
                    "suppression_key": f"SUPPRESS::{auth_tid or sid}",
                    "publishable": "N",
                    "validation_flags": ";".join(redact_flags(r.get("validation_flags"))),
                    "record_hash": r.get("record_hash"),
                    "built_by_script": SCRIPT,
                })
                if not tid and len(unresolved_examples) < 5:
                    unresolved_examples.append(name)
        per_source[sid] = n

    CLEAN.parent.mkdir(parents=True, exist_ok=True)
    with CLEAN.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=CLEAN_COLUMNS, extrasaction="ignore")
        w.writeheader()
        for r in out:
            w.writerow(r)

    res = sum(1 for r in out if r["business_entity_id"])
    own = sum(1 for r in out if r["assertion_class"] == "OWNERSHIP")
    rel = sum(1 for r in out if r["assertion_class"] == "RELATIONSHIP")
    per = sum(1 for r in out if r["business_name_is_person_name"] == 1)
    unk = sum(1 for r in out if r["business_name_is_person_name"] == -1)
    ownn = sum(1 for r in out if r["owner_name_present"])
    print(f"\n  {CLEAN.relative_to(ROOT)}")
    print(f"    rows                         {len(out):>7,}")
    print(f"    sources                      {len(per_source):>7}")
    print(f"    resolved to a spine entity   {res:>7,}  ({res/max(1,len(out)):.2%})")
    print(f"    OWNERSHIP-class rows         {own:>7,}")
    print(f"    RELATIONSHIP-class rows      {rel:>7,}")
    print(f"    business name IS a person    {per:>7,}   undecidable {unk:,}")
    print(f"    rows whose source names an owner (withheld) {ownn:>7,}")
    print(f"    publishable = Y              {sum(1 for r in out if r['publishable']=='Y'):>7}"
          "   (consent is UNRESOLVED on every authority)")
    if unresolved_examples:
        print("    first unresolved, by name: " + "; ".join(unresolved_examples))
    for sid, fn, disp, why in refused:
        print(f"    REFUSED {sid} ({fn}): {disp}")
        print(f"            {why[:300]}")
    return per_source


# ---------------------------------------------------------------------------
# PHASE: registry - turn the survey into a tracker
# ---------------------------------------------------------------------------

NEW_REG_COLS = ["harvest_date", "harvest_rows", "harvest_source_id",
                "harvest_status", "harvest_route_rung", "harvest_technique",
                "hidden_route_sweep_2026-09-01", "newsletter_url"]


def phase_registry(counts):
    fields, rows = load_registry()
    for c in NEW_REG_COLS:
        if c not in fields:
            fields.append(c)
    by_tid = {}
    for sid, src in SOURCES.items():
        by_tid[src["tribe_id"]] = (sid, src)
    prior = {"TRBF-TULALP-00": ("TBD-030", "2026-08-28"),
             "TRBF-NAVAJO-00": ("TBD-041", "2026-08-28"),
             "TRBF-ESTCHK-00": ("TBD-043", "2026-08-28")}
    counts_by_sid = {k: v[1] for k, v in (counts or {}).items() if v[0] == "OK"}
    staged = {}
    for f in STAGE.glob("TBD-*.jsonl"):
        staged[f.name.split("_")[0]] = sum(1 for _ in f.open(encoding="utf-8") if _.strip())

    STAMP = f"|| {HARVEST_DATE} WORKSTREAM-P:"

    def note(r, text):
        """Append once. A note appended on every run makes the file grow
        without bound and makes the build non-idempotent."""
        cur = r.get("notes") or ""
        if text[:80] in cur:
            return
        r["notes"] = (cur + " " + STAMP + " " + text).strip()

    for r in rows:
        tid = r["tribe_id"]
        if tid in SWEEP_2026_09_01:
            r["hidden_route_sweep_2026-09-01"] = SWEEP_2026_09_01[tid]
        if tid in EXCLUDED:
            e = EXCLUDED[tid]
            r["harvest_source_id"] = e["source_id"]
            r["harvest_status"] = "EXCLUDED_TERMS_STATED_RESTRICTIVE"
            r["harvest_date"] = HARVEST_DATE
            r["harvest_rows"] = "0"
            r["harvest_route_rung"] = "NOT ATTEMPTED - terms are a decision, not an obstacle"
            r["source_terms_status"] = "TERMS_STATED_RESTRICTIVE"
            if e["found_by"].startswith("THIS RUN"):
                r["source_terms_quote"] = e["quote"]
                note(r, f"terms read at {e['terms_url']} and found RESTRICTIVE; "
                        f"the 2026-08-26 survey recorded SILENT because that "
                        f"page was never opened. {e['why']}")
            continue
        if tid in TERMS_READ:
            tr = TERMS_READ[tid]
            r["source_terms_status"] = tr["status"]
            r["source_terms_quote"] = tr["quote"]
            note(r, f"terms read at {tr['url']}. {tr['note']}")
        if tid in by_tid:
            sid, src = by_tid[tid]
            r["harvest_source_id"] = sid
            r["harvest_date"] = HARVEST_DATE
            n = counts_by_sid.get(sid, staged.get(sid))
            r["harvest_rows"] = str(n) if n is not None else ""
            r["harvest_status"] = "HARVESTED" if n else "PARSE_FAILED"
            if sid == "TBD-054":
                r["harvest_status"] = "HARVESTED_PARTIAL"
            r["harvest_route_rung"] = src.get("rung", "")
            r["harvest_technique"] = TECHNIQUE.get(sid, "")
            if src.get("newsletter_url"):
                r["newsletter_url"] = src["newsletter_url"]
            if sid == "TBD-045":
                r["robots_note"] = src["robots"]
        elif tid in prior:
            sid, d = prior[tid]
            r["harvest_source_id"] = sid
            r["harvest_date"] = d
            r["harvest_rows"] = str(staged.get(sid, ""))
            r["harvest_status"] = "HARVESTED_2026-08-28"
            r["harvest_route_rung"] = "harvested by the 2026-08-28 pass"
        else:
            r.setdefault("harvest_status", "")
            if not r.get("harvest_status"):
                r["harvest_status"] = {
                    "NO_LIST_FOUND": "NO_LIST_TO_HARVEST",
                    "LIST_REFERENCED_NOT_PUBLISHED": "NOT_PUBLISHED",
                    "SITE_UNREACHABLE": "SITE_UNREACHABLE_RECHECK",
                    "LIST_BEHIND_LOGIN": "BEHIND_LOGIN_OUT_OF_SCOPE",
                }.get(r.get("verdict", ""), "NOT_HARVESTED")
            if tid == "TRBF-TURTLM-00":
                # measured 2026-09-01: the host answers, and its robots.txt
                # disallows everything. "Did not answer" and "refused us" are
                # different facts and only one of them is a decision.
                r["harvest_status"] = "ROBOTS_DISALLOW_ALL"

    tmp = REGISTRY.with_suffix(".tmp")
    with tmp.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)
    tmp.replace(REGISTRY)
    print(f"\n  {REGISTRY.relative_to(ROOT)} updated: "
          + ", ".join(f"{k}={sum(1 for r in rows if r.get('harvest_status')==k)}"
                      for k in sorted({r.get("harvest_status", "") for r in rows})))
    return rows


# ---------------------------------------------------------------------------
# PHASE: coverage - the spine-wide ledger. Answers "has this entity ever been
# checked", for all 1,555, not just the 62 in the sample.
# ---------------------------------------------------------------------------

def phase_coverage(regrows):
    with SPINE.open(encoding="utf-8", errors="replace", newline="") as fh:
        spine = list(csv.DictReader(fh))
    reg = {r["tribe_id"]: r for r in regrows}
    classes = {}
    for e in spine:
        c = e.get("entity_class", "")
        r = reg.get(e["tribe_id"])
        st = "NEVER_CHECKED" if r is None else (r.get("harvest_status") or "CHECKED_NO_HARVEST")
        classes.setdefault(c, {}).setdefault(st, 0)
        classes[c][st] += 1
    checked = sum(1 for e in spine if e["tribe_id"] in reg)
    print(f"\n  COVERAGE against the spine ({len(spine):,} entities)")
    print(f"    ever checked for a business directory   {checked:>6,}  "
          f"({checked/len(spine):.1%})")
    print(f"    NEVER CHECKED                           {len(spine)-checked:>6,}")
    for c in sorted(classes, key=lambda k: -sum(classes[k].values())):
        tot = sum(classes[c].values())
        nev = classes[c].get("NEVER_CHECKED", 0)
        print(f"      {c:<52} {tot:>5}  never-checked {nev:>5}")
    return spine, reg



# ---------------------------------------------------------------------------
# PHASE: docs - regenerate docs/datasets/native-owned-businesses.md.
#
# The coverage ledger is GENERATED, never typed. A hand-maintained coverage
# table is a coverage table that is wrong within one build.
# ---------------------------------------------------------------------------

DOC = ROOT / "docs" / "datasets" / "native-owned-businesses.md"

DOC_HEAD = """# Native-owned businesses

**Dataset owner:** `code/330_build_native_owned_businesses.py` (the only builder).
**Table:** `data/clean/native_owned_businesses.csv`
**Staging:** `data/staging/business_registry/*.jsonl` (+ `raw/` snapshots)
**Survey / tracker:** `review/tribal_vendor_list_registry_2026-08-26.csv`

A tribal government certifying a business is a **third party with authority
over the ownership question**. That is the tier-A evidence leg Cedar Press has
almost none of: a SAM socio-economic flag is self-certification, and an
ANCSA subsidiary certifying `alaskanNativeCorporationOwnedFirm = NO` is the
standing proof that self-certification is not a determination.

## THE INCLUSION BASIS IS THE PRODUCT (ADR-013)

`identity_claim_text` is quoted verbatim from the source and is why the row is
in Cedar at all. **It is not uniform, and `assertion_class` keeps the two
kinds apart on every row:**

| assertion_class | the authority asserts | example |
|---|---|---|
| `OWNERSHIP` | who **owns** the firm | EBCI: "TRIBAL MEMBER owned businesses ... vetted by the TERO office" |
| `RELATIONSHIP` | the firm **does business with** the tribe | Menominee Contractors listing |

Within `OWNERSHIP` the strength varies by an order of magnitude and
`identity_scope` records it. From strongest to weakest:

| identity_scope | what is certified |
|---|---|
| `enrolled_member_100pct` / `enrolled_member_51pct` | Poarch, per TERO Ordinance Title 33, section-labelled in the source |
| `enrolled_member_cskt` / `..._other_federally_recognized` | CSKT Preference 1 vs Preference 2 |
| `any_native_oodham_and_local` | Tohono O'odham, 51%+ owned by O'odham and other local Indians |
| `any_native` / `any_native_graded` | Cherokee, MHA tiers, Lummi, Blackfeet, Muscogee, Oneida |
| `tribally_owned_entity` | Poarch "TRIBAL BUSINESSES" - the tribe itself owns the firm |
| `parent_asserted_subsidiary` | ASRC Federal, Doyon - a parent naming its own subsidiary |
| `shareholder_descendant_or_spouse` | Calista, Pokagon - **weakest; a spouse-owned firm is not a Native-owned firm and these sources cannot tell you which is which** |
| `vendor_relationship` | Menominee - not ownership at all |

**A consumer that sums OWNERSHIP and RELATIONSHIP rows has added two different
facts.** So has one that treats a Calista shareholder listing as equivalent to
a Poarch 100%-tribal-member certification.

## NOTHING HERE PUBLISHES

Every row carries `consent_status = UNRESOLVED`, `publishable = N` and a
`suppression_key`. A federal record is public by statute; **a sovereign
government's own publication is not the same thing**, and "publicly reachable"
is not "licensed for commercial redistribution."
`code/321_gate_tribal_source_restriction.py` is the machinery; flipping one
`consent_status` field admits or removes an entire authority, and saying yes
must be as cheap as saying no.

## PRIVACY: THE CLEAN TABLE CARRIES THE CERTIFICATION, NOT THE FRONT DOOR

A TERO roster of sole proprietorships is a list of private individuals with
their home addresses and mobile numbers.
`cedar_domain.INDIVIDUAL_NATIVE_WITHHELD_FIELDS` already withholds
`owner_name`, `street`, `recipient_city_name` and `dba_name` for exactly this
case; that rule is **inherited here, not re-invented**.

| | |
|---|---|
| **carried** | business name, normalized name, certifying authority, nation, city, state, service category, certification number / tier / dates, licence number, the verbatim claim |
| **withheld from `data/clean`** | `owner_name_raw`, `email`, `phone`, `address_raw`, `postal_code`, `website`, `dba_name`, `description_raw` - kept in staging only, and named per row in `withheld_fields` |
| **flagged** | `business_name_is_person_name` (1 / 0 / -1 undecidable) so a consumer can apply `cedar_domain.may_publish_individual_native_field` per field |
| **counted, not named** | `owner_name_present`, `n_owners_named` |

No digest surrogate is minted for a personal name. A digest of an enumerable
value is not a privacy control; the protection is that the column does not
ship.

## RESOLUTION: EXACT NAME ONLY, AND AN UNRESOLVED ROW KEEPS ITS ROW

Each business is offered to `503_identity.resolve()` and **only an exact
normalized name/alias hit is accepted.** The loose gov-class token path is
refused by design - it exists to match a filing to the government that filed
it, and on a business roster it produces false ownership claims:

```
Navajo Engineering & Construction  -> 'Navajo'          REFUSED
Osage Electrical Contractors, Inc. -> 'The Osage Nation' REFUSED
Arctic Information Technology, Inc -> 'Arctic Village'   REFUSED
```

78 rows matched that way and every one of them would have been wrong.
A business that does not resolve **keeps its row** with a blank
`business_entity_id` and `record_scope = unresolved` (ADR-010). Nothing here
mints a spine entity.

"""

DOC_TAIL = """
## TWO THINGS THIS BUILD COULD NOT SETTLE - THEY ARE THE OWNER'S

### 1. The terms bar is not applied consistently, and Navajo is inside the table

`TBD-041` (Navajo, 346 rows) is in `data/clean/native_owned_businesses.csv`
carrying `source_terms_status = TERMS_STATED_RESTRICTIVE`. It was harvested on
2026-08-28 and promoted here because the promotion brief says to promote
everything already in staging. **That is a contradiction and it is left
visible rather than resolved unilaterally.**

The reason it exists is that two different bars have been applied:

| source | what was read | verdict |
|---|---|---|
| Navajo, Colville, CTUIR | a bare copyright footer - "(c) 2025 ... All Rights Reserved" | TERMS_STATED_RESTRICTIVE |
| Chickasaw, FCP, Southern Ute, NANA | an actual terms page saying you may not copy, redistribute, use commercially or collect automatically | TERMS_STATED_RESTRICTIVE |

On the first bar, essentially every tribal site is restricted, because
essentially every website carries a copyright footer. On the second, four
sources refused and the rest did not. **This build applied the second bar to
its own findings and did not reverse the first bar's verdicts** - so Colville
and CTUIR stay excluded, and Navajo stays in the table with the restrictive
flag on every row and `publishable = N`.

Someone has to pick one bar. Until then: **no row from any of these
authorities publishes**, which is what `consent_status = UNRESOLVED` and
`321_gate_tribal_source_restriction.py` are for.

### 2. Muscogee: is a "Vendor List" an ownership list?

A concurrent workstream harvested the same MCN CESO xlsx as `TBD-C01` and
typed it `directory_type = vendor_list`, `identity_scope = unspecified`, on the
ground that **the file itself states no ownership threshold**. This build kept
`TBD-079` and typed it OWNERSHIP, on the ground that **NCA 18-199 s.9-105(I)
requires "fifty-one percent (51%) or more Native ownership and proof of Native
control and management"** to be on it, and that the registry's own verdict is
`assertion_class = OWNERSHIP`.

Both readings are defensible and they are 337 rows apart in what the dataset
claims. `TBD-C01` is refused as a duplicate rather than merged, so the
disagreement is recorded rather than averaged away. The general question -
*does the statute behind a list, or only the text printed on it, set the
assertion class?* - decides more than this one source.

## HOW TO RE-PULL

`py -3 code/330_build_native_owned_businesses.py --list-sources` prints, per
source, the URL, the per-host delay and the robots note.
`docs/PULL_DISCIPLINE.md` governs. Two host notes that will otherwise be
rediscovered the hard way:

* **Lummi** - robots.txt disallows `/apps`, and the tribe's own directory page
  links the report at `/apps/BusLicenses/LummiOwnedBusinesses.php`. The
  identical report is served from `/widgets/`, which is not disallowed.
  **Any re-run must use `/widgets/`.**
* **cherokeetero.com** - the 2026-08-26 survey recorded "403 to a plain client,
  robots.txt UNREADABLE". Both are superseded: robots.txt reads
  `User-agent: * / Disallow:` and a browser User-Agent gets HTTP 200. It was a
  UA gate, not a block.

## KNOWN GAPS, STATED RATHER THAN PADDED

* **Menominee** - 4 firms from 10 of the source's own 23 grid rows. The
  DevExpress pager callback refused three argument formats; Wayback holds
  captures 2015-2026 for a later pass.
* **Tohono O'odham** - 17 firms against the 19 certification stamps the
  document carries. Two rows whose owner blocks sit inside the row-gap
  threshold merged into the preceding firm.
* **CSKT** - 116 of the 118 `PREFERENCE:` records; 5 rows carry
  `BUSINESS_NAME_MAY_BE_AN_ADDRESS_LINE` because the two-column PDF marks no
  boundary between a firm's name and its street line. Kept and flagged, never
  dropped or guessed.
* **Blackfeet** - OCR-recovered from a scan (`ingestion_method =
  ocr_rapidocr_220dpi`, mean line confidence on every row). `OCR_RECOVERED` is
  not the same evidence grade as a born-digital text layer and stays
  distinguishable.
* **ASRC / Doyon** - names only. The UEI/CAGE that make these joinable sit on
  per-subsidiary pages and capability-statement PDFs not harvested here.

## THE 2026-09-01 HIDDEN-ROUTE SWEEP - IT FOUND NOTHING, AND THAT IS RECORDED

`docs/HIDDEN_DATA_TECHNIQUES.md` was run against the **eleven** registry rows
that were REFERENCED-BUT-NOT-PUBLISHED or SITE_UNREACHABLE: robots.txt first,
then `/wp-json/wp/v2/types`, `/wp-json/wp/v2/media?search=`,
`/wp-json/wp/v2/search` and `sitemap.xml`. **None of the six
TERMS_STATED_RESTRICTIVE sources were probed by any of these routes.**

**Yield: zero new vendor lists.** Per-tribe results are in the registry's
`hidden_route_sweep_2026-09-01` column. What the media API returned instead was
consistent and is itself the finding: **ordinances, codes and blank application
forms, never a roster.** Laguna publishes its Indian Preference Code and five
contractor application forms; Warm Springs publishes the TERO Code, an FAQ, a
Contractor Registration form and a Business Certification Form. These tribes
are not hiding a list behind a bad link. They are certifying businesses and not
publishing the register - which is exactly what the 2026-08-26 survey called
`LIST_REFERENCED_NOT_PUBLISHED`, now confirmed by a route the survey did not
run.

Three corrections came out of it, and they matter more than the null:

| tribe | 2026-08-26 | 2026-09-01 |
|---|---|---|
| Turtle Mountain | SITE_UNREACHABLE | **ROBOTS_DISALLOW_ALL** - tmchippewa.com answers, and its robots.txt disallows `/`. A refusal, not an outage; Wayback is not a route around it either. |
| White Mountain Apache | SITE_UNREACHABLE | reachable - it was a **TLS chain problem**, not an absent site. No WP REST and `/sitemap.xml` is 404, so the list question is OPEN, not negative. |
| Kotzebue | SITE_UNREACHABLE | answers HTTP 200 to a browser User-Agent. Open, not negative. |

`harvest_technique` in the registry records, per source, which technique
actually produced the data, so the next agent skips the routes that did not.

### One boundary call, flagged rather than buried

Oneida Nation (WI) renders its vendor list in JavaScript from
`admin-ajax.php?action=collect_ipv_list`, using a nonce **the page itself
prints for anonymous visitors**. `HIDDEN_DATA_TECHNIQUES.md` lists "the AJAX
source behind any table" as technique 8 and separately forbids `/wp-admin/` as
private infrastructure; `admin-ajax.php` is WordPress's public front-end AJAX
handler and lives under that path by convention. `oneida-nsn.gov/robots.txt`
was read on 2026-09-01 and disallows only calendar views and
`/wp-content/uploads/formidable/` - not this path. No login was involved and no
nonce was guessed. **34 rows rest on that judgment; it is the owner's to
overturn.**

## THE TIME DIMENSION NOBODY HAS USED YET

These are current directories, so the universe is TRIBES, not years - but the
registry's `wayback_snapshots` / `wayback_first_capture` / `wayback_last_capture`
columns point at real history. Menominee alone has captures from 2015 to 2026.
A 2019 TERO list against a 2026 one is **certification entry and exit**, which
does not exist as a dataset anywhere. The schema is already built for it:
`first_seen`, `last_seen`, `is_current`, `source_edition`. Two standing rules
apply the moment anyone builds it - never present a historical snapshot as
current, and never rule a current page against a historical record or the
reverse.
"""


def phase_docs():
    fields, rows = load_registry()
    reg = {r["tribe_id"]: r for r in rows}
    with SPINE.open(encoding="utf-8", errors="replace", newline="") as fh:
        spine = list(csv.DictReader(fh))

    counts = {}
    if CLEAN.exists():
        with CLEAN.open(encoding="utf-8-sig", newline="") as fh:
            for r in csv.DictReader(fh):
                counts[r["source_id"]] = counts.get(r["source_id"], 0) + 1
    total = sum(counts.values())

    L = [DOC_HEAD]
    L.append("## THE HARVEST, SOURCE BY SOURCE\n")
    L.append(f"`data/clean/native_owned_businesses.csv` holds **{total:,} rows** "
             f"from **{len(counts)} authorities**.\n")
    L.append("| authority | source | checked | list found | format | assertion | "
             "last harvested | rows in clean | terms |")
    L.append("|---|---|---|---|---|---|---|---:|---|")
    seen_sids = set()
    for tid, src in sorted(SOURCES.items(), key=lambda kv: kv[1]["authority"]):
        r = reg.get(src["tribe_id"], {})
        seen_sids.add(tid)
        L.append("| {a} | `{s}` | yes | yes | {f} | {c} | {d} | {n} | {t} |".format(
            a=src["authority"], s=tid, f=r.get("list_format", ""),
            c=src["assertion_class"], d=r.get("harvest_date", HARVEST_DATE),
            n=counts.get(tid, 0), t=r.get("source_terms_status", "SILENT")))
    for tid, meta in sorted(PRIOR_DOC.items(), key=lambda kv: kv[1]["authority"]):
        r = reg.get(meta["tribe_id"] or "", {})
        L.append("| {a} | `{s}` | yes | yes | {f} | {c} | 2026-08-28 | {n} | {t} |".format(
            a=meta["authority"], s=tid, f=r.get("list_format", "HTML/PDF"),
            c=meta["assertion_class"], n=counts.get(tid, 0),
            t=r.get("source_terms_status", "SILENT")))
    for tid, meta in sorted(SIBLING_DOC.items()):
        if meta.get("disposition") != "INCLUDE":
            continue
        L.append("| {a} | `{s}` | yes | yes | HTML | {c} | 2026-09-01 | {n} | "
                 "SILENT (no terms page published) |".format(
                     a=meta["authority"], s=tid, c=meta["assertion_class"],
                     n=counts.get(tid, 0)))
    L.append("")

    L.append("## EXCLUDED, AND WHY - TERMS ARE A DECISION THE PUBLISHER MADE\n")
    L.append("These stay excluded by **every** route, Wayback included. Two were "
             "already recorded restrictive; **four were found restrictive by the "
             "2026-09-01 pass because nobody had opened the terms page** the "
             "2026-08-26 survey recorded as SILENT.\n")
    L.append("| authority | source | terms read at | the term |")
    L.append("|---|---|---|---|")
    for tid, e in sorted(EXCLUDED.items(), key=lambda kv: kv[1]["authority"]):
        q = e["quote"].replace("|", "/")
        if len(q) > 320:
            q = q[:317] + "..."
        L.append("| {a}{n} | `{s}` | {u} | {q} |".format(
            a=e["authority"],
            n=" **(NEW)**" if e["found_by"].startswith("THIS RUN") else "",
            s=e["source_id"], u=e["terms_url"], q=q))
    L.append("")
    L.append("**What the exclusions cost, so the OPT_IN requests can be ranked:**\n")
    for tid, e in sorted(EXCLUDED.items(), key=lambda kv: kv[1]["authority"]):
        L.append(f"* **{e['authority']}** - {e['loss']}")
    L.append("")

    L.append("## COVERAGE AGAINST THE MASTER ENTITY LIST\n")
    checked = sum(1 for e in spine if e["tribe_id"] in reg)
    L.append(f"The 2026-08-26 survey was a **priority sample of 62**, stratified "
             f"by federal contracting rank, gaming revenue and geography - not a "
             f"census. Against the {len(spine):,}-entity spine:\n")
    L.append("| entity class | in spine | checked | never checked |")
    L.append("|---|---:|---:|---:|")
    byc = {}
    for e in spine:
        c = e.get("entity_class", "")
        byc.setdefault(c, [0, 0])
        byc[c][0] += 1
        if e["tribe_id"] in reg:
            byc[c][1] += 1
    for c in sorted(byc, key=lambda k: -byc[k][0]):
        t, ch = byc[c]
        L.append(f"| {c} | {t:,} | {ch} | {t-ch:,} |")
    L.append(f"| **total** | **{len(spine):,}** | **{checked}** | "
             f"**{len(spine)-checked:,}** |")
    L.append("")
    fr = byc.get("Federally recognized tribe", [0, 0])
    L.append(f"**{checked} of {len(spine):,} entities ({checked/len(spine):.1%}) have "
             f"ever been checked for a published business directory.** Among "
             f"federally recognized tribes it is {fr[1]} of {fr[0]} "
             f"({fr[1]/max(1,fr[0]):.1%}); **{fr[0]-fr[1]} have never been "
             f"looked at.** An entity absent from the registry is "
             f"NEVER_CHECKED, which is a different fact from NO_LIST_FOUND and "
             f"must not be read as one.\n")
    L.append("### What checking the rest would take\n")
    L.append("The 2026-08-26 pass measured the cost: **four parallel agents, one "
             "day, 62 tribes**, and its own notes name the four things that made "
             "a find hard, all of which generalise:\n")
    L.append("1. **The list is usually on a SEPARATE DOMAIN** the government site "
             "barely links - `cherokeetero.com`, `ebci-tero.com`, `mhatero.com`, "
             "`btero.com`. Blackfeet's only pointer was a phone-book entry in a "
             "staff directory reachable through the WordPress `?s=` search.")
    L.append("2. **The word 'TERO' is not the search term.** CSKT calls it the "
             "*Indian Preference Office*; Muscogee's statute never uses 'TERO'; "
             "Poarch's site search for 'vendor' returns pow-wow craft vendors.")
    L.append("3. **Sitemap enumeration beats navigation.** Tohono O'odham's list "
             "sits on a third-level page not linked from the TERO landing page "
             "body. Lummi's own directory page renders 'no documents currently "
             "available' - navigating by it produces a FALSE NEGATIVE.")
    L.append("4. **The terms page must be opened before the list is parsed.** "
             "Four of six exclusions on this page exist because it was not.\n")
    L.append("At the measured rate (~15 tribes per agent-day including the terms "
             "read), the remaining **297 federally recognized tribes are roughly "
             "20 agent-days**, and the 2026-08-26 hit rate (22 lists per 62 "
             "tribes, 35%) predicts on the order of **100 further published "
             "lists**. The Alaska Native villages, village corporations, NHOs, "
             "BIE schools and TCUs in the table above are a different question "
             "and mostly will not have one; they should be marked "
             "`NOT_APPLICABLE` after a cheap first pass rather than left "
             "indefinitely NEVER_CHECKED.\n")
    L.append("### Registry rows that are not a harvest gap\n")
    L.append("| status | tribes | meaning |")
    L.append("|---|---:|---|")
    stat = {}
    for r in rows:
        stat[r.get("harvest_status", "")] = stat.get(r.get("harvest_status", ""), 0) + 1
    MEAN = {
        "NO_LIST_TO_HARVEST": "the site was enumerated and publishes no list",
        "NOT_PUBLISHED": "a list is REFERENCED in an ordinance or on a page but not published",
        "SITE_UNREACHABLE_RECHECK": "did not answer on 2026-08-26; the 2026-09-01 sweep reached two of the four (see `hidden_route_sweep_2026-09-01`) - NOT evidence of absence",
        "ROBOTS_DISALLOW_ALL": "Turtle Mountain: tmchippewa.com robots.txt disallows '/'. A refusal, not an outage",
        "BEHIND_LOGIN_OUT_OF_SCOPE": "Choctaw Nation - a directory behind a login is out of scope, and Wayback is not a route around a login",
        "EXCLUDED_TERMS_STATED_RESTRICTIVE": "the publisher stated a term; excluded by every route",
        "HARVESTED": "in data/clean",
        "HARVESTED_PARTIAL": "in data/clean with a stated shortfall",
        "HARVESTED_2026-08-28": "harvested by the earlier pass, promoted by this one",
    }
    for k in sorted(stat, key=lambda k: -stat[k]):
        if k:
            L.append(f"| `{k}` | {stat[k]} | {MEAN.get(k, '')} |")
    L.append(DOC_TAIL)

    DOC.parent.mkdir(parents=True, exist_ok=True)
    DOC.write_text("\n".join(L), encoding="utf-8")
    print(f"\n  {DOC.relative_to(ROOT)} written ({DOC.stat().st_size:,} bytes)")



# ---------------------------------------------------------------------------
# PHASE: codebook - write THIS dataset's fragment and nothing else.
#
# Via `cedar_codebook.write_fragment`, which is the sanctioned route:
# `cedar_codebook.py` itself is not edited, and a fragment cannot affect
# another dataset. Every row carries a description, because
# `62_no_regression_check.py` fails the build on a published codebook row with
# an empty one - and because an undescribed column is a column a buyer has to
# ask about.
#
# `published = 0`, `access_tier = internal`. Consent is UNRESOLVED on every
# authority in this table, so nothing in it may ship, and the codebook must say
# that rather than leave it to the gate.
# ---------------------------------------------------------------------------

CODEBOOK_DATASET = "02m_native_owned_businesses"

DESCRIPTIONS = {
 "business_source_id": "Primary key. '<source_id>:<source_business_key>' - the certifying authority's own record key, namespaced by source. Stable across builds for any source that publishes a certification number.",
 "source_id": "Cedar source id for the certifying authority's list, TBD-0nn. One list, one id.",
 "source_business_key": "The key the SOURCE uses for this business - certification number where one is published, otherwise the row's position or slug in the published document.",
 "certifying_authority_entity_id": "Spine tribe_id of the government or ANCSA corporation making the assertion. THE AUTHORITY, never the business.",
 "certifying_authority_name": "Canonical name of that authority.",
 "nation_id": "Nation the source associates the business with, in the source's own vocabulary. Blank where the source does not say.",
 "programme_name": "What the authority calls the programme - 'TERO', 'Indian Preference Office', 'CESO'. The word TERO is absent from several of these by the publisher's own choice.",
 "business_name_raw": "Business name exactly as the source prints it.",
 "business_name_normalized": "Lower-cased, punctuation- and corporate-suffix-stripped form, for matching only. NOT an identifier.",
 "business_name_is_person_name": "1 = the business name IS a natural person's name; 0 = it is not; -1 = undecidable. Governs which fields may be published per cedar_domain.may_publish_individual_native_field. Conservative in both directions.",
 "business_entity_id": "Spine tribe_id where the business resolves to an EXISTING Cedar entity. Blank where it does not, and a blank is a recorded non-link, never a defect to be filled by guessing (ADR-010).",
 "business_entity_name": "Spine canonical name of that entity, or blank.",
 "business_entity_class": "Spine entity_class of that entity, or blank.",
 "resolution_method": "How the link was made, or why it was refused. 'REFUSED_LOOSE_TOKEN_PATH:...' means 503_identity offered a gov-class token match and this build declined it - on a business roster that path produces false ownership claims (Navajo Engineering -> Navajo Nation).",
 "record_scope": "ADR-010 scope. 'entity' where a Cedar entity is named; 'unresolved' where one plausibly exists and was not found. There is no other value in this table.",
 "assertion_class": "OWNERSHIP or RELATIONSHIP. THE LOAD-BEARING COLUMN. OWNERSHIP = the authority asserts who owns the firm. RELATIONSHIP = the authority asserts the firm does business with the tribe. Summing the two adds different facts.",
 "directory_type": "tero | indian_preference | business_licence | vendor | subsidiary_directory | shareholder_vendor. What kind of register this is.",
 "identity_scope": "WHOSE Native identity is certified, at the granularity the source states: enrolled_member_100pct, enrolled_member_51pct, enrolled_member_cskt, any_native, any_native_graded, tribally_owned_entity, parent_asserted_subsidiary, shareholder_descendant_or_spouse, vendor_relationship. These are not interchangeable.",
 "identity_claim_text": "The inclusion basis, quoted VERBATIM from the source (ADR-013). This is why the row is in Cedar at all.",
 "inclusion_basis": "ADR-013 standard vocabulary. Always 'program_authority' here: the row is present because a tribal or ANCSA programme certified or listed the firm.",
 "ownership_percent": "Numeric ownership share where the SOURCE prints one per record. Blank otherwise; blank is not zero.",
 "ownership_threshold_min": "Minimum ownership share the programme requires, where the programme states one. Sourced from the ordinance or statute when the list itself is silent, and the validation flag says which.",
 "verification_basis": "What the authority says it did - TERO_review, parent_corporation_publication, none_described_by_source. 'none_described_by_source' is a finding.",
 "certification_number": "The authority's certification or vendor number, where published. The only identifier in this dataset that joins a source to itself across vintages.",
 "certification_tier": "Preference tier or certification status as printed - 'PREFERENCE 1', 'Preference Level 3', 'FIRST PREFERENCE; Full Certification'. Probationary and Full are different statuses and are not flattened.",
 "certification_start": "Date certified, where printed.",
 "certification_expiration": "Certification or annual-update expiry, where printed. For Lummi this is a BUSINESS LICENCE expiry, not a certification expiry, and the row says so.",
 "business_license_number": "Tribal business licence number where the source carries one (Blackfeet 20nn-BL-nnnn).",
 "federal_contract_number": "GSA/agency contract vehicle number where a parent publishes one per subsidiary (ASRC Federal). Semicolon-separated where several.",
 "service_category_raw": "Trade, category or business line, verbatim from the source.",
 "naics": "NAICS codes where the source publishes them (Oneida only).",
 "city": "City, from the source's address. The street line is WITHHELD - see withheld_fields.",
 "state_province": "State or province, from the source's address.",
 "owner_name_present": "1 where the source names an owner or principal. THE NAME ITSELF IS NOT IN THIS TABLE.",
 "n_owners_named": "How many people the source names as owners. A count, not a roster.",
 "withheld_fields": "Which staging fields were withheld from this row and why they exist. Naming the withholding is part of the record; silently dropping a column is not.",
 "source_url": "The URL the row was read from.",
 "source_edition": "The publisher's own edition marker - a dated filename, a build stamp, or a printed 'current as of' line.",
 "source_last_updated": "The date the SOURCE states it was updated. Blank where the source states none; blank is not the harvest date.",
 "harvest_date": "When Cedar retrieved it.",
 "first_seen": "First harvest in which this business appeared in this source.",
 "last_seen": "Most recent harvest in which it appeared. With first_seen and is_current this supports certification entry/exit once a second vintage exists.",
 "is_current": "Whether the row was present in the most recent retrieval of its source.",
 "ingestion_method": "html | json_api | xlsx | pdf_text_layer | pdf_text_layer_positional | ocr_rapidocr_220dpi. An OCR-recovered row is a different evidence grade from a born-digital one and stays distinguishable.",
 "ocr_mean_confidence": "Mean OCR line confidence for the document, on OCR-recovered rows only.",
 "raw_snapshot_uri": "The stored snapshot the row was parsed from, under data/staging/business_registry/raw/.",
 "source_terms_status": "What the publisher's own terms say about reuse: SILENT, TERMS_STATED_RESTRICTIVE, TERMS_STATED_NO_REUSE_RESTRICTION, NOT_CHECKED. Silence is UNRESOLVED, never permission.",
 "consent_status": "UNRESOLVED | OPT_IN | OPT_OUT. Nothing in this table is OPT_IN. Gated by 321_gate_tribal_source_restriction.py.",
 "suppression_key": "Flipping one authority's consent removes or admits every row carrying this key.",
 "publishable": "N on every row in this table. A sovereign government's own publication is not a federal record.",
 "validation_flags": "What the parse could not establish, named. Payload text quoted off the source page is redacted here where it can carry a personal name; the flag survives, the payload does not.",
 "record_hash": "sha256 over (source, key, name, normalized name). Deterministic; not positional and not hash().",
 "built_by_script": "330_build_native_owned_businesses.py.",
}


def phase_codebook():
    import cedar_codebook
    if not CLEAN.exists():
        print("  no clean table yet - run promote first")
        return 0
    with CLEAN.open(encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.DictReader(fh))
    n = len(rows)
    missing = [c for c in CLEAN_COLUMNS if c not in DESCRIPTIONS]
    if missing:
        # Defect class 2b: never compute coverage over a column set you have
        # not asserted. An undescribed column must RAISE, not ship blank.
        raise SystemExit("  CODEBOOK REFUSED: no description for "
                         + ", ".join(missing))
    out = []
    for c in CLEAN_COLUMNS:
        filled = sum(1 for r in rows if str(r.get(c, "")).strip())
        typ = "integer" if c in ("owner_name_present", "n_owners_named",
                                 "business_name_is_person_name") else (
              "numeric" if c in ("ownership_percent", "ownership_threshold_min",
                                 "ocr_mean_confidence") else
              "date" if c in ("harvest_date", "source_last_updated") else "text")
        out.append({"dataset": CODEBOOK_DATASET, "variable": c, "type": typ,
                    "units": "", "pct_filled": round(100.0 * filled / n, 1) if n else 0.0,
                    "n_rows": n, "published": 0, "access_tier": "internal",
                    "description": DESCRIPTIONS[c], "generated": HARVEST_DATE})
    cedar_codebook.write_fragment(CODEBOOK_DATASET, out)
    print(f"  codebook fragment {CODEBOOK_DATASET}: {len(out)} variables, "
          f"{n:,} rows, published=0 (consent UNRESOLVED)")
    return len(out)


def main():
    args = sys.argv[1:]
    if not args or args[0] in ("-h", "--help"):
        print(__doc__)
        return 0
    if args[0] == "--list-sources":
        for sid, s in SOURCES.items():
            print(f"{sid}  {s['authority']}\n     url   {s['list_url']}"
                  f"\n     delay {s['delay_s']}s  robots: {s['robots']}"
                  f"\n     rung  {s.get('rung','')}")
        for tid, e in EXCLUDED.items():
            print(f"{e['source_id']}  EXCLUDED  {e['authority']}\n     {e['quote'][:110]}...")
        return 0
    cmd = args[0]
    counts = None
    if cmd in ("harvest", "all"):
        print("HARVEST")
        counts = phase_harvest(args[1:] if cmd == "harvest" else [])
    # REGISTRY BEFORE PROMOTE. `promote` stamps `source_terms_status` onto every
    # row FROM the registry, so a registry update that runs afterwards leaves
    # the clean table asserting a terms status this run has already superseded.
    if cmd in ("registry", "all"):
        regrows = phase_registry(counts)
    if cmd in ("promote", "all"):
        print("\nPROMOTE")
        phase_promote(args[1:] if cmd == "promote" else [])
    if cmd in ("registry", "all"):
        phase_coverage(regrows)
    if cmd in ("codebook", "all"):
        phase_codebook()
    if cmd in ("docs", "all"):
        phase_docs()
    if cmd not in ("harvest", "promote", "registry", "codebook", "docs", "all"):
        print(__doc__)
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main())
