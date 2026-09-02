"""SHARD-G: build the institutional REGISTRY CROSSWALK for the 315-entity slice.

Consumes the bulk registry objects pulled by code/shard_g_registry_pull.py plus the
2026-08-06 caches, and emits durable registry keys per Cedar entity.

  NCES/CCD  ccd_sch_029_2425 directory  -> NCESSCH (12-digit), LEAID, SY_STATUS,
            SCH_TYPE, grades, WEBSITE, address
            ccd_sch_052_2425 membership -> enrollment (Education Unit Total only)
  IPEDS     HD2023 (TRIBAL=1) -> UNITID, OPEID, EIN, UEI, WEBADDR, HLOFFER,
            LANDGRNT, INSTSIZE ; DRVEF2023 -> total enrollment
  CDFI Fund List of Certified CDFIs (2026-08-14) -> Cert Control Num, Date
            Certified, FI Type, RSSD ID, Native flag, website
  NCUA      FOICU.txt 2026Q1 -> CU_NUMBER (charter), RSSD, IsMDI, LIMITED_INC
  FDIC      api.fdic.gov active institutions -> CERT, FED_RSSD, WEBADDR, CLASS
  CICD      Fed Minneapolis NAFI map (support) -> bank_cert / cu_number / rssd

MATCHING. Registry names and Cedar canonical names differ. Every link records
`match_method` and, where fuzzy, `match_score`. Rules, strictest first:
  exact_name            normalised names identical
  exact_alias           normalised name equals one of the spine aliases
  name_and_place        name similarity >= 0.86 AND same state AND same city
  name_and_state        name similarity >= 0.92 AND same state
  core_and_place        after dropping institution-generic words (school, day,
                        community, high, junior, boarding, district, ...) the two
                        residual token sets are equal or one contains the other,
                        AND state AND city agree, AND the residual is not empty.
                        Uniqueness is enforced: if two slice rows survive this
                        rule for one registry row, the row is left unmatched.
Anything weaker is written to _unmatched_* and left for a human. NOTHING is
guessed: an entity with no registry row gets a row saying so.

NEVER writes to the spine. NEVER mints. Reads the spine; writes only under
data/staging/institution_registry/ and data/staging/tribe_harvest/shard_g/.
"""
from __future__ import annotations

import csv, difflib, io, json, re, subprocess, sys, unicodedata, zipfile
from datetime import date, datetime, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "data" / "staging" / "tribe_harvest" / "shard_g" / "raw"
OUTREG = ROOT / "data" / "staging" / "institution_registry"
OUTH = ROOT / "data" / "staging" / "tribe_harvest" / "shard_g"
EXT = ROOT / "data" / "raw" / "external"
SCRATCH = Path(r"C:\Users\esm247\AppData\Local\Temp\claude\C--Users-esm247"
               r"\6f0cc363-573d-4f3a-97b1-c84e32f43c8b\scratchpad\shard_g")
TODAY = date.today().isoformat()
csv.field_size_limit(10_000_000)

# Words that carry no discriminating power between two institutions in the SAME
# town: they describe the kind of institution, not which one it is.
GENERIC = {"school", "schools", "day", "boarding", "district", "the", "of", "and",
           "inc", "incorporated", "board", "bureau", "education", "educational",
           "center", "centre", "dormitory", "dorm", "residential", "hall",
           "academy", "learning", "public", "consolidated", "no", "number",
           # NCUA's FOICU.txt stores charter names WITHOUT the words "Federal
           # Credit Union" - e.g. Lakota Federal Credit Union is "LAKOTA".
           "federal", "credit", "union", "fcu"}
# Words that DO discriminate and must never be dropped (grade span, tribal name):
LEVEL = {"elementary", "middle", "high", "junior", "senior", "jr", "sr",
         "primary", "secondary", "preparatory", "prep"}
SUFFIX = re.compile(r"\b(inc|incorporated|llc|l l c|ltd|corp|corporation|co)\b")


ABBR = {"jr": "junior", "sr": "senior", "elem": "elementary", "hs": "high",
        "co": "county", "st": "saint", "univ": "university", "coll": "college",
        "fed": "federal", "cu": "credit union", "assn": "association"}


def norm(s):
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    s = s.lower().replace("&", " and ")
    s = re.sub(r"[\u2018\u2019'`]", "", s)
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = SUFFIX.sub(" ", s)
    s = " ".join(ABBR.get(t, t) for t in s.split())
    return re.sub(r"\s+", " ", s).strip()


def core(s):
    """Residual token set after dropping institution-generic words."""
    return frozenset(t for t in norm(s).split() if t and t not in GENERIC)


def sim(a, b):
    return difflib.SequenceMatcher(None, a, b).ratio()


def rows(p, enc="utf-8-sig"):
    with open(p, encoding=enc, newline="") as f:
        return list(csv.DictReader(f))


# ------------------------------------------------------------------ slice
SLICE = rows(OUTREG / "_slice.csv")
assert len(SLICE) == 315, len(SLICE)
for r in SLICE:
    r["_n"] = norm(r["canonical_name"])
    r["_alias_n"] = {norm(a) for a in (r["aliases"] or "").split("|") if a.strip()}
    r["_core"] = core(r["canonical_name"])
    r["_city"] = norm(r["city"])
    r["_st"] = (r["state"] or "").strip().upper()

BIE = [r for r in SLICE if r["entity_class"] == "BIE School"]
TCU = [r for r in SLICE if r["entity_class"] == "Tribal College or University"]
FIN = [r for r in SLICE if r["entity_class"] in
       ("Native Community Development Financial Institution",
        "Native Financial Institution")]

XW, FACTS, UNMATCHED = [], [], []


def add_id(row, system, value, field, url, quote, method, score=""):
    if value is None:
        return
    v = str(value).strip()
    if not v or v.lower() in ("none", "nan", "not applicable", "unavailable",
                              "-2", "-1", "0", "n/a"):
        return
    XW.append({"cedar_uid": row["cedar_uid"], "tribe_id": row["tribe_id"],
               "canonical_name": row["canonical_name"],
               "entity_class": row["entity_class"], "id_system": system,
               "id_value": v, "id_field": field, "source_url": url,
               "source_quote": quote[:700], "match_method": method,
               "match_score": score, "captured_date": TODAY})


def add_fact(row, attr, value, url, quote, method, score=""):
    if value is None:
        return
    v = str(value).strip()
    if not v or v.lower() in ("none", "nan", "unavailable", "not applicable",
                              "-2", "-1", "n/a", ""):
        return
    FACTS.append({"cedar_uid": row["cedar_uid"], "tribe_id": row["tribe_id"],
                  "canonical_name": row["canonical_name"],
                  "entity_class": row["entity_class"], "attribute": attr,
                  "value": v, "source_url": url, "source_quote": quote[:700],
                  "match_method": method, "match_score": score,
                  "captured_date": TODAY})


def link(pool, name, state="", city=""):
    """Return (row, method, score) or (None, reason, score)."""
    n = norm(name)
    st = (state or "").strip().upper()
    ct = norm(city)
    hits = [r for r in pool if r["_n"] == n]
    if len(hits) == 1:
        return hits[0], "exact_name", "1.000"
    if len(hits) > 1:
        h = [r for r in hits if r["_st"] == st] if st else hits
        if len(h) == 1:
            return h[0], "exact_name_state_tiebreak", "1.000"
        return None, "ambiguous_exact_name", ""
    hits = [r for r in pool if n and n in r["_alias_n"]]
    if len(hits) == 1:
        return hits[0], "exact_alias", "1.000"
    best, bs = None, 0.0
    for r in pool:
        s = sim(n, r["_n"])
        if s > bs:
            best, bs = r, s
    if best is None:
        return None, "no_candidate", ""
    if st and best["_st"] == st and ct and best["_city"] == ct and bs >= 0.86:
        return best, "name_and_place", f"{bs:.3f}"
    if st and best["_st"] == st and bs >= 0.92:
        return best, "name_and_state", f"{bs:.3f}"
    if bs >= 0.95:
        return best, "name_only_high", f"{bs:.3f}"

    # core_and_place: residual token sets equal or nested, place agrees, unique.
    rc = core(name)
    if rc and len("".join(sorted(rc))) >= 4:
        cands = []
        for r in pool:
            if st and r["_st"] and r["_st"] != st:
                continue
            pc = r["_core"]
            if not pc:
                continue
            nested = rc <= pc or pc <= rc
            if not nested:
                # tolerate one mangled/transliterated character in the residual
                a, b = "".join(sorted(rc)), "".join(sorted(pc))
                if not (sim("".join(sorted(rc)), "".join(sorted(pc))) >= 0.90
                        and abs(len(rc) - len(pc)) <= 1):
                    continue
            if ct and r["_city"] and r["_city"] != ct and sim(ct, r["_city"]) < 0.85:
                continue
            cands.append(r)
        if len(cands) == 1:
            r = cands[0]
            return r, "core_and_place", f"{sim(n, r['_n']):.3f}"
        if len(cands) > 1:
            return None, ("ambiguous_core_and_place(" +
                          "; ".join(c["canonical_name"] for c in cands[:4]) + ")"), ""
    return None, f"below_threshold(best={best['canonical_name']})", f"{bs:.3f}"


# ------------------------------------------------------------------ NCES CCD
CCD_DIR_URL = "https://nces.ed.gov/ccd/Data/zip/ccd_sch_029_2425_w_1a_073025.zip"
CCD_MEM_URL = "https://nces.ed.gov/ccd/Data/zip/ccd_sch_052_2425_l_1a_073025.zip"
ccd = []
z = zipfile.ZipFile(RAW / "ccd_sch_directory_2425.zip")
with io.TextIOWrapper(z.open("ccd_sch_029_2425_w_1a_073025.csv"),
                      encoding="utf-8-sig", errors="replace") as f:
    for d in csv.DictReader(f):
        if d.get("FIPST") == "59":
            ccd.append(d)
print(f"CCD 2024-25 BIE (FIPST=59) schools: {len(ccd)}", file=sys.stderr)

# enrollment: Education Unit Total rows only (institution-level; no demographics)
enroll = {}
memf = SCRATCH / "ccd_membership_bie_2425.csv"
if memf.exists():
    for d in rows(memf):
        if (d.get("TOTAL_INDICATOR") == "Education Unit Total"
                and d.get("NCESSCH")):
            v = (d.get("STUDENT_COUNT") or "").strip()
            if v not in ("", "-1", "-2", "-9"):
                enroll[d["NCESSCH"]] = v

nces_linked = 0
for d in ccd:
    r, meth, sc = link(BIE, d["SCH_NAME"], d.get("LSTATE") or d.get("MSTATE"),
                       d.get("LCITY") or d.get("MCITY"))
    if r is None:
        UNMATCHED.append({"registry": "NCES_CCD_2024-25", "registry_name": d["SCH_NAME"],
                          "registry_id": d["NCESSCH"], "state": d.get("LSTATE", ""),
                          "city": d.get("LCITY", ""), "reason": meth, "score": sc})
        continue
    nces_linked += 1
    q = (f"NCES Common Core of Data, Public Elementary/Secondary School Universe "
         f"Survey SY2024-25 v.2a, file ccd_sch_029_2425_w_1a_073025.csv: "
         f"NCESSCH={d['NCESSCH']}; SCH_NAME={d['SCH_NAME']}; LEAID={d['LEAID']}; "
         f"ST_LEAID={d['ST_LEAID']}; STATENAME={d['STATENAME']}; "
         f"SY_STATUS_TEXT={d['SY_STATUS_TEXT']}; SCH_TYPE_TEXT={d['SCH_TYPE_TEXT']}")
    add_id(r, "NCES_SCHOOL_ID", d["NCESSCH"], "NCESSCH", CCD_DIR_URL, q, meth, sc)
    add_id(r, "NCES_LEA_ID", d["LEAID"], "LEAID", CCD_DIR_URL, q, meth, sc)
    add_id(r, "NCES_ST_LEAID", d["ST_LEAID"], "ST_LEAID", CCD_DIR_URL, q, meth, sc)
    for a, k in (("nces_school_name", "SCH_NAME"), ("nces_lea_name", "LEA_NAME"),
                 ("nces_operating_status", "SY_STATUS_TEXT"),
                 ("nces_school_type", "SCH_TYPE_TEXT"),
                 ("nces_school_level", "LEVEL"),
                 ("grade_span_low", "GSLO"), ("grade_span_high", "GSHI"),
                 ("website_registry_nces", "WEBSITE"),
                 ("street_address", "LSTREET1"), ("city", "LCITY"),
                 ("state_code", "LSTATE"), ("zip", "LZIP"), ("phone", "PHONE")):
        add_fact(r, a, d.get(k), CCD_DIR_URL, q, meth, sc)
    if d["NCESSCH"] in enroll:
        add_fact(r, "enrollment_total", enroll[d["NCESSCH"]], CCD_MEM_URL,
                 f"NCES CCD SY2024-25 membership file, NCESSCH={d['NCESSCH']}, "
                 f"TOTAL_INDICATOR='Education Unit Total', "
                 f"STUDENT_COUNT={enroll[d['NCESSCH']]}", meth, sc)
        add_fact(r, "enrollment_year", "2024-2025", CCD_MEM_URL,
                 "NCES CCD SY2024-25 membership file", meth, sc)

# ------------------------------------------------------------------ BIE directory
BIE_URL = ("https://biamaps.geoplatform.gov/BIE-Schools-Directory -> "
           "https://services1.arcgis.com/UxqqIfhng71wUT9x/arcgis/rest/services/"
           "BIE_Schools_Directory/FeatureServer/0/query (retrieved 2026-09-01)")
bd = json.load(open(RAW / "bie_schools_featureserver_2026-09-01.json",
                    encoding="utf-8"))["features"]
bie_linked = 0
for ft in bd:
    a = ft["attributes"]
    r, meth, sc = link(BIE, a.get("School_Name", ""), "", a.get("City"))
    if r is None:
        # BIE also directly operates two postsecondary institutions (Haskell
        # Indian Nations University and SIPI); Cedar classes those as TCUs.
        r, meth, sc = link(TCU, a.get("School_Name", ""), "", a.get("City"))
        if r is not None:
            meth = meth + "_postsecondary_pool"
    if r is None:
        UNMATCHED.append({"registry": "BIE_directory_2026-09-01",
                          "registry_name": a.get("School_Name", ""),
                          "registry_id": str(a.get("OBJECTID")),
                          "state": a.get("State", ""), "city": a.get("City", ""),
                          "reason": meth, "score": sc})
        continue
    bie_linked += 1
    q = (f"Bureau of Indian Education school directory row: "
         f"School_Name={a.get('School_Name')}; Operation_Type={a.get('Operation_Type')}; "
         f"Navajo_Operation={a.get('Navajo_Operation')}; "
         f"Grades_Served={a.get('Grades_Served')}; "
         f"Education_Resource_Center={a.get('Education_Resource_Center')}")
    for at, k in (("bie_operation_type_current", "Operation_Type"),
                  ("bie_navajo_operation", "Navajo_Operation"),
                  ("grades_served", "Grades_Served"),
                  ("bie_education_resource_center", "Education_Resource_Center"),
                  ("website_registry_bie", "website"),
                  ("street_address", "Street_Address"), ("city", "City"),
                  ("state_name", "State"), ("zip", "Zip_Code"),
                  ("phone", "telephone")):
        add_fact(r, at, a.get(k), BIE_URL, q, meth, sc)

# ------------------------------------------------------------------ IPEDS
IPEDS_HD = "https://nces.ed.gov/ipeds/datacenter/data/HD2023.zip"
IPEDS_EF = "https://nces.ed.gov/ipeds/datacenter/data/DRVEF2023.zip"
HLOFFER = {"1": "Award of less than one academic year", "2": "At least 1 but less "
           "than 2 academic yrs", "3": "Associate's degree", "4": "At least 2 but "
           "less than 4 academic yrs", "5": "Bachelor's degree",
           "6": "Postbaccalaureate certificate", "7": "Master's degree",
           "8": "Post-master's certificate", "9": "Doctor's degree"}
LANDG = {"1": "Land Grant Institution", "2": "Not a Land Grant Institution"}
z = zipfile.ZipFile(RAW / "ipeds_hd2023.zip")
with io.TextIOWrapper(z.open("HD2023.csv"), encoding="latin-1") as f:
    hd = list(csv.DictReader(f))
# IPEDS writes UTF-8 BOM BYTES into a latin-1 file, so the first header key
# arrives as the three characters \xef\xbb\xbf - NOT as U+FEFF. lstrip("\ufeff")
# looks correct and does nothing, UNITID then reads as None, and the whole IPEDS
# UnitID column silently fails to land while EIN and OPEID (later columns) work.
# That is exactly the shape of defect a coverage count must never hide.
BOMS = "\ufeff\xef\xbb\xbf"
hd = [{(k.lstrip(BOMS) if k else k): v for k, v in d.items()} for d in hd]
assert "UNITID" in hd[0], f"IPEDS HD header not de-BOMed: {list(hd[0])[:2]}"
tribal = [d for d in hd if d.get("TRIBAL") == "1"]
print(f"IPEDS HD2023 TRIBAL=1: {len(tribal)}", file=sys.stderr)

z = zipfile.ZipFile(RAW / "ipeds_drvef2023.zip")
drv = {}
with io.TextIOWrapper(z.open("drvef2023.csv"), encoding="latin-1") as f:
    for _d in csv.DictReader(f):
        _d = {(k.lstrip(BOMS) if k else k): v for k, v in _d.items()}
        _u = (_d.get("UNITID") or "").strip()
        if _u:
            drv[_u] = _d

# TCUs are matched against IPEDS by name; also try the full HD universe for the two
# TCUs IPEDS does not flag TRIBAL (e.g. new / non-Title-III institutions).
ipeds_linked = 0
matched_tcu = set()
for pool_name, pool in (("TRIBAL=1", tribal), ("full HD2023", hd)):
    for d in pool:
        uid = d.get("UNITID") or ""
        avail = [t for t in TCU if t["cedar_uid"] not in matched_tcu]
        r, meth, sc = link(avail, d.get("INSTNM", ""), d.get("STABBR"), d.get("CITY"))
        if r is None and (d.get("IALIAS") or "").strip():
            for al in re.split(r"[|;,]", d["IALIAS"]):
                al = al.strip()
                if len(al) < 4:
                    continue
                r2, m2, s2 = link(avail, al, d.get("STABBR"), d.get("CITY"))
                if r2 is not None:
                    r, meth, sc = r2, m2 + "_via_ipeds_ialias", s2
                    break
        if r is None:
            if pool_name == "TRIBAL=1":
                UNMATCHED.append({"registry": "IPEDS_HD2023_TRIBAL",
                                  "registry_name": d.get("INSTNM", ""),
                                  "registry_id": uid, "state": d.get("STABBR", ""),
                                  "city": d.get("CITY", ""), "reason": meth,
                                  "score": sc})
            continue
        matched_tcu.add(r["cedar_uid"])
        ipeds_linked += 1
        q = (f"IPEDS Institutional Characteristics directory HD2023: UNITID={uid}; "
             f"INSTNM={d.get('INSTNM')}; STABBR={d.get('STABBR')}; "
             f"TRIBAL={d.get('TRIBAL')} (1=Tribal college); "
             f"LANDGRNT={d.get('LANDGRNT')}; HLOFFER={d.get('HLOFFER')}; "
             f"OPEID={(d.get('OPEID') or '').strip()}")
        add_id(r, "IPEDS_UNITID", uid, "UNITID", IPEDS_HD, q, meth, sc)
        add_id(r, "OPEID", (d.get("OPEID") or "").strip(), "OPEID",
               IPEDS_HD, q, meth, sc)
        add_id(r, "EIN", (d.get("EIN") or "").strip(), "EIN", IPEDS_HD, q, meth, sc)
        add_id(r, "UEI", (d.get("UEIS") or "").strip(), "UEIS", IPEDS_HD, q, meth, sc)
        add_fact(r, "tribal_college_flag_ipeds",
                 "Yes (TRIBAL=1)" if d.get("TRIBAL") == "1" else
                 f"No (TRIBAL={d.get('TRIBAL')})", IPEDS_HD, q, meth, sc)
        add_fact(r, "land_grant_status", LANDG.get(d.get("LANDGRNT", ""), ""),
                 IPEDS_HD, q, meth, sc)
        add_fact(r, "highest_degree_offered", HLOFFER.get(d.get("HLOFFER", ""), ""),
                 IPEDS_HD, q, meth, sc)
        for at, k in (("website_registry_ipeds", "WEBADDR"),
                      ("chief_executive", "CHFNM"), ("chief_executive_title", "CHFTITLE"),
                      ("street_address", "ADDR"), ("city", "CITY"),
                      ("state_code", "STABBR"), ("zip", "ZIP"),
                      ("county_name", "COUNTYNM"), ("phone", "GENTELE"),
                      ("ipeds_institution_name", "INSTNM")):
            add_fact(r, at, (d.get(k) or "").strip(), IPEDS_HD, q, meth, sc)
        e = drv.get(uid)
        if e:
            tot = (e.get("ENRTOT") or "").strip()
            if tot:
                add_fact(r, "enrollment_total", tot, IPEDS_EF,
                         f"IPEDS DRVEF2023 derived enrollment, UNITID={uid}, "
                         f"ENRTOT={tot}, FTE={e.get('FTE','')}", meth, sc)
                add_fact(r, "enrollment_fte", e.get("FTE", ""), IPEDS_EF,
                         f"IPEDS DRVEF2023, UNITID={uid}, FTE={e.get('FTE','')}",
                         meth, sc)
                add_fact(r, "enrollment_year", "2023-24 (IPEDS DRVEF2023)",
                         IPEDS_EF, f"IPEDS DRVEF2023, UNITID={uid}", meth, sc)
    if pool_name == "TRIBAL=1":
        continue

# ------------------------------------------------------------------ CDFI Fund
CDFI_URL = ("https://www.cdfifund.gov/programs-training/certification/cdfi -> "
            "https://www.cdfifund.gov/media/8018681/download?inline "
            "(List of Currently Certified CDFIs, as of August 14, 2026)")
import openpyxl
wb = openpyxl.load_workbook(RAW / "cdfi_certified_list_2026-08-14.xlsx",
                            read_only=True, data_only=True)
ws = wb["List of Certified CDFIs"]
rr = list(ws.iter_rows(values_only=True))
wb.close()
hi = next(i for i, x in enumerate(rr) if x and "Organization Name" in
          [str(c) for c in x if c])
hdr = [str(c).strip() if c else "" for c in rr[hi]]
cdfi_linked = 0
cdfi_asof = next((str(c).strip() for x in rr[:hi] for c in x
                  if c and "Native CDFIs as of" in str(c)), "")
for raw in rr[hi + 1:]:
    d = dict(zip(hdr, raw))
    nm = (str(d.get("Organization Name") or "")).strip().rstrip("`")
    if not nm:
        continue
    r, meth, sc = link(FIN, nm, str(d.get("State") or ""), str(d.get("City") or ""))
    if r is None:
        if str(d.get("Native CDFI (Y/N)") or "").strip().upper() == "Y":
            UNMATCHED.append({"registry": "CDFI_Fund_certified_2026-08-14",
                              "registry_name": nm,
                              "registry_id": str(d.get("Cert Control Num") or ""),
                              "state": str(d.get("State") or ""),
                              "city": str(d.get("City") or ""),
                              "reason": meth + " [Native CDFI=Y]", "score": sc})
        continue
    cdfi_linked += 1
    q = (f"CDFI Fund, List of Currently Certified CDFIs (as of August 14, 2026), "
         f"row: Organization Name={nm}; Cert Control Num={d.get('Cert Control Num')}; "
         f"Date Certified={d.get('Date Certified')}; "
         f"Financial Institution Type={d.get('Financial Institution Type')}; "
         f"Native CDFI (Y/N)={d.get('Native CDFI (Y/N)')}; RSSD ID={d.get('RSSD ID')}")
    add_id(r, "CDFI_FUND_CERT", d.get("Cert Control Num"), "Cert Control Num",
           CDFI_URL, q, meth, sc)
    add_id(r, "FRB_RSSD", d.get("RSSD ID"), "RSSD ID", CDFI_URL, q, meth, sc)
    dc = d.get("Date Certified")
    if dc is not None and str(dc).strip():
        if isinstance(dc, datetime):
            dcv = dc.date().isoformat()
        else:
            try:
                dcv = (datetime(1899, 12, 30) + timedelta(days=int(dc))).date().isoformat()
            except Exception:
                dcv = str(dc)[:10]
        add_fact(r, "cdfi_date_certified", dcv, CDFI_URL, q, meth, sc)
    add_fact(r, "cdfi_certification_status",
             "Certified - appears on the CDFI Fund list of currently certified "
             "CDFIs as of 2026-08-14", CDFI_URL, q, meth, sc)
    add_fact(r, "financial_institution_type", d.get("Financial Institution Type"),
             CDFI_URL, q, meth, sc)
    add_fact(r, "native_cdfi_flag", d.get("Native CDFI (Y/N)"), CDFI_URL, q, meth, sc)
    add_fact(r, "website_registry_cdfi", d.get("Organization Website"),
             CDFI_URL, q, meth, sc)
    for at, k in (("city", "City"), ("state_code", "State"), ("zip", "Zipcode"),
                  ("street_address", "Address")):
        add_fact(r, at, d.get(k), CDFI_URL, q, meth, sc)

# ------------------------------------------------------------------ known ids
# The Fed Minneapolis CICD NAFI map already carries FDIC certificate numbers and
# NCUA charter numbers for the depositories in this slice. Those are STRONG keys:
# joining the regulator's own file on the number is exact, where joining on the
# name is not - the FIN pool contains "First State Bank", "Legacy Bank",
# "Pinnacle Bank", each of which collides with dozens of unrelated institutions
# in the FDIC universe. So: join on the number where we have one, and fall back
# to a UNIQUENESS-ENFORCED name+state match only where we do not.
NAFI_XLSX = EXT / "tcu_cdfi" / "cicd_nafi_map_data_2026-08-06.xlsx"
NAFI_SRC = ("https://github.com/frb-mpls-cde/nafi-map -> data/"
            "nafi-map-data_current.xlsx (Federal Reserve Bank of Minneapolis, "
            "Center for Indian Country Development, Native American Financial "
            "Institutions map; cached 2026-08-06 by code/73_add_tcu_and_cdfi.py. "
            "NOT re-fetched 2026-09-01: github.com/robots.txt disallows the "
            "/*/raw/ path for UA *.)")
KNOWN, nafi_rows = {}, []
if NAFI_XLSX.exists():
    wbn = openpyxl.load_workbook(NAFI_XLSX, read_only=True, data_only=True)
    rrn = list(wbn["data"].iter_rows(values_only=True))
    wbn.close()
    hn = [str(c).strip() if c else "" for c in rrn[0]]
    nafi_rows = [dict(zip(hn, x)) for x in rrn[1:] if x and x[1]]
for d in nafi_rows:
    r, meth, sc = link(FIN, str(d.get("name") or ""), str(d.get("state") or ""),
                       str(d.get("city") or ""))
    if r is None:
        continue
    q = (f"Federal Reserve Bank of Minneapolis CICD Native American Financial "
         f"Institutions map, row: name={d.get('name')}; type={d.get('type')}; "
         f"regulator={d.get('regulator')}; rssd_id={d.get('rssd_id')}; "
         f"bank_cert={d.get('bank_cert')}; cu_number={d.get('cu_number')}; "
         f"nmdi={d.get('nmdi')}; ncdfi={d.get('ncdfi')}")
    for sysn, key in (("FDIC_CERT", "bank_cert"), ("NCUA_CHARTER", "cu_number"),
                      ("FRB_RSSD", "rssd_id")):
        v = str(d.get(key) or "").strip()
        if v and v.lower() not in ("none", "not applicable", "unavailable", ""):
            add_id(r, sysn, v, key, NAFI_SRC, q, meth, sc)
            KNOWN.setdefault(sysn, {})[v.lstrip("0")] = {"cedar_uid": r["cedar_uid"]}
    for at, key in (("financial_institution_type", "type"),
                    ("federal_regulator", "regulator"),
                    ("native_mdi_flag", "nmdi"), ("native_cdfi_flag", "ncdfi"),
                    ("cu_low_income_designated", "cu_lowinc"),
                    ("website_registry_cicd", "website"),
                    ("year_established", "yearest"), ("employees", "employees"),
                    ("credit_union_members", "cu_members"),
                    ("total_assets", "total_assets"), ("total_loans", "total_loans"),
                    ("cdfi_award_2021", "award_2021"),
                    ("cdfi_award_2022", "award_2022"),
                    ("cdfi_award_2024", "award_2024"),
                    ("financials_source", "detail_source"),
                    ("financials_period", "data_period")):
        add_fact(r, at, d.get(key), NAFI_SRC, q, meth, sc)
SLICE_BY_UID = {r["cedar_uid"]: r for r in SLICE}

# A regulator only charters its own kind. Gate the NAME fallback on the declared
# institution type so a loan fund cannot pick up a credit union's charter number
# (it did: "Lakota Fund, Inc., The" nested into NCUA charter 24847, which belongs
# to Lakota Federal Credit Union - a different entity in the same town).
FI_TYPE = {}
for f in FACTS:
    if f["attribute"] == "financial_institution_type":
        FI_TYPE.setdefault(f["cedar_uid"], set()).add(f["value"].strip().lower())


def type_is(uid, kind):
    return any(kind in t for t in FI_TYPE.get(uid, set()))


def unique_name_match(pool, reg_rows, name_of, state_of, city_of):
    """slice_uid -> (reg_row, method, score). Only ONE candidate may survive."""
    out = {}
    for s in pool:
        st = s["_st"]
        cands = [d for d in reg_rows
                 if not st or not state_of(d) or (state_of(d) or "").upper() == st]
        tiers = {"exact_name": [], "exact_alias": [], "core_nested": [],
                 "high_sim": []}
        for d in cands:
            rn = norm(name_of(d))
            if not rn:
                continue
            if rn == s["_n"]:
                tiers["exact_name"].append(d)
            elif rn in s["_alias_n"]:
                tiers["exact_alias"].append(d)
            else:
                rc, pc = core(name_of(d)), s["_core"]
                sm = sim(rn, s["_n"])
                # a nested residual only counts when the SHORTER residual is at
                # least two tokens or >= 6 characters - "BANK" must not nest into
                # "Pinnacle Bank".
                short = rc if len(rc) <= len(pc) else pc
                strong = len(short) >= 2 or len("".join(short)) >= 6
                if (rc <= pc or pc <= rc) and strong and sm >= 0.55:
                    tiers["core_nested"].append(d)
                elif sm >= 0.93:
                    tiers["high_sim"].append(d)
        for tname in ("exact_name", "exact_alias", "core_nested", "high_sim"):
            t = tiers[tname]
            if len(t) == 1:
                out[s["cedar_uid"]] = (t[0], tname + "_unique_in_state",
                                       f"{sim(norm(name_of(t[0])), s['_n']):.3f}")
                break
            if len(t) > 1:
                UNMATCHED.append({
                    "registry": "ambiguity", "registry_name": s["canonical_name"],
                    "registry_id": "", "state": st, "city": "",
                    "reason": (f"{len(t)} candidates tie at tier {tname}: " +
                               "; ".join(norm(name_of(x)) for x in t[:5])),
                    "score": ""})
                break
    return out


# ------------------------------------------------------------------ NCUA
NCUA_URL = ("https://ncua.gov/analysis/credit-union-corporate-call-report-data/"
            "quarterly-data -> https://ncua.gov/files/publications/analysis/"
            "call-report-data-2026-03.zip (FOICU.txt, cycle 2026-03-31)")
ncua_linked = 0
zf = SCRATCH / "ncua_call_report_2026q1.zip"
foicu = []
if zf.exists():
    p = subprocess.run([r"C:\Program Files\7-Zip\7z.exe", "e", "-so", str(zf),
                        "FOICU.txt"], capture_output=True)
    foicu = list(csv.DictReader(io.StringIO(p.stdout.decode("latin-1"))))
if foicu:
    by_charter = {(d.get("CU_NUMBER") or "").strip().lstrip("0"): d for d in foicu}
    pairs = {}
    for cu, x in KNOWN.get("NCUA_CHARTER", {}).items():
        if cu in by_charter:
            pairs[x["cedar_uid"]] = (by_charter[cu], "known_charter_number_join",
                                     "1.000")
    remaining = [s for s in FIN if s["cedar_uid"] not in pairs
                 and type_is(s["cedar_uid"], "credit union")]
    for uid, v in unique_name_match(remaining, foicu,
                                    lambda d: d.get("CU_NAME", ""),
                                    lambda d: d.get("STATE"),
                                    lambda d: d.get("CITY")).items():
        pairs[uid] = v
    for uid, (d, meth, sc) in pairs.items():
        r = SLICE_BY_UID[uid]
        nm = (d.get("CU_NAME") or "").strip()
        ncua_linked += 1
        q = (f"NCUA quarterly call report data 2026-03-31, FOICU.txt row: "
             f"CU_NUMBER={d.get('CU_NUMBER')}; CU_NAME={nm}; RSSD={d.get('RSSD')}; "
             f"CU_TYPE={d.get('CU_TYPE')}; CharterState={d.get('CharterState')}; "
             f"LIMITED_INC={d.get('LIMITED_INC')}; IsMDI={d.get('IsMDI')}")
        add_id(r, "NCUA_CHARTER", d.get("CU_NUMBER"), "CU_NUMBER", NCUA_URL, q, meth, sc)
        add_id(r, "FRB_RSSD", d.get("RSSD"), "RSSD", NCUA_URL, q, meth, sc)
        for at, k in (("ncua_charter_state", "CharterState"),
                      ("ncua_low_income_designated", "LIMITED_INC"),
                      ("ncua_minority_depository_institution", "IsMDI"),
                      ("year_opened", "YEAR_OPENED"), ("city", "CITY"),
                      ("state_code", "STATE"), ("zip", "ZIP_CODE"),
                      ("street_address", "STREET"),
                      ("ncua_regulator_name", "CU_NAME")):
            add_fact(r, at, d.get(k), NCUA_URL, q, meth, sc)

# ------------------------------------------------------------------ FDIC
FDIC_URL = ("https://api.fdic.gov/banks/institutions?filters=ACTIVE:1 "
            "(retrieved 2026-09-01)")
fdic_linked = 0
fj = RAW / "fdic_active_institutions_2026-09-01.json"
if fj.exists():
    fd = [it["data"] for it in json.load(open(fj, encoding="utf-8"))["data"]]
    by_cert = {str(d.get("CERT")).lstrip("0"): d for d in fd}
    pairs = {}
    for cert, x in KNOWN.get("FDIC_CERT", {}).items():
        if cert in by_cert:
            pairs[x["cedar_uid"]] = (by_cert[cert], "known_fdic_cert_join", "1.000")
    remaining = [s for s in FIN if s["cedar_uid"] not in pairs
                 and type_is(s["cedar_uid"], "bank")]
    for uid, v in unique_name_match(remaining, fd, lambda d: d.get("NAME", ""),
                                    lambda d: d.get("STALP"),
                                    lambda d: d.get("CITY")).items():
        pairs[uid] = v
    for uid, (d, meth, sc) in pairs.items():
        r = SLICE_BY_UID[uid]
        fdic_linked += 1
        q = (f"FDIC BankFind institutions API, active institution row: "
             f"NAME={d.get('NAME')}; CERT={d.get('CERT')}; "
             f"FED_RSSD={d.get('FED_RSSD')}; BKCLASS={d.get('BKCLASS')}; "
             f"CITY={d.get('CITY')}, {d.get('STALP')}; "
             f"ESTYMD={d.get('ESTYMD')}; MINRTY={d.get('MINRTY')}")
        add_id(r, "FDIC_CERT", d.get("CERT"), "CERT", FDIC_URL, q, meth, sc)
        add_id(r, "FRB_RSSD", d.get("FED_RSSD"), "FED_RSSD", FDIC_URL, q, meth, sc)
        for at, k in (("fdic_bank_class", "BKCLASS"),
                      ("fdic_minority_status_code", "MINRTY"),
                      ("fdic_established_date", "ESTYMD"),
                      ("fdic_total_assets_thousands", "ASSET"),
                      ("fdic_domestic_offices", "OFFDOM"),
                      ("website_registry_fdic", "WEBADDR"),
                      ("city", "CITY"), ("state_code", "STALP"), ("zip", "ZIP")):
            add_fact(r, at, d.get(k), FDIC_URL, q, meth, sc)

# ------------------------------------------------------------------ AIHEC + spine
AIHEC_SRC = "https://www.aihec.org/tcu-roster-and-profiles/ (cached 2026-08-06)"
ap = EXT / "tcu_cdfi" / "_aihec_parsed.json"
aihec_linked = 0
if ap.exists():
    for d in json.load(open(ap, encoding="utf-8")):
        r, meth, sc = link(TCU, d.get("name", ""), d.get("state"), "")
        if r is None:
            UNMATCHED.append({"registry": "AIHEC_roster", "registry_name":
                              d.get("name", ""), "registry_id": "",
                              "state": d.get("state", ""), "city": "",
                              "reason": meth, "score": sc})
            continue
        aihec_linked += 1
        ev = " ".join(d.get("evidence") or [])[:500]
        q = (f"AIHEC TCU Roster and Profiles, section={d.get('section')}: "
             f"{d.get('name')}. {ev}")
        add_fact(r, "aihec_membership", d.get("section"), AIHEC_SRC, q, meth, sc)
        add_fact(r, "acronym", d.get("acronym"), AIHEC_SRC, q, meth, sc)
        add_fact(r, "chartered_year", d.get("chartered_year"), AIHEC_SRC, q, meth, sc)

# websites the spine already holds, carried into the fact table so the web map
# has a fallback where no registry names a site
for r in SLICE:
    note = r["spine_recon_note"] or ""
    src = r["spine_entity_source_url"] or r["spine_source_url"] or \
        "data/spine/cedar_entity_spine.csv"
    m = re.search(r"website[= ](\S+)", note)
    if m and m.group(1).startswith("http"):
        add_fact(r, "website_spine_note", m.group(1).rstrip(";,"), src,
                 f"cedar_entity_spine.csv reconciliation_note: {note[:300]}",
                 "spine_field")
    if r["spine_entity_website"]:
        add_fact(r, "website_spine", r["spine_entity_website"],
                 r["spine_source_url"] or "data/spine/cedar_entity_spine.csv",
                 f"cedar_entity_spine.csv entity_website for {r['canonical_name']}",
                 "spine_field")

# ------------------------------------------------------------------ coverage
have = {}
for x in XW:
    have.setdefault(x["cedar_uid"], set()).add(x["id_system"])
SYSTEMS_FOR = {
    "BIE School": ["NCES_SCHOOL_ID", "NCES_LEA_ID"],
    "Tribal College or University": ["IPEDS_UNITID", "OPEID", "EIN", "UEI"],
    "Native Community Development Financial Institution":
        ["CDFI_FUND_CERT", "FRB_RSSD", "FDIC_CERT", "NCUA_CHARTER"],
    "Native Financial Institution":
        ["FDIC_CERT", "NCUA_CHARTER", "CDFI_FUND_CERT", "FRB_RSSD"],
}
# PRIMARY is the key that, on its own, is a durable external identifier for that
# class. A Native Financial Institution is by construction NOT a certified CDFI,
# so demanding CDFI_FUND_CERT of it would report a definition as a gap.
PRIMARY_FOR = {
    "BIE School": {"NCES_SCHOOL_ID"},
    "Tribal College or University": {"IPEDS_UNITID"},
    "Native Community Development Financial Institution": {"CDFI_FUND_CERT"},
    "Native Financial Institution": {"FDIC_CERT", "NCUA_CHARTER",
                                     "CDFI_FUND_CERT", "FRB_RSSD"},
}
COV = []
for r in SLICE:
    h = have.get(r["cedar_uid"], set())
    exp = SYSTEMS_FOR[r["entity_class"]]
    COV.append({
        "cedar_uid": r["cedar_uid"], "tribe_id": r["tribe_id"],
        "canonical_name": r["canonical_name"], "entity_class": r["entity_class"],
        "state": r["state"],
        "primary_registry_id_found":
            "1" if h & PRIMARY_FOR[r["entity_class"]] else "0",
        "primary_registry_id_system":
            "|".join(sorted(h & PRIMARY_FOR[r["entity_class"]])),
        "id_systems_found": "|".join(sorted(h)),
        "id_systems_expected": "|".join(exp),
        "id_systems_absent": "|".join(s for s in exp if s not in h),
        "note": ("" if h & PRIMARY_FOR[r["entity_class"]] else
                 "no row for this entity in the authoritative registry searched; "
                 "left blank rather than guessed"),
        "checked_date": TODAY,
    })


def write(path, recs, fields):
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(recs)
    print(f"wrote {path.relative_to(ROOT)} rows={len(recs)}", file=sys.stderr)


XW.sort(key=lambda r: (r["entity_class"], r["canonical_name"], r["id_system"]))
FACTS.sort(key=lambda r: (r["entity_class"], r["canonical_name"], r["attribute"]))
write(OUTREG / "registry_crosswalk.csv", XW,
      ["cedar_uid", "tribe_id", "canonical_name", "entity_class", "id_system",
       "id_value", "id_field", "source_url", "source_quote", "match_method",
       "match_score", "captured_date"])
write(OUTREG / "institution_facts.csv", FACTS,
      ["cedar_uid", "tribe_id", "canonical_name", "entity_class", "attribute",
       "value", "source_url", "source_quote", "match_method", "match_score",
       "captured_date"])
write(OUTREG / "registry_coverage.csv", COV, list(COV[0].keys()))
write(OUTREG / "_registry_rows_unmatched_to_slice.csv", UNMATCHED,
      ["registry", "registry_name", "registry_id", "state", "city", "reason",
       "score"])

st = {
    "script": "code/shard_g_build_crosswalk.py", "run_date": TODAY,
    "slice_entities": len(SLICE),
    "ccd_bie_rows_in_registry": len(ccd),
    "ccd_linked_to_slice": nces_linked,
    "bie_directory_features": len(bd), "bie_directory_linked": bie_linked,
    "ipeds_tribal_flag_rows": len(tribal), "ipeds_linked_to_slice": ipeds_linked,
    "cdfi_fund_linked": cdfi_linked, "cdfi_fund_asof_line": cdfi_asof,
    "ncua_linked": ncua_linked, "fdic_linked": fdic_linked,
    "aihec_linked": aihec_linked, "cicd_nafi_rows": len(nafi_rows),
    "enrollment_rows_bie": len(enroll),
    "ids_by_system": {s: len({x["cedar_uid"] for x in XW if x["id_system"] == s})
                      for s in sorted({x["id_system"] for x in XW})},
    "entities_with_primary_registry_id":
        sum(1 for c in COV if c["primary_registry_id_found"] == "1"),
    "unmatched_registry_rows": len(UNMATCHED),
}
(OUTH / "_crosswalk_state.json").write_text(json.dumps(st, indent=2), encoding="utf-8")
print(json.dumps(st, indent=2))
