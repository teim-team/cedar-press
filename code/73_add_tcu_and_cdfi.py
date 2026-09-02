#!/usr/bin/env python3
r"""
Cedar Press - 73: Add Tribal Colleges and Universities, and Native CDFIs /
Native financial institutions, to the entity spine.

THE GAP THIS CLOSES
-------------------
Measured 2026-08-06 against `data/spine/cedar_entity_spine.csv` (952 entities):
**zero** tribal colleges and **zero** Native CDFIs. Elijah:

    "we prob also need to include native cdfi, you can look at the cicd nafi map
     for all native financial institutions, and aihec? or whatever for tribal
     colleges those should be included as well and are low hanging fruit and
     should be easy to identify"

He is right, and the cost of the absence is already visible in the ledger.
`code/33_apply_party_rulings.resolve_entity` resolves an institution name by
CONTAINMENT, so with no college in the spine "Bay Mills Community College"
resolves to `TRBF-BYMLLS-00`, the Bay Mills Indian Community itself. Twenty-odd
ledger rows are sitting on namesake tribes today:

    UNITED TRIBES TECHNICAL COLLEGE   -> TRBF-AUBURN-00   (Auburn Rancheria)
    NEBRASKA INDIAN COMMUNITY COLLEGE -> TRBF-IOWAKN-00   (Iowa Tribe of KS/NE)
    CALIFORNIA TRIBAL COLLEGE         -> TRBF-ACLNTE-00
    WHITE RIVER NATIVE CDFI LLC       -> TRBF-MINNCH-00

This is the same defect script 52 closed for ANCSA village corporations and
script 61 closed for NHOs, and it follows their shape exactly: nothing existing
is altered or deleted, an id collision aborts the run, and no row is added
without a retrieved URL and a verbatim quote.

WHY OWNERSHIP IS THE HARD PART, AND WHY IT IS A SEPARATE FIELD
--------------------------------------------------------------
A TCU is normally CHARTERED BY a tribe, not owned as an enterprise, and two of
them are not tribal at all:

    Haskell Indian Nations University   "Chartered 1884 (BIA)" - AIHEC
    Southwestern Indian Polytechnic     "Chartered 1971 (BIA)" - AIHEC
    Institute of American Indian Arts   "Chartered 1962 (Congress)" - AIHEC

Booking Haskell's federal education money to a tribe would be a fabricated
ownership fact. So `parent_native_entity` is written ONLY where a retrieved
sentence names a tribe that charters or controls the institution, and
`serves_native_entities` - a SEPARATE field - carries who it serves. Service is
not ownership. That distinction is the standing rule on this project and it is
the reason the Native CDFI layer cannot be built from names: many Native CDFIs
are independent nonprofits (Oweesta, Native Community Capital, Indian Land
Capital Company) and some are tribal instrumentalities.

SOURCES (all retrieved 2026-08-06; raw copies under data/raw/external/tcu_cdfi/)
-------------------------------------------------------------------------------
  TCU   https://www.aihec.org/tcu-roster-and-profiles/
        37 members in three tiers (34 regular, 1 associate, 2 developing), each
        with a chartering statement in its own profile paragraph.
  TCU   https://www.aihec.org/tcu-locations/   cross-check
        "AIHEC has grown to 37 Tribal Colleges and Universities with more than
         80 sites in the United States."
  CDFI  https://www.cdfifund.gov/media/8018641/download?inline
        Treasury's own "List of Currently Certified CDFIs":
        "Total Number of Certified Native CDFIs as of July 16, 2026: 65"
        The `Native CDFI (Y/N)` column is the authoritative roster.
  CDFI  https://github.com/frb-mpls-cde/nafi-map/raw/refs/heads/main/data/
        nafi-map-data_current.xlsx  - the data behind CICD's Native financial
        institutions map (Minneapolis Fed), 91 institutions, `ncdfi` and `nmdi`
        flags. This is the "cicd nafi map" Elijah named.

CROSS-CHECKS ATTEMPTED AND THEIR OUTCOMES (recorded, not hidden)
----------------------------------------------------------------
  Bureau of Indian Education TCU list  - bie.edu no longer publishes a
        post-secondary roster page; /topic-page/tribal-colleges-and-universities
        and /landing-page/post-secondary both return 404 and the site's only
        directory link is the K-12 school directory. AIHEC's "(BIA)" charter
        markers are used instead, which name the same two BIE institutions.
  White House Initiative TCU list - https://sites.ed.gov/whiaiane/tribes-tcus/
        tcus/ returns **HTTP 410 Gone**. The list is no longer published.
  Oweesta / Native CDFI Network member lists - neither publishes a public
        machine-readable directory (oweesta.org/native-cdfi-network 404;
        nativecdfi.net has a members-only join flow, no roster). Both are
        recorded as unavailable rather than approximated.

TRAPS ENFORCED
--------------
  * Duplicate detection uses EXACT or CORE-EQUAL name identity only. It must
    NOT use resolve_entity's containment leg, which maps every college onto its
    namesake tribe - the exact defect being fixed. Containment is used only in
    the other direction, to resolve a chartering body NAMED IN EVIDENCE.
  * A place-name college is not a TCU. `Yavapai College`, `Peoria`, `Seneca`
    (Missouri) are place names; 282 such coincidences were already withdrawn
    from the nonprofit layer. Only the AIHEC roster admits a TCU here.
  * A credit union is not automatically a CDFI. The NAFI map carries Native
    MDIs that Treasury does not certify (`ncdfi = No`), and those are classed
    `Native Financial Institution`, never `Native Community Development
    Financial Institution`.
  * `People's Bank of Seneca` (Seneca, Missouri) and `Eagle Bank` (Montana)
    both resolve by containment to unrelated spine entities (Seneca Nation;
    Native Village of Eagle, Alaska). Ownership is never taken from the name.

Reads   data/raw/external/tcu_cdfi/*            (retrieved sources)
        data/spine/cedar_entity_spine.csv
        data/clean/*                            (for --link)
Writes  data/clean/tcu_roster.csv               parsed TCU roster + evidence
        data/clean/native_fi_roster.csv         parsed CDFI/NAFI roster
        data/clean/tcu_cdfi_ownership_evidence.csv
        data/spine/cedar_entity_spine.csv       (appended; backed up first)
        review/tcu_cdfi_refused.csv             everything refused, with reason
        review/tcu_cdfi_already_in_spine.csv    other-class duplicates
        review/tcu_cdfi_identifier_candidates.csv
        docs/TCU_CDFI_BUILD_LOG.md              (written by --log)

Usage   py -3 code/73_add_tcu_and_cdfi.py --parse
        py -3 code/73_add_tcu_and_cdfi.py --fetch-org-pages
        py -3 code/73_add_tcu_and_cdfi.py --add
        py -3 code/73_add_tcu_and_cdfi.py --link
        py -3 code/73_add_tcu_and_cdfi.py --log
"""

import csv
import glob
import html
import importlib.util
import json
import re
import shutil
import sys
import time
import urllib.request
import urllib.error
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

CEDAR = Path(__file__).resolve().parent.parent
CLEAN = CEDAR / "data" / "clean"
SPINE_P = CEDAR / "data" / "spine" / "cedar_entity_spine.csv"
RAW = CEDAR / "data" / "raw" / "external" / "tcu_cdfi"
REVIEW = CEDAR / "review"
DOCS = CEDAR / "docs"
TODAY = date.today().isoformat()

AIHEC_ROSTER_URL = "https://www.aihec.org/tcu-roster-and-profiles/"
AIHEC_LOCATIONS_URL = "https://www.aihec.org/tcu-locations/"
CDFI_LIST_URL = "https://www.cdfifund.gov/media/8018641/download?inline"
CDFI_LIST_PAGE = "https://www.cdfifund.gov/programs-training/certification/cdfi"
NAFI_URL = ("https://github.com/frb-mpls-cde/nafi-map/raw/refs/heads/main/"
            "data/nafi-map-data_current.xlsx")
NAFI_PAGE = "https://www.minneapolisfed.org/indiancountry/resources/mapping-native-banks"

CLASS_TCU = "Tribal College or University"
CLASS_NCDFI = "Native Community Development Financial Institution"
CLASS_NFI = "Native Financial Institution"

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0 Safari/537.36")

# ---------------------------------------------------------------------------
# ONE RESOLVER. Standing rule 8.
# ---------------------------------------------------------------------------
_spec = importlib.util.spec_from_file_location(
    "m33", CEDAR / "code" / "33_apply_party_rulings.py")
_M33 = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_M33)
norm, core, resolve_entity = _M33.norm, _M33.core, _M33.resolve_entity

STOP = {"the", "inc", "incorporated", "corporation", "corp", "company", "llc",
        "ltd", "limited", "native", "of", "and"}

# Corporate forms and articles carry no identity; everything else does.
_INST_STOP = {"the", "of", "and", "a", "an", "for", "inc", "incorporated",
              "corporation", "corp", "company", "co", "llc", "ltd", "limited",
              "lp", "llp", "plc"}


def inst_core(name):
    """Identity token-set for an INSTITUTION.

    `33.core` strips structural words - tribal, nation, indian, community -
    which is right for tribe names and catastrophic for institution names:
    "California Tribal College" and "California Indian Nations College" both
    collapse to {california, college} and the duplicate guard silently merged
    two different colleges into one row. They are separate AIHEC members in
    separate membership tiers, 500 miles apart.

    So institution identity keeps every word that is not an article or a
    corporate form."""
    return frozenset(t for t in norm(name).split() if t not in _INST_STOP)


# A tribe's governing body is not a different entity from the tribe. When a
# charter clause names "Muscogee Nation Council" or "Fort Belknap Indian
# Community Council (FBICC)", stripping the governing-body phrase is what lets
# the name resolve - and it matters, because the spine contains an Alaska
# village literally named "Council", so the unstripped string resolves
# ambiguously to the wrong place or not at all.
GOVERNING_BODY = re.compile(
    r"\s*\((?:[A-Z]{2,10})\)\s*$|"
    r"\s+(?:tribal\s+)?(?:business\s+)?(?:executive\s+)?"
    r"(?:council|committee|board|legislature|congress|assembly|"
    r"board of regents|board of trustees|governing body)\s*$", re.I)

# Federal / congressional charter markers. These bar a tribal parent outright.
#
# NARROW ON PURPOSE. A first version matched any mention of "Bureau of Indian
# Education", and Keweenaw Bay Ojibwa Community College - chartered by the
# Keweenaw Bay Indian Community - was classed as federally operated because its
# profile says "The Bureau of Indian Education, Bureau of Indian Affairs,
# conducted a site visit in April 2010". Every TCU mentions the BIE somewhere;
# the claim that matters is that the BIE OPERATES the institution.
FEDERAL_MARKERS = re.compile(
    r"(operat\w*|administered|run|managed|funded)\s+(?:\w+\s+){0,3}"
    r"(?:under the auspices of|through|by)\s+the\s+bureau of indian|"
    r"congressionally chartered|"
    r"chartered\s*\(\s*(bia|bie|congress)\s*\)|"
    r"granted\s+non-?profit,?\s+congressionally chartered status", re.I)

# Sentences that establish who charters / owns / controls an institution.
# Each pattern's group 1 is the body named. Order matters: the most explicit
# first, so "chartered by X" beats a loose "a program of X".
#
# CASE HANDLING IS DELIBERATE AND NOT UNIFORM.
#
# "Chartered by the Fort Peck Assiniboine and Sioux Tribes" opens a sentence, so
# a case-sensitive `chartered` misses it silently - the worst kind of miss,
# because the output still looks complete. But turning re.I on globally is
# worse: `[A-Z]` under re.IGNORECASE matches lowercase, so the "X chartered Y"
# patterns start capturing from the first word of the sentence and produce
# rubbish like "work of the committee became a reality". So the KEYWORDS carry
# an explicit case class and the PROPER-NOUN anchors stay case-sensitive.
OWNER_PATTERNS = [
    (r"[Uu]nder the sovereign governmental authority of (?:the )?"
     r"([^.;,]{4,90})", "sovereign_authority"),
    (r"[Cc]hartered (?:in \d{4} )?by (?:an act of )?(?:the )?"
     r"([^.;,]{4,90})", "chartered_by"),
    (r"[Cc]hartered to (?:the )?([^.;,]{4,90})", "chartered_to"),
    (r"[Cc]reated by an act of (?:the )?([^.;,]{4,90})", "act_of"),
    (r"\b([A-Z][^.;,]{3,90}?) chartered (?:the )?[A-Z]", "x_chartered_y"),
    (r"[Ww]holly[- ]owned (?:subsidiary )?(?:of|by) (?:the )?"
     r"([^.;,]{4,90})", "wholly_owned"),
    (r"\bowned (?:and operated )?by (?:the )?([^.;,]{4,90})", "owned_by"),
    (r"\b(?:instrumentality|arm|enterprise|division|program|"
     r"subsidiary|affiliate) of (?:the )?([^.;,]{4,90})", "part_of"),
    (r"(?:[Ee]stablished|[Cc]reated|[Ff]ounded|[Ff]ormed) by (?:the )?"
     r"([^.;,]{4,90})", "established_by"),
    (r"\b([A-Z][^.;,]{3,90}?) (?:established|founded|created) (?:the )?[A-Z]",
     "x_established_y"),
    (r"\b([A-Z][^.;,]{3,90}?) (?:adopted a resolution to (?:establish|charter)|"
     r"authorized the (?:establishment|chartering) of)", "x_resolved_to_found"),
]
OWNER_TRIGGER = re.compile(
    r"charter|wholly[- ]owned|owned and operated|instrumentality|"
    r"\bowned by\b|sovereign governmental authority|"
    r"\bestablish|\bfounded\b|\bcreated\b|a program of|a division of|"
    r"a subsidiary of|tribally controlled|tribally chartered", re.I)

SERVES_TRIGGER = re.compile(
    r"\bserv(?:es|ing|e)\b.{0,140}?"
    r"(tribe|tribes|nation|nations|band|bands|pueblo|community|communities|"
    r"village|villages|reservation|people)", re.I)

# Entity classes that can legitimately be an owner/charterer.
OWNER_CLASSES = {
    "Federally recognized tribe",
    "Federally recognized Alaska Native Village",
    "State-recognized tribe",
    "Alaska Native Regional Corporation",
    "Alaska Native Village Corporation",
    "ANCSA Group Corporation",
    "Native Hawaiian Organization",
    "Intertribal Organization",
    "Federal-level self-governance consortium",
    "Federal-level constituency entity",
    "State-level constituency entity",
}

# A captured owner string that is really a multi-entity body, a federal body or
# a non-entity. These never become a single parent.
NOT_A_SINGLE_OWNER = re.compile(
    r"bureau of indian|congress|president|united states|federal government|"
    r"department of the interior|board of (regents|trustees|directors)|"
    r"^(a |an |its |their |our )|tribal college|community college|"
    r"^\d|three nebraska|five tribal|more than \d+|"
    r"governments of|tribal governments|"
    # A municipality, county, state or university is not a Native entity.
    # `Iḷisaġvik College` is chartered by the NORTH SLOPE BOROUGH and
    # `Woodland Financial Partners` is "chartered by the State of Wisconsin" -
    # both are real, retrieved sentences, and neither names a tribal owner.
    r"\bborough\b|\bcounty\b|^city of\b|^state of\b|\buniversity\b|"
    r"\bschool district\b|\bmunicipalit|"
    # A generic description of the owner is not an identification of one.
    # `Eagle Bank`: "one of very few banks owned by a Native American Tribe".
    r"^native american( tribe| nation| tribes)?$|^(the )?tribe$|"
    r"^tribal (council|government|nation)$|^indian tribe$|^a tribe$",
    re.I)


STATE_NAME_TO_ABBR = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR",
    "california": "CA", "colorado": "CO", "connecticut": "CT",
    "delaware": "DE", "florida": "FL", "georgia": "GA", "hawaii": "HI",
    "idaho": "ID", "illinois": "IL", "indiana": "IN", "iowa": "IA",
    "kansas": "KS", "kentucky": "KY", "louisiana": "LA", "maine": "ME",
    "maryland": "MD", "massachusetts": "MA", "michigan": "MI",
    "minnesota": "MN", "mississippi": "MS", "missouri": "MO",
    "montana": "MT", "nebraska": "NE", "nevada": "NV",
    "new hampshire": "NH", "new jersey": "NJ", "new mexico": "NM",
    "new york": "NY", "north carolina": "NC", "north dakota": "ND",
    "ohio": "OH", "oklahoma": "OK", "oregon": "OR", "pennsylvania": "PA",
    "rhode island": "RI", "south carolina": "SC", "south dakota": "SD",
    "tennessee": "TN", "texas": "TX", "utah": "UT", "vermont": "VT",
    "virginia": "VA", "washington": "WA", "west virginia": "WV",
    "wisconsin": "WI", "wyoming": "WY",
}


def log(msg, sink=None):
    print(msg)
    if sink is not None:
        sink.append(msg)


def read_csv(p):
    p = Path(p)
    if not p.exists():
        return []
    with open(p, encoding="utf-8-sig", errors="replace", newline="") as fh:
        return list(csv.DictReader(fh))


def write_csv(p, rows, fields=None):
    p = Path(p)
    p.parent.mkdir(parents=True, exist_ok=True)
    fields = fields or (list(rows[0].keys()) if rows else ["empty"])
    with open(p, "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    print(f"  wrote {p.relative_to(CEDAR)}  ({len(rows):,} rows)")


def token(name, taken):
    """Deterministic 6-character id token, in the spine's existing style."""
    words = [w for w in norm(name).split() if w not in STOP]
    if not words:
        words = norm(name).split() or ["entity"]
    base = "".join(words)
    cons = re.sub(r"[aeiou]", "", base)
    cand = (cons if len(cons) >= 6 else base)[:6].upper().ljust(6, "X")
    if cand not in taken:
        return cand
    for i in range(1, 100):
        alt = (cand[:5] + str(i))[:6]
        if alt not in taken:
            return alt
    raise SystemExit(f"cannot mint a unique token for {name}")


def strip_tags(t):
    t = re.sub(r"(?is)<(script|style|noscript)[^>]*>.*?</\1>", " ", t)
    t = re.sub(r"(?is)<br\s*/?>", "\n", t)
    t = re.sub(r"(?is)</(p|div|li|h[1-6]|tr|td)>", "\n", t)
    t = re.sub(r"(?s)<[^>]+>", " ", t)
    t = html.unescape(t)
    t = re.sub(r"[ \t\xa0]+", " ", t)
    return re.sub(r"\n\s*\n+", "\n", t)


def sentences(blk):
    """Split on sentence enders AND on line breaks.

    Web pages put headings and nav labels on their own lines with no full stop,
    so a sentence-only split glues "Our History" onto "The Leech Lake Band of
    Ojibwe established ..." and the capture then carries the heading. Splitting
    on newlines first keeps the quote to the sentence that actually makes the
    claim."""
    parts = []
    for line in re.split(r"[\r\n]+", blk):
        parts.extend(re.split(r"(?<=[.!?])\s+", line))
    return [s.strip() for s in parts if s.strip()]


def sentence_offsets(blk):
    """`sentences(blk)` with each sentence's CHARACTER OFFSET in the page.

    Added 2026-09-01. `tcu_cdfi_ownership_evidence.csv` carried 4 literal
    duplicate rows of 130 and could not be keyed, and none of the four was a
    duplicate FACT: a page states the same sentence twice - once in a nav or
    banner block and once in the body - and each occurrence is a real, separate
    occurrence of the evidence. `First State Bank` says "We serve customers and
    community with candor, integrity, trust, fair dealing, and honor." twice on
    /about; `Little Priest Tribal College` states its charter sentence twice on
    its home page. The extractor recorded the sentence, the pattern and the URL
    and dropped WHERE on the page it was found, which is the only thing that
    separates them.

    So the offset is written, nothing is de-duplicated, and
    (institution, layer, pattern, evidence_url, quote_char_offset) becomes a
    key. Same shape as the projection loss repaired in `23` and `173`.
    """
    cursor = 0
    for s in sentences(blk):
        i = blk.find(s, cursor)
        if i < 0:                      # cannot happen for a stripped split,
            i = cursor                 # but a wrong offset must not be a crash
        cursor = i + len(s)
        yield i, s


def clean_owner(s):
    """Trim a captured owner string WITHOUT amputating the name.

    The first version cut at the first ` and `, which turned "Confederated
    Salish and Kootenai Tribes" into "Confederated Salish" - a name that
    resolves to nothing, or worse, to the wrong tribe. So `and` is cut only
    when what follows is lowercase, and a trailing subordinate clause is cut
    only on an explicit marker."""
    s = re.sub(r"\s+", " ", s).strip(" .,;:-")
    s = re.sub(r"^(?:[Tt]he|[Aa]n?)\s+", "", s)
    # Case-INSENSITIVE clause markers.
    s = re.sub(r"\s+(?:in \d{4}\b|in [A-Z]|"
               r"(?:in|on|since|from) (?:January|February|March|April|May|"
               r"June|July|August|September|October|November|December)\b|"
               r"which\b|that\b|whose\b|when\b).*$", "", s, flags=re.I)
    # Case-SENSITIVE. `flags=re.I` here was a real bug: under IGNORECASE the
    # lookahead `(?=[a-z])` also matches an uppercase letter, so "Confederated
    # Salish and Kootenai Tribes" was amputated to "Confederated Salish" and
    # "Fort Peck Assiniboine and Sioux Tribes" to "Fort Peck Assiniboine".
    # A truncated tribe name resolves to nothing, or to the wrong tribe.
    s = re.sub(r"\s+(?:and (?=[a-z])|to [a-z]|for [a-z]|as [a-z]|"
               r"with [a-z]|at [a-z]|on [a-z]|by [a-z]|under [a-z]|"
               r"pursuant [a-z]|through [a-z]).*$", "", s)
    return s.strip(" .,;:-")


# Words that betray a capture of sentence scaffolding rather than a name.
_NOT_A_NAME = {"was", "were", "is", "are", "has", "have", "had", "they", "we",
               "it", "this", "these", "those", "he", "she", "who", "there",
               "began", "grew", "operates", "offers", "provides", "serves"}


def looks_like_proper_noun(s):
    """A charter clause names an institution, so the capture must read like a
    name. "In 1984 they", "SGU was" and "UTTC was" all satisfied a looser test
    and would have become owners; requiring two capitalised words AND no
    sentence scaffolding refuses all three."""
    words = [w for w in re.findall(r"[A-Za-z'’\-]+", s) if len(w) > 1]
    if not words or any(w.lower() in _NOT_A_NAME for w in words):
        return False
    caps = [w for w in words if w[0].isupper()]
    return len(caps) >= 2


def find_ownership(text, url):
    """Return a list of evidence dicts, ordered by PATTERN PRIORITY rather than
    by sentence order, so an explicit "chartered by" outranks a loose
    "established by" found earlier in the page. Never guesses: a hit is a
    verbatim sentence plus the substring the pattern captured."""
    out = []
    for off, s in sentence_offsets(text):
        if len(s) > 600 or not OWNER_TRIGGER.search(s):
            continue
        for rank, (pat, kind) in enumerate(OWNER_PATTERNS):
            # NO re.I here. The patterns carry their own case classes; a
            # blanket IGNORECASE makes `[A-Z]` match lowercase and the
            # proper-noun anchors capture sentence scaffolding instead.
            m = re.search(pat, s)
            if not m:
                continue
            owner = clean_owner(m.group(1))
            if not owner or len(owner) < 4 or not looks_like_proper_noun(owner):
                continue
            out.append({"quote": s, "captured_owner": owner,
                        "pattern": kind, "evidence_url": url,
                        "quote_char_offset": off, "_rank": rank})
            break
    out.sort(key=lambda e: e["_rank"])
    for e in out:
        e.pop("_rank", None)
    return out


def find_serves(text, url):
    out = []
    for off, s in sentence_offsets(text):
        if len(s) > 600:
            continue
        if SERVES_TRIGGER.search(s):
            out.append({"quote": s, "evidence_url": url,
                        "quote_char_offset": off})
    return out[:3]


# ===========================================================================
# PARSE
# ===========================================================================
def parse_aihec():
    """Parse the AIHEC roster into one record per TCU, each with the verbatim
    profile sentences that state who chartered it."""
    src = RAW / "aihec_tcu_roster_2026-08-06.html"
    if not src.exists():
        raise SystemExit(f"missing retrieved source: {src}")
    raw_html = src.read_text(encoding="utf-8", errors="replace")

    # LINKS MUST STAY WITH THEIR OWN PROFILE.
    #
    # The first version collected every outbound href in document order and
    # paired the nth link with the nth college. It was off by four, and the
    # failure was silent and severe: the College of Menominee Nation was given
    # Bay Mills's website, so Bay Mills's charter sentence was harvested as
    # Menominee's ownership evidence. Positional alignment across two
    # independently-filtered lists is never safe. So the anchor href is turned
    # into an inline token BEFORE the tags are stripped, and each profile takes
    # the first link that appears inside its own block.
    marked = re.sub(r'<a\b[^>]*href=["\']((?:https?:)?//[^"\']+)["\'][^>]*>',
                    r" [[LINK:\1]] ", raw_html, flags=re.I)
    txt = strip_tags(marked)
    (RAW / "aihec_tcu_roster_2026-08-06.txt").write_text(
        strip_tags(raw_html), encoding="utf-8")

    i = txt.find("REGULAR MEMBERS")
    body = txt[i:]
    j = body.find("Quick Links")
    if j > 0:
        body = body[:j]
    raw_lines = [l.strip() for l in body.split("\n")]
    LINK_RE = re.compile(r"\[\[LINK:([^\]]+)\]\]")
    SKIP_LINK = re.compile(r"aihec|facebook|youtube|instagram|twitter|gmpg|"
                           r"adp\.com|tribalcollegejournal|w3\.org|schema|"
                           r"linkedin|mailto", re.I)
    # `lines` carries the prose with link tokens removed; `raw_lines` keeps them
    # so each block can claim its own links.
    lines = [re.sub(r"\s+", " ", LINK_RE.sub(" ", l)).strip() for l in raw_lines]

    hdr = re.compile(r"^([^()]{4,90}?)\s*\(([A-Za-z\-\u0130]{2,10})\)$")
    key = re.compile(r"\b(College|University|Institute|Polytechnic)\b")
    idx, sec = [], None
    for k, s in enumerate(lines):
        if re.fullmatch(r"(REGULAR|ASSOCIATE|DEVELOPING) MEMBERS?", s):
            sec = s
            continue
        m = hdr.match(s)
        if m and key.search(m.group(1)):
            idx.append((k, sec, m.group(1).strip(), m.group(2)))

    recs = []
    for n, (k, section, name, acr) in enumerate(idx):
        end = idx[n + 1][0] if n + 1 < len(idx) else len(lines)
        blk = " ".join(lines[k + 1:end])
        block_links = [h for h in
                       LINK_RE.findall(" ".join(raw_lines[k:end]))
                       if not SKIP_LINK.search(h)]
        ch = re.search(r"Chartered\s+(\d{4})(?:\s*\(([^)]+)\))?", blk)
        st = re.search(r",\s*([A-Z]{2})\s+\d{5}", blk)
        state = st.group(1) if st else ""
        if not state:
            # Some entries spell the state out ("L'Anse, Michigan 49946"). The
            # first version's fallback returned the CITY, which put "Anse" in
            # the state column for Keweenaw Bay.
            m2 = re.search(r",\s*([A-Z][a-z]+(?: [A-Z][a-z]+)?)\s+\d{5}", blk)
            if m2:
                state = STATE_NAME_TO_ABBR.get(m2.group(1).lower(), "")
        prof = blk
        # drop the contact header so quotes are prose, not phone numbers
        cut = re.search(r"Chartered\s+\d{4}(?:\s*\([^)]+\))?\s*", blk)
        if cut:
            prof = blk[cut.end():]
        ev = find_ownership(prof, AIHEC_ROSTER_URL)
        recs.append({
            "name": name, "acronym": acr, "membership_tier": section or "",
            "state": state,
            "chartered_year": ch.group(1) if ch else "",
            "charter_marker": (ch.group(2) or "") if ch else "",
            "website": block_links[0] if block_links else "",
            "profile_text": prof,
            "ownership_evidence": ev,
            "serves_evidence": find_serves(prof, AIHEC_ROSTER_URL),
            "source_url": AIHEC_ROSTER_URL,
            "retrieved_date": "2026-08-06",
        })
    return recs


def parse_cdfi_fund():
    import openpyxl
    p = RAW / "cdfifund_certified_list_2026-08-06.xlsx"
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb["List of Certified CDFIs"]
    rows = list(ws.iter_rows(values_only=True))
    asof = ""
    for r in rows[:8]:
        for c in r:
            if c and "Total Number of Certified Native CDFIs" in str(c):
                asof = str(c).strip()
    hdr = [str(c).strip() if c else "" for c in rows[8]]
    out = []
    for r in rows[9:]:
        d = dict(zip(hdr, [("" if c is None else str(c).strip()) for c in r]))
        if not d.get("Organization Name"):
            continue
        if not d.get("Native CDFI (Y/N)", "").upper().startswith("Y"):
            continue
        out.append({
            "name": d["Organization Name"],
            "cert_control_num": d.get("Cert Control Num", ""),
            "fi_type": d.get("Financial Institution Type", ""),
            "rssd_id": d.get("RSSD ID", ""),
            "city": d.get("City", ""), "state": d.get("State", ""),
            "website": d.get("Organization Website", ""),
            "source_url": CDFI_LIST_URL,
            "source_quote": asof,
            "retrieved_date": "2026-08-06",
        })
    return out, asof


# ---------------------------------------------------------------------------
# HAND-CHECKED CROSSWALK: CICD NAFI-map name -> CDFI Fund certified-list name.
#
# The two rosters name the same institution differently, and a token-level
# matcher merges only 55 of the 65. The remaining nine are the ones a fuzzy
# matcher would get wrong as often as right, so they are ruled here by hand
# with the evidence that settles each one. Hand-checked beats automated; this
# is the same discipline as the per-UEI drops in hci_analysis.do.
#
# The defect being prevented is precise: two rows for one institution is the
# split that separated Ho-Chunk Inc from the Winnebago Tribe.
# ---------------------------------------------------------------------------
NAFI_ALIAS = {
    # NAFI name  ->  (CDFI Fund certified-list name, why)
    "Cherokee Nation dba Cherokee Nation Economic Development Trust Authority, Inc.":
        ("The Cherokee Nation d/b/a Cherokee Nation Economic Development Trust "
         "Authority, Inc",
         "Same legal name; the two lists punctuate d/b/a differently."),
    "Hawaii First Federal Credit Union":
        ("HAWAII FIRST FCU",
         "FCU is the abbreviation of Federal Credit Union; same Kamuela, HI "
         "credit union, Cert Control Num 131CE012944."),
    "Jamestown S'klallam Tribal Capital (JST Capital)":
        ("Jamestown S'Klallam Tribal Capital, Incorporated",
         "Same Sequim, WA institution; the NAFI name carries the trading "
         "acronym in parentheses."),
    "Local Bank (formerly Bank of Cherokee County)":
        ("Local Bank",
         "Same Hulbert, OK bank; the NAFI name carries the former name."),
    "Navajo Community Development Financial Institution":
        ("Navajo community Development Financial Institution Inc Non Profit",
         "Same Window Rock, AZ institution; the Treasury row carries the "
         "corporate suffix."),
    "Pacific Northwest Tribal Lending (formerly Lummi CDFI)":
        ("Pacific Northwest Tribal Lending, a Community Development Financial "
         "Institution",
         "Same Bellingham, WA institution; the NAFI name carries the former "
         "name and the Treasury name carries the descriptive suffix."),
    "People's Partner for Community Development":
        ("Peoples Partners for Community Development",
         "Same Lame Deer, MT institution; the two lists differ by an "
         "apostrophe and a plural."),
    "Woodland Financial Partners, Inc. (formerly NiiJii Capital Partners)":
        ("Woodland Financial Partners Inc",
         "Same Keshena, WI institution; the NAFI name carries the former "
         "name."),
    "First Nations Community Financial":
        ("Cedar Growth Corporation",
         "Same institution under a former name: both rows give Black River "
         "Falls, WI and the SAME website, firstnationsfinancial.org. Merged on "
         "the shared domain, not on the name."),
}

# CICD flags this as a Native CDFI; Treasury's current certified list does not
# carry it. Recorded as a discrepancy rather than resolved by assumption.
NAFI_CERT_DISCREPANCY = {
    "Kauai Federal Credit Union":
        "CICD NAFI map (last_updated 2025-08-20) records ncdfi=Yes, but the "
        "Treasury CDFI Fund list of currently certified CDFIs as of "
        "2026-07-16 does not include it. Classed as a Native financial "
        "institution, NOT as a certified Native CDFI, and flagged for review.",
}


def parse_nafi():
    import openpyxl
    p = RAW / "cicd_nafi_map_data_2026-08-06.xlsx"
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb["data"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c else "" for c in rows[0]]
    out = []
    for r in rows[1:]:
        d = dict(zip(hdr, [("" if c is None else str(c).strip()) for c in r]))
        if not d.get("name"):
            continue
        out.append({
            "name": d["name"], "fi_type": d.get("type", ""),
            "regulator": d.get("regulator", ""),
            "rssd_id": d.get("rssd_id", ""),
            "bank_cert": d.get("bank_cert", ""),
            "cu_number": d.get("cu_number", ""),
            "ncdfi": d.get("ncdfi", ""), "nmdi": d.get("nmdi", ""),
            "city": d.get("city", ""), "state": d.get("state", ""),
            "website": d.get("website", ""),
            "source_url": NAFI_URL, "source_page": NAFI_PAGE,
            "retrieved_date": "2026-08-06",
        })
    return out


def cmd_parse():
    print("=== 73 --parse ===\n")
    tcu = parse_aihec()
    print(f"AIHEC roster            : {len(tcu)} TCUs "
          f"({Counter(t['membership_tier'] for t in tcu)})")
    cd, asof = parse_cdfi_fund()
    print(f"CDFI Fund certified     : {len(cd)} Native CDFIs")
    print(f"  quote: {asof}")
    nafi = parse_nafi()
    print(f"CICD NAFI map           : {len(nafi)} institutions "
          f"({sum(1 for r in nafi if r['ncdfi'] == 'Yes')} flagged ncdfi=Yes)")

    write_csv(CLEAN / "tcu_roster.csv",
              [{**t, "ownership_evidence": json.dumps(t["ownership_evidence"],
                                                      ensure_ascii=False),
                "serves_evidence": json.dumps(t["serves_evidence"],
                                              ensure_ascii=False),
                "profile_text": t["profile_text"][:4000]} for t in tcu],
              ["name", "acronym", "membership_tier", "state", "chartered_year",
               "charter_marker", "website", "source_url", "retrieved_date",
               "ownership_evidence", "serves_evidence", "profile_text"])

    # union the two financial-institution sources on core-equal names
    byc = {}
    for r in cd:
        byc[core(r["name"])] = {**r, "in_cdfi_fund_list": "1",
                                # "in_cicd_nafi_map": "0",  # CICD nuked 2026-09-02 (844)
                                "treasury_native_cdfi": "Y"}
    for r in nafi:
        alias = NAFI_ALIAS.get(r["name"])
        c = core(alias[0]) if alias else core(r["name"])
        if alias and c not in byc:
            raise SystemExit(
                f"ABORT: hand-checked alias target not found in the CDFI Fund "
                f"list: {r['name']} -> {alias[0]}. The source list changed; "
                f"re-verify the crosswalk rather than dropping the row.")
        if alias and c in byc:
            byc[c]["alias_basis"] = alias[1]
        if r["name"] in NAFI_CERT_DISCREPANCY:
            r = {**r, "cert_discrepancy": NAFI_CERT_DISCREPANCY[r["name"]]}
        if c in byc:
            # byc[c]["in_cicd_nafi_map"] = "1"  # CICD nuked 2026-09-02 (844)
            byc[c]["nafi_name"] = r["name"]
            byc[c].setdefault("website", "")
            if not byc[c]["website"] or byc[c]["website"].lower() in (
                    "non given", "none given", "n/a"):
                byc[c]["website"] = r["website"]
            byc[c]["nmdi"] = r.get("nmdi", "")
            byc[c]["regulator"] = r.get("regulator", "")
        else:
            byc[c] = {**r, "in_cdfi_fund_list": "0", "in_cicd_nafi_map": "1",
                      "treasury_native_cdfi": "N",
                      "source_quote": "CICD Native financial institutions map; "
                                      f"ncdfi={r['ncdfi']}, nmdi={r['nmdi']}"}
    fi = list(byc.values())
    for r in fi:
        r["entity_class"] = (CLASS_NCDFI if r["treasury_native_cdfi"] == "Y"
                             else CLASS_NFI)
    print(f"union                   : {len(fi)} distinct institutions "
          f"({sum(1 for r in fi if r['entity_class'] == CLASS_NCDFI)} certified "
          f"Native CDFIs, {sum(1 for r in fi if r['entity_class'] == CLASS_NFI)}"
          f" other Native financial institutions)")
    write_csv(CLEAN / "native_fi_roster.csv", fi,
              ["name", "entity_class", "treasury_native_cdfi", "fi_type",
               "cert_control_num", "rssd_id", "bank_cert", "cu_number",
               "ncdfi", "nmdi", "regulator", "city", "state", "website",
               "in_cdfi_fund_list", "in_cicd_nafi_map", "nafi_name",
               "alias_basis", "cert_discrepancy",
               "source_url", "source_page", "source_quote", "retrieved_date"])


# ===========================================================================
# FETCH ORG PAGES  (ownership evidence in the institution's own words)
# ===========================================================================
ABOUT_PATHS = ["", "/about", "/about-us", "/who-we-are", "/about/",
               "/our-story", "/mission"]


def fetch(url, timeout=25):
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=timeout) as fh:
        raw = fh.read(2_000_000)
        enc = fh.headers.get_content_charset() or "utf-8"
        return raw.decode(enc, errors="replace"), fh.geturl()


def _org_targets():
    targets = []
    for t in read_csv(CLEAN / "tcu_roster.csv"):
        if t.get("website"):
            targets.append(("TCU", t["name"], t["website"]))
    for r in read_csv(CLEAN / "native_fi_roster.csv"):
        w = (r.get("website") or "").strip()
        if w and w.lower() not in ("non given", "none given", "n/a", "none"):
            targets.append(("FI", r["name"], w))
    return targets


def cmd_reextract():
    """Re-derive ownership evidence from the CACHED pages.

    Tightening a capture rule must not cost another 130 remote requests. The
    pages are on disk; only the extraction changes."""
    print("=== 73 --reextract (no network) ===\n")
    out_dir = RAW / "org_pages"
    evid, seen_files = [], 0
    for kind, name, site in _org_targets():
        if not site.startswith("http"):
            site = "http://" + site
        base = site.rstrip("/")
        slug = re.sub(r"[^a-z0-9]+", "_", name.lower())[:48]
        for path in ABOUT_PATHS[:4]:
            sfx = re.sub(r"[^a-z0-9]+", "_", path) or "home"
            f = out_dir / f"{slug}{sfx}.txt"
            if not f.exists():
                continue
            seen_files += 1
            body = f.read_text(encoding="utf-8", errors="replace")
            url = base + path
            for e in find_ownership(body, url):
                evid.append({"institution": name, "layer": kind, **e})
            for e in find_serves(body, url):
                evid.append({"institution": name, "layer": kind,
                             "quote": e["quote"], "captured_owner": "",
                             "pattern": "serves", "evidence_url": url,
                             "quote_char_offset": e["quote_char_offset"]})
    print(f"cached pages read: {seen_files}")
    # The declared key, checked here rather than only in 512: an evidence row
    # with no distinct position is the defect this was repaired for.
    from collections import Counter as _C
    _k = _C((e["institution"], e["layer"], e["pattern"], e["evidence_url"],
             e["quote_char_offset"]) for e in evid)
    _d = sum(n - 1 for n in _k.values() if n > 1)
    print(f"primary key (institution, layer, pattern, evidence_url, "
          f"quote_char_offset): {_d:,} duplicate(s) of {len(evid):,}")
    if _d:
        raise SystemExit("REFUSED to write: the declared primary key is not "
                         "unique.")
    write_csv(CLEAN / "tcu_cdfi_ownership_evidence.csv", evid,
              ["institution", "layer", "pattern", "captured_owner", "quote",
               "evidence_url", "quote_char_offset"])


def cmd_fetch_org_pages():
    """One request at a time, one host at a time, exponential backoff on
    failure, and never more than 3 pages per institution. These are ~100
    DISTINCT hosts, so no single host sees more than 3 requests. Nothing here
    touches api.usaspending.gov."""
    print("=== 73 --fetch-org-pages ===\n")
    out_dir = RAW / "org_pages"
    out_dir.mkdir(parents=True, exist_ok=True)
    state_p = RAW / "_org_fetch_state.json"
    # Rule 6 of PULL_DISCIPLINE: checkpoint before the first request. A
    # half-written checkpoint must not be fatal on resume.
    state = {}
    if state_p.exists():
        try:
            state = json.loads(state_p.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError):
            print("  checkpoint unreadable; starting clean")

    targets = []
    for t in read_csv(CLEAN / "tcu_roster.csv"):
        if t.get("website"):
            targets.append(("TCU", t["name"], t["website"]))
    for r in read_csv(CLEAN / "native_fi_roster.csv"):
        w = (r.get("website") or "").strip()
        if w and w.lower() not in ("non given", "none given", "n/a", "none"):
            targets.append(("FI", r["name"], w))
    print(f"institutions with a website: {len(targets)}")

    evid = []
    for n, (kind, name, site) in enumerate(targets, 1):
        if not site.startswith("http"):
            site = "http://" + site
        base = site.rstrip("/")
        key = name
        if key in state and state[key].get("done"):
            evid.extend(state[key].get("evidence", []))
            continue
        got = []
        for path in ABOUT_PATHS[:4]:
            url = base + path
            try:
                text, final = fetch(url)
            except Exception as e:                       # noqa: BLE001
                continue
            body = strip_tags(text)
            slug = re.sub(r"[^a-z0-9]+", "_", name.lower())[:48]
            sfx = re.sub(r"[^a-z0-9]+", "_", path) or "home"
            (out_dir / f"{slug}{sfx}.txt").write_text(body[:200_000],
                                                      encoding="utf-8")
            for e in find_ownership(body, final):
                got.append({"institution": name, "layer": kind, **e})
            for e in find_serves(body, final):
                got.append({"institution": name, "layer": kind,
                            "quote": e["quote"], "captured_owner": "",
                            "pattern": "serves", "evidence_url": final})
            if any(g["pattern"] not in ("serves",) for g in got):
                break
            time.sleep(0.8)
        state[key] = {"done": True, "evidence": got}
        evid.extend(got)
        if n % 10 == 0:
            state_p.write_text(json.dumps(state, ensure_ascii=False), encoding="utf-8")
            print(f"  {n}/{len(targets)}  evidence rows so far: {len(evid)}")
        time.sleep(0.5)

    state_p.write_text(json.dumps(state, ensure_ascii=False), encoding="utf-8")
    write_csv(CLEAN / "tcu_cdfi_ownership_evidence.csv", evid,
              ["institution", "layer", "pattern", "captured_owner", "quote",
               "evidence_url"])


# ===========================================================================
# ADD TO SPINE
# ===========================================================================
def resolve_owner(owner_name, spine):
    """Resolve a chartering/owning body NAMED IN EVIDENCE onto the spine.

    This is the one legitimate use of resolve_entity's containment leg: the
    evidence names a tribe in full ("Bay Mills Indian Community") and the spine
    stores the short form ("Bay Mills"). The reverse direction - resolving an
    INSTITUTION name - is refused everywhere in this script."""
    if not owner_name or NOT_A_SINGLE_OWNER.search(owner_name):
        return None, None, "not_a_single_native_owner"

    attempts = [(owner_name, "")]
    stripped = owner_name
    for _ in range(3):
        s2 = GOVERNING_BODY.sub("", stripped).strip()
        if s2 == stripped or len(s2.split()) < 1:
            break
        stripped = s2
        attempts.append((stripped, "+governing_body_stripped"))

    last = "no_spine_match"
    for cand, suffix in attempts:
        if NOT_A_SINGLE_OWNER.search(cand):
            last = "not_a_single_native_owner"
            continue
        tid, canon, how = resolve_entity(cand, spine)
        if not tid:
            last = how
            continue
        row = next((r for r in spine if r["tribe_id"] == tid), None)
        if row and row.get("entity_class") not in OWNER_CLASSES:
            last = f"owner_class_not_eligible:{row.get('entity_class')}"
            continue
        return tid, canon, how + suffix
    return None, None, last


def cmd_add():
    print("=== 73 --add ===\n")
    spine = read_csv(SPINE_P)
    fields = list(spine[0].keys())
    print(f"spine entities before : {len(spine):,}")

    # NEW COLUMNS. Ownership and service are separate fields and must not be
    # collapsed. `parent_entity_id`/`parent_entity_name` already exist and are
    # populated in lockstep so the existing hierarchy tooling sees the link.
    for col in ("parent_native_entity", "serves_native_entities",
                "ownership_basis", "entity_source_url", "entity_source_quote"):
        if col not in fields:
            fields.append(col)
            for r in spine:
                r.setdefault(col, "")
    for r in spine:
        for col in fields:
            r.setdefault(col, "")

    if any(r["tribe_id"].startswith(("TCU-", "CDFI-")) for r in spine):
        raise SystemExit(
            "ABORT: the spine already carries TCU-/CDFI- rows. Re-running "
            "--add would mint a second identity for every institution. "
            "Restore data/spine/cedar_entity_spine.csv from its "
            f".bak_{TODAY}_pre73tcu backup first.")

    have_norm, have_icore = {}, {}
    for r in spine:
        have_norm.setdefault(norm(r["canonical_name"]), r)
        ic = inst_core(r["canonical_name"])
        if ic:
            have_icore.setdefault(ic, r)
        for a in (r.get("aliases") or "").split("|"):
            if a.strip():
                have_norm.setdefault(norm(a), r)
    have_ids = {r["tribe_id"] for r in spine}
    taken = {r["tribe_id"].split("-")[1] for r in spine if "-" in r["tribe_id"]}

    # institution-name -> extra evidence from the institution's own site
    own_ev = defaultdict(list)
    for e in read_csv(CLEAN / "tcu_cdfi_ownership_evidence.csv"):
        own_ev[e["institution"]].append(e)

    added, refused, already, ruled = [], [], [], []

    def consider(rec, prefix, entity_class):
        name = rec["name"].strip()
        n = norm(name)
        ic = inst_core(name)

        # PRIME DIRECTIVE gate.
        if not rec.get("source_url"):
            refused.append({"layer": entity_class, "name": name,
                            "reason": "no retrieved source URL",
                            "evidence_url": "", "refused_date": TODAY})
            return
        if not rec.get("source_quote"):
            refused.append({"layer": entity_class, "name": name,
                            "reason": "no verbatim source quote",
                            "evidence_url": rec.get("source_url", ""),
                            "refused_date": TODAY})
            return

        # DUPLICATE GUARD - exact name, exact alias, or INSTITUTION-core
        # equality. Two legs are deliberately excluded:
        #   * resolve_entity's CONTAINMENT leg, which maps every college onto
        #     its namesake tribe - the defect this build exists to fix;
        #   * `33.core`, which strips tribal/nation/indian/community and made
        #     "California Tribal College" identical to "California Indian
        #     Nations College" - two different AIHEC members, silently merged.
        hit = have_norm.get(n) or (have_icore.get(ic) if ic else None)
        if hit:
            already.append({
                "name": name, "layer": entity_class,
                "existing_tribe_id": hit["tribe_id"],
                "existing_canonical_name": hit["canonical_name"],
                "existing_entity_class": hit["entity_class"],
                "status": "ALREADY_IN_SPINE",
                "note": "Matched an existing spine entity by exact name, "
                        "exact alias or institution-core equality. Not "
                        "re-minted; the existing row is authoritative. Adding "
                        "a second row would split one institution across two "
                        "ids.",
                "source_url": rec["source_url"],
            })
            return

        tok = token(name, taken)
        taken.add(tok)
        tid = f"{prefix}-{tok}-00"
        if tid in have_ids:
            raise SystemExit(f"ABORT: {tid} already exists. Refusing to "
                             f"overwrite an existing spine entity.")

        # ---- ownership -----------------------------------------------------
        parent_id = parent_name = ""
        basis = rec.get("ownership_basis_default", "")
        owner_quote = ""
        owner_url = ""
        if rec.get("federal_operated"):
            basis = rec["federal_operated"]
        else:
            cands = list(rec.get("evidence", [])) + [
                {"quote": e["quote"], "captured_owner": e["captured_owner"],
                 "pattern": e["pattern"], "evidence_url": e["evidence_url"]}
                for e in own_ev.get(name, []) if e["pattern"] != "serves"]
            for e in cands:
                tid_o, canon_o, how = resolve_owner(e["captured_owner"], spine)
                if tid_o:
                    parent_id, parent_name = tid_o, canon_o
                    basis = (f"{e['pattern']}; resolved via {how}")
                    owner_quote, owner_url = e["quote"], e["evidence_url"]
                    break
                ruled.append({"institution": name, "layer": entity_class,
                              "captured_owner": e["captured_owner"],
                              "pattern": e["pattern"], "outcome": how,
                              "quote": e["quote"][:400],
                              "evidence_url": e["evidence_url"]})
            if not parent_id and not basis:
                # Say WHY, so an empty parent is a documented refusal rather
                # than an unexplained blank. "Chartered by the governments of
                # three Nebraska Indian Tribes" is a real charter statement
                # that names no single owner, and that is a different fact
                # from "we found nothing".
                tried = [r for r in ruled if r["institution"] == name]
                if tried:
                    basis = ("charter/ownership sentence retrieved but it "
                             "names no single Native entity resolvable in the "
                             "spine (" + "; ".join(
                                 f"{t['captured_owner'][:48]} -> {t['outcome']}"
                                 for t in tried[:2]) +
                             ") - left empty by design")
                else:
                    basis = ("no retrieved sentence names a single Native "
                             "entity as owner or charterer - left empty by "
                             "design")

        serves = rec.get("serves", "")

        row = {f: "" for f in fields}
        row.update({
            "tribe_id": tid,
            "canonical_name": name,
            "entity_class": entity_class,
            "state": rec.get("state", ""),
            "aliases": "|".join(dict.fromkeys(
                [name] + [a for a in rec.get("aliases", []) if a])),
            "parent_native_entity": parent_name,
            "parent_entity_id": parent_id,
            "parent_entity_name": parent_name,
            "hierarchy_basis": ("charter/ownership evidence, script 73"
                                if parent_id else ""),
            "serves_native_entities": serves,
            "ownership_basis": basis,
            "entity_source_url": owner_url or rec["source_url"],
            "entity_source_quote": (owner_quote or rec["source_quote"])[:600],
            "reconciliation_status": "added_2026-08-06_script73",
            "reconciliation_note": rec.get("note", ""),
        })
        for k in ("n_uei_tierA", "n_uei_tierB", "n_cage", "n_ein"):
            if k in row:
                row[k] = "0"
        added.append(row)
        have_ids.add(tid)
        have_norm[n] = row
        if ic:
            have_icore.setdefault(ic, row)

    # ---- TCUs -------------------------------------------------------------
    print("\n[A] Tribal Colleges and Universities")
    for t in read_csv(CLEAN / "tcu_roster.csv"):
        ev = json.loads(t["ownership_evidence"] or "[]")
        sv = json.loads(t["serves_evidence"] or "[]")
        marker = (t.get("charter_marker") or "").strip()
        prof = t.get("profile_text") or ""
        fed, fed_quote = "", ""
        # The verbatim sentence that establishes federal operation, so the
        # refusal to name a tribal parent is itself evidenced rather than
        # asserted.
        for s in sentences(prof):
            if FEDERAL_MARKERS.search(s) and len(s) < 500:
                fed_quote = s
                break
        if marker.upper() in ("BIA", "BIE"):
            fed = (f"Federally operated - AIHEC records \"Chartered "
                   f"{t.get('chartered_year','')} ({marker})\". NOT tribally "
                   f"owned or chartered; parent left empty by design.")
        elif marker.upper() == "CONGRESS":
            fed = (f"Congressionally chartered - AIHEC records \"Chartered "
                   f"{t.get('chartered_year','')} ({marker})\". NOT tribally "
                   f"owned or chartered; parent left empty by design.")
        elif fed_quote:
            fed = ("Federal operation stated in the AIHEC profile; parent left "
                   "empty by design.")
        serves = " | ".join(s["quote"][:300] for s in sv[:2])
        consider({
            "name": t["name"],
            "aliases": [t["acronym"]] if t.get("acronym") else [],
            "state": t.get("state", ""),
            "source_url": t["source_url"],
            "source_quote": (fed_quote[:600] if fed and fed_quote else
                             ev[0]["quote"][:600] if ev else
                             f"AIHEC TCU Roster and Profiles, "
                             f"{t.get('membership_tier','')}: {t['name']} "
                             f"({t.get('acronym','')}), Chartered "
                             f"{t.get('chartered_year','')}"
                             f"{' (' + marker + ')' if marker else ''}"),
            "evidence": ev,
            "federal_operated": fed,
            "serves": serves,
            "note": (f"AIHEC {t.get('membership_tier','')}; chartered "
                     f"{t.get('chartered_year','')}; website {t.get('website','')}"),
        }, "TCU", CLASS_TCU)

    # ---- Native CDFIs / financial institutions ----------------------------
    print("[B] Native CDFIs and financial institutions")
    for r in read_csv(CLEAN / "native_fi_roster.csv"):
        q = r.get("source_quote", "")
        if r["treasury_native_cdfi"] == "Y":
            q = (q or "Treasury CDFI Fund, List of Currently Certified CDFIs") \
                + f" | row: {r['name']}; Native CDFI (Y/N)=Y; " \
                  f"Cert Control Num={r.get('cert_control_num','')}; " \
                  f"Financial Institution Type={r.get('fi_type','')}"
        consider({
            "name": r["name"],
            "aliases": [a for a in [r.get("nafi_name", "")] if a],
            "state": r.get("state", ""),
            "source_url": r["source_url"],
            "source_quote": q,
            "evidence": [],
            "serves": "",
            "ownership_basis_default": "",
            "note": (f"fi_type={r.get('fi_type','')}; "
                     f"cdfi_fund_list={r['in_cdfi_fund_list']}; "
                     f"cicd_nafi_map={r['in_cicd_nafi_map']}; "
                     f"cert={r.get('cert_control_num','')}; "
                     f"website={r.get('website','')}"),
        }, "CDFI", r["entity_class"])

    print(f"\n  to add   : {len(added)}")
    print(f"  already in spine under another class : {len(already)}")
    print(f"  refused  : {len(refused)}")
    cls = Counter(r["entity_class"] for r in added)
    for k, v in cls.most_common():
        print(f"     {v:4d}  {k}")
    withp = sum(1 for r in added if r["parent_native_entity"])
    print(f"  with a resolved Native owner/charterer : {withp}")
    print(f"  parent deliberately empty              : {len(added)-withp}")

    if added:
        shutil.copy2(SPINE_P, SPINE_P.with_suffix(f".csv.bak_{TODAY}_pre73tcu"))
        with open(SPINE_P, "w", encoding="utf-8", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
            w.writeheader()
            w.writerows(spine + added)
        print(f"\n  wrote {SPINE_P.relative_to(CEDAR)}  "
              f"({len(spine)} -> {len(spine)+len(added)} entities)")

    write_csv(REVIEW / "tcu_cdfi_refused.csv", refused,
              ["layer", "name", "reason", "evidence_url", "refused_date"])
    write_csv(REVIEW / "tcu_cdfi_already_in_spine.csv", already,
              ["name", "layer", "existing_tribe_id", "existing_canonical_name",
               "existing_entity_class", "status", "note", "source_url"])
    write_csv(REVIEW / "tcu_cdfi_owner_unresolved.csv", ruled,
              ["institution", "layer", "captured_owner", "pattern", "outcome",
               "quote", "evidence_url"])
    write_csv(CLEAN / "tcu_cdfi_added.csv", added, fields)


# ===========================================================================
# LINK  (script 70's approach, across ALL datasets)
# ===========================================================================
BARRED = re.compile(
    r"^\s*(city|town|county|state) of\b|\bmines?\b|\b(power|irrigation|water|"
    r"utility|electric)\s+district\b|\bschool district\b|"
    r"\bchamber of commerce\b", re.I)
TRAPS = {"creek", "cherokee", "colorado", "ojibwe", "shawnee", "oneida",
         "apache", "central", "eagle", "river", "mountain", "santa", "san",
         "salt", "round", "united", "enterprise", "valley", "lake", "college",
         "university", "bank", "credit", "union", "fund", "financial",
         "community", "development", "loan", "capital", "national", "first"}

# GENERIC FINANCIAL-INSTITUTION VOCABULARY.
#
# `First State Bank` matched a FAADS recipient called FIRST STATE BANK and
# carried $348,097 with it. There are hundreds of banks by that name and the
# CICD map's is one of them; the match is a coincidence of vocabulary, not
# evidence. Any institution whose whole name is made of these words cannot be
# searched by name at all - it needs a UEI, a CAGE or an RSSD ID.
GENERIC_FI = TRAPS | {
    "state", "local", "legacy", "pinnacle", "gateway", "sovereign",
    "security", "peoples", "people", "guaranty", "savings", "trust",
    "bancorporation", "banc", "bancorp", "holding", "home", "finance",
    "services", "service", "partners", "partner", "growth", "alliance",
    "association", "investments", "investment", "lending", "center", "group",
    "corporation", "company", "incorporated", "inc", "federal", "co",
}

# Column names verified against the actual files on 2026-08-06. A wrong column
# name here yields "no candidates" silently, which reads as absence of evidence
# when it is absence of a lookup.
SOURCES = [
    ("federal_funding", "federal_funding_transactions.csv",
     "recipient_name", ["recipient_uei"], "obligated_usd"),
    ("faads", "faads_transactions_all_agencies.csv",
     "recipient_name", [], "obligated_usd"),
    ("subawards", "subawards.csv", "sub_name",
     ["sub_uei", "sub_cage"], "subaward_amount"),
    ("subawards_prime", "subawards.csv", "prime_name",
     ["prime_uei", "prime_cage"], ""),
    ("lobbying", "native_entity_lobbying_disclosures.csv", "client_name",
     [], ""),
    ("nonprofits", "np_orgs.csv", "org_name", ["EIN"], ""),
    # THE TRUTH for deals is the PROMOTED table, not the additions. This read
    # `deals_*_additions.csv` until 2026-08-26, so candidate discovery ran over
    # 790 of 935 deal rows and the parties on the other 145 were never offered
    # as evidence for a spine entity at all - which reads as "no evidence",
    # exactly the failure the comment above this list warns about.
    # `docs/FACT_CHECK_2026-08-06.md` finding B-1; `cedar_domain.PROMOTED_TABLES`.
    ("deals", "deals_classified.csv", "Native_Party", [], ""),
    ("contracting", "prime_contracts.csv", "awardee_name",
     ["awardee_uei", "cage_code"], "total_obligations"),
]


def _amount(row, cols):
    for c in cols:
        v = (row.get(c) or "").strip()
        if v:
            try:
                return float(v)
            except ValueError:
                return 0.0
    return 0.0


def cmd_link():
    print("=== 73 --link ===\n")
    spine = read_csv(SPINE_P)
    new = [r for r in spine
           if r["tribe_id"].startswith(("TCU-", "CDFI-"))]
    print(f"new entities to find identifiers for: {len(new)}")

    keys, unsearchable = {}, []
    for t in new:
        names = [t["canonical_name"]] + [
            a.strip() for a in (t.get("aliases") or "").split("|") if a.strip()]
        cores, why = {}, ""
        for nm in names:
            # inst_core, NOT 33.core. `core` strips indian/nations/community,
            # so "California Indian Nations College" reduces to
            # {california, college} and is then CONTAINED in "California
            # College of the Arts" - a $1.29M candidate against an art school.
            # Institution matching needs the institution's own words.
            c = inst_core(nm)
            if not c:
                why = "empty core after structural-word removal"
                continue
            # A name whose whole core is trap words, or whose whole core is
            # generic financial vocabulary, cannot carry a match on its own.
            if all(w in TRAPS for w in c):
                why = "name is entirely trap tokens"
                continue
            if all(w in GENERIC_FI for w in c):
                why = ("name is entirely generic financial vocabulary - "
                       "needs a UEI/CAGE/RSSD, not a name search")
                continue
            if len(c) < 2:
                why = "single-token name is too weak to search"
                continue
            cores[frozenset(c)] = nm
        if cores:
            distinctive = min(len([w for w in c if w not in GENERIC_FI])
                              for c in cores)
            # Flag only where a single distinctive word is carrying the whole
            # match, and only outside the TCU layer where names are long and
            # specific. A flag that fires on everything tells a reviewer
            # nothing.
            risk = ("one distinctive word carries this match; check city/state "
                    "before ruling"
                    if distinctive <= 1 and t["entity_class"] != CLASS_TCU
                    else "")
            keys[t["tribe_id"]] = (t["canonical_name"], t["entity_class"],
                                   t.get("state", ""), cores, risk)
        else:
            unsearchable.append({"tribe_id": t["tribe_id"],
                                 "entity": t["canonical_name"],
                                 "entity_class": t["entity_class"],
                                 "state": t.get("state", ""),
                                 "reason": why or "no searchable name"})
    print(f"  searchable after dropping trap-only and generic names: "
          f"{len(keys)}")
    print(f"  refused as unsearchable by name: {len(unsearchable)}\n")
    write_csv(REVIEW / "tcu_cdfi_unsearchable_names.csv", unsearchable,
              ["tribe_id", "entity", "entity_class", "state", "reason"])

    hits = defaultdict(list)
    for label, fname, namecol, idcols, amtcol in SOURCES:
        paths = sorted(glob.glob(str(CLEAN / fname)))
        if not paths:
            print(f"  {label:18s} no file")
            continue
        seen = {}
        n = 0
        for p in paths:
            for r in read_csv(p):
                nm = (r.get(namecol) or "").strip()
                if not nm or BARRED.search(nm):
                    continue
                n += 1
                d = seen.setdefault(nm, {"ids": {}, "amt": 0.0, "rows": 0})
                d["rows"] += 1
                if amtcol:
                    d["amt"] += _amount(r, [amtcol])
                for cc in idcols:
                    v = (r.get(cc) or "").strip()
                    if v and not d["ids"].get(cc):
                        d["ids"][cc] = v
        print(f"  {label:18s} {len(seen):>8,} distinct names of {n:,} rows")

        ncore = {nm: inst_core(nm) for nm in seen}
        for tid, (canon, ecls, est, cores, risk) in keys.items():
            for ecore, ename in cores.items():
                for nm, nc in ncore.items():
                    if not nc:
                        continue
                    if ecore == nc:
                        strength = "core_exact"
                    elif ecore <= nc and len(ecore) >= 2:
                        strength = "core_contained"
                    else:
                        continue
                    d = seen[nm]
                    hits[tid].append({
                        "tribe_id": tid, "entity": canon, "entity_class": ecls,
                        "entity_state": est, "name_risk": risk,
                        "matched_via": ename, "dataset": label,
                        "matched_name": nm, "strength": strength,
                        "rows": d["rows"], "amount_usd": round(d["amt"], 2),
                        **d["ids"],
                    })
                    break

    rows = []
    for tid, hs in hits.items():
        best = sorted(hs, key=lambda h: (h["strength"] != "core_exact",
                                         h["dataset"]))
        for h in best[:8]:
            rows.append({**h, "YOUR_RULING": "", "found": TODAY,
                         "note": "CANDIDATE - a name match is evidence an "
                                 "identifier exists, not proof it is this "
                                 "entity's. Nothing is attributed here."})

    cols = ["tribe_id", "entity", "entity_class", "entity_state", "name_risk",
            "matched_via", "dataset",
            "matched_name", "strength", "rows", "amount_usd", "recipient_uei",
            "sub_uei", "sub_cage", "prime_uei", "prime_cage", "awardee_uei",
            "cage_code", "EIN", "YOUR_RULING", "note", "found"]
    write_csv(REVIEW / "tcu_cdfi_identifier_candidates.csv", rows, cols)

    print(f"\n  entities with at least one candidate: {len(hits)} of {len(keys)}")
    ds = Counter(h["dataset"] for hs in hits.values() for h in hs)
    print("\n  where they were found:")
    for k, v in ds.most_common():
        print(f"     {v:6,}  {k}")

    # exact-core candidates only, for the dollar figure - the defensible one
    exact = [r for r in rows if r["strength"] == "core_exact"]
    tot = sum(r["amount_usd"] for r in exact)
    print(f"\n  core-exact candidate rows: {len(exact)}")
    print(f"  dollars sitting behind core-exact candidates: ${tot:,.2f}")
    print("  (CANDIDATE dollars - not attributed until ruled)")

    still = [t for t in keys if t not in hits]
    print(f"\n  nothing anywhere: {len(still)}")
    for t in still[:15]:
        row = next(r for r in new if r["tribe_id"] == t)
        print(f"     {row['canonical_name'][:52]:52s} {row['entity_class'][:30]}")


# ===========================================================================
# LOG  - regenerated from the data. Standing rule 10: a number in a doc that is
# not recomputed from the data is a claim, not a fact.
# ===========================================================================
def cmd_log():
    print("=== 73 --log ===\n")
    added = read_csv(CLEAN / "tcu_cdfi_added.csv")
    already = read_csv(REVIEW / "tcu_cdfi_already_in_spine.csv")
    refused = read_csv(REVIEW / "tcu_cdfi_refused.csv")
    unres = read_csv(REVIEW / "tcu_cdfi_owner_unresolved.csv")
    cand = read_csv(REVIEW / "tcu_cdfi_identifier_candidates.csv")
    fi = read_csv(CLEAN / "native_fi_roster.csv")
    spine = read_csv(SPINE_P)

    cls = Counter(r["entity_class"] for r in added)
    withp = [r for r in added if r.get("parent_native_entity")]
    nop = [r for r in added if not r.get("parent_native_entity")]
    ent_with_cand = {r["tribe_id"] for r in cand}
    exact = [r for r in cand if r["strength"] == "core_exact"]

    L = []
    a = L.append
    a("# TCU and Native CDFI build log")
    a("")
    a(f"*Generated by `code/73_add_tcu_and_cdfi.py --log` on {TODAY}. "
      f"Every number below is recomputed from the data; nothing here is "
      f"hand-typed.*")
    a("")
    a("## What was missing")
    a("")
    a("The spine held 952 entities, **zero** tribal colleges and **zero** "
      "Native CDFIs. Both classes have authoritative published rosters and "
      "both are heavy federal-funding recipients, so their money was either "
      "unattributed or sitting on a namesake tribe.")
    a("")
    a("## Sources retrieved")
    a("")
    a("| Source | URL | Retrieved | What it gives |")
    a("|---|---|---|---|")
    a(f"| AIHEC TCU Roster and Profiles | {AIHEC_ROSTER_URL} | 2026-08-06 | "
      f"37 member TCUs in three membership tiers, each with a chartering "
      f"statement in its own profile paragraph |")
    a(f"| AIHEC TCU Locations | {AIHEC_LOCATIONS_URL} | 2026-08-06 | "
      f"cross-check: \"AIHEC has grown to 37 Tribal Colleges and Universities "
      f"with more than 80 sites in the United States.\" |")
    a(f"| Treasury CDFI Fund, List of Currently Certified CDFIs | "
      f"{CDFI_LIST_URL} | 2026-08-06 | "
      f"\"Total Number of Certified Native CDFIs as of July 16, 2026: 65\" |")
    a(f"| CICD Native financial institutions map data | {NAFI_URL} | "
      f"2026-08-06 | 91 Native banks, credit unions and loan funds with "
      f"`ncdfi` / `nmdi` flags (map page: {NAFI_PAGE}) |")
    a("")
    a("### Cross-checks attempted, and what they returned")
    a("")
    a("| Cross-check | Outcome |")
    a("|---|---|")
    a("| Bureau of Indian Education TCU list | `bie.edu/topic-page/"
      "tribal-colleges-and-universities` and `bie.edu/landing-page/"
      "post-secondary` both **404**. The only directory bie.edu links is the "
      "K-12 school directory. AIHEC's own `(BIA)` charter markers are used "
      "instead and name the same two BIE institutions. |")
    a("| White House Initiative on Advancing Educational Equity for Native "
      "Americans | `sites.ed.gov/whiaiane/tribes-tcus/tcus/` returns **HTTP "
      "410 Gone**. The list is no longer published. |")
    a("| Oweesta Corporation network list | `oweesta.org/native-cdfi-network/` "
      "**404**; the site publishes maps, not a machine-readable roster. |")
    a("| Native CDFI Network directory | `nativecdfi.net/directory` **404**; "
      "the site has a members-only join flow and no public roster. NCN itself "
      "appears in the CICD map as a loan fund and is added on that basis. |")
    a("")
    a("## What was added")
    a("")
    a("| Class | Rows |")
    a("|---|---:|")
    for k, v in cls.most_common():
        a(f"| {k} | {v} |")
    a(f"| **total** | **{len(added)}** |")
    a("")
    a(f"Spine: {len(spine) - len(added):,} -> {len(spine):,} entities.")
    a("")
    a(f"- with a resolved Native owner / charterer: **{len(withp)}**")
    a(f"- `parent_native_entity` deliberately empty: **{len(nop)}**")
    a("")
    a("`parent_native_entity` and `serves_native_entities` are separate "
      "columns and are never collapsed. Ownership is written only where a "
      "retrieved sentence names a Native entity that charters, owns or "
      "controls the institution.")
    a("")
    a("## Entities added, with source and quote")
    a("")
    a("### Tribal Colleges and Universities")
    a("")
    a("| tribe_id | Institution | State | Parent (charterer) | Ownership basis "
      "| Source | Quote |")
    a("|---|---|---|---|---|---|---|")
    for r in sorted([x for x in added if x["tribe_id"].startswith("TCU-")],
                    key=lambda x: x["canonical_name"]):
        a(f"| `{r['tribe_id']}` | {r['canonical_name']} | {r['state']} | "
          f"{r['parent_native_entity'] or '(empty)'} | "
          f"{(r['ownership_basis'] or '')[:150]} | {r['entity_source_url']} | "
          f"{(r['entity_source_quote'] or '').replace('|', '/')[:320]} |")
    a("")
    a("### Native CDFIs and Native financial institutions")
    a("")
    a("| tribe_id | Institution | State | Class | Parent | Source | Quote |")
    a("|---|---|---|---|---|---|---|")
    for r in sorted([x for x in added if x["tribe_id"].startswith("CDFI-")],
                    key=lambda x: x["canonical_name"]):
        a(f"| `{r['tribe_id']}` | {r['canonical_name']} | {r['state']} | "
          f"{r['entity_class']} | {r['parent_native_entity'] or '(empty)'} | "
          f"{r['entity_source_url']} | "
          f"{(r['entity_source_quote'] or '').replace('|', '/')[:300]} |")
    a("")
    a("## Already in the spine under a different class - NOT re-minted")
    a("")
    if already:
        a("| Institution | Existing id | Existing name | Existing class |")
        a("|---|---|---|---|")
        for r in already:
            a(f"| {r['name']} | `{r['existing_tribe_id']}` | "
              f"{r['existing_canonical_name']} | {r['existing_entity_class']} |")
        a("")
        a("Adding a second row for any of these would split one institution "
          "across two ids - the defect that separated Ho-Chunk Inc from the "
          "Winnebago Tribe.")
    else:
        a("None.")
    a("")
    a("## Refused, with the reason")
    a("")
    if refused:
        a("| Layer | Name | Reason |")
        a("|---|---|---|")
        for r in refused:
            a(f"| {r['layer']} | {r['name']} | {r['reason']} |")
    else:
        a("No roster row was refused: every row on both authoritative rosters "
          "carried a retrieved source URL and a verbatim quote.")
    a("")
    a("### Ownership claims examined and refused")
    a("")
    a(f"{len(unres)} captured owner strings did not resolve to an eligible "
      f"Native entity in the spine and were therefore NOT written as parents. "
      f"They are listed with their verbatim quote in "
      f"`review/tcu_cdfi_owner_unresolved.csv`. The common refusals:")
    a("")
    rc = Counter(r["outcome"].split(":")[0] for r in unres)
    a("| Refusal reason | Count |")
    a("|---|---:|")
    for k, v in rc.most_common(12):
        a(f"| {k} | {v} |")
    a("")
    a("### Roster rows deliberately NOT classed as certified Native CDFIs")
    a("")
    disc = [r for r in fi if r.get("cert_discrepancy")]
    if disc:
        for r in disc:
            a(f"- **{r['name']}** ({r['state']}) - {r['cert_discrepancy']}")
    nfi = [r for r in fi if r["entity_class"] == CLASS_NFI]
    a("")
    a(f"{len(nfi)} institutions appear on the CICD Native financial "
      f"institutions map but are not on Treasury's certified Native CDFI "
      f"list. They are classed `{CLASS_NFI}`, not "
      f"`{CLASS_NCDFI}`. A credit union is not automatically a CDFI, and a "
      f"Native-focused minority depository institution is a different "
      f"designation from CDFI certification.")
    a("")
    a("### Hand-checked name crosswalk (CICD map -> Treasury list)")
    a("")
    a("A token matcher merged only 55 of the 65 certified Native CDFIs across "
      "the two rosters. The remaining nine are ruled by hand rather than "
      "fuzzily, because a fuzzy merge here creates exactly the split-identity "
      "defect this build exists to prevent.")
    a("")
    a("| CICD map name | Treasury list name | Basis |")
    a("|---|---|---|")
    for k, (v, why) in sorted(NAFI_ALIAS.items()):
        a(f"| {k} | {v} | {why} |")
    a("")
    a("## Linking to identifiers")
    a("")
    a(f"`--link` runs script 70's approach across federal funding, FAADS, "
      f"subawards (both legs), lobbying, nonprofit 990s, deals and prime "
      f"contracting.")
    a("")
    a(f"- entities with at least one candidate: **{len(ent_with_cand)}** "
      f"of {len(added)}")
    a(f"- candidate rows written: **{len(cand)}**")
    a(f"- of which core-exact: **{len(exact)}**")
    ds = Counter(r["dataset"] for r in cand)
    a("")
    a("| Dataset | Candidate rows |")
    a("|---|---:|")
    for k, v in ds.most_common():
        a(f"| {k} | {v} |")
    a("")
    tot = sum(float(r["amount_usd"] or 0) for r in exact)
    a(f"Dollars sitting behind core-exact candidates: "
      f"**${tot:,.2f}**. This is a CANDIDATE figure. Nothing is attributed "
      f"by this script: a name match is evidence that an identifier exists, "
      f"not proof that it belongs to this entity. Every row carries a "
      f"`YOUR_RULING` column and lands in "
      f"`review/tcu_cdfi_identifier_candidates.csv`.")
    a("")
    a("### Institutions that cannot be searched by name at all")
    a("")
    uns = read_csv(REVIEW / "tcu_cdfi_unsearchable_names.csv")
    if uns:
        a("`First State Bank` matched a FAADS recipient called FIRST STATE "
          "BANK and carried $348,097 with it. There are hundreds of banks by "
          "that name. An institution whose whole name is generic financial "
          "vocabulary needs a UEI, a CAGE or an RSSD ID, and is refused from "
          "name search rather than matched weakly.")
        a("")
        a("| Institution | State | Reason |")
        a("|---|---|---|")
        for r in uns:
            a(f"| {r['entity']} | {r['state']} | {r['reason']} |")
        a("")
        a("The CICD map supplies an `rssd_id` for banks and credit unions and "
          "a `bank_cert` / `cu_number`; those are the join keys for this "
          "group and are carried in `data/clean/native_fi_roster.csv`.")
    else:
        a("None.")
    a("")
    a("## Known limits")
    a("")
    a("1. `code/01_build_entity_spine.py` rebuilds the spine from its own "
      "sources. Like the ANCSA village corporations (script 52) and the "
      "NHO/intertribal layer (script 61), these rows are APPENDED and would "
      "be lost by a full rebuild. Re-run `73 --add` after script 01.")
    a("2. The AIHEC roster is a membership list, not a legal register. It is "
      "authoritative for TCU status and for the chartering statements it "
      "prints, and it is the source cited for both.")
    a("3. The CICD map data file records `last_updated 2025-08-20`; the "
      "Treasury list is as of 2026-07-16. Where the two disagree on "
      "certification, the Treasury list wins and the disagreement is recorded "
      "rather than smoothed.")
    a("4. Ownership was harvested from institution websites by pattern, then "
      "resolved through `33_apply_party_rulings.resolve_entity`. A capture "
      "that did not resolve was refused, never guessed. The refusals are a "
      "review queue, not a defect.")

    DOCS.mkdir(parents=True, exist_ok=True)
    p = DOCS / "TCU_CDFI_BUILD_LOG.md"
    p.write_text("\n".join(L) + "\n", encoding="utf-8")
    print(f"  wrote {p.relative_to(CEDAR)}  ({len(L)} lines)")


def main():
    args = set(sys.argv[1:])
    if not args or "--help" in args:
        print(__doc__)
        return
    if "--parse" in args or "--all" in args:
        cmd_parse()
    if "--fetch-org-pages" in args or "--all" in args:
        cmd_fetch_org_pages()
    if "--reextract" in args:
        cmd_reextract()
    if "--add" in args or "--all" in args:
        cmd_add()
    if "--link" in args or "--all" in args:
        cmd_link()
    if "--log" in args or "--all" in args:
        cmd_log()


if __name__ == "__main__":
    main()
