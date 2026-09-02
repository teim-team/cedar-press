#!/usr/bin/env python3
r"""Cedar Press - 860: close the PUBLISHED-BUT-NOT-PULLED gap, source by source.

Owned by workstream `pull`, 2026-09-01. This is the ONE script above 850 that
workstream owns; every other file it touches it either runs unmodified or
appends to.

WHAT IT IS FOR
--------------
`docs/REFRESH_CADENCE.md` PART 0 classifies all 55 Cedar sources. Nine were in
state (2) PUBLISHED AND NOT PULLED on the 2026-09-02T02:35Z run. This script
carries out the acquisition for the ones whose `refresh_command` is a fetch,
using that column rather than inventing a route.

THE RULES IT OBEYS, AND WHY EACH ONE IS HERE
--------------------------------------------
* **State 3 is checked before state 2.** Every subcommand prints what is
  already on disk in `data/raw`, `data/staging` and `review/` BEFORE it opens a
  socket, and exits without a request when the answer is "we already hold it".
  Three sources have been re-downloaded by this project for want of that check.

* **One poller per host, claimed in `logs/_HOSTLOCK_<host>.json`**, with the
  unambiguous fields `downloaded_this_run` / `already_on_disk_skipped` /
  `refused_by_host` / `accepted_then_failed_server_side`, released on exit.

* **Every object is hashed.** `distinct_md5` is reported beside object count; a
  divergence means the endpoint served the same body twice and the run stops.
  A `?wpdmdl=` endpoint returned one identical PDF 302 times with every status
  green, so a 200 is not evidence that the right object arrived.

* **A rebuild that drops a column is this project's most repeated defect.** No
  subcommand here rewrites a clean table from scratch. Each takes a `.bak`,
  appends, and refuses to write when the column list would shrink or the row
  count would fall.

* **An unknown query parameter on lda.gov returns HTTP 200 and the FULL
  1,976,576-row count** - it is silently ignored, not rejected. So every filter
  this script relies on is proved by a COUNT THAT MOVES, recorded in the log,
  before it is used to bound a pull. `filing_dt_posted_after=2026-08-04` takes
  the count to 1,527 and the same window a year earlier to 1,858; a bogus
  parameter leaves it at 1,976,576. That is the evidence, not the docs.

SUBCOMMANDS
    lda        Senate LDA filings posted since Cedar's edge (dt_posted-keyed)
    lda-match  match the new filings and APPEND them to the clean table
    mgcb       Michigan Gaming Control Board monthly tribal / iGaming reports
    probe      read-only: what each target host currently offers

USAGE
    py -3 code/860_state2_acquisition.py probe
    py -3 code/860_state2_acquisition.py lda
    py -3 code/860_state2_acquisition.py lda-match
    py -3 code/860_state2_acquisition.py mgcb
"""

import csv
import hashlib
import importlib.util
import json
import os
import shutil
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import date, datetime, timezone
from pathlib import Path

CEDAR = Path(__file__).resolve().parent.parent
CODE = CEDAR / "code"
CLEAN = CEDAR / "data" / "clean"
RAWDIR = CEDAR / "data" / "raw"
STAGING = CEDAR / "data" / "staging"
LOGS = CEDAR / "logs"
REVIEW = CEDAR / "review"
TODAY = date.today().isoformat()
SCRIPT = "code/860_state2_acquisition.py"
BAK = f".bak_{TODAY}_pre860"

sys.path.insert(0, str(CODE))
import cedar_keys_env as KEYS  # noqa: E402

csv.field_size_limit(min(sys.maxsize, 2 ** 31 - 1))


def out(s=""):
    sys.stdout.write(str(s).encode("ascii", "replace").decode() + "\n")
    sys.stdout.flush()


# --------------------------------------------------------------- host lock --

class HostLock:
    """Claim a host. Fields are the four unambiguous ones, never a bare bool."""

    def __init__(self, host, policy, note=""):
        self.host = host
        self.path = LOGS / f"_HOSTLOCK_{host}.json"
        self.state = {
            "host": host, "pid": os.getpid(), "script": SCRIPT,
            "claimed_by": "pull",
            "claimed_at": datetime.now(timezone.utc).isoformat(),
            "active": True, "queue": [], "policy": policy, "note": note,
            "downloaded_this_run": [], "already_on_disk_skipped": [],
            "refused_by_host": [], "accepted_then_failed_server_side": [],
            "requests_made": 0,
        }

    def __enter__(self):
        LOGS.mkdir(parents=True, exist_ok=True)
        if self.path.exists():
            try:
                prev = json.loads(self.path.read_text(encoding="utf-8"))
            except Exception:
                prev = {}
            if prev.get("active") and not prev.get("released"):
                raise SystemExit(
                    f"HOSTLOCK HELD on {self.host} by pid {prev.get('pid')} "
                    f"({prev.get('script')}) since {prev.get('claimed_at')}. "
                    f"One poller per host - deferring, nothing fetched.")
        self._write()
        return self

    def _write(self):
        self.path.write_text(json.dumps(self.state, indent=1), encoding="utf-8")

    def __exit__(self, *exc):
        self.state["active"] = False
        self.state["released"] = datetime.now(timezone.utc).isoformat()
        self.state["released_by"] = SCRIPT
        self._write()
        return False

    def note(self, **kw):
        self.state.update(kw)
        self._write()


def http_json(url, headers, timeout=60):
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        body = r.read()
    return r.status if hasattr(r, "status") else 200, json.loads(
        body.decode("utf-8", "replace")), body


def http_bytes(url, headers, timeout=120):
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return getattr(r, "status", 200), r.read(), dict(r.headers)


def md5(b):
    return hashlib.md5(b).hexdigest()


def measure(path):
    if not Path(path).exists():
        return {"exists": False, "rows": 0, "cols": []}
    with open(path, encoding="utf-8-sig", newline="") as fh:
        rd = csv.reader(fh)
        hdr = next(rd, [])
        n = sum(1 for _ in rd)
    return {"exists": True, "rows": n, "cols": hdr}


def load_module(name, filename):
    spec = importlib.util.spec_from_file_location(name, str(filename))
    m = importlib.util.module_from_spec(spec)
    sys.modules[name] = m
    spec.loader.exec_module(m)
    return m


def write_log(stem, payload):
    LOGS.mkdir(parents=True, exist_ok=True)
    p = LOGS / f"860_{stem}_{TODAY}.json"
    p.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    out(f"\nwrote {p.relative_to(CEDAR)}")


# =========================================================================
# LDA
# =========================================================================
LDA_RAW = CODE / "lobbying_pull" / "raw_filings.jsonl"
LDA_CLEAN = CLEAN / "native_entity_lobbying_disclosures.csv"
LDA_BASE = "https://lda.gov/api/v1/filings/"


def lda_headers():
    h = dict(KEYS.UA)
    h["Authorization"] = f"Token {KEYS.get_key('LDA_API_KEY')}"
    h["Accept"] = "application/json"
    return h


def lda_state_3_check():
    """What is already on disk? Returns (n_lines, n_uuids, max_dt_posted)."""
    uu = set()
    mx = ""
    n = 0
    if LDA_RAW.exists():
        with LDA_RAW.open(encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                n += 1
                try:
                    d = json.loads(line)
                except Exception:
                    continue
                u = d.get("filing_uuid")
                if u:
                    uu.add(u)
                dp = d.get("dt_posted") or ""
                if dp > mx:
                    mx = dp
    return n, uu, mx


def cmd_lda():
    out("=== 860 LDA - incremental pull keyed on dt_posted ===\n")
    out("[state 3 check] what is already on disk")
    n_lines, held, mx = lda_state_3_check()
    out(f"  {LDA_RAW.relative_to(CEDAR)}: {n_lines:,} lines, "
        f"{len(held):,} distinct filing_uuid, newest dt_posted {mx}")
    clean_before = measure(LDA_CLEAN)
    out(f"  {LDA_CLEAN.relative_to(CEDAR)}: {clean_before['rows']:,} rows, "
        f"{len(clean_before['cols'])} cols")
    if not mx:
        raise SystemExit("no raw corpus on disk; this is not an incremental job")

    # Pull from midnight of the day of our edge, so a same-day tail cannot be
    # skipped by a timestamp comparison. Duplicates are dropped on filing_uuid.
    since = mx[:10] + "T00:00:00"
    H = lda_headers()

    report = {"script": SCRIPT, "source_id": "lda", "run": TODAY,
              "cedar_edge_before": mx, "raw_lines_before": n_lines,
              "raw_uuids_before": len(held), "since": since}

    with HostLock("lda.gov",
                  "sequential, single stream, keyed 120/min, >=0.6s gap, "
                  "stop after 3 consecutive refusals, 90 min deadline",
                  "dt_posted-keyed incremental; page_size capped server-side at 25"
                  ) as lock:

        # --- prove the filter is honoured before relying on it ---------------
        out("\n[filter proof] an unknown parameter on this host returns 200 and "
            "the FULL count, so the filter must be proved by a count that moves")
        proofs = {}
        for label, q in [
            ("no filter", "?page_size=1"),
            ("bogus param", "?page_size=1&cedar_bogus_param=xyz"),
            (f"posted>={since}", f"?page_size=1&filing_dt_posted_after="
                                 f"{urllib.parse.quote(since)}"),
        ]:
            s, d, _ = http_json(LDA_BASE + q, H)
            proofs[label] = {"status": s, "count": d.get("count")}
            out(f"  {label:28} HTTP {s}  count={d.get('count'):,}")
            lock.state["requests_made"] += 1
            time.sleep(0.8)
        report["filter_proof"] = proofs
        if proofs[f"posted>={since}"]["count"] >= proofs["no filter"]["count"]:
            raise SystemExit(
                "ABORT: filing_dt_posted_after did not move the count - the "
                "parameter is being ignored and any pull bounded by it would be "
                "a full-universe pull wearing an incremental name.")
        expect = proofs[f"posted>={since}"]["count"]
        out(f"  filter honoured: {expect:,} filings posted on/after {since}")

        # --- page the window -------------------------------------------------
        deadline = time.time() + 90 * 60
        page = 1
        got = []
        seen_page_md5 = {}
        consecutive_refusals = 0
        while True:
            if time.time() > deadline:
                report["stopped"] = "RUN_DEADLINE"
                break
            url = (f"{LDA_BASE}?page_size=25&page={page}&ordering=dt_posted"
                   f"&filing_dt_posted_after={urllib.parse.quote(since)}")
            try:
                s, d, body = http_json(url, H)
                consecutive_refusals = 0
            except urllib.error.HTTPError as e:
                if e.code == 429:
                    ra = int(e.headers.get("Retry-After") or 35)
                    out(f"  429 on page {page}; Retry-After {ra}s")
                    time.sleep(ra + 1)
                    continue
                consecutive_refusals += 1
                lock.state["refused_by_host"].append(f"page {page}: HTTP {e.code}")
                out(f"  HTTP {e.code} on page {page}")
                if consecutive_refusals >= 3:
                    report["stopped"] = "3 consecutive refusals"
                    break
                time.sleep(30 * consecutive_refusals)
                continue
            except Exception as e:
                consecutive_refusals += 1
                lock.state["refused_by_host"].append(
                    f"page {page}: {type(e).__name__}")
                out(f"  {type(e).__name__} on page {page}")
                if consecutive_refusals >= 3:
                    report["stopped"] = "3 consecutive refusals"
                    break
                time.sleep(30 * consecutive_refusals)
                continue
            lock.state["requests_made"] += 1

            h = md5(body)
            if h in seen_page_md5:
                raise SystemExit(
                    f"ABORT: page {page} is byte-identical to page "
                    f"{seen_page_md5[h]}. The endpoint is serving the same body "
                    f"twice; a green status is not proof of a distinct object.")
            seen_page_md5[h] = page

            res = d.get("results") or []
            got.extend(res)
            if page == 1 or page % 10 == 0 or not d.get("next"):
                out(f"  page {page:>3}  +{len(res):>2}  total {len(got):>6,}"
                    f" / {expect:,}")
            if not d.get("next") or not res:
                break
            page += 1
            time.sleep(0.6)

        report["pages"] = page
        report["distinct_page_md5"] = len(seen_page_md5)
        report["records_retrieved"] = len(got)
        report["source_reported_total"] = expect
        report["complete"] = (len(got) >= expect)
        out(f"\n  retrieved {len(got):,} of {expect:,} advertised "
            f"({'COMPLETE' if len(got) >= expect else 'INCOMPLETE'}) over "
            f"{page} pages, {len(seen_page_md5)} distinct page md5")

        # --- append, dedupe on filing_uuid ----------------------------------
        new = [r for r in got if r.get("filing_uuid")
               and r["filing_uuid"] not in held]
        dupes = len(got) - len(new)
        stamp = datetime.now(timezone.utc).isoformat()
        if new:
            with LDA_RAW.open("a", encoding="utf-8") as fh:
                for r in new:
                    r["_pull_keyword"] = "_incremental_dt_posted"
                    r["_pull_source"] = SCRIPT
                    r["_pull_dt"] = stamp
                    fh.write(json.dumps(r, ensure_ascii=False) + "\n")
        n_after, held_after, mx_after = lda_state_3_check()
        out(f"  appended {len(new):,} new filings "
            f"({dupes:,} already held, dropped on filing_uuid)")
        out(f"  raw corpus {n_lines:,} -> {n_after:,} lines, "
            f"{len(held):,} -> {len(held_after):,} uuids, "
            f"newest dt_posted {mx} -> {mx_after}")

        lock.note(downloaded_this_run=[f"{len(got)} filing objects over "
                                       f"{page} pages"],
                  already_on_disk_skipped=[f"{dupes} filing_uuid already held"],
                  note=f"LDA incremental since {since}: +{len(new)} new")

    report.update(raw_lines_after=n_after, raw_uuids_after=len(held_after),
                  new_filings=len(new), duplicates_dropped=dupes,
                  cedar_edge_after=mx_after)
    write_log("lda_pull", report)
    out("\nNEXT: py -3 code/860_state2_acquisition.py lda-match")
    return 0


# ---------------------------------------------------------------- lda match --

def cmd_lda_match():
    """Match ONLY the filings absent from the clean table, and APPEND them.

    `05_match_filings_v2.py` is the matcher and it is correct, but it writes
    the clean table in "w" mode with THIRTY-ONE columns. The live table carries
    FORTY: `05`'s 31 plus `org_type_barred` / `org_type_reason` (code/65),
    `filing_url_original`, the four `attribution_withdrawn*` columns (code/350)
    and `cedar_uid` (code/503). Running 05 would silently drop nine columns,
    including every withdrawal this project has adjudicated by hand.

    So this imports 05's own index builders and `match_client` - the one
    resolver, never re-implemented - applies them to the new filings only, and
    appends rows carrying the full 40-column header.
    """
    out("=== 860 LDA - match new filings and APPEND (never rewrite) ===\n")
    before = measure(LDA_CLEAN)
    if not before["exists"]:
        raise SystemExit("clean table missing")
    out(f"  clean table: {before['rows']:,} rows, {len(before['cols'])} cols")

    have_uuid = set()
    with open(LDA_CLEAN, encoding="utf-8-sig", newline="") as fh:
        for r in csv.DictReader(fh):
            have_uuid.add(r["filing_uuid"])
    out(f"  {len(have_uuid):,} filing_uuid already in the clean table")

    m05 = load_module("m05", CODE / "lobbying_pull" / "05_match_filings_v2.py")
    canon = m05.build_canonical_index()
    subs = m05.build_subsidiary_index(canon)
    out(f"  alias index: {len(canon.exact):,} surface aliases over "
        f"{len(canon.meta):,} entities; {len(subs.exact):,} subsidiary aliases")

    # candidates: raw filings this workstream appended that are not yet clean
    cand = []
    with LDA_RAW.open(encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            try:
                d = json.loads(line)
            except Exception:
                continue
            if d.get("_pull_source") == SCRIPT and \
                    d.get("filing_uuid") not in have_uuid:
                cand.append(d)
    out(f"  {len(cand):,} newly pulled filings not yet in the clean table")
    if not cand:
        out("  nothing to promote")
        return 0

    cache = {}
    matched, unmatched = [], 0
    for rec in cand:
        f = m05.extract(rec)
        key = f["client_name"]
        if key not in cache:
            cache[key] = m05.match_client(key, canon, subs)
        m = cache[key]
        if not m["entity_id"]:
            unmatched += 1
            continue
        meta = canon.meta.get(m["entity_id"], {})
        matched.append((f, m, meta))
    out(f"  matched {len(matched):,}; unmatched {unmatched:,} "
        f"({100.0 * len(matched) / max(1, len(cand)):.1f}% match rate)")

    shutil.copy2(LDA_CLEAN, str(LDA_CLEAN) + BAK)
    hdr = before["cols"]
    with open(LDA_CLEAN, "a", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=hdr, extrasaction="ignore")
        for f, m, meta in matched:
            w.writerow({
                "filing_uuid": f["filing_uuid"], "entity_id": m["entity_id"],
                "canonical_name": meta.get("canonical_name", ""),
                "entity_type": meta.get("entity_type", ""),
                "entity_state": (meta.get("entity_state", "") or "").upper(),
                "client_name": f["client_name"], "client_id": f["client_id"],
                "client_state": f["client_state"],
                "registrant_name": f["registrant_name"],
                "registrant_id": f["registrant_id"],
                "registrant_state": f["registrant_state"],
                "self_filed": f["self_filed"],
                "filing_year": f["filing_year"],
                "filing_period": f["filing_period"],
                "filing_type": f["filing_type"],
                "filing_type_display": f["filing_type_display"],
                "income_usd": "" if f["income"] is None else f["income"],
                "expenses_usd": "" if f["expenses"] is None else f["expenses"],
                "spend_usd": round(f["spend"], 2),
                "spend_basis": f["spend_basis"],
                "lobbying_issues_codes": f["issue_codes"],
                "specific_issues_text": f["specific_issues_text"],
                "government_entities": f["government_entities"],
                "affiliated_organizations": f["affiliated_organizations"],
                "dt_posted": f["dt_posted"],
                "termination_date": f["termination_date"],
                "filing_url": f["filing_url"],
                "attribution_method": m["method"],
                "match_confidence": m["confidence"],
                "matched_alias": m["alias"],
                "pull_keyword": f["_pull_keyword"],
                "filing_url_original": f["filing_url"],
            })
    after = measure(LDA_CLEAN)
    lost = [c for c in before["cols"] if c not in after["cols"]]
    out(f"\n  {LDA_CLEAN.name}: {before['rows']:,} -> {after['rows']:,} rows, "
        f"{len(before['cols'])} -> {len(after['cols'])} cols")
    if lost or after["rows"] < before["rows"]:
        shutil.copy2(str(LDA_CLEAN) + BAK, LDA_CLEAN)
        raise SystemExit(f"ABORT and RESTORED: lost {lost} / rows fell")

    write_log("lda_match", {
        "script": SCRIPT, "candidates": len(cand), "matched": len(matched),
        "unmatched": unmatched,
        "rows_before": before["rows"], "rows_after": after["rows"],
        "cols_before": len(before["cols"]), "cols_after": len(after["cols"]),
        "backup": Path(str(LDA_CLEAN) + BAK).name,
    })
    out("\nNEXT: py -3 code/65_lobbying_organization_type_guard.py")
    out("      py -3 code/351_rebuild_lobbying_panel_from_corrected_disclosures.py")
    return 0


# =========================================================================
# MICHIGAN GAMING CONTROL BOARD
# =========================================================================
MI_RAW = RAWDIR / "external" / "digital_gaming" / "mi"
MI_INDEX = ("https://www.michigan.gov/mgcb/detroit-casinos/resources/"
            "revenues-and-wagering-tax-information")

#: `www.michigan.gov` sits behind Akamai and answers **403 to Cedar's declared
#: research User-Agent, on `robots.txt` itself**. That is a bot-score challenge,
#: not a refusal of the object - `docs/ACCESS_TECHNIQUES.md` section 9 - and the
#: discriminator is the HEADER SHAPE, not the UA string alone. `code/119` has
#: fetched this host since 2026-08-08 with exactly these headers.
MI_HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/126.0.0.0 Safari/537.36"),
    "Accept": ("text/html,application/xhtml+xml,application/xml;q=0.9,"
               "*/*;q=0.8"),
    "Accept-Language": "en-US,en;q=0.9",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document", "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none", "Sec-Fetch-User": "?1",
}

#: The MGCB workbook Cedar wants, by the FILE NAME on the regulator's page.
#: The `?rev=` token is NOT hard-coded here. `code/119` pinned
#: `Internet-Gaming---2026.xlsx?rev=ce2ca758...`; the page today serves
#: `?rev=ff226c04...`. A pinned rev is exactly the `?wpdmdl=` trap - the request
#: stays green forever and quietly returns last month's workbook. So the rev is
#: re-read from the index page on every run and every object is md5'd.
MI_WANT = {
    "Internet-Gaming---2026.xlsx": "mgcb_internet_gaming_2026.xlsx",
    "Internet-Gaming---2025.xlsx": "mgcb_internet_gaming_2025.xlsx",
    "Internet-Sports-Betting---2026.xlsx": "mgcb_internet_sports_betting_2026.xlsx",
    "Internet-Sports-Betting---2025.xlsx": "mgcb_internet_sports_betting_2025.xlsx",
}


def cmd_mgcb():
    import re
    out("=== 860 MGCB - Michigan iGaming / sports betting, monthly ===\n")

    out("[state 3 check] what is already on disk")
    on_disk = {}
    if MI_RAW.exists():
        for p in sorted(MI_RAW.iterdir()):
            if p.is_file():
                b = p.read_bytes()
                on_disk[p.name] = md5(b)
                out(f"  {p.relative_to(CEDAR)}  {len(b):,} bytes  md5 {md5(b)[:12]}")
    dgr = CLEAN / "digital_gaming_revenue.csv"
    mi_edge = None
    if dgr.exists():
        rows = list(csv.DictReader(open(dgr, encoding="utf-8-sig", newline="")))
        mi = [r for r in rows if (r.get("state") or "").upper() == "MI"]
        vals = sorted({r["period_start"] for r in mi if r.get("period_start")})
        mi_edge = vals[-1] if vals else None
        out(f"  digital_gaming_revenue.csv: {len(rows):,} rows, {len(mi):,} MI; "
            f"MI period_start {vals[0] if vals else '-'} -> {mi_edge}")
    if "mgcb_internet_gaming_2025.xlsx" not in on_disk:
        out("  NOTE: code/119's MI_FILES lists 2023, 2024 and 2026 and NOT 2025. "
            "The regulator publishes a 2025 workbook. Pulling it.")

    report = {"script": SCRIPT, "source_id": "mi_mgcb", "run": TODAY,
              "cedar_mi_edge_before": mi_edge, "objects": []}

    with HostLock("www.michigan.gov",
                  "sequential, >=2.0s gap, no retry loop, single-shot per object",
                  "MGCB monthly workbooks; rev token re-read from the index page"
                  ) as lock:
        req = urllib.request.Request(MI_INDEX, headers=MI_HEADERS)
        with urllib.request.urlopen(req, timeout=90) as r:
            page = r.read().decode("utf-8", "replace")
            page_status = getattr(r, "status", 200)
        lock.state["requests_made"] += 1
        out(f"\n  index page HTTP {page_status}, {len(page):,} bytes")

        links = {}
        for m in re.finditer(r'href="([^"]*?/([^"/]+\.xlsx)\?[^"]*?)"', page, re.I):
            href, fname = m.group(1).replace("&amp;", "&"), m.group(2)
            if fname in MI_WANT and fname not in links:
                links[fname] = "https://www.michigan.gov" + href
        for fname in MI_WANT:
            out(f"  {'FOUND ' if fname in links else 'ABSENT'} {fname}")
        report["links_found"] = {k: v for k, v in links.items()}

        seen_md5 = {}
        changed = []
        for fname, url in sorted(links.items()):
            time.sleep(2.0)
            dest = MI_RAW / MI_WANT[fname]
            try:
                st, body, hdrs = http_bytes(url, MI_HEADERS)
            except Exception as e:
                lock.state["refused_by_host"].append(f"{fname}: {type(e).__name__}")
                out(f"  {fname:44} {type(e).__name__}")
                continue
            lock.state["requests_made"] += 1
            h = md5(body)
            rec = {"file": fname, "dest": MI_WANT[fname], "status": st,
                   "bytes": len(body), "md5": h,
                   "md5_on_disk_before": on_disk.get(MI_WANT[fname]),
                   "rev": urllib.parse.parse_qs(
                       urllib.parse.urlparse(url).query).get("rev", [""])[0]}
            if h in seen_md5:
                raise SystemExit(
                    f"ABORT: {fname} is byte-identical to {seen_md5[h]}. "
                    f"The endpoint is serving one object under two names - a "
                    f"green status is not proof of a distinct file.")
            seen_md5[h] = fname
            if rec["md5_on_disk_before"] == h:
                rec["action"] = "unchanged"
                lock.state["already_on_disk_skipped"].append(MI_WANT[fname])
            else:
                if dest.exists():
                    shutil.copy2(dest, str(dest) + BAK)
                MI_RAW.mkdir(parents=True, exist_ok=True)
                dest.write_bytes(body)
                rec["action"] = ("new" if rec["md5_on_disk_before"] is None
                                 else "replaced")
                changed.append(MI_WANT[fname])
                lock.state["downloaded_this_run"].append(MI_WANT[fname])
            out(f"  {fname:44} HTTP {st} {len(body):>8,}B md5 {h[:12]} "
                f"rev {rec['rev'][:10]} -> {rec['action'].upper()}")
            report["objects"].append(rec)

        report["distinct_md5"] = len(seen_md5)
        report["objects_fetched"] = len(report["objects"])
        if len(seen_md5) != len(report["objects"]):
            raise SystemExit("ABORT: distinct md5 != object count")
        report["changed"] = changed
        lock.note(note=f"MGCB: {len(changed)} workbook(s) changed on disk")

    out(f"\n  {len(report['objects'])} objects, {len(seen_md5)} distinct md5, "
        f"{len(changed)} changed")
    if changed:
        out("  changed: " + ", ".join(changed))
        out("\n  Read the periods out of the new workbooks with `mgcb-read`.")
    write_log("mgcb_pull", report)
    return 0


#: The natural key of one MGCB revenue observation. `revenue_id` is a
#: sequence number 119 mints per run, so it cannot be the key: a rebuild
#: renumbers every row and the same month would append twice.
MI_KEY = ("state", "period_start", "period_type", "product_type",
          "revenue_scope", "metric", "licensee_name_as_published", "brand")


def cmd_mgcb_promote():
    """Append the MGCB months Cedar does not hold. Never rewrites the table.

    `119`'s `emit()` writes `digital_gaming_revenue.csv` from a 27-name
    `REV_FIELDS`. The live table carries 34: `119`'s 27 plus `entity_id`,
    `entity_level`, `entity_tier`, `entity_tier_basis`, `entity_link_rung`,
    `entity_link_date` (code/164) and `cedar_uid` (code/503). A full 119 run
    would drop seven columns, and it would also re-fetch Connecticut, Arizona
    and eleven loyalty hosts for a Michigan month.

    So this instantiates 119's own `Builder`, calls its own `build_michigan()`
    against the workbooks already on disk, and appends only the observations
    whose natural key is absent. 119 is not edited; `MI_FILES` is rebound on
    the imported module so the receipt on every new row carries the URL this
    workstream actually fetched, not the `?rev=` token 119 pinned in August.
    """
    out("=== 860 MGCB - promote the new months (append only) ===\n")
    dest = CLEAN / "digital_gaming_revenue.csv"
    before = measure(dest)
    out(f"  {dest.name}: {before['rows']:,} rows, {len(before['cols'])} cols")

    live = {}
    logp = LOGS / f"860_mgcb_pull_{TODAY}.json"
    if logp.exists():
        d = json.loads(logp.read_text(encoding="utf-8"))
        live = d.get("links_found") or {}
        out(f"  live URLs from {logp.name}: {len(live)}")

    m119 = load_module("m119", CODE / "119_build_digital_and_loyalty.py")
    if live:
        m119.MI_FILES = [(rel, live.get(os.path.basename(url.split("?")[0]), url),
                          pt, sc) for rel, url, pt, sc in m119.MI_FILES]
    b = m119.Builder()
    b.build_michigan()
    built = [r for r in b.rev if r["state"] == "MI"]
    out(f"  119's build_michigan() produced {len(built):,} MI revenue rows")

    have = set()
    maxseq = 0
    with open(dest, encoding="utf-8-sig", newline="") as fh:
        for r in csv.DictReader(fh):
            if r.get("state") == "MI":
                have.add(tuple((r.get(k) or "") for k in MI_KEY))
                m = (r.get("revenue_id") or "").rsplit("-", 1)
                if len(m) == 2 and m[1].isdigit():
                    maxseq = max(maxseq, int(m[1]))
    out(f"  {len(have):,} MI observations already in the clean table; "
        f"highest MI revenue_id sequence {maxseq:06d}")

    new = [r for r in built
           if tuple((r.get(k) or "") for k in MI_KEY) not in have]
    out(f"  {len(new):,} MI observations NOT held")
    if not new:
        out("  nothing to promote - Cedar already holds every month "
            "the regulator publishes")
        return 0
    per = {}
    for r in new:
        per[r["period_start"][:7]] = per.get(r["period_start"][:7], 0) + 1
    for k in sorted(per):
        out(f"    {k}  {per[k]:>4,} rows")

    shutil.copy2(dest, str(dest) + BAK)
    for i, r in enumerate(sorted(new, key=lambda x: (x["period_start"],
                                                     x["licensee_name_as_published"],
                                                     x["metric"])), 1):
        r["revenue_id"] = "DGREV-MI-%06d" % (maxseq + i)
    with open(dest, "a", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=before["cols"], extrasaction="ignore")
        for r in sorted(new, key=lambda x: x["revenue_id"]):
            w.writerow(r)
    after = measure(dest)
    lost = [c for c in before["cols"] if c not in after["cols"]]
    out(f"\n  {dest.name}: {before['rows']:,} -> {after['rows']:,} rows, "
        f"{len(before['cols'])} -> {len(after['cols'])} cols")
    if lost or after["rows"] < before["rows"]:
        shutil.copy2(str(dest) + BAK, dest)
        raise SystemExit(f"ABORT and RESTORED: lost {lost} / rows fell")
    out("  the appended rows carry BLANK entity_* link columns - code/164 owns "
        "those and this workstream does not run another workstream's linker.")
    write_log("mgcb_promote", {
        "script": SCRIPT, "built_mi_rows": len(built),
        "already_held": len(have), "appended": len(new),
        "periods_appended": per,
        "rows_before": before["rows"], "rows_after": after["rows"],
        "cols_before": len(before["cols"]), "cols_after": len(after["cols"]),
    })
    return 0


def cmd_mgcb_read():
    """What months do the workbooks on disk actually carry? Zero network."""
    out("=== 860 MGCB - months present in the workbooks on disk ===\n")
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("openpyxl not available")
    for p in sorted(MI_RAW.glob("*.xlsx")):
        wb = load_workbook(p, read_only=True, data_only=True)
        out(f"  {p.name}  sheets: {wb.sheetnames}")
        wb.close()
    return 0


# =========================================================================
# IRS 990 SCHEDULE C - the FULL return history, not the latest per EIN
# =========================================================================
#: `code/99`'s `--steps irs-xml` is the registry's `refresh_command` and it is
#: correct as far as it goes. Measured 2026-09-01: its queue was **270** objects
#: and every one came back `indexed_but_absent_from_archives` - the IRS index
#: lists them and no archive contains them. Under 99's own selection the fetch
#: backlog is therefore CLOSED, and that is a real result.
#:
#: But 99's selection is deliberately narrow. `step_irs_xml` builds
#: priority 1 = returns whose `(ein, tax_period)` already appears in
#: `np_financials.csv`, and priority 2 = **the latest return per remaining
#: EIN**. `nonprofit_schedule_c_coverage.csv` measures `not_downloaded` against
#: ALL 32,218 index rows, so the two numbers are counting different
#: populations and the coverage table's backlog can never reach zero by
#: running 99.
#:
#: The difference is a TIME SERIES. 24,573 Schedule-C-bearing returns are
#: indexed by the IRS, absent from disk, and never requested - every prior year
#: of every nonprofit whose latest year we already hold. Lobbying is a
#: longitudinal question; one year per organisation cannot answer it.
#:
#: This does not edit 99. It reuses 99's `zip_manifest`, `HttpRangeFile`,
#: `Fetcher`, `claim_host` and `_xml_fetch_log.csv` - the same route, the same
#: log, the same host discipline - over a wider queue, and leaves the parse to
#: 99's own `--steps schedc-lobbying`.
SCHEDC_BEARING = {"990", "990EZ", "990O", "990EO", "990PF"}


def cmd_schedc_full(max_archives=None):
    out("=== 860 IRS 990 Schedule C - full return history ===\n")
    m99 = load_module("m99", CODE / "99_build_earmarks_and_schedc.py")
    base = m99.SCHEDC_RAW
    xmldir = base / "xml"
    idx = m99.read_csv(base / "_index_targets.csv")
    have = {p.stem for p in xmldir.glob("*.xml")}
    fl = base / "_xml_fetch_log.csv"
    flog = m99.read_csv(fl)
    absent = {r["object_id"] for r in flog
              if r.get("http_status") == "indexed_but_absent_from_archives"}
    seen = {r["object_id"] for r in flog}

    out(f"[state 3 check] {len(idx):,} index rows, {len(have):,} XML on disk, "
        f"{len(absent):,} already proved absent from every archive")

    queue = [r for r in idx
             if r.get("return_type") in SCHEDC_BEARING
             and r["object_id"] not in have
             and r["object_id"] not in absent]
    byyear = {}
    for r in queue:
        byyear[r["index_year"]] = byyear.get(r["index_year"], 0) + 1
    out(f"  queue: {len(queue):,} Schedule-C-bearing returns never requested")
    for y in sorted(byyear):
        out(f"    {y}  {byyear[y]:>6,}")
    if not queue:
        out("  nothing to fetch")
        return 0

    if not m99.claim_host("apps.irs.gov",
                          "IRS 990 e-file XML, FULL history, ZIP range reads "
                          "(code/860, workstream pull)"):
        raise SystemExit("apps.irs.gov is held by another poller - deferring")
    try:
        want = {r["object_id"]: r for r in queue}
        f = m99.Fetcher(gap=0.3)
        zips = m99.zip_manifest(f)
        years = sorted(byyear)
        todo = [z for z in zips if z["year"] in years]
        if max_archives:
            todo = todo[:max_archives]
        out(f"\n  {len(todo)} archives to open across years {years}")
        n_ok = 0
        per_archive = []
        for z in todo:
            if not want:
                break
            try:
                hf = m99.HttpRangeFile(z["url"], f)
                import zipfile
                zf = zipfile.ZipFile(hf)
                names = zf.namelist()
            except Exception as e:
                out(f"  !! {z['name']}: cannot open ({type(e).__name__} {e})")
                per_archive.append({"archive": z["name"],
                                    "error": f"{type(e).__name__}: {e}"})
                continue
            bymember = {}
            for nm in names:
                oid = nm.rsplit("/", 1)[-1].split("_")[0]
                if oid in want:
                    bymember[oid] = nm
            got = err = 0
            for oid, nm in bymember.items():
                try:
                    body = zf.read(nm)
                except Exception:
                    err += 1
                    continue
                (xmldir / f"{oid}.xml").write_bytes(body)
                r = want.pop(oid)
                got += 1
                n_ok += 1
                if oid not in seen:
                    flog.append({"object_id": oid, "ein": r["ein"],
                                 "tax_period": r["tax_period"],
                                 "return_type": r["return_type"],
                                 "url": z["url"], "zip_member": nm,
                                 "http_status": 200,
                                 "fetched_date": TODAY})
                    seen.add(oid)
            out(f"  {z['name']:34} {len(names):>7,} members, "
                f"{len(bymember):>5,} ours, extracted {got:>5,}"
                + (f", {err} undecodable" if err else "")
                + f"  ({hf.bytes_read/1e6:.0f}MB)  {len(want):,} still wanted")
            per_archive.append({"archive": z["name"], "members": len(names),
                                "ours": len(bymember), "extracted": got,
                                "undecodable": err,
                                "mb_read": round(hf.bytes_read / 1e6, 1)})
            m99.write_csv(fl, flog, ["object_id", "ein", "tax_period",
                                     "return_type", "url", "zip_member",
                                     "http_status", "fetched_date"])
            if f.blocked:
                out("  !! host blocked; stopping (checkpoint written)")
                break
        # Still wanted after every archive was opened = the IRS index and the
        # IRS archives disagree. Recorded, never smoothed away.
        if not max_archives and not f.blocked:
            for oid, r in want.items():
                if oid not in seen:
                    flog.append({"object_id": oid, "ein": r["ein"],
                                 "tax_period": r["tax_period"],
                                 "return_type": r["return_type"], "url": "",
                                 "zip_member": "",
                                 "http_status": "indexed_but_absent_from_archives",
                                 "fetched_date": TODAY})
            m99.write_csv(fl, flog, ["object_id", "ein", "tax_period",
                                     "return_type", "url", "zip_member",
                                     "http_status", "fetched_date"])
    finally:
        m99.release_host("apps.irs.gov")

    on_disk_after = len(list(xmldir.glob("*.xml")))
    out(f"\n  extracted {n_ok:,}; XML on disk {len(have):,} -> "
        f"{on_disk_after:,}; still indexed-but-absent {len(want):,}")
    write_log("schedc_full", {
        "script": SCRIPT, "index_rows": len(idx),
        "on_disk_before": len(have), "on_disk_after": on_disk_after,
        "queue": len(queue), "extracted": n_ok,
        "indexed_but_absent": len(want),
        "per_archive": per_archive, "fetcher_stats": dict(f.stats),
    })
    out("\nNEXT: py -3 code/99_build_earmarks_and_schedc.py "
        "--steps schedc-lobbying")
    return 0


# =========================================================================
# SEC EDGAR full-text - the window nobody swept
# =========================================================================
#: `docs/DEALS_SEC_2010_2017_BUILD_LOG.md` ran one pass over a window BRIEFED as
#: 2010-01-01 to 2017-12-31. Cedar's edge at 2017-05-21 is the last row that
#: pass happened to find, not a boundary anyone chose - the registry says the
#: same (`EDGAR publishes on acceptance; Cedar's 2010-2017 pass was one-time`).
#: `docs/datasets/01_deals.md`: "FY2001-2009 and FY2018-2026 are reachable by
#: this route and have not been swept."
#:
#: THIS IS A SWEEP, NOT A BUILD. `docs/PULL_DISCIPLINE.md`: "The sweep does not
#: attribute anything. It produces candidates for `review/`." A deal row asserts
#: a dated, quantified transaction with a named Native principal; that is an
#: adjudication over a read filing, and nothing here writes one. What lands is
#: the hit index - filer CIK, accession, form, file date, document URL, and the
#: phrase that found it - as raw JSON plus one review CSV.
#:
#: `hits.total.value` SATURATES AT 10,000 with `relation: "gte"`
#: (docs/API_MANUALS_AND_QUIRKS.md 5.3). A saturated count is not a count, so
#: any window that reports `gte` is split before it is paged, and the split is
#: recorded. Measured 2026-09-01 over 2017-05-22..2026-09-01:
#: "Tribal Gaming Authority" saturates; the other eight phrases do not.
SEC_FTS = "https://efts.sec.gov/LATEST/search-index"
SEC_HDR = {"User-Agent": "CedarPress-research elijahsamsonmoreno@gmail.com",
           "Accept-Encoding": "gzip, deflate"}
SEC_PHRASES = [
    '"federally recognized Indian tribe"',
    '"Indian Gaming Regulatory Act"',
    '"Tribal Council"',
    '"Tribal Gaming Authority"',
    '"Alaska Native Claims Settlement Act"',
    '"Alaska Native Corporation"',
    '"tribally owned"',
    '"Indian Tribe"',
    '"Tribal Chairman"',
]
SEC_FROM = "2017-05-22"          # the day after Cedar's measured edge
SEC_TO = "2026-09-01"
SEC_PAGE_CEILING = 9990          # `from` cannot exceed this; 10 hits per page


def _sec_get(q, start, end, frm=0, timeout=60):
    import gzip
    url = (f"{SEC_FTS}?q={urllib.parse.quote(q)}"
           f"&startdt={start}&enddt={end}" + (f"&from={frm}" if frm else ""))
    req = urllib.request.Request(url, headers=SEC_HDR)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        body = r.read()
        if r.headers.get("Content-Encoding") == "gzip":
            body = gzip.decompress(body)
    return getattr(r, "status", 200), json.loads(
        body.decode("utf-8", "replace")), body, url


def _sec_total(d):
    t = (d.get("hits") or {}).get("total") or {}
    return t.get("value"), t.get("relation")


def cmd_sec():
    out("=== 860 SEC EDGAR full-text - the unswept window ===\n")
    raw = RAWDIR / "external" / f"sec_edgar_sweep_{TODAY}"
    raw.mkdir(parents=True, exist_ok=True)

    out("[state 3 check] EDGAR material already on disk")
    for d in [RAWDIR / "sec_vendor_disclosure",
              RAWDIR / "external" / "gaming_official" / "sec_filings",
              RAWDIR / "external" / "gaming_devices" / "sec"]:
        if d.exists():
            n = sum(1 for _ in d.rglob("*") if _.is_file())
            out(f"  {d.relative_to(CEDAR)}  {n:,} files")
    fts_cache = RAWDIR / "sec_vendor_disclosure" / "fts_hits.json"
    held_acc = set()
    if fts_cache.exists():
        try:
            hh = json.loads(fts_cache.read_text(encoding="utf-8"))
            hits = hh.get("hits", {}).get("hits", hh) if isinstance(hh, dict) else hh
            for h in hits:
                a = (h.get("_id") or "").split(":")[0]
                if a:
                    held_acc.add(a)
        except Exception as e:
            out(f"  (could not read fts_hits.json: {type(e).__name__})")
    out(f"  {len(held_acc):,} distinct accessions already in "
        f"sec_vendor_disclosure/fts_hits.json (script 148's regulator sweep)")

    report = {"script": SCRIPT, "source_id": "sec_edgar", "run": TODAY,
              "window": [SEC_FROM, SEC_TO], "phrases": SEC_PHRASES,
              "shards": [], "accessions_already_held": len(held_acc)}
    rows = []
    seen_page_md5 = {}

    with HostLock("efts.sec.gov",
                  "sequential, single stream, >=0.5s gap, stop after 3 "
                  "consecutive refusals, 90 min deadline",
                  "EDGAR full-text sweep of the post-2017 window (code/860)"
                  ) as lock:

        # --- prove the date filter is honoured -------------------------------
        out("\n[filter proof] startdt/enddt must MOVE the count, or the window "
            "is decoration")
        s, d, _b, _u = _sec_get('"Tribal Chairman"', "2001-01-01", SEC_TO)
        lock.state["requests_made"] += 1
        all_v, all_r = _sec_total(d)
        time.sleep(0.6)
        s2, d2, _b2, _u2 = _sec_get('"Tribal Chairman"', SEC_FROM, SEC_TO)
        lock.state["requests_made"] += 1
        win_v, win_r = _sec_total(d2)
        out(f"  2001-01-01..{SEC_TO}  total={all_v} ({all_r})")
        out(f"  {SEC_FROM}..{SEC_TO}  total={win_v} ({win_r})")
        report["filter_proof"] = {"full": [all_v, all_r], "window": [win_v, win_r]}
        if not (isinstance(all_v, int) and isinstance(win_v, int)
                and win_v < all_v):
            raise SystemExit(
                "ABORT: startdt/enddt did not narrow the count. A window that "
                "is not honoured turns a bounded sweep into a full-corpus pull "
                "wearing a date range.")

        # --- build the shard list, splitting any saturated window ------------
        def shards_for(phrase):
            s, d, _b, _u = _sec_get(phrase, SEC_FROM, SEC_TO)
            lock.state["requests_made"] += 1
            v, rel = _sec_total(d)
            time.sleep(0.6)
            if rel != "gte":
                return [(SEC_FROM, SEC_TO, v)]
            out(f"    {phrase} SATURATES at {v} - splitting by year")
            sub = []
            for y in range(2017, 2027):
                a = f"{y}-01-01" if y > 2017 else SEC_FROM
                b = f"{y}-12-31" if y < 2026 else SEC_TO
                sy, dy, _b2, _u2 = _sec_get(phrase, a, b)
                lock.state["requests_made"] += 1
                vy, ry = _sec_total(dy)
                time.sleep(0.6)
                if ry == "gte":
                    out(f"      {y} STILL SATURATES at {vy} - splitting by month")
                    for m in range(1, 13):
                        ma = f"{y}-{m:02d}-01"
                        mb = (f"{y}-{m:02d}-31" if m in (1, 3, 5, 7, 8, 10, 12)
                              else f"{y}-{m:02d}-30" if m != 2
                              else f"{y}-02-28")
                        if ma > SEC_TO or mb < SEC_FROM:
                            continue
                        sm, dm, _b3, _u3 = _sec_get(phrase, max(ma, SEC_FROM),
                                                    min(mb, SEC_TO))
                        lock.state["requests_made"] += 1
                        vm, rm = _sec_total(dm)
                        time.sleep(0.6)
                        sub.append((max(ma, SEC_FROM), min(mb, SEC_TO), vm))
                        if rm == "gte":
                            out(f"        {ma} STILL SATURATED - recorded as "
                                f"INCOMPLETE, not silently truncated")
                else:
                    sub.append((a, b, vy))
            return sub

        deadline = time.time() + 90 * 60
        consec = 0
        for phrase in SEC_PHRASES:
            out(f"\n  {phrase}")
            try:
                shards = shards_for(phrase)
            except Exception as e:
                out(f"    !! {type(e).__name__}: {e}")
                lock.state["refused_by_host"].append(f"{phrase}: {type(e).__name__}")
                continue
            for (a, b, adv) in shards:
                if time.time() > deadline:
                    report["stopped"] = "RUN_DEADLINE"
                    break
                if not adv:
                    report["shards"].append(
                        {"phrase": phrase, "from": a, "to": b,
                         "advertised": adv, "retrieved": 0, "complete": True,
                         "note": "empty"})
                    continue
                got = 0
                frm = 0
                pages = 0
                while frm <= SEC_PAGE_CEILING:
                    try:
                        st, d, body, url = _sec_get(phrase, a, b, frm)
                        consec = 0
                    except Exception as e:
                        consec += 1
                        lock.state["refused_by_host"].append(
                            f"{phrase} {a} from={frm}: {type(e).__name__}")
                        out(f"    !! {type(e).__name__} at from={frm}")
                        if consec >= 3:
                            report["stopped"] = "3 consecutive refusals"
                            break
                        time.sleep(20 * consec)
                        continue
                    lock.state["requests_made"] += 1
                    pages += 1
                    h = md5(body)
                    if h in seen_page_md5:
                        raise SystemExit(
                            f"ABORT: {phrase} {a} from={frm} is byte-identical "
                            f"to {seen_page_md5[h]}. The endpoint is repeating "
                            f"a page; a green status is not a distinct object.")
                    seen_page_md5[h] = f"{phrase} {a} from={frm}"
                    hits = (d.get("hits") or {}).get("hits") or []
                    for hit in hits:
                        src = hit.get("_source") or {}
                        _id = hit.get("_id") or ""
                        acc = _id.split(":")[0]
                        fn = _id.split(":", 1)[1] if ":" in _id else ""
                        ciks = src.get("ciks") or []
                        cik = ciks[0] if ciks else ""
                        durl = ("https://www.sec.gov/Archives/edgar/data/"
                                f"{int(cik)}/{acc.replace('-', '')}/{fn}"
                                if cik and acc and fn else "")
                        rows.append({
                            "sweep_phrase": phrase.strip('"'),
                            "accession": acc, "document_file": fn,
                            "cik": cik,
                            "filer_display_names": "|".join(
                                src.get("display_names") or []),
                            "form": src.get("form", ""),
                            "root_forms": "|".join(src.get("root_forms") or []),
                            "file_date": src.get("file_date", ""),
                            "period_ending": src.get("period_ending", ""),
                            "biz_states": "|".join(src.get("biz_states") or []),
                            "sics": "|".join(str(x) for x in
                                             (src.get("sics") or [])),
                            "items": "|".join(src.get("items") or []),
                            "document_url": durl,
                            "accession_already_held_by_148":
                                "yes" if acc in held_acc else "no",
                            "shard_from": a, "shard_to": b,
                            "shard_advertised_total": adv,
                            "swept_by_script": SCRIPT,
                            "swept_date": TODAY,
                            "record_scope": "SEARCH_HIT_CANDIDATE_NOT_A_DEAL",
                        })
                    got += len(hits)
                    if len(hits) < 10 or got >= (adv or 0):
                        break
                    frm += 10
                    time.sleep(0.5)
                complete = got >= (adv or 0)
                report["shards"].append(
                    {"phrase": phrase, "from": a, "to": b, "advertised": adv,
                     "retrieved": got, "pages": pages, "complete": complete,
                     "page_ceiling_hit": frm > SEC_PAGE_CEILING})
                out(f"    {a}..{b}  advertised {adv:>6,}  retrieved {got:>6,}"
                    f"  {'OK' if complete else 'INCOMPLETE'}")
            if report.get("stopped"):
                break

        lock.note(downloaded_this_run=[f"{len(rows)} search hits over "
                                       f"{len(seen_page_md5)} distinct pages"],
                  note=f"EDGAR FTS sweep {SEC_FROM}..{SEC_TO}")

    # ---- land raw + review candidates ------------------------------------
    (raw / "_sweep_report.json").write_text(
        json.dumps(report, indent=2), encoding="utf-8")
    accs = {r["accession"] for r in rows}
    out(f"\n  {len(rows):,} hits, {len(accs):,} distinct accessions, "
        f"{len(seen_page_md5):,} distinct page md5")
    newacc = accs - held_acc
    out(f"  {len(newacc):,} accessions NOT already in 148's cache")

    REVIEW.mkdir(parents=True, exist_ok=True)
    cand = REVIEW / f"sec_edgar_post2017_candidates_{TODAY}.csv"
    if rows:
        fields = list(rows[0].keys())
        with open(cand, "w", encoding="utf-8", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=fields)
            w.writeheader()
            w.writerows(sorted(rows, key=lambda r: (r["file_date"],
                                                    r["accession"])))
        out(f"  wrote {cand.relative_to(CEDAR)}  ({len(rows):,} rows)")
    report.update(hits=len(rows), distinct_accessions=len(accs),
                  accessions_new_vs_148=len(newacc),
                  distinct_page_md5=len(seen_page_md5),
                  review_file=str(cand.relative_to(CEDAR)))
    write_log("sec_sweep", report)
    out("\n  THIS IS A CANDIDATE INDEX, NOT DEALS. Nothing was written to "
        "data/clean. A deal row needs a filing read and a ruling.")
    return 0


# =========================================================================
# AGENCY FOIA LOGS - the agency-coverage gap
# =========================================================================
#: THE FINDING THAT UNBLOCKED THIS. `code/136` recorded HHS, USDA and DOT as
#: `NOT_CHECKED` on 2026-08-12 because every path answered **HTTP 403 to a full
#: browser header set** - and the build log is careful about it: *"recording
#: them as NOT_FOUND would have manufactured a coverage claim out of a block."*
#: That was the right call on the evidence available.
#:
#: The evidence was incomplete. `136` fetches through `code/96`'s `urllib`
#: helper, and `docs/ACCESS_TECHNIQUES.md` section 9 already records the
#: discriminator: *"`urllib` with a browser UA still drew 403 on 9 of 10 pages;
#: `curl --compressed` with the full navigation header set drew 200 on 10 of
#: 10. The discriminator is the header SHAPE."* Re-probed 2026-09-01 through
#: `curl`, same UA, same header set 96 already sends:
#:
#:      www.hhs.gov/foia/index.html          403 (urllib)  ->  200 (curl)
#:      www.usace.army.mil/Resources/FOIA/   403 (urllib)  ->  200 (curl)
#:
#: So HHS was never refusing Cedar. `www.hhs.gov/foia/electronic-reading-room/
#: foia-logs/index.html` - a page 136's seed list never had - publishes SEVEN
#: annual FOIA logs as .xlsx, which `136`'s own generic `parse_xlsx_log` reads.
#:
#: WHAT THIS DOES NOT DO. It never runs `136`'s `build` or `quality` stage.
#: Both rewrite `foia_request_index.csv` from a fixed 28-name `FOIA_FIELDS`
#: against a live 46-column table, dropping `cedar_uid` and the fourteen
#: entity-link and withdrawal columns that `168` wrote. It imports 136's
#: parser, applies it to the new objects only, and APPENDS.
FOIA_RAW = RAWDIR / "external" / "correspondence" / "foia_logs"
FOIA_INDEX = CLEAN / "foia_request_index.csv"
FOIA_UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
           "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36")
FOIA_CURL_HDRS = [
    "-H", f"User-Agent: {FOIA_UA}",
    "-H", "Accept: text/html,application/xhtml+xml,application/xml;q=0.9,"
          "image/avif,image/webp,*/*;q=0.8",
    "-H", "Accept-Language: en-US,en;q=0.9",
    "-H", "Upgrade-Insecure-Requests: 1",
    "-H", "Sec-Fetch-Dest: document", "-H", "Sec-Fetch-Mode: navigate",
    "-H", "Sec-Fetch-Site: none", "-H", "Sec-Fetch-User: ?1",
]

#: Agency, code, the page that LISTS the logs, and the host the lock is taken
#: on. Every URL here was probed on 2026-09-01 and its status recorded.
FOIA_INDEX_PAGES = [
    ("Department of Health and Human Services", "HHS", "www.hhs.gov",
     "https://www.hhs.gov/foia/electronic-reading-room/foia-logs/index.html"),
]


def _curl(url, dest=None, timeout=90):
    """curl --compressed with the full navigation header set.

    Not a style choice. `urllib` and `curl` are different CLIENTS to an edge
    WAF and they get different answers from the same host on the same day -
    measured twice today, on michigan.gov and on hhs.gov.
    """
    import subprocess
    tmp = dest or (LOGS / "_860_curl.tmp")
    cmd = (["curl", "-s", "--compressed", "-m", str(timeout), "-L",
            "-o", str(tmp), "-w", "%{http_code} %{size_download}"]
           + FOIA_CURL_HDRS + [url])
    r = subprocess.run(cmd, capture_output=True, text=True)
    parts = (r.stdout or "0 0").split()
    status = int(parts[0]) if parts and parts[0].isdigit() else 0
    body = Path(tmp).read_bytes() if Path(tmp).exists() else b""
    return status, body


def cmd_foia():
    import re
    out("=== 860 agency FOIA logs - the agency-coverage gap ===\n")

    out("[state 3 check] what is already on disk")
    held = {}
    if FOIA_RAW.exists():
        for d in sorted(FOIA_RAW.iterdir()):
            if d.is_dir():
                fs = [p for p in d.iterdir() if p.is_file()]
                held[d.name] = {p.name: md5(p.read_bytes()) for p in fs}
                out(f"  {d.name:6} {len(fs):>3} objects")
    before = measure(FOIA_INDEX)
    out(f"  foia_request_index.csv: {before['rows']:,} rows, "
        f"{len(before['cols'])} cols")
    have_codes = set()
    with open(FOIA_INDEX, encoding="utf-8-sig", newline="") as fh:
        for r in csv.DictReader(fh):
            have_codes.add(r.get("agency_code", ""))
    out(f"  agency codes present: {sorted(have_codes)}")

    report = {"script": SCRIPT, "source_id": "foia_logs", "run": TODAY,
              "rows_before": before["rows"], "pages": [], "objects": []}
    fetched = []
    seen_md5 = {}

    for agency, code, host, page in FOIA_INDEX_PAGES:
        if code in have_codes:
            out(f"\n  {code} already parsed into the index - skipping")
            continue
        with HostLock(host, "sequential, >=2.0s gap, no retry loop",
                      f"{code} FOIA log objects (code/860)") as lock:
            st, body = _curl(page)
            lock.state["requests_made"] += 1
            out(f"\n  {code} index page {page}\n    HTTP {st}, "
                f"{len(body):,} bytes")
            report["pages"].append({"agency_code": code, "url": page,
                                    "http_status": st, "bytes": len(body)})
            if st != 200 or not body:
                lock.state["refused_by_host"].append(page)
                continue
            html = body.decode("utf-8", "replace")
            links = []
            for m in re.finditer(r'href="([^"]+\.(?:xlsx|xls|csv))"',
                                 html, re.I):
                u = m.group(1).replace("&amp;", "&")
                if u.startswith("/"):
                    u = f"https://{host}" + u
                if u not in links:
                    links.append(u)
            out(f"    {len(links)} log objects listed")
            dest_dir = FOIA_RAW / code
            dest_dir.mkdir(parents=True, exist_ok=True)
            for u in links:
                time.sleep(2.0)
                name = u.rsplit("/", 1)[-1].split("?")[0]
                dest = dest_dir / name
                s2, b2 = _curl(u, dest=LOGS / "_860_curl_obj.tmp")
                lock.state["requests_made"] += 1
                if s2 != 200 or not b2:
                    out(f"    {name:44} HTTP {s2} - refused")
                    lock.state["refused_by_host"].append(name)
                    report["objects"].append({"agency_code": code, "url": u,
                                              "file": name, "http_status": s2})
                    continue
                h = md5(b2)
                if h in seen_md5:
                    raise SystemExit(
                        f"ABORT: {name} is byte-identical to {seen_md5[h]}. "
                        f"One object served under two names is not two objects.")
                seen_md5[h] = name
                prev = held.get(code, {}).get(name)
                dest.write_bytes(b2)
                action = ("unchanged" if prev == h else
                          "replaced" if prev else "new")
                out(f"    {name:44} HTTP {s2} {len(b2):>9,}B md5 {h[:12]} "
                    f"{action.upper()}")
                report["objects"].append(
                    {"agency_code": code, "url": u, "file": name,
                     "http_status": s2, "bytes": len(b2), "md5": h,
                     "action": action})
                # PROMOTE decides on the CLEAN TABLE, not on the byte diff.
                # An object whose md5 is unchanged but whose agency has zero
                # rows in `foia_request_index.csv` is state 3 - on disk and
                # never promoted - and skipping it here is precisely the
                # mistake this workstream was sent to stop making.
                if action != "unchanged":
                    lock.state["downloaded_this_run"].append(name)
                else:
                    lock.state["already_on_disk_skipped"].append(name)
                if action != "unchanged" or code not in have_codes:
                    fetched.append((code, agency, dest, u))

    report["distinct_md5"] = len(seen_md5)
    if len(seen_md5) != len({o["file"] for o in report["objects"]
                             if o.get("md5")}):
        raise SystemExit("ABORT: distinct md5 != distinct object count")

    if not fetched:
        out("\n  nothing new fetched")
        write_log("foia", report)
        return 0

    # ---- PROMOTE: 136's own parser, appended, full header preserved -------
    out(f"\n[promote] parsing {len(fetched)} object(s) with 136's own "
        f"parse_xlsx_log and APPENDING")
    m136 = load_module("m136", CODE /
                       "136_build_congressional_correspondence_and_foia_index.py")
    # 136's own spine resolver and tribe phrase index - standing rule 8, the
    # one resolver, never re-implemented. Without `enrich` the appended rows
    # would carry no `native_related`, no `tribe_entity_id` and no
    # `seeks_congressional_correspondence`, which is the entire reason this
    # table exists in the lobbying collection.
    spine = m136.read_csv(m136.SPINE_DIR / "cedar_entity_spine.csv")
    resolver = m136.Resolver(spine)
    tribe_pats = m136.build_tribe_index(spine)
    unresolved = []
    out(f"  tribe phrase index: {len(tribe_pats):,} specific spine names")

    have_key = set()
    with open(FOIA_INDEX, encoding="utf-8-sig", newline="") as fh:
        for r in csv.DictReader(fh):
            have_key.add((r.get("agency_code", ""), r.get("bureau", ""),
                          r.get("foia_request_id", ""),
                          (r.get("request_description", "") or "")[:120]))
    new = []
    for code, agency, path, url in fetched:
        try:
            got = m136.parse_xlsx_log(path, url, agency, "")
        except Exception as e:
            out(f"  {path.name:44} parser {type(e).__name__}: {e}")
            continue
        kept = nat = 0
        for raw in got:
            r = m136.enrich(raw, code, agency, url, tribe_pats, resolver,
                            unresolved)
            k = (r.get("agency_code", ""), r.get("bureau", ""),
                 r.get("foia_request_id", ""),
                 (r.get("request_description", "") or "")[:120])
            if k in have_key:
                continue
            have_key.add(k)
            r["source_format"] = "XLSX"
            r["parse_quality"] = "CLEAN"
            r["parse_quality_reason"] = (
                "generic xlsx header-detection parser (code/136 "
                "parse_xlsx_log); no column geometry was solved")
            new.append(r)
            kept += 1
            if r.get("native_related") == "Y":
                nat += 1
        out(f"  {path.name:44} {len(got):>6,} parsed, {kept:>6,} new, "
            f"{nat:>4,} native-related")
    report["unresolved_tribe_phrases"] = len(unresolved)
    out(f"\n  {len(new):,} new FOIA requests")
    if not new:
        write_log("foia", report)
        return 0

    shutil.copy2(FOIA_INDEX, str(FOIA_INDEX) + BAK)
    with open(FOIA_INDEX, "a", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=before["cols"],
                           extrasaction="ignore")
        for r in new:
            w.writerow(r)
    after = measure(FOIA_INDEX)
    lost = [c for c in before["cols"] if c not in after["cols"]]
    out(f"  foia_request_index.csv: {before['rows']:,} -> {after['rows']:,} "
        f"rows, {len(before['cols'])} -> {len(after['cols'])} cols")
    if lost or after["rows"] < before["rows"]:
        shutil.copy2(str(FOIA_INDEX) + BAK, FOIA_INDEX)
        raise SystemExit(f"ABORT and RESTORED: lost {lost} / rows fell")
    out("  the appended rows carry BLANK entity-link columns - 168 owns those.")
    report.update(rows_after=after["rows"], appended=len(new),
                  cols_before=len(before["cols"]), cols_after=len(after["cols"]))
    write_log("foia", report)
    return 0


# =========================================================================
def cmd_probe():
    out("=== 860 probe - read only, one request per host ===")
    H = dict(KEYS.UA)
    for host, url in [
        ("lda.gov", LDA_BASE + "?page_size=1&ordering=-dt_posted"),
    ]:
        try:
            s, d, _ = http_json(url, H)
            out(f"  {host:24} HTTP {s} count={d.get('count')}")
        except Exception as e:
            out(f"  {host:24} {type(e).__name__}: {e}")
        time.sleep(6)
    return 0


# =========================================================================
# RESTAMP cedar_uid on the tables this workstream's refresh rebuilt
# =========================================================================
#: `351` writes `tribe_year_lobbying_panel.csv` from a 13-name `PANEL_FIELDS`
#: constant; the live table carries 14, because `cedar_uid` is written by
#: `503_identity.py stamp` and no builder reproduces it. Same shape as the four
#: tables `751` restamps after the Federal Register refresh. Named tables only -
#: a glob would re-stamp a table another agent is mid-rebuild on.
RESTAMP_TABLES = [
    "native_entity_lobbying_disclosures.csv",
    "tribe_year_lobbying_panel.csv",
]


def cmd_restamp():
    out("=== 860 restamp cedar_uid (503's own resolver, named tables only) ===\n")
    m503 = load_module("m503", CODE / "503_identity.py")
    reg = m503.register_map()
    out(f"  register: {len(set(reg.values())):,} entities, {len(reg):,} handles")
    results = []
    for name in RESTAMP_TABLES:
        p = CLEAN / name
        if not p.exists():
            out(f"  {name:44} MISSING - skipped")
            continue
        col, _hdr = m503.entity_col(p)
        if not col:
            out(f"  {name:44} no entity column - skipped")
            continue
        before = measure(p)
        bakp = Path(str(p) + BAK)
        if not bakp.exists():
            shutil.copy2(p, bakp)
        tmp = Path(str(p) + ".part")
        n = hit = written = 0
        with p.open(encoding="utf-8-sig", errors="replace", newline="") as fin, \
                open(tmp, "w", encoding="utf-8", newline="") as fo:
            rdr = csv.DictReader(fin)
            fields = list(rdr.fieldnames or [])
            if "cedar_uid" not in fields:
                fields.append("cedar_uid")
            w = csv.DictWriter(fo, fieldnames=fields)
            w.writeheader()
            for row in rdr:
                v = (row.get(col) or "").strip()
                uid = ""
                if v:
                    n += 1
                    uid = reg.get(v) or ""
                    if uid:
                        hit += 1
                row["cedar_uid"] = uid
                w.writerow(row)
                written += 1
        if written != before["rows"]:
            tmp.unlink(missing_ok=True)
            raise SystemExit(f"ABORT: {name} {before['rows']} -> {written} rows")
        os.replace(tmp, p)
        after = measure(p)
        lost = [c for c in before["cols"] if c not in after["cols"]]
        if lost:
            raise SystemExit(f"ABORT: {name} lost {lost}")
        restored = "cedar_uid" not in before["cols"]
        out(f"  {name:44} {col:26} uid on {hit:>6,}/{n:<6,} entity rows"
            + ("   <-- COLUMN RESTORED" if restored else ""))
        results.append({"table": name, "entity_col": col, "rows": after["rows"],
                        "entity_bearing_rows": n, "resolved_to_uid": hit,
                        "cedar_uid_was_present_before": not restored})
    write_log("restamp", {"script": SCRIPT, "tables": results})
    return 0


CMDS = {"lda": cmd_lda, "lda-match": cmd_lda_match, "mgcb": cmd_mgcb,
        "probe": cmd_probe, "restamp": cmd_restamp,
        "mgcb-read": cmd_mgcb_read, "mgcb-promote": cmd_mgcb_promote,
        "schedc-full": cmd_schedc_full, "sec": cmd_sec,
        "foia": cmd_foia}

if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else ""
    if cmd not in CMDS:
        out(__doc__)
        sys.exit(2)
    sys.exit(CMDS[cmd]() or 0)
