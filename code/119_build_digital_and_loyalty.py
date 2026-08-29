#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
Cedar Press - digital gaming and loyalty layer.
=================================================================

    code/119_build_digital_and_loyalty.py     built 2026-08-07

WHAT THIS IS
------------
Digital gaming is a SEPARATE LINKED UNIVERSE, not a casino amenity. A tribe may
hold any combination of a retail sportsbook, mobile betting confined to its own
premises, statewide mobile rights, an outside sportsbook designee, iGaming
rights, its own digital brand, a platform provider, and a former provider that
has ceased. None of those maps one-to-one onto a physical property, so
`facility_id` is nullable and is blank far more often than it is filled.

THE ONE RULE THAT ORGANISES THE WHOLE BUILD
-------------------------------------------
**A compact authorisation is not an operation.** The compacts say who MAY. The
state regulator's monthly file says who DOES. Those are two different facts and
they live in two different columns: `compact_authority_cite` and `launch_date`.
A launch is NEVER inferred from a right. The gap between the two is the finding
this dataset exists to expose.

WRITES
------
    data/clean/digital_gaming_relationships.csv
    data/clean/digital_gaming_revenue.csv
    data/clean/loyalty_programs.csv
    data/clean/loyalty_program_property.csv
    data/clean/codebook/16_digital_gaming.csv      (fragment; master untouched)
    review/digital_gaming_unresolved_<date>.csv
    data/raw/external/digital_gaming/**            + _SOURCE_MANIFEST.csv (md5)
    data/interim/119_run_summary.txt

READS ONLY (never written)
--------------------------
    data/clean/compact_structured_terms.csv        (2,887 terms, 27 states)
    data/clean/gaming_facilities.csv               (774 properties)
    data/clean/nigc_declination_letters.csv        (327 letters)
    data/clean/gaming_financing_events.csv
    data/spine/cedar_entity_spine.csv

FOUR REFUSALS BUILT INTO THE CODE, NOT INTO THE PROSE
-----------------------------------------------------
1. **Online revenue is never merged with physical casino GGR.** Every revenue
   row carries `revenue_scope`, and an assertion refuses the file if any row
   carries a scope outside the online / retail-sportsbook / fantasy set. The
   Michigan and Connecticut physical-casino series are on the same regulator
   pages and are deliberately not fetched.
2. **A platform provider is not a manager.** Michigan's own table is headed
   "Platform Provider". A technology vendor supplying a skin is recorded in
   `technology_provider`; `operator_entity_id` stays the tribe. Where an NIGC
   declination letter characterises the contractual role, its legal
   characterisation wins over any trade-press word - so the declination file is
   read and any tribe/vendor pair it names is reported rather than overwritten.
3. **A brand is an alias, not a property.** "Bay Mills Resort & Casino" on the
   MGCB sheet is the existing `CCP-`/`VP-`/`TPL-` property; a new facility ID is
   never minted here.
4. **An operator brand is not a tribe.** Arizona publishes event-wagering
   revenue by brand (FanDuel, DraftKings, BetMGM). Ten of Arizona's twenty
   event-wagering licences are tribal and ten are pro-sports-franchise, and the
   monthly report does not say which licence each brand sits under. Those rows
   are written with `is_tribe_attributable = no` and a blank `tribe_id`, and
   they are NEVER summed into a tribal total.

PULL DISCIPLINE
---------------
One lock per host in `logs/_HOSTLOCK_<host>.json`, sequential, >= 1.6s gap,
single-shot fetches with no retry loop, idempotent skip-if-present.
`web.archive.org` is held by another agent (lock younger than the 6h takeover
threshold), so the Wayback leg is APPENDED TO THAT LOCK'S QUEUE and not run.
`files.usaspending.gov`, `api.usaspending.gov`, `apps.nd.gov`,
`www.treasurer.nd.gov` and `www.nigc.gov` are not touched.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import html
import importlib.util
import json
import os
import re
import subprocess
import sys
import time
from collections import Counter, defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CODE = os.path.join(ROOT, "code")
RAW = os.path.join(ROOT, "data", "raw", "external", "digital_gaming")
CLEAN = os.path.join(ROOT, "data", "clean")
CODEBOOK = os.path.join(CLEAN, "codebook")
INTERIM = os.path.join(ROOT, "data", "interim")
REVIEW = os.path.join(ROOT, "review")
LOGS = os.path.join(ROOT, "logs")
for _d in (RAW, CLEAN, CODEBOOK, INTERIM, REVIEW, LOGS):
    os.makedirs(_d, exist_ok=True)

TODAY = dt.date.today().isoformat()
SCRIPT = "code/119_build_digital_and_loyalty.py"
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36")
GAP = 1.6

# Hosts another agent owns, or that this build is forbidden to touch.
FORBIDDEN_HOSTS = {
    "files.usaspending.gov", "api.usaspending.gov", "apps.nd.gov",
    "www.treasurer.nd.gov", "www.nigc.gov", "nigc.gov",
}

STATE_ABBR = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT",
    "Florida": "FL", "Idaho": "ID", "Indiana": "IN", "Iowa": "IA",
    "Kansas": "KS", "Louisiana": "LA", "Massachusetts": "MA",
    "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS",
    "Montana": "MT", "Nebraska": "NE", "Nevada": "NV", "New Mexico": "NM",
    "New York": "NY", "North Carolina": "NC", "North Dakota": "ND",
    "Oklahoma": "OK", "Oregon": "OR", "Rhode Island": "RI",
    "South Dakota": "SD", "Texas": "TX", "Washington": "WA",
    "Wisconsin": "WI", "Wyoming": "WY",
}


def st_abbr(s):
    s = clean_text(s)
    if len(s) == 2:
        return s.upper()
    return STATE_ABBR.get(s, s[:2].upper() if s else "")


sys.path.insert(0, CODE)
from cedar_domain import Tier                                    # noqa: E402

_spec = importlib.util.spec_from_file_location(
    "party_rulings", os.path.join(CODE, "33_apply_party_rulings.py"))
_pr = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_pr)
resolve_entity = _pr.resolve_entity          # the ONE resolver. Never re-write.

PRODUCT_TYPES = {
    "RETAIL_SPORTSBOOK", "ONLINE_SPORTSBOOK", "ONLINE_CASINO",
    "ONLINE_POKER", "FANTASY", "OTHER_DIGITAL",
}
# Every revenue row must be one of these. A physical casino GGR figure has no
# legal scope here, which is what stops the two from ever being added together.
REVENUE_SCOPES = {
    "ONLINE_CASINO_ONLY", "ONLINE_SPORTS_WAGERING_ONLY",
    "RETAIL_SPORTS_WAGERING_ONLY", "FANTASY_CONTESTS_ONLY",
    "ONLINE_CASINO_AND_SPORTS_COMBINED", "NO_REVENUE_OBSERVATION",
}
METRICS = {
    "HANDLE", "GROSS_GAMING_REVENUE", "ADJUSTED_GROSS_REVENUE",
    "TAX_OR_PAYMENT", "PROMOTIONAL_DEDUCTION", "PATRON_WINNINGS",
    "CANCELLED_WAGERS", "FEDERAL_EXCISE_TAX", "ENTRY_FEES",
    "NET_FANTASY_CONTEST_REVENUE", "MONTHLY_RESETTLEMENTS", "AMOUNT_WAGERED",
}

REVIEW_ROWS = []
MANIFEST = []
NOTES = Counter()


def note(k, n=1):
    NOTES[k] += n


def queue_review(item_key, kind, state, subject, evidence, reason,
                 source_url="", proposed=""):
    REVIEW_ROWS.append(dict(
        item_key=item_key, kind=kind, state=state, subject=subject,
        evidence=(evidence or "")[:900], reason=reason, source_url=source_url,
        proposed_resolution=proposed, YOUR_RULING="", raised_by=SCRIPT,
        raised_date=TODAY))


# ===========================================================================
# I/O helpers
# ===========================================================================
def read_csv(path):
    if not os.path.exists(path):
        return []
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path, rows, fields):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fields})
    os.replace(tmp, path)
    return len(rows)


def clean_text(s):
    s = re.sub(r"\s+", " ", (s or "")).strip()
    return s


# ===========================================================================
# Pull discipline
# ===========================================================================
def lock_path(host):
    return os.path.join(LOGS, "_HOSTLOCK_%s.json" % host)


def read_lock(host):
    p = lock_path(host)
    if not os.path.exists(p):
        return None
    try:
        return json.loads(open(p, encoding="utf-8").read())
    except Exception:
        return None


def pid_alive(pid):
    try:
        out = subprocess.run(
            ["powershell", "-NoProfile", "-Command",
             "if (Get-Process -Id %d -ErrorAction SilentlyContinue) "
             "{'Y'} else {'N'}" % int(pid)],
            capture_output=True, text=True, timeout=25).stdout
        return "Y" in out
    except Exception:
        return False


def claim_host(host, purpose):
    """Return True if we may poll `host`. Never starts a second poller."""
    if host in FORBIDDEN_HOSTS:
        note("host_forbidden:%s" % host)
        return False
    cur = read_lock(host)
    if cur and cur.get("active") and not cur.get("released"):
        holder = cur.get("pid")
        age_h = 99.0
        try:
            t = cur.get("claimed_at") or cur.get("started") or ""
            t0 = dt.datetime.fromisoformat(t.replace("Z", "+00:00"))
            if t0.tzinfo is None:
                t0 = t0.replace(tzinfo=dt.timezone.utc)
            age_h = (dt.datetime.now(dt.timezone.utc) - t0).total_seconds() / 3600
        except Exception:
            pass
        if holder and (pid_alive(holder) or age_h < 6):
            # Append and exit. This is the whole of rule 1.
            cur.setdefault("queue", []).append(
                {"script": SCRIPT, "purpose": purpose,
                 "queued_at": dt.datetime.now(dt.timezone.utc).isoformat()})
            open(lock_path(host), "w", encoding="utf-8").write(
                json.dumps(cur, indent=1))
            note("host_deferred:%s" % host)
            return False
    open(lock_path(host), "w", encoding="utf-8").write(json.dumps({
        "host": host, "pid": os.getpid(), "script": SCRIPT,
        "claimed_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        "active": True, "queue": [],
        "policy": "single-shot fetches, >=1.6s gap, no retry loop",
        "note": purpose}, indent=1))
    return True


def release_host(host, note_text=""):
    cur = read_lock(host) or {"host": host}
    cur["active"] = False
    cur["released"] = dt.datetime.now(dt.timezone.utc).isoformat()
    if note_text:
        cur["note"] = note_text
    open(lock_path(host), "w", encoding="utf-8").write(json.dumps(cur, indent=1))


_LAST = {"t": 0.0}


def fetch(url, relpath, timeout=90, force=False):
    """Single-shot GET. Saves bytes under RAW/relpath. Returns (status, bytes).

    Idempotent: an already-downloaded file is reused and re-manifested, which
    is what makes it safe to re-run this build without re-hitting a regulator.
    """
    dest = os.path.join(RAW, relpath)
    os.makedirs(os.path.dirname(dest), exist_ok=True)
    if os.path.exists(dest) and not force and os.path.getsize(dest) > 0:
        body = open(dest, "rb").read()
        manifest_add(relpath, url, 200, body, cached=True)
        return 200, body
    dtime = time.time() - _LAST["t"]
    if dtime < GAP:
        time.sleep(GAP - dtime)
    _LAST["t"] = time.time()
    cmd = ["curl", "-s", "-L", "-A", UA,
           "-H", "Accept: text/html,application/xhtml+xml,application/pdf,"
                 "application/vnd.openxmlformats-officedocument."
                 "spreadsheetml.sheet,application/json,*/*;q=0.8",
           "-H", "Accept-Language: en-US,en;q=0.9",
           "--max-time", str(timeout),
           "-w", "\n__HTTPSTATUS__%{http_code}", url]
    try:
        p = subprocess.run(cmd, capture_output=True, timeout=timeout + 30)
    except subprocess.TimeoutExpired:
        note("fetch_timeout")
        return 0, b""
    out = p.stdout
    m = re.search(rb"\n__HTTPSTATUS__(\d+)$", out)
    status = int(m.group(1)) if m else 0
    body = out[:m.start()] if m else out
    # CHECK THE STATUS, NOT THE FILE - a 404 body still has content.
    if status == 200 and body:
        with open(dest, "wb") as f:
            f.write(body)
    manifest_add(relpath, url, status, body)
    note("fetch_%d" % status)
    return status, body


def manifest_add(relpath, url, status, body, cached=False):
    MANIFEST.append(dict(
        file=relpath.replace("\\", "/"), url=url, http_status=status,
        bytes=len(body), md5=hashlib.md5(body).hexdigest(),
        cached="yes" if cached else "no", fetched_date=TODAY, script=SCRIPT))


def save_manifest():
    """Merge, never replace. A build run with --skip-fetch has an empty
    in-memory manifest, and writing that over the file would erase the md5 of
    every document already retrieved."""
    path = os.path.join(RAW, "_SOURCE_MANIFEST.csv")
    fields = ["file", "url", "http_status", "bytes", "md5", "cached",
              "fetched_date", "script"]
    rows, seen = [], set()
    for r in MANIFEST + read_csv(path):
        f = r.get("file", "")
        if not f or f in seen:
            continue
        seen.add(f)
        rows.append(r)
    rows.sort(key=lambda r: r["file"])
    return write_csv(path, rows, fields)


def strip_html(b):
    t = b.decode("utf-8", "replace") if isinstance(b, bytes) else b
    t = re.sub(r"<(script|style|noscript)\b.*?</\1>", " ", t, flags=re.S | re.I)
    t = re.sub(r"<br\s*/?>", "\n", t, flags=re.I)
    t = re.sub(r"</(p|div|li|tr|h\d|section)>", "\n", t, flags=re.I)
    t = html.unescape(re.sub(r"<[^>]+>", " ", t))
    return "\n".join(clean_text(l) for l in t.split("\n") if clean_text(l))


# ===========================================================================
# Entity resolution
# ===========================================================================
SPINE_PATH = os.path.join(ROOT, "data", "spine", "cedar_entity_spine.csv")


# A gaming licensee is a GOVERNMENT, or a government's enterprise. It is never
# a college, a CDFI, a school or a clinic. This is the guard AGENTS.md records
# as one that WORKS - restrict to government-class rows - and it is what stops
# `Keweenaw Bay Indian Community` landing on *Keweenaw Bay Ojibwa Community
# College*, which containment does on its own.
GOVERNMENT_CLASSES = frozenset({
    "Federally recognized tribe",
    "Federally recognized Alaska Native Village",
    "State-recognized tribe",
})


class Resolver:
    """The ONE resolver, plus two guards it does not carry itself.

    Containment has failed ten ways in this repository. Every resolution here
    must additionally (a) agree on state with the publishing regulator, because
    a state gaming regulator can only be publishing about a tribe in its own
    state, and (b) land on a government-class entity, because a gaming licensee
    is a government. Where either fails the row is refused to `review/`.
    """

    def __init__(self):
        spine = read_csv(SPINE_PATH)
        for r in spine:
            extra = [r.get("fr_official_name", "")]
            r["aliases"] = "|".join(
                a for a in ([r.get("aliases", "")] + extra) if a and a.strip())
        self.spine = spine
        # A government-only VIEW of the same spine. Handing the one resolver a
        # narrower view is not a second matcher; it is the same matcher told
        # which universe the question is about. Without it, containment sends
        # `Keweenaw Bay Indian Community` to that tribe's community COLLEGE,
        # which is a longer name in the same state.
        self.gov_spine = [r for r in spine
                          if clean_text(r.get("entity_class", ""))
                          in GOVERNMENT_CLASSES]
        self.by_id = {r["tribe_id"]: r for r in spine}
        self.cache = {}
        self.reasons = Counter()

    def resolve(self, raw, state):
        key = (clean_text(raw), state)
        if key in self.cache:
            return self.cache[key]
        name = clean_text(raw)
        res = dict(tribe_id="", tribe_canonical_name="", match_method="",
                   entity_tier="", refusal="")
        tid, canon, how = resolve_entity(name, self.spine)
        if tid and clean_text(
                self.by_id[tid].get("entity_class", "")) not in \
                GOVERNMENT_CLASSES:
            tid2, canon2, how2 = resolve_entity(name, self.gov_spine)
            if tid2:
                tid, canon, how = tid2, canon2, how2 + "_gov_view"
        res["match_method"] = how
        if not tid:
            res["refusal"] = "unresolved:%s" % how
        else:
            ent = self.by_id[tid]
            st = st_abbr(ent.get("state", ""))
            cls = clean_text(ent.get("entity_class", ""))
            if state and st and st != st_abbr(state):
                res["refusal"] = "state_disagreement:spine=%s,source=%s" % (
                    st, state)
            elif cls not in GOVERNMENT_CLASSES:
                res["refusal"] = "refused_entity_class:%s" % cls
            else:
                res.update(tribe_id=tid,
                           tribe_canonical_name=ent.get("canonical_name", ""),
                           entity_tier=Tier.B.value)
        self.reasons[(res["refusal"] or "resolved_" + how).split(":")[0]] += 1
        self.cache[key] = res
        return res


# ===========================================================================
# STAGE 1 - the compacts define the RIGHTS. No network.
# ===========================================================================
DIGITAL_TERM_FIELDS = ("sports_wagering_authorized",
                       "internet_wagering_authorized",
                       "mobile_wagering_scope")


def stage_compact_rights():
    """Return {(tribe_id, state): {right -> [rows]}} from the parsed compacts.

    Read FIRST, before anything is fetched, because these rows say which tribes
    hold which rights and therefore what an operation would even mean.
    """
    terms = read_csv(os.path.join(CLEAN, "compact_structured_terms.csv"))
    rights = defaultdict(lambda: defaultdict(list))
    for r in terms:
        if r["term_field"] not in DIGITAL_TERM_FIELDS:
            continue
        tid = clean_text(r.get("tribe_id"))
        if not tid:
            queue_review(
                "COMPACT-RIGHT-UNKEYED-%s" % r.get("term_id", ""),
                "compact_right_unkeyed", r.get("state", ""),
                r.get("tribe", ""), r.get("source_quote", ""),
                "digital right parsed but the compact row carries no tribe_id",
                r.get("source_url", ""))
            continue
        rights[(tid, clean_text(r.get("state", "")))][r["term_field"]].append(r)
    note("compact_digital_term_rows",
         sum(len(v) for d in rights.values() for v in d.values()))
    return rights


def authority_cite(rows):
    """A compact citation, assembled from the compact row's own fields."""
    if not rows:
        return "", "", ""
    # Latest effective instrument wins as the operative citation.
    rs = sorted(rows, key=lambda r: (r.get("effective_from") or ""))
    r = rs[-1]
    cite = "%s; %s p.%s (%s)" % (
        r.get("compact_id", ""), r.get("source_pdf", ""),
        r.get("source_page", ""), r.get("effective_from", ""))
    return cite, r.get("source_url", ""), r.get("source_quote", "")


# ===========================================================================
# STAGE 2 - MICHIGAN. The regulator publishes the relationship AND the launch.
# ===========================================================================
MI_HOST = "www.michigan.gov"
MI_MEDIA = ("https://www.michigan.gov/mgcb/-/media/Project/Websites/mgcb/"
            "Detroit-Casino-Revenue-Files")
MI_PLATFORMS_URL = ("https://www.michigan.gov/mgcb/internet-gaming-and-"
                    "fantasy-contests/authorized-online-gaming-and-sports-"
                    "betting-platform-providers-in-michigan")
MI_FILES = [
    # (relpath, url, product_type, revenue_scope)
    ("mi/mgcb_internet_gaming_2026.xlsx",
     MI_MEDIA + "/Internet-Gaming---2026.xlsx"
     "?rev=ce2ca758c0814feda44a368f4ef853d1",
     "ONLINE_CASINO", "ONLINE_CASINO_ONLY"),
    ("mi/mgcb_internet_gaming_2024.xlsx",
     MI_MEDIA + "/Internet-Gaming---2024.xlsx"
     "?rev=fcc0b988a4614fcd80deb876435f79a2",
     "ONLINE_CASINO", "ONLINE_CASINO_ONLY"),
    ("mi/mgcb_internet_gaming_2023.xlsx",
     MI_MEDIA + "/Internet-Gaming-January-2023.xlsx"
     "?rev=bde3a40b2fe04eba8acf21b0f9cd3010",
     "ONLINE_CASINO", "ONLINE_CASINO_ONLY"),
    ("mi/mgcb_internet_sports_betting_2026.xlsx",
     MI_MEDIA + "/Internet-Sports-Betting---2026.xlsx"
     "?rev=acf63aff5c034a44a79f6a934b2155a8",
     "ONLINE_SPORTSBOOK", "ONLINE_SPORTS_WAGERING_ONLY"),
    ("mi/mgcb_internet_sports_betting_2024.xlsx",
     MI_MEDIA + "/Internet-Sports-Betting---2024.xlsx"
     "?rev=f553eb08e45843298495efac636020a9",
     "ONLINE_SPORTSBOOK", "ONLINE_SPORTS_WAGERING_ONLY"),
    ("mi/mgcb_internet_sports_betting_2023.xlsx",
     MI_MEDIA + "/Internet-Sports-Betting-January-2023.xlsx"
     "?rev=c1408798556c469ea40ba91289034eb2",
     "ONLINE_SPORTSBOOK", "ONLINE_SPORTS_WAGERING_ONLY"),
]

MI_METRIC_MAP = [
    (r"total handle", "HANDLE"),
    (r"^gross\b.*receipts", "GROSS_GAMING_REVENUE"),
    (r"adjusted gross", "ADJUSTED_GROSS_REVENUE"),
    (r"state (tax|payment)", "TAX_OR_PAYMENT"),
    (r"city wagering", "TAX_OR_PAYMENT"),
    (r"governing body", "TAX_OR_PAYMENT"),
]
MONTHS = ["January", "February", "March", "April", "May", "June", "July",
          "August", "September", "October", "November", "December"]

# Aggregate columns on the MGCB sheets. They are totals, not an operator, and
# writing them as one would double every tribal figure in the state.
MI_AGGREGATE = re.compile(
    r"all internet|total|commercial operators|tribal operators", re.I)

# (product_type, year, month, metric) -> {"published": x, "summed": y}
# A per-operator grid read out of a wide merged spreadsheet is exactly the kind
# of thing that produces a plausible wrong number, so every month is footed
# against the sheet's own printed total before anything ships.
MI_FOOTING = defaultdict(dict)
MI_SHEETS_SEEN = set()


def mi_provider_key(name):
    """Group a platform provider across MGCB's own spelling drift.

    MGCB writes "Hard Rock" and "Hard Rock Bet", "Caesars Horshoe" and
    "Caesars Horseshoe", "Golden Nugget Casino" and "Golden Nugget Online
    Gaming" for the same vendor on different sheets. An eight-character
    normalised head collapses all three pairs and separates every distinct
    vendor in the file. The VERBATIM spellings are kept on the row.
    """
    return re.sub(r"[^a-z0-9]", "", clean_text(name).lower())[:8]


def mi_metric(label):
    lab = clean_text(label).lower().replace("  ", " ")
    for pat, met in MI_METRIC_MAP:
        if re.search(pat, lab):
            return met
    return ""


def fetch_michigan():
    if not claim_host(MI_HOST, "MGCB platform providers + monthly iGaming / "
                               "internet sports betting operator files"):
        return False
    try:
        fetch(MI_PLATFORMS_URL, "mi/mgcb_platform_providers.html")
        for rel, url, _pt, _sc in MI_FILES:
            fetch(url, rel)
    finally:
        release_host(MI_HOST, "MGCB digital gaming files retrieved")
    return True


def parse_mi_platform_table():
    """tribe/casino name -> (brand, product urls) straight from the regulator."""
    p = os.path.join(RAW, "mi/mgcb_platform_providers.html")
    if not os.path.exists(p):
        return []
    t = open(p, encoding="utf-8", errors="replace").read()
    tabs = re.findall(r"<table.*?</table>", t, re.S | re.I)
    out = []
    for tb in tabs:
        for row in re.findall(r"<tr.*?</tr>", tb, re.S | re.I):
            cells = [clean_text(html.unescape(re.sub(r"<[^>]+>", " ", c)))
                     for c in re.findall(r"<t[dh].*?</t[dh]>", row, re.S | re.I)]
            urls = re.findall(r"https?://[^\s\"'<>]+", html.unescape(row))
            if len(cells) >= 2 and cells[0] and cells[1]:
                # some rows carry the brand and the product links in ONE cell
                brand = clean_text(re.split(
                    r"Casino Games?|Sports Betting|Poker\s*:|https?://",
                    cells[1])[0]).strip(" :|")
                out.append(dict(licensee=cells[0], brand=brand or cells[1],
                                detail=" | ".join(cells[2:]), urls=urls))
    return out


def parse_mi_workbook(relpath, product_type, revenue_scope, url):
    """Return (operator_meta, revenue_rows_raw) from one MGCB workbook.

    The sheet is a wide grid: one block of columns per operator, one row per
    month. Row 2 is the operator, row 3 the casino, row 4 the platform
    provider, row 5 the initial date of operation - the REGULATOR'S OWN launch
    date, which is the only launch date in this build that is not a refusal.
    """
    import openpyxl
    p = os.path.join(RAW, relpath)
    if not os.path.exists(p):
        return {}, []
    wb = openpyxl.load_workbook(p, data_only=True)
    meta, rows = {}, []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        if len(grid) < 8:
            continue
        yr = re.search(r"(20\d\d)", sheet)
        year = int(yr.group(1)) if yr else None
        # Each MGCB workbook carries TWO year sheets and the ranges OVERLAP -
        # the 2024 workbook holds 2024 and 2023, the 2023 workbook holds 2023
        # and 2022. Parsing both would double every 2023 figure in the file.
        if (product_type, year) in MI_SHEETS_SEEN:
            note("mi_duplicate_year_sheet_skipped:%s" % sheet)
            continue
        MI_SHEETS_SEEN.add((product_type, year))
        # locate the header rows by their own labels
        idx = {}
        for i, row in enumerate(grid[:12]):
            lab = clean_text(str(row[0] or "")).lower()
            if lab.startswith("operator"):
                idx["operator"] = i
            elif lab.startswith("casino name"):
                idx["casino"] = i
            elif lab.startswith("platform provider"):
                idx["platform"] = i
            elif lab.startswith("initial date"):
                idx["launch"] = i
            elif lab == "month":
                idx["metric"] = i
        if "operator" not in idx or "metric" not in idx:
            note("mi_sheet_unrecognised:%s" % sheet)
            continue
        orow, mrow = grid[idx["operator"]], grid[idx["metric"]]
        crow = grid[idx["casino"]] if "casino" in idx else [None] * len(orow)
        prow = grid[idx["platform"]] if "platform" in idx else [None] * len(orow)
        lrow = grid[idx["launch"]] if "launch" in idx else [None] * len(orow)
        # forward-fill the merged operator header across its column block
        cur = {}
        colop = {}
        for c in range(1, len(mrow)):
            for key, src in (("op", orow), ("casino", crow),
                             ("plat", prow), ("launch", lrow)):
                v = src[c] if c < len(src) else None
                if v is not None and clean_text(str(v)):
                    cur[key] = v
                    if key == "op":
                        cur["casino"] = crow[c] if c < len(crow) else None
                        cur["plat"] = prow[c] if c < len(prow) else None
                        cur["launch"] = lrow[c] if c < len(lrow) else None
            colop[c] = dict(cur)
        for c in range(1, len(mrow)):
            opname = clean_text(str(colop.get(c, {}).get("op") or ""))
            metric = mi_metric(mrow[c] if c < len(mrow) else "")
            if not opname or not metric:
                continue
            if MI_AGGREGATE.search(opname):
                # Not written as an operator - but kept, because the sheet's
                # own total is the only way to prove the per-operator columns
                # were read correctly. See MI_FOOTING.
                note("mi_aggregate_column_kept_for_footing")
                if re.match(r"(?i)all internet|total", opname):
                    for i, row in enumerate(grid):
                        mlabel = clean_text(str(row[0] or ""))
                        if mlabel not in MONTHS or year is None:
                            continue
                        v = row[c] if c < len(row) else None
                        try:
                            fv = float(v)
                        except (TypeError, ValueError):
                            continue
                        MI_FOOTING[(product_type, year,
                                    MONTHS.index(mlabel) + 1,
                                    metric)]["published"] = fv
                continue
            info = colop[c]
            key = opname
            if key not in meta:
                lv = info.get("launch")
                launch = ""
                if isinstance(lv, dt.datetime):
                    launch = lv.date().isoformat()
                elif isinstance(lv, dt.date):
                    launch = lv.isoformat()
                elif lv:
                    mm = re.search(r"(\d{4})-(\d{2})-(\d{2})", str(lv))
                    launch = mm.group(0) if mm else ""
                meta[key] = dict(
                    operator=opname,
                    casino=clean_text(str(info.get("casino") or "")),
                    platform=clean_text(str(info.get("plat") or "")),
                    launch=launch, sheet=sheet, url=url)
            for i, row in enumerate(grid):
                mlabel = clean_text(str(row[0] or ""))
                if mlabel not in MONTHS or year is None:
                    continue
                v = row[c] if c < len(row) else None
                if v is None or isinstance(v, str) and not clean_text(v):
                    continue
                try:
                    val = float(v)
                except (TypeError, ValueError):
                    continue
                mi = MONTHS.index(mlabel) + 1
                last = [31, 29 if mi == 2 and year % 4 == 0 else 28, 31, 30, 31,
                        30, 31, 31, 30, 31, 30, 31][mi - 1]
                fk = (product_type, year, mi, metric)
                MI_FOOTING[fk]["summed"] = \
                    MI_FOOTING[fk].get("summed", 0.0) + val
                rows.append(dict(
                    operator=opname, casino=meta[key]["casino"],
                    platform=meta[key]["platform"],
                    period_start="%04d-%02d-01" % (year, mi),
                    period_end="%04d-%02d-%02d" % (year, mi, last),
                    metric=metric, value=val, product_type=product_type,
                    revenue_scope=revenue_scope, sheet=sheet, url=url,
                    header=clean_text(str(mrow[c])), month=mlabel,
                    title=clean_text(str(grid[0][1] if len(grid[0]) > 1 else ""))))
    return meta, rows


# ===========================================================================
# STAGE 3 - CONNECTICUT. Two tribal nations, four monthly series, an open API.
# ===========================================================================
CT_HOST = "data.ct.gov"
CT_SETS = [
    # (id, slug, product_type, revenue_scope, human title)
    ("imqd-at3c", "ct_online_casino_gaming", "ONLINE_CASINO",
     "ONLINE_CASINO_ONLY", "Selected Online Casino Gaming Data"),
    ("xf6g-659c", "ct_online_sports_wagering", "ONLINE_SPORTSBOOK",
     "ONLINE_SPORTS_WAGERING_ONLY", "Selected Online Sport Wagering Data"),
    ("yb54-t38r", "ct_retail_sports_wagering", "RETAIL_SPORTSBOOK",
     "RETAIL_SPORTS_WAGERING_ONLY",
     "Schedule of Selected Retail Sports Wagering Data"),
    ("cnz5-ik5d", "ct_fantasy_sports", "FANTASY", "FANTASY_CONTESTS_ONLY",
     "Schedule of Selected Fantasy Sports Wagering"),
]
CT_METRIC_MAP = [
    (r"^wagers$", "HANDLE"),
    (r"^entry_fees_from_all", "ENTRY_FEES"),
    (r"^patron_winnings", "PATRON_WINNINGS"),
    (r"^cancelled_wagers$", "CANCELLED_WAGERS"),
    (r"^monthly_resettlements", "MONTHLY_RESETTLEMENTS"),
    (r"online_casino_gaming_win|online_sports_wagering_win", "GROSS_GAMING_REVENUE"),
    (r"^net_fantasy_contest_revenue$", "NET_FANTASY_CONTEST_REVENUE"),
    (r"^federal_excise_tax", "FEDERAL_EXCISE_TAX"),
    (r"^unadjusted_monthly_gaming", "GROSS_GAMING_REVENUE"),
    (r"^promotional_coupons_or_credits$", "PROMOTIONAL_DEDUCTION"),
    (r"^promotional_deduction", "PROMOTIONAL_DEDUCTION"),
    (r"^total_gross_gaming", "ADJUSTED_GROSS_REVENUE"),
    (r"^in_state_total_fantasy_contest", "ADJUSTED_GROSS_REVENUE"),
    (r"^tax_payment|^payment_\d", "TAX_OR_PAYMENT"),
]


def ct_metric(field, product_type=""):
    """Sports HANDLE and online-casino AMOUNT WAGERED are not the same measure.

    Connecticut labels both columns `wagers`. In a sportsbook that is handle -
    money staked once. In an online casino it is coin-in, which recycles every
    spin, so it runs an order of magnitude larger and means something else
    entirely. They are given DIFFERENT metric names so that nobody can add
    them, which a shared label makes trivially easy to do by accident.
    """
    for pat, met in CT_METRIC_MAP:
        if re.search(pat, field):
            if met == "HANDLE" and product_type == "ONLINE_CASINO":
                return "AMOUNT_WAGERED"
            return met
    return ""


def fetch_connecticut():
    if not claim_host(CT_HOST, "CT DCP monthly online casino / online sports / "
                               "retail sports / fantasy series (Socrata)"):
        return False
    try:
        for sid, slug, _pt, _sc, _t in CT_SETS:
            fetch("https://data.ct.gov/resource/%s.json?$limit=50000" % sid,
                  "ct/%s.json" % slug)
            fetch("https://data.ct.gov/api/views/%s.json" % sid,
                  "ct/%s_metadata.json" % slug)
    finally:
        release_host(CT_HOST, "CT digital gaming series retrieved")
    return True


# ===========================================================================
# STAGE 4 - ARIZONA. Revenue by BRAND, not by licence holder.
# ===========================================================================
AZ_HOST = "gaming.az.gov"
AZ_DOCS = [
    ("az/adg_event_wagering_may2026.pdf",
     "https://gaming.az.gov/sites/default/files/"
     "EW%20Website%20Report-May%202026%20UNAUDITED.pdf",
     "ONLINE_SPORTSBOOK"),
    ("az/adg_fantasy_sports_may2026.pdf",
     "https://gaming.az.gov/sites/default/files/"
     "FS%20Website%20Report-May%202026%20UNAUDITED.pdf",
     "FANTASY"),
    ("az/adg_tribal_gaming_status_20260701.pdf",
     "https://gaming.az.gov/sites/default/files/"
     "Gaming%20Status%20Report%2007012026_0.pdf", ""),
]


def fetch_arizona():
    if not claim_host(AZ_HOST, "ADG monthly event wagering + fantasy sports "
                               "revenue reports"):
        return False
    try:
        fetch("https://gaming.az.gov/resources/reports", "az/adg_reports.html")
        for rel, url, _pt in AZ_DOCS:
            fetch(url, rel)
    finally:
        release_host(AZ_HOST, "ADG event wagering reports retrieved")
    return True


AZ_MONEY = re.compile(r"^-?[\d,]+\.\d{2}$")
# The six printed metric columns of the ADG Event Wagering Revenue Report, in
# the order the report prints them. Read off the header, not assumed.
AZ_EW_METRICS = ["HANDLE", "PATRON_WINNINGS", "GROSS_GAMING_REVENUE",
                 "PROMOTIONAL_DEDUCTION", "ADJUSTED_GROSS_REVENUE",
                 "TAX_OR_PAYMENT"]
AZ_EW_HEADERS = [
    "Gross Event Wagering Receipts (Wagers)",
    "Winnings Paid to Players (Payouts)",
    "Adjusted Gross Event Wagering Receipts prior to Free Bets Allowable "
    "Deduction (Wagers minus Payouts minus Federal Excise Tax)",
    "Free Bets / Promotional Credits Deduction Allowed and Taken",
    "Adjusted Gross Event Wagering Receipts Subject to Privilege Fees",
    "Privilege Fees",
]


def parse_az_event_wagering(relpath, url, period_label):
    """ADG prints one row per event wagering OPERATOR BRAND.

    The page is rotated 90 degrees, so a naive text extraction interleaves the
    columns and silently produces a plausible, wrong grid. Every value here is
    placed by its own coordinate: the operator comes from the x-band, the
    Retail/Mobile leg and the metric come from the y-distance to the report's
    own printed `Retail` / `Mobile` header words. An operator that publishes
    only one leg therefore lands in the RIGHT leg instead of being guessed.

    ADG does not print the licence holder, and ten of Arizona's twenty event
    wagering licences are tribal while ten are pro-sports franchises. So no row
    from here is attributed to a tribe on the strength of a brand.
    """
    p = os.path.join(RAW, relpath)
    if not os.path.exists(p):
        return []
    try:
        import fitz
    except ImportError:
        return []
    doc = fitz.open(p)
    out = []
    for page in doc:
        words = [(x0, y0, t) for x0, y0, x1, y1, t, *_ in page.get_text("words")]
        bands = defaultdict(list)
        for x0, y0, t in words:
            bands[round(x0 / 4) * 4].append((y0, t))
        header = None
        for x in sorted(bands):
            ws = bands[x]
            if any(t == "Operator:" for _y, t in ws):
                header = [(y, t) for y, t in ws if t in ("Retail", "Mobile")]
                break
        if not header or len(header) != 12:
            note("az_header_not_found")
            continue
        header = sorted(header, key=lambda z: -z[0])   # top of page first
        for x in sorted(bands):
            ws = sorted(bands[x], key=lambda z: -z[0])
            nums = [(y, t) for y, t in ws if AZ_MONEY.match(t)]
            if len(nums) < 3:
                continue
            name = clean_text(" ".join(
                t for y, t in ws
                if not AZ_MONEY.match(t) and t not in ("$", "-", "*")))
            if not name or re.match(
                    r"^(Total|All |Privilege|Adjusted|Gross|Net|Fiscal|"
                    r"Calendar|Limited)", name):
                continue
            cells = []
            for y, t in nums:
                hi = min(range(12), key=lambda i: abs(header[i][0] - y))
                cells.append((AZ_EW_METRICS[hi // 2], header[hi][1].lower(),
                              float(t.replace(",", ""))))
            out.append(dict(brand=name, cells=cells, url=url,
                            period=period_label, relpath=relpath))
    return out


# ===========================================================================
# STAGE 5 - LOYALTY. Free, scrapeable, and structurally revealing.
# ===========================================================================
# Seeded on ENTERPRISES, not properties: the question is whether one programme
# spans several properties, which can only be asked of a multi-property
# operator. Each entry is only a site root; nothing about the programme is
# asserted here, it is read off the page.
LOYALTY_SITES = [
    # (host, root url, state, tribe name as the site's own operator)
    ("kewadin.com", "https://kewadin.com/", "MI", "Sault Ste. Marie"),
    ("www.fourwindscasino.com", "https://www.fourwindscasino.com/", "MI",
     "Pokagon"),
    ("www.gunlakecasino.com", "https://www.gunlakecasino.com/", "MI",
     "Match-E-Be-Nash-She-Wish"),
    ("www.firekeeperscasino.com", "https://www.firekeeperscasino.com/", "MI",
     "Nottawaseppi Potawatomi"),
    ("www.soaringeaglecasino.com", "https://www.soaringeaglecasino.com/", "MI",
     "Saginaw Chippewa"),
    ("www.turtlecreekcasino.com", "https://www.turtlecreekcasino.com/", "MI",
     "Grand Traverse"),
    ("www.odawacasino.com", "https://www.odawacasino.com/", "MI",
     "Little Traverse Bay"),
    ("www.islandresortandcasino.com",
     "https://www.islandresortandcasino.com/", "MI", "Hannahville"),
    ("www.baymillscasino.com", "https://www.baymillscasino.com/", "MI",
     "Bay Mills"),
    ("www.foxwoods.com", "https://www.foxwoods.com/", "CT",
     "Mashantucket Pequot"),
    ("mohegansun.com", "https://mohegansun.com/", "CT", "Mohegan"),
    ("www.choctawcasinos.com", "https://www.choctawcasinos.com/", "OK",
     "The Choctaw Nation of Oklahoma"),
    ("www.cherokeecasino.com", "https://www.cherokeecasino.com/", "OK",
     "Cherokee Nation"),
    ("osagecasino.com", "https://osagecasino.com/", "OK", "The Osage Nation"),
    ("www.riverspirittulsa.com", "https://www.riverspirittulsa.com/", "OK",
     "The Muscogee (Creek) Nation"),
    ("www.winstar.com", "https://www.winstar.com/", "OK",
     "The Chickasaw Nation"),
    ("www.ho-chunkgaming.com", "https://www.ho-chunkgaming.com/", "WI",
     "Ho-Chunk"),
    ("www.oneidacasino.net", "https://www.oneidacasino.net/", "WI",
     "Oneida Nation (Wisconsin)"),
    ("www.paysbig.com", "https://www.paysbig.com/", "WI", "Forest County"),
    ("www.senecacasinos.com", "https://www.senecacasinos.com/", "NY",
     "Seneca"),
    ("www.turningstone.com", "https://www.turningstone.com/", "NY", "Oneida"),
    ("www.ddcaz.com", "https://www.ddcaz.com/", "AZ", "Tohono O'odham"),
    ("www.casinoarizona.com", "https://www.casinoarizona.com/", "AZ",
     "Salt River"),
    ("www.wingilariver.com", "https://www.wingilariver.com/", "AZ",
     "Gila River"),
    ("www.emeraldqueen.com", "https://www.emeraldqueen.com/", "WA",
     "Puyallup"),
    ("www.tulalipresortcasino.com", "https://www.tulalipresortcasino.com/",
     "WA", "Tulalip"),
    ("www.muckleshootcasino.com", "https://www.muckleshootcasino.com/", "WA",
     "Muckleshoot"),
    ("www.pearlriverresort.com", "https://www.pearlriverresort.com/", "MS",
     "Mississippi Choctaw"),
    ("www.palacasino.com", "https://www.palacasino.com/", "CA", "Pala"),
    ("www.seminolehardrocktampa.com",
     "https://www.seminolehardrocktampa.com/", "FL", "Seminole"),
    ("shootingstarcasino.com", "https://shootingstarcasino.com/", "MN",
     "White Earth"),
    ("www.windcreek.com", "https://www.windcreek.com/", "AL", "Poarch Band"),
]

LOYALTY_LINK = re.compile(
    r"reward|players[-_ ]?club|loyalt|momentum|wild[-_ ]?card|club[-_/]|"
    r"one[-_]?club|advantage|passport|player[-_]?s?[-_]?card", re.I)
# Order matters. `Club Osage` must be tried BEFORE the bare `Club`, or
# left-to-right scanning consumes "Club" at position 0, discards it as generic,
# and resumes past the very word that named the programme.
PROGRAM_NAME = re.compile(
    r"\b(Club\s+[A-Z][\w'&+-]+"
    r"|(?:[A-Z][\w'&.+’-]*\s+){0,3}"
    r"(?:Rewards?\s+Club|Players?\s+Club|Rewards?|Club|Momentum|"
    r"Wild\s+Card|Advantage|Passport|Players?\s+Card))\b")
TITLE_SPLIT = re.compile(r"\s*[|»•·–—]\s*|\s+[-]\s+")
# A bare "Rewards" or "Club" is a page heading, not a programme name. A name
# must carry at least one distinguishing word.
GENERIC_NAME = re.compile(
    r"^(?:the\s+)?(?:rewards?|club|players?\s+club|players?\s+card|"
    r"rewards?\s+club|my\s+rewards)$", re.I)
# Tiers, only where the page ENUMERATES them. The sentence-scan version of this
# returned "Please|Points|Copper|Premier" and "Take|Osage|Casino|Hotel" - random
# capitalised words that look exactly like tier names and are not. The tier
# labels carry little analytic weight anyway (a shared programme across
# properties is the finding), so recall is traded away entirely for precision.
TIER_LIST = re.compile(
    r"(?:tiers?|levels?|card levels?)\s*(?:are|include|:)\s*"
    r"([A-Z][\w'&-]*(?:\s*,\s*[A-Z][\w'&-]*){1,5}"
    r"(?:\s*,?\s*and\s+[A-Z][\w'&-]*)?)", re.I)
GENERIC_FACILITY_TOKENS = frozenset({
    "casino", "casinos", "resort", "resorts", "hotel", "hotels", "the", "and",
    "of", "at", "inn", "lodge", "travel", "plaza", "center", "centre", "spa",
    "gaming", "club", "bingo", "tribal", "indian", "nation", "band", "tribe",
    "north", "south", "east", "west", "grand", "casino's",
})

EARN_VERB = (r"earn|redeem|accrue|points?|comp|tier credit|reward credit|"
             r"free play|discount")
ELIGIBILITY = [
    ("slots_eligible", r"\bslots?\b|slot machine|electronic gaming"),
    ("tables_eligible", r"table games?|blackjack|roulette|craps"),
    ("sportsbook_eligible", r"sportsbook|sports bett|sports wager|race book"),
    ("online_gaming_eligible",
     r"online casino|online gaming|igaming|online play|play online"),
    ("hotel_eligible", r"\bhotel\b|\broom\b|\bsuite\b|\blodg"),
    ("dining_eligible", r"dining|restaurant|food and beverage|buffet"),
    ("retail_eligible", r"gift shop|retail shop|\bretail\b|logo shop"),
]
CROSS_PROPERTY = re.compile(
    r"all (?:of our |our )?(?:\w+ )?(?:casinos|locations|properties)|"
    r"any (?:of our )?(?:\w+ )?(?:casino|location|propert)|"
    r"at all (?:\w+ ){0,3}(?:casinos|locations|properties)|"
    r"across (?:all )?(?:\w+ ){0,3}(?:casinos|locations|properties)|"
    r"one card|single card|same card", re.I)
MOBILE_APP = re.compile(
    r"mobile app|download (?:the |our )?app|app store|google play|"
    r"\bios\b.{0,20}android", re.I)
DIGITAL_WALLET = re.compile(r"digital wallet|e-?wallet|mobile wallet", re.I)
CASHLESS = re.compile(r"cashless", re.I)


def sentences(text):
    for para in text.split("\n"):
        for s in re.split(r"(?<=[.!?;])\s+", para):
            s = clean_text(s)
            if 15 <= len(s) <= 400:
                yield s


def fetch_loyalty():
    """Homepage -> discover the loyalty URL -> fetch it. Two GETs per host."""
    got = []
    for host, root, state, tribe in LOYALTY_SITES:
        if not claim_host(host, "loyalty programme page (2 single-shot GETs)"):
            continue
        try:
            st, body = fetch(root, "loyalty/%s/home.html" % host)
            if st != 200 or not body:
                note("loyalty_home_failed:%s" % host)
                continue
            page = body.decode("utf-8", "replace")
            hrefs = set(re.findall(r'href="([^"]+)"', page))
            cands = []
            for h in hrefs:
                if not LOYALTY_LINK.search(h):
                    continue
                if h.startswith("#") or h.lower().startswith(
                        ("mailto:", "tel:", "javascript:")):
                    continue
                if h.startswith("/"):
                    h = root.rstrip("/") + h
                if not h.startswith("http"):
                    continue
                cands.append(h)
            # shortest path wins: the programme's own landing page, not a
            # sub-page about kiosks or a single promotion.
            cands = sorted(set(cands), key=lambda u: (u.count("/"), len(u)))
            if not cands:
                note("loyalty_no_link:%s" % host)
                queue_review("LOYALTY-NOLINK-%s" % host, "loyalty_not_found",
                             state, tribe, "",
                             "homepage retrieved but no loyalty-shaped link",
                             root)
                continue
            target = cands[0]
            st2, body2 = fetch(target, "loyalty/%s/program.html" % host)
            if st2 != 200:
                note("loyalty_page_failed:%s" % host)
                continue
            got.append((host, root, state, tribe, target))
        finally:
            release_host(host, "loyalty page retrieved")
    return got


def parse_loyalty(host, root, state, tribe, target_url, resolver, facilities):
    """Extract only what the page STATES. Silence stays silence.

    Every eligibility flag is set to `yes` on an explicit sentence and left
    BLANK otherwise. It is never set to `no`, because a page that does not
    mention the sportsbook is evidence about the page, not about the programme.
    """
    p = os.path.join(RAW, "loyalty/%s/program.html" % host)
    if not os.path.exists(p):
        return None, []
    raw = open(p, encoding="utf-8", errors="replace").read()
    text = strip_html(raw)
    title = ""
    m = re.search(r"<title[^>]*>(.*?)</title>", raw, re.S | re.I)
    if m:
        title = clean_text(html.unescape(m.group(1)))
    h1 = [clean_text(html.unescape(re.sub(r"<[^>]+>", " ", x)))
          for x in re.findall(r"<h1.*?</h1>", raw, re.S | re.I)]

    # --- programme name.
    # It must come from the page's OWN heading or title. Searching the body
    # text finds the site's navigation menu instead, which is how a first pass
    # produced "Facilities Weddings Reservations Rewards" and "Course Calendar
    # Rates Momentum" - strings that are verbatim on the page and are not
    # anybody's programme. A body-only candidate is recorded, but it goes to
    # `review/` as name-only rather than into the dataset.
    def first_name(strings):
        for src in strings:
            for mm in PROGRAM_NAME.finditer(src):
                cand = clean_text(mm.group(1))
                if cand and not GENERIC_NAME.match(cand):
                    return cand, clean_text(src)[:400]
        return "", ""

    # A title segment that opens with a call to action is marketing prose, not
    # a programme name: "Join the Best Casino Loyalty Rewards | Status+
    # Rewards" names the programme in its SECOND segment.
    title_parts = [p for p in TITLE_SPLIT.split(title)
                   if clean_text(p) and not re.match(
                       r"(?i)^(join|play|get|discover|welcome|learn|sign|"
                       r"become|start|earn more|explore|find)\b", p.strip())]
    name, name_quote = first_name(title_parts)
    name_source = "<title>"
    if not name:
        name, name_quote = first_name([x for x in h1 if x])
        name_source = "<h1>"
        if name:
            # An h1 on a marketing page is often a concatenated banner, so the
            # NAME is queued for confirmation even though the row ships - the
            # programme's existence and its property span are separately
            # evidenced and do not depend on getting the label exactly right.
            queue_review(
                "LOYALTY-NAME-CONFIRM-%s" % host, "loyalty_name_unconfirmed",
                state, "%s (%s)" % (name, tribe), name_quote,
                "programme name taken from an <h1>, not from the page title. "
                "Confirm the label; the programme row and its property map do "
                "not depend on it.", target_url,
                proposed="confirm or correct program_name")
            note("loyalty_name_from_h1")
    if not name:
        body_name, body_quote = first_name(
            [l for l in text.split("\n")[:80] if 4 < len(l) < 120])
        queue_review(
            "LOYALTY-NAMEONLY-%s" % host, "loyalty_name_only", state, tribe,
            "best body-text candidate: %s | title: %s"
            % (body_name or "(none)", title[:160]),
            "no programme name appears in the page's own <h1> or <title>. A "
            "body-text match is the site's navigation menu as often as it is a "
            "programme, so this stays name-only at Tier B and is not written.",
            target_url,
            proposed="confirm the programme name by hand from %s" % target_url)
        note("loyalty_name_only")
        return None, []

    res = resolver.resolve(tribe, state)
    if not res["tribe_id"]:
        queue_review("LOYALTY-UNRESOLVED-%s" % host, "loyalty_tribe_unresolved",
                     state, tribe, name,
                     "operator name did not resolve: %s" % res["refusal"],
                     target_url)

    flags, quotes = {}, {}
    for field, pat in ELIGIBILITY:
        rx = re.compile(pat, re.I)
        for s in sentences(text):
            if rx.search(s) and re.search(EARN_VERB, s, re.I):
                flags[field] = "yes"
                quotes[field] = s
                break

    tiers, tier_quote = [], ""
    for s in sentences(text):
        mm = TIER_LIST.search(s)
        if mm:
            parts = [clean_text(x) for x in
                     re.split(r"\s*,\s*|\s+and\s+", mm.group(1))]
            parts = [p for p in parts if p and len(p.split()) == 1]
            if 2 <= len(parts) <= 6:
                tiers, tier_quote = parts, s
                break

    cross, cross_quote = "", ""
    for s in sentences(text):
        if CROSS_PROPERTY.search(s):
            cross, cross_quote = "yes", s
            break

    app = wallet = cashless = ""
    app_q = wallet_q = cash_q = ""
    for s in sentences(text):
        if not app and MOBILE_APP.search(s):
            app, app_q = "yes", s
        if not wallet and DIGITAL_WALLET.search(s):
            wallet, wallet_q = "yes", s
        if not cashless and CASHLESS.search(s):
            cashless, cash_q = "yes", s

    # The property map. A facility counts only where the page NAMES it, and
    # "names it" is deliberately strict: either the full property name appears,
    # or every DISTINCTIVE token of it does. Without the second test a loose
    # two-word head match put 24 Choctaw properties on one page, because every
    # one of them begins "Choctaw Casino" and so does every sentence about the
    # programme. A token shared with the tribe's own name distinguishes
    # nothing, which is the same lesson as the `core()` folding defect.
    props = []
    tid = res["tribe_id"]
    tribe_tokens = set(re.findall(
        r"[a-z']+", (res["tribe_canonical_name"] or tribe).lower()))
    for f in facilities:
        if f.get("tribe_id") != tid:
            continue
        fname = clean_text(f.get("facility_name", ""))
        if len(fname) < 5:
            continue
        hit, basis = "", ""
        for s in sentences(text):
            if fname.lower() in s.lower():
                hit, basis = s, "the programme page names this property in full"
                break
        if not hit:
            toks = [t for t in re.findall(r"[a-z']+", fname.lower())
                    if t not in GENERIC_FACILITY_TOKENS
                    and t not in tribe_tokens and len(t) > 2]
            if toks:
                for s in sentences(text):
                    sl = s.lower()
                    if all(t in sl for t in toks):
                        hit = s
                        basis = ("the programme page names every distinctive "
                                 "token of this property (%s)"
                                 % ", ".join(toks))
                        break
        if hit:
            props.append((f, hit, basis))

    prog = dict(
        program_name=name, program_name_basis=name_source,
        tribe_id=tid, tribe_canonical_name=res["tribe_canonical_name"],
        state=state, host=host, source_url=target_url,
        source_quote=(name_quote or (props[0][1] if props else ""))[:600],
        tier_names="|".join(tiers), tier_quote=tier_quote,
        cross_property=cross, cross_quote=cross_quote,
        mobile_app=app, app_quote=app_q,
        digital_wallet=wallet, wallet_quote=wallet_q,
        cashless=cashless, cash_quote=cash_q,
        flags=flags, quotes=quotes, n_chars=len(text))
    return prog, props


# ===========================================================================
# THE BUILD
# ===========================================================================
REL_FIELDS = [
    "digital_gaming_id", "tribe_id", "tribe_canonical_name", "facility_id",
    "facility_name", "state", "product_type", "license_type", "brand",
    "operator_entity_id", "designee_entity_id", "technology_provider",
    "retail_available", "mobile_on_premises", "mobile_statewide",
    "launch_date", "launch_date_basis", "cessation_date", "current_status",
    "compact_authority_cite", "compact_authority_url",
    "compact_authority_quote", "authorisation_observed",
    "operation_observed", "source_url", "source_quote", "fetched_date",
    "tier", "confidence", "entity_match_method", "note", "built_by_script",
    "built_date",
]
REV_FIELDS = [
    "revenue_id", "state", "period_start", "period_end", "period_type",
    "digital_gaming_id", "tribe_id", "tribe_canonical_name", "facility_id",
    "licensee_name_as_published", "brand", "product_type", "revenue_scope",
    "metric", "value_usd", "is_tribe_attributable", "attribution_basis",
    "is_online_only", "source_agency", "source_document", "source_url",
    "source_quote", "fetched_date", "confidence_tier", "note",
    "built_by_script", "built_date",
]
LOY_FIELDS = [
    "loyalty_program_id", "program_name", "operator_entity_id", "tribe_id",
    "tribe_canonical_name", "state", "start_date", "end_date",
    "current_status", "program_name_basis", "tier_names", "tier_thresholds",
    "earning_currency",
    "slots_eligible", "tables_eligible", "sportsbook_eligible",
    "online_gaming_eligible", "hotel_eligible", "dining_eligible",
    "retail_eligible", "cross_property_redemption", "mobile_app",
    "digital_wallet", "cashless_gaming", "n_properties_mapped",
    "eligibility_quotes", "source_url", "source_quote", "observation_date",
    "confidence_tier", "built_by_script", "built_date",
]
LOYPROP_FIELDS = [
    "loyalty_program_id", "program_name", "facility_id", "facility_name",
    "tribe_id", "state", "evidence_basis", "source_url", "source_quote",
    "observation_date", "confidence_tier", "built_by_script", "built_date",
]
REVIEW_FIELDS = ["item_key", "kind", "state", "subject", "evidence", "reason",
                 "source_url", "proposed_resolution", "YOUR_RULING",
                 "raised_by", "raised_date"]


class Builder:
    def __init__(self):
        self.resolver = Resolver()
        self.facilities = read_csv(os.path.join(CLEAN, "gaming_facilities.csv"))
        self.fac_by_tribe = defaultdict(list)
        for f in self.facilities:
            if f.get("tribe_id"):
                self.fac_by_tribe[f["tribe_id"]].append(f)
        self.rights = stage_compact_rights()
        self.rel, self.rev, self.loy, self.loyprop = [], [], [], []
        self._n = Counter()
        # NIGC declination letters are read so a vendor named there can be
        # reported. Their legal characterisation WINS over trade press.
        self.declinations = read_csv(
            os.path.join(CLEAN, "nigc_declination_letters.csv"))

    # -- ids -------------------------------------------------------------
    def dgid(self, state):
        self._n[state] += 1
        return "DGR-%s-%04d" % (state or "XX", self._n[state])

    def revid(self, state):
        self._n["REV" + state] += 1
        return "DGREV-%s-%06d" % (state or "XX", self._n["REV" + state])

    # -- relationships ---------------------------------------------------
    def add_rel(self, **kw):
        state = kw.get("state", "")
        row = {k: "" for k in REL_FIELDS}
        row.update(kw)
        row["digital_gaming_id"] = self.dgid(state)
        row["built_by_script"] = SCRIPT
        row["built_date"] = TODAY
        row["fetched_date"] = row.get("fetched_date") or TODAY
        # BUG FIXED 2026-08-26 (see 174_backfill_digital_gaming_tiers.py):
        # this was `row.setdefault("tier", Tier.B.value)`, and `setdefault` is
        # a NO-OP here - `row` is pre-initialised with `{k: "" for k in
        # REL_FIELDS}` two lines above, so the key ALREADY EXISTS with an
        # empty string and the default never fires. Every one of the 154
        # relationship rows shipped with a blank `tier`, which is what raised
        # the `MIRRORED_LINK_CARRIES_NO_TIER` review item over 7,983 rows.
        # `or` on the empty string is the behaviour `setdefault` was reaching
        # for. NOT RE-RUN when the fix was made; the two clean files were
        # backfilled in place by 174 instead, because 119 is a full rebuild.
        row["tier"] = row.get("tier") or Tier.B.value
        assert row["product_type"] in PRODUCT_TYPES, row["product_type"]
        assert row["source_url"] and row["source_quote"], \
            "zero fabrication: every row needs a URL and a verbatim quote"
        self.rel.append(row)
        return row["digital_gaming_id"]

    def add_rev(self, **kw):
        row = {k: "" for k in REV_FIELDS}
        row.update(kw)
        row["revenue_id"] = self.revid(kw.get("state", ""))
        row["built_by_script"] = SCRIPT
        row["built_date"] = TODAY
        row["fetched_date"] = row.get("fetched_date") or TODAY
        # Same dead-`setdefault` defect as add_rel above: `confidence_tier`
        # was blank on all 10,661 revenue rows and `period_type` on 10,660,
        # because both keys already exist (empty) in the pre-initialised row.
        row["confidence_tier"] = row.get("confidence_tier") or Tier.B.value
        row["period_type"] = row.get("period_type") or "month"
        assert row["revenue_scope"] in REVENUE_SCOPES, row["revenue_scope"]
        assert row["metric"] in METRICS, row["metric"]
        assert row["source_url"] and row["source_quote"]
        self.rev.append(row)

    # ------------------------------------------------------------------
    def compact_cite(self, tid, state, fields):
        rows = []
        for f in fields:
            rows += self.rights.get((tid, state), {}).get(f, [])
        return authority_cite(rows)

    # ---------------- MICHIGAN ----------------------------------------
    def build_michigan(self):
        plat = parse_mi_platform_table()
        # key the regulator's platform table by RESOLVED TRIBE, because MGCB
        # styles the same operator two different ways on two of its own pages
        # ("Gun Lake Tribe" vs "Gun Lake Band of Pottawatomi Indians").
        plat_by_op = {}
        for pr in plat:
            r = self._mi_resolve(pr["licensee"])
            if r["tribe_id"]:
                plat_by_op.setdefault(r["tribe_id"], pr)
        # (tribe_id, product_type, provider_key) -> era record. Grouping on the
        # PROVIDER rather than on the operator is what surfaces the former
        # providers: Hannahville ran TwinSpires before Hard Rock, Little
        # Traverse ran FoxBet before bet365, Sault Ste. Marie ran Wynn before
        # Caesars. Each is its own relationship with its own end.
        eras = {}
        latest_sheet_year = 0
        for rel, url, product_type, scope in MI_FILES:
            meta, rows = parse_mi_workbook(rel, product_type, scope, url)
            for opname, info in meta.items():
                res = self._mi_resolve(opname)
                if not res["tribe_id"]:
                    if not re.search(r"MGM Grand|MotorCity|Greektown|Hollywood",
                                     opname, re.I):
                        queue_review(
                            "MI-OP-UNRESOLVED-%s" % opname,
                            "operator_unresolved", "MI", opname,
                            info.get("casino", ""),
                            "MGCB operator did not resolve: %s"
                            % res["refusal"], url)
                        note("mi_operator_unresolved")
                    continue
                yr = int(re.search(r"(20\d\d)", info["sheet"]).group(1))
                latest_sheet_year = max(latest_sheet_year, yr)
                pkey = mi_provider_key(info.get("platform", ""))
                k = (res["tribe_id"], product_type, pkey)
                e = eras.setdefault(k, dict(
                    res=res, product_type=product_type, years=set(),
                    spellings=set(), info=info, url=url, opnames=set()))
                e["years"].add(yr)
                e["spellings"].add(clean_text(info.get("platform", "")))
                e["opnames"].add(opname)
                if yr >= max(int(re.search(r"(20\d\d)", e["info"]["sheet"])
                                 .group(1)), 0):
                    e["info"] = info
                    e["url"] = url
            for r in rows:
                self._mi_rev_row(r, rel)
        # earliest era per (tribe, product) is the one MGCB's own
        # "Initial date of Operation" belongs to
        first_year = {}
        for (tid, pt, pk), e in eras.items():
            y = min(e["years"])
            if (tid, pt) not in first_year or y < first_year[(tid, pt)][0]:
                first_year[(tid, pt)] = (y, pk)
        for k in sorted(eras, key=lambda z: (z[0], z[1], z[2])):
            self._mi_rel_row(k, eras[k], plat_by_op, latest_sheet_year,
                             first_year)

    # Two MGCB strings that no name matcher can bridge, each resolved to a
    # TRIBE NAME (never to an ID by hand) and each carrying its evidence. The
    # resolution itself still runs through the one resolver.
    MI_OPERATOR_ALIASES = {
        "soaring eagle gaming": (
            "Saginaw Chippewa Indian Tribe of Michigan",
            "MGCB's own 'Authorized Online Gaming and Sports Betting Platform "
            "Providers in Michigan' table lists 'Saginaw Chippewa Indian Tribe "
            "of Michigan | GAN | playeagle.com', and the revenue workbook "
            "gives 'Soaring Eagle Gaming' the same platform provider, GAN, and "
            "the casino name 'Soaring Eagle Casino'."),
        "gun lake band of pottawatomi indians": (
            "Match-E-Be-Nash-She-Wish Band of Pottawatomi",
            "MGCB's platform provider table names the same operator 'Gun Lake "
            "Tribe' against BetPARX and playgunlake.com; the federally "
            "recognised name in the spine is Match-E-Be-Nash-She-Wish Band of "
            "Pottawatomi."),
        "gun lake tribe": (
            "Match-E-Be-Nash-She-Wish Band of Pottawatomi",
            "same MGCB table; 'Gun Lake Tribe' is the operator's own style of "
            "the federally recognised Match-E-Be-Nash-She-Wish Band."),
        "gun lake band tribal community": (
            "Match-E-Be-Nash-She-Wish Band of Pottawatomi",
            "MGCB's own earlier-vintage workbook styling for the same operator "
            "column: same casino name (Gun Lake Casino) and same platform "
            "provider as the 'Gun Lake Tribe' row on MGCB's platform table."),
        "nottawaseppi huron band of pottawatomi indians": (
            "Nottawaseppi Huron Band of the Potawatomi",
            "MGCB spells the same operator two ways on its own two pages - "
            "'Pottawatomi Indians' in the revenue workbook, 'the Potawatomi' "
            "on the platform provider table - against the same casino "
            "(FireKeepers Casino) and the same provider (NeoGames)."),
    }

    def _mi_resolve(self, opname):
        # MGCB appends the casino in parentheses on some sheets
        # ("... (FireKeepers Casino)"). Strip it before the alias lookup.
        bare = clean_text(re.sub(r"\s*\([^)]*\)\s*", " ", opname))
        alias = self.MI_OPERATOR_ALIASES.get(bare.lower()) or \
            self.MI_OPERATOR_ALIASES.get(clean_text(opname).lower())
        if alias:
            r = self.resolver.resolve(alias[0], "Michigan")
            if r["tribe_id"]:
                r = dict(r)
                r["match_method"] = "documented_alias(%s)" % r["match_method"]
                r["alias_basis"] = alias[1]
                return r
        res = self.resolver.resolve(opname, "Michigan")
        if res["tribe_id"]:
            return res
        # MGCB writes the full federal name; the spine writes the short one.
        m = re.split(r"\s+(?:Indian|Band|Tribe|Community|Nation)\b", opname)[0]
        if m and m != opname:
            r2 = self.resolver.resolve(m, "Michigan")
            if r2["tribe_id"]:
                r2 = dict(r2)
                r2["match_method"] = r2["match_method"] + "_head"
                return r2
        return res

    def _mi_rel_row(self, key, era, plat_by_op, latest_year, first_year):
        tid, product_type, pkey = key
        res, info = era["res"], era["info"]
        url = era["url"]
        fac_id, fac_name = self.match_facility(tid, info.get("casino", ""))
        pl = plat_by_op.get(tid, {})
        provider = clean_text(info.get("platform", ""))
        current = max(era["years"]) >= latest_year
        # The regulator's brand cell only describes the CURRENT provider.
        brand = provider
        if current and pl and mi_provider_key(pl.get("brand", "")) == pkey:
            brand = clean_text(pl.get("brand", ""))
        product_urls = " ".join(pl.get("urls", [])[:6]) if current else ""
        is_first = first_year.get((tid, product_type), (None, None))[1] == pkey
        cite, curl, cquote = self.compact_cite(
            tid, "Michigan",
            ("internet_wagering_authorized", "sports_wagering_authorized",
             "mobile_wagering_scope"))
        quote = ('MGCB "%s": Operator "%s" | Casino Name "%s" | '
                 'Platform Provider "%s" | Initial date of Operation "%s"'
                 % (info.get("sheet", ""), sorted(era["opnames"])[0],
                    info.get("casino", ""), provider, info.get("launch", "")))
        if current and pl:
            quote += ('. MGCB "Authorized Online Gaming and Sports Betting '
                      'Platform Providers in Michigan": "%s" | "%s"%s'
                      % (pl.get("licensee", ""), pl.get("brand", ""),
                         (" | " + product_urls) if product_urls else ""))
        yrs = sorted(era["years"])
        if current:
            status = "operating"
            cessation = ""
        else:
            status = "ceased_as_platform_provider"
            cessation = ""
        n = ("facility_id deliberately blank: Michigan internet gaming is "
             "statewide and has no property. The bricks-and-mortar casino "
             "MGCB names alongside this operator is %s (Cedar %s). "
             "Provider observed on MGCB workbook sheets %s."
             % (info.get("casino", "") or "not named", fac_id or "unkeyed",
                "-".join(str(y) for y in (yrs[0], yrs[-1]))
                if len(yrs) > 1 else str(yrs[0])))
        if not current:
            n += (" NOT present on the %d sheet: this is a FORMER platform "
                  "provider. cessation_date is blank because MGCB publishes "
                  "no end date - only an absence, which dates the change no "
                  "finer than a year." % latest_year)
        if len(era["spellings"]) > 1:
            n += (" MGCB spells this provider %s across its own sheets."
                  % " / ".join('"%s"' % s for s in sorted(era["spellings"])))
        self.add_rel(
            tribe_id=tid, tribe_canonical_name=res["tribe_canonical_name"],
            facility_id="", facility_name="", state="MI",
            product_type=product_type,
            license_type="internet gaming operator (Michigan Lawful Internet "
                         "Gaming Act 2019 PA 152) / internet sports betting "
                         "operator (2019 PA 149)",
            brand=brand, operator_entity_id=tid,
            technology_provider=provider,
            retail_available="", mobile_on_premises="",
            mobile_statewide="yes",
            launch_date=info.get("launch", "") if is_first else "",
            launch_date_basis=("MGCB column 'Initial date of Operation'"
                               if is_first else
                               "blank by rule: MGCB's initial date of "
                               "operation belongs to the operator's FIRST "
                               "platform era, not to this one"),
            cessation_date=cessation, current_status=status,
            compact_authority_cite=cite, compact_authority_url=curl,
            compact_authority_quote=(cquote or "")[:600],
            authorisation_observed="yes" if cite else "no",
            operation_observed="yes",
            source_url=url, source_quote=quote,
            entity_match_method=res["match_method"],
            confidence=Tier.B.value, note=n)

    def match_facility(self, tid, casino_name):
        """Attach to an EXISTING CCP-/VP-/TPL- id, or return blank. Never mint."""
        if not casino_name:
            return "", ""
        cn = casino_name.lower()
        best = ("", "")
        for f in self.fac_by_tribe.get(tid, []):
            fn = clean_text(f.get("facility_name", "")).lower()
            if not fn:
                continue
            if fn in cn or cn.split("/")[0].strip() in fn:
                best = (f.get("facility_id", ""), f.get("facility_name", ""))
                break
        return best

    def _mi_rev_row(self, r, relpath):
        opname = r["operator"]
        commercial = bool(re.search(
            r"MGM Grand|MotorCity|Greektown|Hollywood", opname, re.I))
        res = None if commercial else self._mi_resolve(opname)
        tid = res["tribe_id"] if res else ""
        quote = ('MGCB "%s" / "%s": Operator "%s" | Casino Name "%s" | '
                 'Platform Provider "%s" | Month "%s" | column "%s"'
                 % (r["title"], r["sheet"], opname, r["casino"], r["platform"],
                    r["month"], r["header"]))
        self.add_rev(
            state="MI", period_start=r["period_start"],
            period_end=r["period_end"],
            tribe_id=tid,
            tribe_canonical_name=res["tribe_canonical_name"] if res else "",
            licensee_name_as_published=opname, brand=r["platform"],
            product_type=r["product_type"], revenue_scope=r["revenue_scope"],
            metric=r["metric"], value_usd="%.2f" % r["value"],
            is_tribe_attributable="yes" if tid else "no",
            attribution_basis=("MGCB names the tribal operator in its own "
                               "column header" if tid else
                               "commercial Detroit operator, not tribal"),
            is_online_only="yes",
            source_agency="Michigan Gaming Control Board",
            source_document=os.path.basename(relpath),
            source_url=r["url"], source_quote=quote,
            note="online only; never to be added to a physical casino GGR "
                 "figure")

    # ---------------- CONNECTICUT --------------------------------------
    # The licensee strings CT publishes, and what each one IS. Every mapping
    # is a name the State prints; nothing is inferred from a brand.
    CT_LICENSEES = {
        "mohegan tribe on-reservation": ("Mohegan", "on_premises"),
        "mohegan digital, llc": ("Mohegan", "statewide"),
        "mohegan digital llc": ("Mohegan", "statewide"),
        "mptn on-reservation": ("Mashantucket Pequot", "on_premises"),
        "mpi master wagering license ct, llc": ("Mashantucket Pequot",
                                                "statewide"),
        "mpi master wagering license ct llc": ("Mashantucket Pequot",
                                               "statewide"),
        "mohegan sun": ("Mohegan", "on_premises"),
        "foxwoods": ("Mashantucket Pequot", "on_premises"),
    }

    def build_connecticut(self):
        seen_rel = set()
        for sid, slug, product_type, scope, title in CT_SETS:
            p = os.path.join(RAW, "ct/%s.json" % slug)
            if not os.path.exists(p):
                continue
            recs = json.loads(open(p, encoding="utf-8").read())
            url = "https://data.ct.gov/resource/%s.json" % sid
            for rec in recs:
                lic = clean_text(rec.get("licensee", ""))
                if not lic:
                    continue
                key = lic.lower().rstrip(".")
                info = self.CT_LICENSEES.get(key)
                if not info:
                    # CT Lottery / XL Center / Sportech are the State's own
                    # retail licensees. Not tribal, and never keyed to one.
                    if re.search(r"lottery|xl center|sportech|clc", lic, re.I):
                        note("ct_state_licensee_not_tribal")
                        self._ct_rev(rec, "", "", lic, "", product_type, scope,
                                     url, title, sid, tribe_attr="no",
                                     basis="Connecticut Lottery Corporation "
                                           "licensee, not tribal")
                        continue
                    queue_review("CT-LICENSEE-%s" % lic, "licensee_unmapped",
                                 "CT", lic, json.dumps(rec)[:400],
                                 "Connecticut licensee string not in the "
                                 "hand-checked mapping", url)
                    note("ct_licensee_unmapped")
                    continue
                tribe_name, scope_kind = info
                res = self.resolver.resolve(tribe_name, "Connecticut")
                if not res["tribe_id"]:
                    queue_review("CT-TRIBE-%s" % tribe_name, "tribe_unresolved",
                                 "CT", tribe_name, lic, res["refusal"], url)
                    continue
                rk = (res["tribe_id"], product_type, scope_kind)
                if rk not in seen_rel:
                    seen_rel.add(rk)
                    self._ct_rel(res, lic, product_type, scope_kind, url,
                                 title, rec)
                self._ct_rev(rec, res["tribe_id"],
                             res["tribe_canonical_name"], lic, "",
                             product_type, scope, url, title, sid,
                             tribe_attr="yes",
                             basis="Connecticut DCP names the tribal licensee")

    def _ct_rel(self, res, lic, product_type, scope_kind, url, title, rec):
        tid = res["tribe_id"]
        cite, curl, cquote = self.compact_cite(
            tid, "Connecticut",
            ("internet_wagering_authorized", "sports_wagering_authorized",
             "mobile_wagering_scope"))
        quote = ('Connecticut Department of Consumer Protection, "%s": '
                 'licensee "%s", month ending "%s"'
                 % (title, lic, rec.get("month_ending", "")))
        self.add_rel(
            tribe_id=tid, tribe_canonical_name=res["tribe_canonical_name"],
            facility_id="", state="CT", product_type=product_type,
            license_type=("master wagering licensee, statewide"
                          if scope_kind == "statewide"
                          else "master wagering licensee, on-reservation"),
            brand=lic, operator_entity_id=tid,
            mobile_on_premises="yes" if scope_kind == "on_premises" else "",
            mobile_statewide="yes" if scope_kind == "statewide" else "",
            retail_available="yes" if product_type == "RETAIL_SPORTSBOOK"
                             else "",
            launch_date="", launch_date_basis="",
            current_status="operating",
            compact_authority_cite=cite, compact_authority_url=curl,
            compact_authority_quote=(cquote or "")[:600],
            authorisation_observed="yes" if cite else "no",
            operation_observed="yes",
            source_url=url, source_quote=quote,
            entity_match_method=res["match_method"], confidence=Tier.B.value,
            note="Connecticut separates ON-RESERVATION online gaming from "
                 "STATEWIDE, and publishes them as different licensees. The "
                 "two are recorded as two relationships, never merged. "
                 "launch_date is blank: the series' earliest month is not a "
                 "launch statement.")

    def _ct_rev(self, rec, tid, canon, lic, fac, product_type, scope, url,
                title, sid, tribe_attr, basis):  # noqa: C901
        me = clean_text(rec.get("month_ending", ""))[:10]
        if not me:
            return
        y, m, d = me.split("-")
        for field, val in rec.items():
            if field in ("licensee", "fiscal_year", "month_ending"):
                continue
            met = ct_metric(field, product_type)
            if not met:
                continue
            try:
                v = float(val)
            except (TypeError, ValueError):
                continue
            quote = ('Connecticut DCP "%s" (data.ct.gov/%s), licensee "%s", '
                     'month_ending "%s", %s = %s'
                     % (title, sid, lic, rec.get("month_ending", ""), field,
                        val))
            self.add_rev(
                state="CT", period_start="%s-%s-01" % (y, m), period_end=me,
                tribe_id=tid, tribe_canonical_name=canon,
                licensee_name_as_published=lic, brand=lic,
                product_type=product_type, revenue_scope=scope, metric=met,
                value_usd="%.2f" % v, is_tribe_attributable=tribe_attr,
                attribution_basis=basis,
                is_online_only="yes" if scope != "RETAIL_SPORTS_WAGERING_ONLY"
                               else "no",
                source_agency="Connecticut Department of Consumer Protection",
                source_document="%s.json" % sid, source_url=url,
                source_quote=quote,
                note="%s only; never to be added to a physical casino GGR "
                     "figure" % scope.replace("_", " ").lower())

    # ---------------- ARIZONA ------------------------------------------
    AZ_EW_COLS = ["HANDLE", "PATRON_WINNINGS", "GROSS_GAMING_REVENUE",
                  "PROMOTIONAL_DEDUCTION", "ADJUSTED_GROSS_REVENUE",
                  "TAX_OR_PAYMENT"]

    # Tribal enterprise brands Arizona prints. A brand is keyed to a tribe ONLY
    # where the brand is that tribe's own enterprise name and Cedar's property
    # universe already carries it under that tribe. Every other brand stays
    # unattributed, which is the honest reading of a by-brand report.
    AZ_TRIBAL_BRANDS = {
        "Desert Diamond Mobile": ("Tohono O'odham",
                                  "Desert Diamond is the Tohono O'odham "
                                  "Nation's own casino enterprise brand and "
                                  "Cedar's Arizona property universe carries "
                                  "Desert Diamond properties under that "
                                  "tribe."),
    }

    def build_arizona(self):
        for rel, url, product_type in AZ_DOCS:
            if product_type != "ONLINE_SPORTSBOOK":
                continue
            recs = parse_az_event_wagering(rel, url, "2026-05")
            if not recs:
                note("az_event_wagering_unparsed")
            for r in recs:
                brand = r["brand"]
                tid = canon = ""
                basis = ("ADG publishes by operator brand, not by licence "
                         "holder; NOT attributable to a tribe")
                hit = None
                for b, (tribe, why) in self.AZ_TRIBAL_BRANDS.items():
                    if b.lower() in brand.lower():
                        hit = (tribe, why)
                        break
                if hit:
                    res = self.resolver.resolve(hit[0], "Arizona")
                    tid, canon = res["tribe_id"], res["tribe_canonical_name"]
                    basis = hit[1]
                else:
                    queue_review(
                        "AZ-EW-BRAND-%s" % brand, "az_brand_not_attributable",
                        "AZ", brand, "monthly event wagering figures, May 2026",
                        "ADG publishes by operator BRAND. Ten of Arizona's "
                        "twenty event wagering licences are tribal and ten are "
                        "pro-sports franchises; the report does not say which "
                        "licence this brand operates under, and no ADG "
                        "licensee list was found at /event-wagering, "
                        "/event-wagering/licensees or /resources/reports.",
                        url,
                        proposed="obtain an ADG event wagering licensee list "
                                 "naming the licence holder behind each brand")
                    note("az_brand_not_attributable")
                for met, leg, val in r["cells"]:
                    self.add_rev(
                        state="AZ", period_start="2026-05-01",
                        period_end="2026-05-31",
                        tribe_id=tid, tribe_canonical_name=canon,
                        licensee_name_as_published=brand, brand=brand,
                        product_type=("RETAIL_SPORTSBOOK" if leg == "retail"
                                      else "ONLINE_SPORTSBOOK"),
                        revenue_scope=("RETAIL_SPORTS_WAGERING_ONLY"
                                       if leg == "retail"
                                       else "ONLINE_SPORTS_WAGERING_ONLY"),
                        metric=met, value_usd="%.2f" % val,
                        is_tribe_attributable="yes" if tid else "no",
                        attribution_basis=basis,
                        is_online_only="no" if leg == "retail" else "yes",
                        source_agency="Arizona Department of Gaming",
                        source_document=os.path.basename(rel),
                        source_url=url,
                        source_quote=(
                            'ADG "Event Wagering Revenue Report", '
                            '"May 1-31, 2026*": Operator "%s", "%s" column, '
                            '"%s". "*These numbers are self-reported by the '
                            'operators; as such these numbers are subject to '
                            'adjustments after audit and review by the Arizona '
                            'Department of Gaming."'
                            % (brand, leg.capitalize(),
                               AZ_EW_HEADERS[AZ_EW_METRICS.index(met)])),
                        note="UNAUDITED and self-reported, as the report "
                             "itself states")
        self._arizona_fantasy_absence()

    def _arizona_fantasy_absence(self):
        """ADG's fantasy sports report names no tribal operator at all.

        Fourteen operators are printed and every one is a commercial fantasy
        contest company. Writing them into a Native dataset would pad it; the
        finding is the absence, so the absence is what is written.
        """
        p = os.path.join(RAW, "az/adg_fantasy_sports_may2026.pdf")
        if not os.path.exists(p):
            return
        try:
            import fitz
        except ImportError:
            return
        txt = "\n".join(pg.get_text() for pg in fitz.open(p))
        ops = re.findall(r"^([A-Z][A-Za-z0-9&(),.' -]{2,40})\s*$",
                         txt, re.M)
        queue_review(
            "AZ-FANTASY-NO-TRIBAL", "documented_absence", "AZ",
            "Arizona fantasy sports contest operators",
            "operators printed in the May 2026 report: %s"
            % "; ".join(sorted({o.strip() for o in ops})[:25]),
            "ADG's Fantasy Sports Contest Revenue Report names no tribal "
            "operator. No rows written - the absence is the finding, not a "
            "gap.",
            "https://gaming.az.gov/sites/default/files/"
            "FS%20Website%20Report-May%202026%20UNAUDITED.pdf",
            proposed="no action; re-check if a tribe licenses a fantasy "
                     "contest operation")
        note("az_fantasy_no_tribal_operator")

    # ---------------- FLORIDA -------------------------------------------
    def build_florida(self):
        """Florida is authority-rich and operation-poor in the public record.

        `docs/FL_GAMING_BUILD_LOG.md` already establishes that EDR publishes
        RECEIPTS, not obligations, and that sports wagering appears there only
        as a separately-labelled FORECAST. That is not re-derived here. What is
        recorded is the compact right itself and an explicit statement that no
        actual monthly digital figure exists to record.
        """
        res = self.resolver.resolve("Seminole", "Florida")
        tid = res["tribe_id"]
        if not tid:
            queue_review("FL-SEMINOLE", "tribe_unresolved", "FL", "Seminole",
                         "", res["refusal"], "")
            return
        for field, product_type, statewide in (
                ("mobile_wagering_scope", "ONLINE_SPORTSBOOK", "yes"),
                ("sports_wagering_authorized", "RETAIL_SPORTSBOOK", "")):
            rows = self.rights.get((tid, "Florida"), {}).get(field, [])
            if not rows:
                continue
            cite, curl, cquote = authority_cite(rows)
            scopes = sorted({clean_text(r.get("value", "")) for r in rows})
            self.add_rel(
                tribe_id=tid, tribe_canonical_name=res["tribe_canonical_name"],
                facility_id="", state="FL", product_type=product_type,
                license_type="2021 Tribal-State Gaming Compact, Part XII "
                             "(sports betting)",
                brand="", operator_entity_id=tid,
                retail_available="yes" if product_type == "RETAIL_SPORTSBOOK"
                                 else "",
                mobile_statewide=statewide if "statewide" in scopes else "",
                mobile_on_premises="yes" if "on_indian_lands" in scopes else "",
                launch_date="", launch_date_basis="",
                current_status="authorised_by_compact; operation not observed "
                               "in any source read by this build",
                compact_authority_cite=cite, compact_authority_url=curl,
                compact_authority_quote=(cquote or "")[:600],
                authorisation_observed="yes", operation_observed="no",
                source_url=curl,
                source_quote=(cquote or "")[:600],
                confidence=Tier.B.value,
                entity_match_method=res["match_method"],
                note="AUTHORISATION ONLY. Florida publishes no monthly digital "
                     "series: EDR's sports wagering figures are a separately-"
                     "labelled FORECAST (docs/FL_GAMING_BUILD_LOG.md), and the "
                     "compact lets the Tribe mark what it gives the State "
                     "Trade Secret. launch_date is deliberately blank - a "
                     "right is not a launch.")
        # An explicit absence row, so a blank Florida reads as a fact.
        self.add_rev(
            state="FL", period_start="", period_end="", period_type="none",
            tribe_id=tid, tribe_canonical_name=res["tribe_canonical_name"],
            licensee_name_as_published="Seminole Tribe of Florida",
            product_type="ONLINE_SPORTSBOOK",
            revenue_scope="NO_REVENUE_OBSERVATION",
            metric="GROSS_GAMING_REVENUE", value_usd="",
            is_tribe_attributable="no",
            attribution_basis="no observation exists to attribute",
            is_online_only="yes",
            source_agency="Florida Office of Economic and Demographic Research",
            source_document="docs/FL_GAMING_BUILD_LOG.md",
            source_url="http://edr.state.fl.us/Content/conferences/"
                       "indiangaming/index.cfm",
            source_quote="Sports wagering separately from physical casino GGR. "
                         "Florida publishes the split only as a forecast, and "
                         "it is recorded only as a forecast. No sports-betting "
                         "figure is merged into any casino figure anywhere in "
                         "this file.",
            note="DOCUMENTED ABSENCE, not an unworked gap. See "
                 "docs/FL_GAMING_BUILD_LOG.md; not re-derived here.")

    # ---------------- COMPACT RIGHTS WITHOUT AN OBSERVED LAUNCH ---------
    def build_authorisation_only(self):
        """Every tribe the compacts authorise, with no operation observed.

        This is the point of the dataset. A row here says: the instrument says
        they may, and nothing this build read says they do.
        """
        observed = {(r["tribe_id"], r["product_type"]) for r in self.rel}
        observed_tribes = {r["tribe_id"] for r in self.rel}
        for (tid, state), fields in sorted(self.rights.items()):
            if not tid:
                continue
            for field, rows in sorted(fields.items()):
                vals = sorted({clean_text(r.get("value", "")) for r in rows})
                # A prohibition is a fact too, but it is not a relationship.
                if vals == ["prohibited"]:
                    note("compact_right_prohibited")
                    continue
                if field == "sports_wagering_authorized":
                    ptype = "RETAIL_SPORTSBOOK"
                elif field == "internet_wagering_authorized":
                    ptype = "ONLINE_CASINO"
                else:
                    ptype = "ONLINE_SPORTSBOOK"
                if (tid, ptype) in observed:
                    continue
                if tid in observed_tribes and state in ("Michigan",
                                                        "Connecticut"):
                    continue
                cite, curl, cquote = authority_cite(rows)
                if not (curl and cquote):
                    continue
                st = STATE_ABBR.get(state, state[:2].upper())
                self.add_rel(
                    tribe_id=tid,
                    tribe_canonical_name=clean_text(
                        rows[-1].get("tribe_canonical_name", "")),
                    facility_id="", state=st, product_type=ptype,
                    license_type="compact authorisation",
                    operator_entity_id=tid,
                    mobile_statewide="yes" if "statewide" in vals else "",
                    mobile_on_premises="yes" if "on_indian_lands" in vals
                                       else "",
                    retail_available="",
                    launch_date="", launch_date_basis="",
                    current_status="authorised_by_compact; no operation "
                                   "observed in any source read by this build",
                    compact_authority_cite=cite, compact_authority_url=curl,
                    compact_authority_quote=(cquote or "")[:600],
                    authorisation_observed="yes", operation_observed="no",
                    source_url=curl, source_quote=(cquote or "")[:600],
                    confidence=Tier.B.value,
                    entity_match_method="inherited from compacts.csv",
                    note="AUTHORISATION ONLY. %s = %s. A compact "
                         "authorisation is not an operation; launch_date is "
                         "blank by rule, not by omission."
                         % (field, "/".join(vals)))

    # ---------------- LOYALTY -------------------------------------------
    def build_loyalty(self, got):
        have = {g[0] for g in got}
        for host, root, state, tribe in LOYALTY_SITES:
            if host in have:
                continue
            queue_review(
                "LOYALTY-NOPAGE-%s" % host, "loyalty_page_not_retrieved",
                state, tribe, root,
                "no loyalty page was retrieved for this enterprise: either "
                "the site root did not return 200 or it carried no "
                "loyalty-shaped link. NOT_FOUND, not evidence that the "
                "enterprise runs no programme.", root,
                proposed="locate the programme page by hand and re-run")
            note("loyalty_site_not_retrieved")
        n = 0
        for host, root, state, tribe, target in got:
            prog, props = parse_loyalty(host, root, state, tribe, target,
                                        self.resolver, self.facilities)
            if not prog:
                note("loyalty_no_program")
                continue
            n += 1
            pid = "LOY-%04d" % n
            row = {k: "" for k in LOY_FIELDS}
            row.update(
                loyalty_program_id=pid, program_name=prog["program_name"],
                operator_entity_id=prog["tribe_id"], tribe_id=prog["tribe_id"],
                tribe_canonical_name=prog["tribe_canonical_name"],
                state=state, current_status="observed_active",
                program_name_basis=prog["program_name_basis"],
                tier_names=prog["tier_names"], tier_thresholds="",
                earning_currency="", n_properties_mapped=str(len(props)),
                cross_property_redemption=prog["cross_property"],
                mobile_app=prog["mobile_app"],
                digital_wallet=prog["digital_wallet"],
                cashless_gaming=prog["cashless"],
                source_url=prog["source_url"],
                source_quote=prog["source_quote"][:600],
                observation_date=TODAY, confidence_tier=Tier.B.value,
                built_by_script=SCRIPT, built_date=TODAY)
            for f, _ in ELIGIBILITY:
                row[f] = prog["flags"].get(f, "")
            row["eligibility_quotes"] = " || ".join(
                "%s: %s" % (k, v[:180]) for k, v in
                sorted(prog["quotes"].items()))
            if prog["tier_quote"]:
                row["eligibility_quotes"] = (
                    row["eligibility_quotes"] +
                    (" || " if row["eligibility_quotes"] else "") +
                    "tiers: " + prog["tier_quote"][:180])
            if prog["cross_quote"]:
                row["eligibility_quotes"] += (
                    " || cross_property: " + prog["cross_quote"][:180])
            if not row["source_quote"]:
                queue_review("LOYALTY-NOQUOTE-%s" % host, "loyalty_no_quote",
                             state, prog["program_name"], "",
                             "programme named but no quotable sentence "
                             "retained; not written", target)
                n -= 1
                continue
            if not prog["tribe_id"]:
                queue_review("LOYALTY-TIERB-%s" % host, "loyalty_name_only",
                             state, prog["program_name"], prog["source_quote"],
                             "programme found but the operator did not resolve "
                             "to the spine - name-only stays in review",
                             target)
                n -= 1
                continue
            self.loy.append(row)
            for f, quote, basis in props:
                self.loyprop.append(dict(
                    loyalty_program_id=pid, program_name=prog["program_name"],
                    facility_id=f.get("facility_id", ""),
                    facility_name=f.get("facility_name", ""),
                    tribe_id=prog["tribe_id"], state=state,
                    evidence_basis=basis,
                    source_url=prog["source_url"], source_quote=quote[:400],
                    observation_date=TODAY, confidence_tier=Tier.B.value,
                    built_by_script=SCRIPT, built_date=TODAY))

    # ---------------- declination cross-check ---------------------------
    def crosscheck_declinations(self):
        """NIGC's legal characterisation wins over any trade-press word.

        We do not overwrite anything from a vendor name. Where a declination
        letter names a tribe that this build has given a `technology_provider`,
        the pair is reported so a human can read the letter's own
        characterisation of the contractual role.
        """
        if not self.declinations:
            note("declination_file_absent")
            return
        by_tribe = defaultdict(list)
        for d in self.declinations:
            t = (d.get("tribe_entity_id") or d.get("tribe_id")
                 or d.get("entity_id") or "")
            if t:
                by_tribe[t].append(d)
        # Most declinations a gaming tribe holds are FINANCING letters naming a
        # bank - PNC, Key Bank, Wells Fargo, Devon Bank. Those say nothing
        # about a platform contract, and surfacing all of them buries the two
        # that matter. Only a contractor that is gaming-shaped, or that shares
        # a name token with the provider we recorded, is raised.
        gaming_shaped = re.compile(
            r"wager|book|sportsbook|gaming|interactive|digital|casino|"
            r"entertainment|technolog|platform", re.I)
        for r in self.rel:
            if not r.get("technology_provider"):
                continue
            ptoks = {t for t in re.findall(
                r"[a-z0-9]+", r["technology_provider"].lower()) if len(t) > 3}
            for d in by_tribe.get(r["tribe_id"], []):
                contractor = clean_text(d.get("index_company_string", ""))
                ctoks = {t for t in re.findall(r"[a-z0-9]+", contractor.lower())
                         if len(t) > 3}
                if not (gaming_shaped.search(contractor) or (ptoks & ctoks)):
                    note("declination_financing_letter_not_raised")
                    continue
                queue_review(
                    "DECL-XCHECK-%s-%s" % (r["tribe_id"],
                                           d.get("cedar_opinion_id", "")),
                    "declination_crosscheck", r["state"],
                    "%s / %s" % (r["tribe_canonical_name"],
                                 r["technology_provider"]),
                    "NIGC declination %s (%s) names contractor \"%s\""
                    % (d.get("cedar_opinion_id", ""),
                       d.get("opinion_date", ""),
                       clean_text(d.get("index_company_string", ""))[:120]),
                    "this build records a platform/technology provider for a "
                    "tribe that also holds an NIGC declination letter. The "
                    "letter's legal characterisation of the contractual role "
                    "wins over any trade-press term; read it before any "
                    "management claim is made.",
                    d.get("source_url", ""),
                    proposed="no change unless the letter characterises the "
                             "same vendor")
                note("declination_crosscheck")
                break

    # ---------------- emit ----------------------------------------------
    def emit(self):
        # Hard gate: online and physical must never be one figure.
        for r in self.rev:
            assert r["revenue_scope"] in REVENUE_SCOPES
            assert "CASINO_GGR" not in r["revenue_scope"]
        missing = [r for r in self.rel
                   if not (r["source_url"] and r["source_quote"])]
        assert not missing, "%d relationship rows without receipts" % len(missing)
        n1 = write_csv(os.path.join(CLEAN, "digital_gaming_relationships.csv"),
                       self.rel, REL_FIELDS)
        n2 = write_csv(os.path.join(CLEAN, "digital_gaming_revenue.csv"),
                       self.rev, REV_FIELDS)
        n3 = write_csv(os.path.join(CLEAN, "loyalty_programs.csv"),
                       self.loy, LOY_FIELDS)
        n4 = write_csv(os.path.join(CLEAN, "loyalty_program_property.csv"),
                       self.loyprop, LOYPROP_FIELDS)
        # one row per DECISION, not per occurrence
        seen, uniq = set(), []
        for r in REVIEW_ROWS:
            if r["item_key"] in seen:
                continue
            seen.add(r["item_key"])
            uniq.append(r)
        n5 = write_csv(
            os.path.join(REVIEW, "digital_gaming_unresolved_%s.csv" % TODAY),
            uniq, REVIEW_FIELDS)
        return n1, n2, n3, n4, n5


# ===========================================================================
# codebook fragment
# ===========================================================================
CODEBOOK_DESCRIPTIONS = {
    "digital_gaming_id": "Cedar row identifier for one digital gaming "
        "relationship. `DGR-<state>-<n>`.",
    "tribe_id": "Cedar Press permanent identifier for the Native entity.",
    "facility_id": "Cedar property ID (`CCP-`/`VP-`/`TPL-`). NULLABLE AND "
        "USUALLY NULL BY DESIGN: statewide mobile wagering has no property. A "
        "blank here is a fact about the product, not a missing value.",
    "product_type": "RETAIL_SPORTSBOOK | ONLINE_SPORTSBOOK | ONLINE_CASINO | "
        "ONLINE_POKER | FANTASY | OTHER_DIGITAL.",
    "license_type": "The licence or instrument the regulator names, in its "
        "own words.",
    "brand": "The consumer-facing brand as the regulator prints it. A brand "
        "using a different name from the property is an ALIAS, not a new "
        "property.",
    "technology_provider": "The platform provider. A technology vendor is NOT "
        "a manager: where an NIGC declination letter characterises the "
        "contractual role, its legal characterisation wins.",
    "mobile_statewide": "`yes` where the right or the licence reaches wagers "
        "placed off Indian lands.",
    "mobile_on_premises": "`yes` where mobile wagering is confined to the "
        "tribe's own premises. Connecticut publishes on-reservation and "
        "statewide as separate licensees and they are never merged here.",
    "launch_date": "Date operation began, ONLY where a source states it. "
        "NEVER inferred from an authorisation.",
    "launch_date_basis": "The source's own words for what the date is.",
    "compact_authority_cite": "The compact instrument, PDF and page that "
        "authorises the product. Recorded separately from `launch_date` "
        "because a right and an operation are different facts.",
    "authorisation_observed": "`yes` where a parsed compact term authorises "
        "this product for this tribe.",
    "operation_observed": "`yes` where a source read by this build shows the "
        "product actually operating. The gap between this and "
        "`authorisation_observed` is the dataset's central finding.",
    "revenue_scope": "What the money figure covers. Constrained to online, "
        "retail sports wagering, or fantasy. A physical casino GGR figure has "
        "no legal value here, which is the mechanism that stops online and "
        "floor revenue ever being summed.",
    "metric": "HANDLE | GROSS_GAMING_REVENUE | ADJUSTED_GROSS_REVENUE | "
        "TAX_OR_PAYMENT | PROMOTIONAL_DEDUCTION | PATRON_WINNINGS | "
        "CANCELLED_WAGERS | FEDERAL_EXCISE_TAX | ENTRY_FEES | "
        "NET_FANTASY_CONTEST_REVENUE | MONTHLY_RESETTLEMENTS | "
        "AMOUNT_WAGERED. HANDLE is sports money staked once; AMOUNT_WAGERED "
        "is online-casino coin-in, which recycles on every spin. Connecticut "
        "labels both columns `wagers`; they are separated here so they cannot "
        "be added.",
    "is_tribe_attributable": "`no` where the regulator publishes by operator "
        "BRAND rather than by licence holder. Arizona is the whole of this "
        "class; those rows must never be summed into a tribal total.",
    "is_online_only": "`yes` where the figure covers online play only.",
    "cross_property_redemption": "`yes` only where the programme page itself "
        "states the card works across properties.",
    "n_properties_mapped": "How many Cedar properties the programme page "
        "names. A programme spanning several properties is evidence of "
        "ENTERPRISE-LEVEL INTEGRATION, which is the analytic value here - not "
        "the tier names.",
    "tier_names": "Tier labels as printed, pipe-separated. Blank where the "
        "page does not enumerate them.",
    "source_quote": "Verbatim text from the source supporting the row. No row "
        "exists without one.",
}


def write_codebook_fragment(counts):
    """A FRAGMENT under data/clean/codebook/. codebook_master.csv is untouched."""
    rows = []
    for dataset, path, fields in (
            ("16_digital_gaming",
             os.path.join(CLEAN, "digital_gaming_relationships.csv"),
             REL_FIELDS),
            ("16b_digital_gaming_revenue",
             os.path.join(CLEAN, "digital_gaming_revenue.csv"), REV_FIELDS),
            ("16c_loyalty_programs",
             os.path.join(CLEAN, "loyalty_programs.csv"), LOY_FIELDS),
            ("16d_loyalty_program_property",
             os.path.join(CLEAN, "loyalty_program_property.csv"),
             LOYPROP_FIELDS)):
        data = read_csv(path)
        n = len(data)
        for f in fields:
            filled = sum(1 for r in data if clean_text(r.get(f, "")))
            rows.append(dict(
                dataset=dataset, variable=f, type="text",
                units="", pct_filled=("%.1f" % (100.0 * filled / n)) if n else "0.0",
                n_rows=n, published=1, access_tier="public",
                description=CODEBOOK_DESCRIPTIONS.get(
                    f, f.replace("_", " ").capitalize() + "."),
                generated=TODAY))
    return write_csv(
        os.path.join(CODEBOOK, "16_digital_gaming.csv"), rows,
        ["dataset", "variable", "type", "units", "pct_filled", "n_rows",
         "published", "access_tier", "description", "generated"])


# ===========================================================================
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--skip-fetch", action="store_true")
    ap.add_argument("--only", default="")
    a = ap.parse_args()

    t0 = time.time()
    print("=" * 74)
    print("119  digital gaming and loyalty layer")
    print("=" * 74)

    got_loyalty = []
    if not a.skip_fetch:
        only = {x.strip() for x in a.only.split(",") if x.strip()}
        if not only or "mi" in only:
            print("[fetch] Michigan (MGCB)")
            fetch_michigan()
        if not only or "ct" in only:
            print("[fetch] Connecticut (data.ct.gov)")
            fetch_connecticut()
        if not only or "az" in only:
            print("[fetch] Arizona (ADG)")
            fetch_arizona()
        if not only or "loyalty" in only:
            print("[fetch] loyalty programme pages (%d enterprises)"
                  % len(LOYALTY_SITES))
            got_loyalty = fetch_loyalty()
        # Wayback is held by another agent; queue rather than start a second
        # poller. Rule 1 of PULL_DISCIPLINE.
        claim_host("web.archive.org",
                   "historic loyalty programme pages - programme changes over "
                   "time (119_build_digital_and_loyalty.py)")
    else:
        for host, root, state, tribe in LOYALTY_SITES:
            p = os.path.join(RAW, "loyalty/%s/program.html" % host)
            if os.path.exists(p):
                got_loyalty.append((host, root, state, tribe, ""))

    b = Builder()
    print("[build] Michigan");      b.build_michigan()
    print("[build] Connecticut");   b.build_connecticut()
    print("[build] Arizona");       b.build_arizona()
    print("[build] Florida");       b.build_florida()
    print("[build] authorisation-only rows"); b.build_authorisation_only()
    print("[build] loyalty");       b.build_loyalty(got_loyalty)
    b.crosscheck_declinations()

    # --- footing: does our per-operator read reproduce MGCB's own totals? ---
    foot_ok = foot_bad = foot_nopair = 0
    foot_rows = []
    for (pt, yr, mo, met), d in sorted(MI_FOOTING.items()):
        if "published" not in d or "summed" not in d:
            foot_nopair += 1
            continue
        ok = abs(d["published"] - d["summed"]) <= max(
            1.0, abs(d["published"]) * 1e-6)
        foot_ok += ok
        foot_bad += (not ok)
        foot_rows.append(dict(product_type=pt, year=yr, month=mo, metric=met,
                              published="%.2f" % d["published"],
                              summed="%.2f" % d["summed"],
                              foots="yes" if ok else "NO"))
        if not ok:
            queue_review("MI-FOOTING-%s-%d-%02d-%s" % (pt, yr, mo, met),
                         "footing_failure", "MI",
                         "%s %d-%02d %s" % (pt, yr, mo, met),
                         "published %.2f vs summed %.2f"
                         % (d["published"], d["summed"]),
                         "the sum of MGCB's per-operator columns does not "
                         "reproduce MGCB's own printed total for this month",
                         MI_MEDIA)
    write_csv(os.path.join(INTERIM, "119_mi_footing.csv"), foot_rows,
              ["product_type", "year", "month", "metric", "published",
               "summed", "foots"])

    n1, n2, n3, n4, n5 = b.emit()
    nm = save_manifest()
    ncb = write_codebook_fragment((n1, n2, n3, n4))

    # ---- summary ----------------------------------------------------------
    auth = {r["tribe_id"] for r in b.rel if r["authorisation_observed"] == "yes"}
    oper = {r["tribe_id"] for r in b.rel if r["operation_observed"] == "yes"}
    rights_tribes = {t for (t, s) in b.rights if t}
    states_pub = sorted({r["state"] for r in b.rev
                         if r["revenue_scope"] != "NO_REVENUE_OBSERVATION"})
    multi = Counter(r["loyalty_program_id"] for r in b.loyprop)
    lines = []
    lines.append("119 digital gaming + loyalty  run %s" % TODAY)
    lines.append("relationships                    %6d" % n1)
    lines.append("revenue rows                     %6d" % n2)
    lines.append("loyalty programmes               %6d" % n3)
    lines.append("loyalty programme-property rows  %6d" % n4)
    lines.append("review items                     %6d" % n5)
    lines.append("raw files manifested             %6d" % nm)
    lines.append("codebook fragment variables      %6d" % ncb)
    lines.append("")
    lines.append("tribes with a digital right in the compacts   %4d"
                 % len(rights_tribes))
    lines.append("tribes with an OBSERVED operation             %4d"
                 % len(oper))
    lines.append("  ... of which the compacts carry ANY digital term %4d"
                 % len(rights_tribes & oper))
    for t in sorted(rights_tribes & oper):
        vals = sorted({clean_text(r.get("value", ""))
                       for f in b.rights[(t, [s for (x, s) in b.rights
                                              if x == t][0])].values()
                       for r in f})
        lines.append("        %s -> %s" % (t, "/".join(vals)))
    lines.append("AUTHORISED BUT NO OPERATION OBSERVED          %4d"
                 % len(rights_tribes - oper))
    lines.append("OPERATING WITH NO COMPACT AUTHORISATION PARSED %3d"
                 % len(oper - rights_tribes))
    lines.append("states publishing digital revenue we hold: %s"
                 % ", ".join(states_pub))
    lines.append("loyalty programmes spanning >1 property: %d of %d"
                 % (sum(1 for v in multi.values() if v > 1), n3))
    lines.append("MGCB month x metric footings: %d foot, %d fail, %d with no "
                 "published total to foot against" % (foot_ok, foot_bad,
                                                      foot_nopair))
    lines.append("")
    lines.append("resolver outcomes: %s" % dict(b.resolver.reasons))
    lines.append("notes: %s" % dict(NOTES))
    txt = "\n".join(lines)
    open(os.path.join(INTERIM, "119_run_summary.txt"), "w",
         encoding="utf-8").write(txt + "\n")
    print(txt)
    print("\ndone in %.1fs" % (time.time() - t0))


if __name__ == "__main__":
    main()
